#!/usr/bin/env python3
"""
Cell Operations - Clean Cell Management
Single responsibility: Direct cell value and property operations
"""

from typing import Any, Dict, Optional, List
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import datetime


class CellOperations:
    """Clean, focused cell operations without the utility mess"""
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
    
    def set_cell_value(
        self, 
        worksheet: Worksheet, 
        row: int, 
        col: int, 
        value: Any,
        format_as_date: bool = False
    ) -> Cell:
        """
        Set cell value with proper type handling
        
        Args:
            worksheet: Target worksheet
            row: Row number (1-based)
            col: Column number (1-based) 
            value: Value to set
            format_as_date: Whether to format as date
            
        Returns:
            The modified cell
        """
        cell = worksheet.cell(row=row, column=col)
        
        if format_as_date and value:
            self._set_date_value(cell, value)
        else:
            cell.value = self._normalize_value(value)
        
        return cell
    
    def find_and_replace_in_sheet(self, sheet: Worksheet, replacements: Dict[str, Any]):
        """
        Finds and replaces placeholder strings in all cells of a worksheet.

        Args:
            sheet: The worksheet to process.
            replacements: A dictionary where keys are placeholders (e.g., "{{invoice_id}}")
                          and values are the replacement values.
        """
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for placeholder, value in replacements.items():
                        if placeholder in cell.value:
                            # Using str() to handle potential non-string types gracefully
                            cell.value = cell.value.replace(placeholder, str(value))

    def set_cell_range_values(
        self,
        worksheet: Worksheet,
        start_row: int,
        start_col: int,
        values: List[List[Any]]
    ) -> None:
        """Set values for a range of cells efficiently"""
        for row_offset, row_values in enumerate(values):
            for col_offset, value in enumerate(row_values):
                if value is not None:  # Skip None values
                    self.set_cell_value(
                        worksheet, 
                        start_row + row_offset, 
                        start_col + col_offset, 
                        value
                    )
    
    def get_cell_value(self, worksheet: Worksheet, row: int, col: int) -> Any:
        """Get cell value with proper type handling"""
        cell = worksheet.cell(row=row, column=col)
        return self._normalize_value(cell.value)
    
    def clear_cell_range(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int
    ) -> None:
        """Clear a range of cells"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=col).value = None
    
    def copy_cell_properties(self, source_cell: Cell, target_cell: Cell) -> None:
        """Copy formatting and properties from source to target cell"""
        # Copy number format
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        
        # Copy font properties
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()
        
        # Copy border
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        
        # Copy fill
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
    
    def _normalize_value(self, value: Any) -> Any:
        """Normalize values for Excel cells"""
        if value is None:
            return None
        elif isinstance(value, str):
            return value.strip() if value.strip() else None
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, datetime.datetime):
            return value
        else:
            return str(value)
    
    def _set_date_value(self, cell: Cell, value: Any) -> None:
        """Set cell value as a properly formatted date"""
        try:
            if isinstance(value, str):
                # Try to parse string as date
                date_value = datetime.datetime.strptime(value, "%Y-%m-%d")
                cell.value = date_value
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, (int, float)):
                # Excel serial date number
                date_value = self._excel_number_to_datetime(value)
                if date_value:
                    cell.value = date_value
                    cell.number_format = "yyyy-mm-dd"
                else:
                    cell.value = value
            elif isinstance(value, datetime.datetime):
                cell.value = value
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value
        except Exception:
            # Fallback to raw value if date parsing fails
            cell.value = value
    
    def _excel_number_to_datetime(self, excel_num: Any) -> Optional[datetime.datetime]:
        """Convert Excel serial number to datetime"""
        try:
            if isinstance(excel_num, (int, float)) and excel_num > 0:
                # Excel epoch starts at 1900-01-01, but has a leap year bug
                base_date = datetime.datetime(1899, 12, 30)
                return base_date + datetime.timedelta(days=excel_num)
        except Exception:
            pass
        return None