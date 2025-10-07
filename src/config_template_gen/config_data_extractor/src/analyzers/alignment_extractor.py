"""
Alignment Extractor for Excel Analysis Tool.

This module provides functionality to extract cell alignment information
from Excel worksheets for use in configuration generation.
"""

from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from models.data_models import AlignmentInfo, HeaderMatch


class AlignmentExtractor:
    """Extracts alignment information from Excel worksheets."""

    def __init__(self):
        """Initialize the AlignmentExtractor."""
        pass

    def extract_alignments(self, worksheet: Worksheet, header_positions: List[HeaderMatch],
                          start_row: int, mapping_config: Optional[Dict[str, Any]] = None) -> List[AlignmentInfo]:
        """
        Extract alignment information from data rows in the worksheet.

        Args:
            worksheet: The openpyxl worksheet to analyze
            header_positions: List of detected header positions
            start_row: The row where data starts
            mapping_config: Optional mapping configuration to map column indices to IDs

        Returns:
            List of AlignmentInfo objects for each column with non-default alignment
        """
        alignments = []

        if not header_positions:
            return alignments

        # Find the maximum column index used in headers
        max_col = max(pos.column for pos in header_positions)

        # Sample a few data rows to determine alignment patterns
        sample_rows = []
        for row_idx in range(start_row, min(start_row + 5, worksheet.max_row + 1)):
            if any(worksheet.cell(row_idx, col).value for col in range(1, max_col + 1)):
                sample_rows.append(row_idx)

        if not sample_rows:
            return alignments

        # For each column that has headers, check alignment in data rows
        for header_pos in header_positions:
            col_idx = header_pos.column
            column_id = self._generate_column_id(col_idx, header_pos, mapping_config)

            # Check alignment in sample data rows
            alignment_counts = {}
            
            for row_idx in sample_rows:
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None:  # Only count cells with values
                    alignment = cell.alignment
                    if alignment:
                        horiz = alignment.horizontal or 'center'
                        vert = alignment.vertical or 'center'
                        align_key = f"{horiz}_{vert}"
                        alignment_counts[align_key] = alignment_counts.get(align_key, 0) + 1
                    else:
                        # Default alignment
                        align_key = "center_center"
                        alignment_counts[align_key] = alignment_counts.get(align_key, 0) + 1

            # Determine the most common alignment
            if alignment_counts:
                most_common_alignment = max(alignment_counts, key=alignment_counts.get)
                horizontal, vertical = most_common_alignment.split('_')
                
                # Extract alignment for this column (include all alignments, not just non-default)
                alignments.append(AlignmentInfo(
                    column_id=column_id,
                    horizontal=horizontal,
                    vertical=vertical
                ))

        return alignments

    def _generate_column_id(self, col: int, header_pos: HeaderMatch, mapping_config: Optional[Dict[str, Any]] = None) -> str:
        """Generate a column ID based on column position and header information."""
        # First try to use the mapping config if available
        if mapping_config:
            header_mappings = mapping_config.get('header_text_mappings', {}).get('mappings', {})
            header_text = header_pos.keyword.strip()

            # Try exact match first
            if header_text in header_mappings:
                return header_mappings[header_text]

            # Try case-insensitive match
            for mapped_header, column_id in header_mappings.items():
                if mapped_header.lower() == header_text.lower():
                    return column_id

            # Try partial matching for common patterns
            header_lower = header_text.lower()
            if 'sqft' in header_lower or 'sf' in header_lower:
                return 'col_qty_sf'
            elif 'pcs' in header_lower:
                return 'col_qty_pcs'
            elif 'quantity' in header_lower:
                return 'col_qty_sf'  # Default to SF for quantity
            elif 'amount' in header_lower or 'total' in header_lower:
                return 'col_amount'
            elif 'price' in header_lower or 'unit' in header_lower:
                return 'col_unit_price'
            elif 'gross' in header_lower and 'weight' in header_lower:
                return 'col_gross'
            elif 'net' in header_lower and 'weight' in header_lower:
                return 'col_net'
            elif 'cbm' in header_lower:
                return 'col_cbm'
            elif 'description' in header_lower or 'desc' in header_lower:
                return 'col_desc'
            elif 'po' in header_lower:
                return 'col_po'
            elif 'item' in header_lower:
                return 'col_item'
            elif 'mark' in header_lower and 'no' in header_lower:
                return 'col_static'

        # Fallback to the original logic if no mapping config
        header_cell = header_pos.worksheet.cell(header_pos.row, col) if hasattr(header_pos, 'worksheet') else None
        if header_cell and header_cell.value:
            header_text = str(header_cell.value).strip().lower()
            # Map common header texts to column IDs
            if 'description' in header_text:
                return 'description'
            elif 'quantity' in header_text or 'qty' in header_text:
                return 'quantity'
            elif 'price' in header_text or 'unit price' in header_text:
                return 'unit_price'
            elif 'amount' in header_text or 'total' in header_text:
                return 'amount'
            elif 'pcs' in header_text:
                return 'pcs'
            elif 'net' in header_text and 'weight' in header_text:
                return 'net_weight'
            elif 'gross' in header_text and 'weight' in header_text:
                return 'gross_weight'

        # Final fallback to generic column ID
        return f"col_{col}"
