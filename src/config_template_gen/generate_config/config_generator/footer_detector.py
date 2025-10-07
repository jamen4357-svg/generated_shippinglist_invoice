"""
FooterDetector component for detecting footer rows with SUM formulas.

This module provides functionality to detect footer rows in Excel sheets by looking for
SUM formulas, similar to how the xlsx_generator/row_processor identifies table boundaries.
"""

import re
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from .models import FooterInfo, FontInfo
from pathlib import Path
import tempfile


class FooterDetectorError(Exception):
    """Custom exception for FooterDetector errors."""
    pass


class FooterDetector:
    """Detects footer rows with SUM formulas in Excel sheets."""
    
    def __init__(self):
        """Initialize FooterDetector with formula patterns."""
        # Same patterns used by row_processor for consistency
        self.formula_patterns = [
            r'=sum\(',
            r'=SUM\(',
            r'=Sum\('
        ]
    
    def detect_footer_from_file(self, file_path: str, sheet_name: str, header_row: int) -> Optional[FooterInfo]:
        """
        Detect footer information from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to analyze
            header_row: Row number where the header is located
            
        Returns:
            FooterInfo object if footer is found, None otherwise
            
        Raises:
            FooterDetectorError: If file cannot be read or analyzed
        """
        try:
            # Load the workbook
            workbook = load_workbook(file_path, data_only=False)  # Keep formulas
            
            if sheet_name not in workbook.sheetnames:
                print(f"[FOOTER_DETECTOR] Sheet '{sheet_name}' not found in workbook")
                return None
            
            worksheet = workbook[sheet_name]
            
            # Detect footer using the same logic as row_processor
            footer_info = self._detect_footer_in_worksheet(worksheet, header_row)
            
            workbook.close()
            return footer_info
            
        except Exception as e:
            raise FooterDetectorError(f"Failed to detect footer in {file_path}:{sheet_name}: {str(e)}") from e
    
    def _detect_footer_in_worksheet(self, worksheet: Worksheet, header_row: int) -> Optional[FooterInfo]:
        """
        Detect footer information in a worksheet.
        
        Args:
            worksheet: The worksheet to analyze
            header_row: Row number where the header is located
            
        Returns:
            FooterInfo object if footer is found, None otherwise
        """
        # Step 1: Find formula columns (any column with SUM formulas)
        formula_columns = self._find_formula_columns(worksheet, header_row)
        
        if not formula_columns:
            print(f"[FOOTER_DETECTOR] No formula columns found after header row {header_row}")
            return None
        
        # Step 2: Find the formula row (where SUM formulas appear)
        formula_row = self._find_formula_row(worksheet, header_row, formula_columns[0])
        
        if formula_row is None:
            print(f"[FOOTER_DETECTOR] No formula row found after header row {header_row}")
            # Try to find footer without formulas by looking for TOTAL text
            return self._detect_footer_without_formulas(worksheet, header_row)
        
        # Step 3: Find total text column
        total_text_column, total_text_value = self._find_total_text_column(worksheet, formula_row)
        
        # Step 4: Find pallet count column
        pallet_count_column, pallet_count_value = self._find_pallet_count_column(worksheet, formula_row)
        
        # Step 5: Extract font information from the formula row
        footer_font = self._extract_footer_font(worksheet, formula_row, formula_columns)
        
        # Step 6: Scan for all merged cells in the footer area
        footer_merged_cells = self._scan_footer_area_for_merged_cells(worksheet, formula_row)
        
        # Step 7: Create FooterInfo object
        footer_info = FooterInfo(
            row=formula_row,
            font=footer_font,
            has_formulas=True,
            formula_columns=formula_columns,
            total_text_column=total_text_column,
            total_text_value=total_text_value,
            pallet_count_column=pallet_count_column,
            pallet_count_value=pallet_count_value,
            merged_cells=footer_merged_cells
        )
        
        print(f"[FOOTER_DETECTOR] Footer detected at row {formula_row} with {len(formula_columns)} formula columns")
        if total_text_column:
            print(f"[FOOTER_DETECTOR] Total text '{total_text_value}' found in column {total_text_column}")
        if pallet_count_column:
            print(f"[FOOTER_DETECTOR] Pallet count '{pallet_count_value}' found in column {pallet_count_column}")
        
        return footer_info
    
    def _find_total_text_column(self, worksheet: Worksheet, formula_row: int) -> tuple[Optional[int], Optional[str]]:
        """
        Find the column containing total text (TOTAL: or TOTAL OF:) in the footer row.
        
        Args:
            worksheet: The worksheet to search
            formula_row: The row containing the footer
            
        Returns:
            Tuple of (column_number, total_text) or (None, None) if not found
        """
        total_text_patterns = [
            "TOTAL:",
            "TOTAL OF:",
            "TOTAL",
            "Total:",
            "Total of:",
            "Total"
        ]
        
        # Search in the formula row and a few rows before/after for total text
        search_rows = [formula_row - 1, formula_row, formula_row + 1]
        
        for row in search_rows:
            if row < 1:
                continue
                
            for col in range(1, min(worksheet.max_column + 1, 15)):  # Check first 15 columns
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    
                    # Check if cell contains total text
                    for pattern in total_text_patterns:
                        if pattern.lower() in cell_text.lower():
                            print(f"[FOOTER_DETECTOR] Found total text '{cell_text}' at row {row}, column {col}")
                            return col, cell_text
        
        print(f"[FOOTER_DETECTOR] No total text found in footer area around row {formula_row}")
        return None, None
    
    def _find_pallet_count_column(self, worksheet: Worksheet, formula_row: int) -> tuple[Optional[int], Optional[str]]:
        """
        Find the column containing pallet count text (e.g., "8 PALLETS", "10 PALLETS") in the footer row.
        
        Args:
            worksheet: The worksheet to search
            formula_row: The row containing the footer
            
        Returns:
            Tuple of (column_number, pallet_text) or (None, None) if not found
        """
        pallet_text_patterns = [
            r'\d+\s*PALLETS?',      # "8 PALLETS", "10 PALLET"
            r'\d+\s*PALLETS?:',     # "8 PALLETS:"
            r'PALLETS?:\s*\d+',     # "PALLETS: 8"
            r'\d+\s*PCS',           # "8 PCS" (sometimes used for pallet count)
            r'\d+\s*CTNS',          # "8 CTNS" (cartons)
            r'\d+\s*BOXES',         # "8 BOXES"
        ]
        
        # Search in the formula row and a few rows before/after for pallet count
        search_rows = [formula_row - 1, formula_row, formula_row + 1]
        
        for row in search_rows:
            if row < 1:
                continue
                
            for col in range(1, min(worksheet.max_column + 1, 15)):  # Check first 15 columns
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value).strip().upper()
                    
                    # Check if cell contains pallet count text using regex
                    import re
                    for pattern in pallet_text_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            print(f"[FOOTER_DETECTOR] Found pallet count '{cell.value}' at row {row}, column {col}")
                            return col, str(cell.value)
        
        print(f"[FOOTER_DETECTOR] No pallet count found in footer area around row {formula_row}")
        return None, None
    
    def _find_formula_columns(self, worksheet: Worksheet, header_row: int) -> List[int]:
        """
        Find all columns that contain SUM formulas below the header row.
        
        Args:
            worksheet: The worksheet to search
            header_row: The header row number
            
        Returns:
            List of column numbers that contain formulas
        """
        formula_columns = []
        
        # Look for SUM formulas in the next 50 rows after the header
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value)
                    
                    # Check if cell contains a SUM formula
                    for pattern in self.formula_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            if col not in formula_columns:
                                formula_columns.append(col)
                                print(f"[FOOTER_DETECTOR] Found formula column {col} at row {row}: {cell_text}")
        
        return sorted(formula_columns)
    
    def _find_formula_row(self, worksheet: Worksheet, header_row: int, formula_col: int) -> Optional[int]:
        """
        Find the formula row (SUM formula) in the specified column.
        
        Args:
            worksheet: The worksheet to search
            header_row: The header row number
            formula_col: The column number to search
            
        Returns:
            Row number of formula row, or None if not found
        """
        # Search downward from header row
        for row in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=formula_col)
            
            if cell.value is not None:
                cell_text = str(cell.value)
                
                # Check if cell contains a SUM formula
                for pattern in self.formula_patterns:
                    if re.search(pattern, cell_text, re.IGNORECASE):
                        print(f"[FOOTER_DETECTOR] Found formula row {row}: {cell_text}")
                        return row
        
        return None
    
    def _extract_footer_font(self, worksheet: Worksheet, formula_row: int, formula_columns: List[int]) -> FontInfo:
        """
        Extract font information from the footer row.
        
        Args:
            worksheet: The worksheet to analyze
            formula_row: The row containing formulas
            formula_columns: List of columns with formulas
            
        Returns:
            FontInfo object with footer font information
        """
        # Try to get font from the first formula cell
        if formula_columns:
            cell = worksheet.cell(row=formula_row, column=formula_columns[0])
            
            if cell.font:
                font_name = cell.font.name or "Times New Roman"
                font_size = cell.font.size or 12.0
                
                print(f"[FOOTER_DETECTOR] Footer font: {font_name}, size: {font_size}")
                return FontInfo(name=font_name, size=float(font_size))
        
        # Fallback to default font
        print(f"[FOOTER_DETECTOR] Using default footer font")
        return FontInfo(name="Times New Roman", size=12.0)
    
    def _scan_footer_area_for_merged_cells(self, worksheet: Worksheet, footer_row: int) -> Dict[int, Dict[str, int]]:
        """
        Scan for merged cells specifically on the footer row to avoid conflicts.
        
        Args:
            worksheet: The worksheet to scan
            footer_row: The row number of the footer
            
        Returns:
            Dictionary mapping Excel column number (1-based) to span info {'colspan': X, 'rowspan': Y, 'start_row': Y, 'end_row': Y}
        """
        merged_cells = {}
        
        print(f"[FOOTER_DETECTOR] Scanning for merged cells specifically on footer row {footer_row}...")
        
        # Check all merged cell ranges in the worksheet
        for merged_range in worksheet.merged_cells.ranges:
            # Only include merges that START on the footer row to avoid conflicts
            # This prevents overlapping merge rules from different rows
            if merged_range.min_row == footer_row:
                
                start_col = merged_range.min_col  # Excel column (1-based)
                start_row = merged_range.min_row
                end_col = merged_range.max_col
                end_row = merged_range.max_row
                
                colspan = end_col - start_col + 1
                rowspan = end_row - start_row + 1
                
                # Store the merge information
                merged_cells[start_col] = {
                    'colspan': colspan,
                    'rowspan': rowspan,
                    'start_row': start_row,
                    'end_row': end_row
                }
                
                print(f"[FOOTER_DETECTOR] Found footer row merge: Excel col {start_col} (row {start_row}-{end_row}) spans {colspan} columns, {rowspan} rows")
        
        if merged_cells:
            print(f"[FOOTER_DETECTOR] Found {len(merged_cells)} merged cells specifically on footer row {footer_row}")
        else:
            print(f"[FOOTER_DETECTOR] No merged cells found on footer row {footer_row}")
        
        return merged_cells
    
    def _detect_footer_without_formulas(self, worksheet: Worksheet, header_row: int) -> Optional[FooterInfo]:
        """
        Detect footer when there are no SUM formulas by looking for TOTAL text patterns.
        
        Args:
            worksheet: The worksheet to analyze
            header_row: Row number where the header is located
            
        Returns:
            FooterInfo object if footer is found, None otherwise
        """
        print(f"[FOOTER_DETECTOR] Attempting to detect footer without formulas after header row {header_row}")
        
        total_text_patterns = [
            "TOTAL:",
            "TOTAL OF:",
            "TOTAL",
            "Total:",
            "Total of:",
            "Total"
        ]
        
        # Search for TOTAL text in a wider range
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            for col in range(1, min(worksheet.max_column + 1, 15)):  # Check first 15 columns
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    
                    # Check if cell contains total text
                    for pattern in total_text_patterns:
                        if pattern.lower() in cell_text.lower():
                            print(f"[FOOTER_DETECTOR] Found total text '{cell_text}' at row {row}, column {col}")
                            
                            # Found TOTAL text, treat this as the footer row
                            footer_row = row
                            total_text_column = col
                            total_text_value = cell_text
                            
                            # Find pallet count column
                            pallet_count_column, pallet_count_value = self._find_pallet_count_column(worksheet, footer_row)
                            
                            # Extract font information
                            footer_font = self._extract_footer_font(worksheet, footer_row, [col])
                            
                            # Scan for merged cells
                            footer_merged_cells = self._scan_footer_area_for_merged_cells(worksheet, footer_row)
                            
                            # Create FooterInfo object
                            footer_info = FooterInfo(
                                row=footer_row,
                                font=footer_font,
                                has_formulas=False,  # No formulas found
                                formula_columns=[],  # No formula columns
                                total_text_column=total_text_column,
                                total_text_value=total_text_value,
                                pallet_count_column=pallet_count_column,
                                pallet_count_value=pallet_count_value,
                                merged_cells=footer_merged_cells
                            )
                            
                            print(f"[FOOTER_DETECTOR] Footer detected at row {footer_row} without formulas")
                            if total_text_column:
                                print(f"[FOOTER_DETECTOR] Total text '{total_text_value}' found in column {total_text_column}")
                            if pallet_count_column:
                                print(f"[FOOTER_DETECTOR] Pallet count '{pallet_count_value}' found in column {pallet_count_column}")
                            
                            return footer_info
        
        print(f"[FOOTER_DETECTOR] No footer found without formulas after header row {header_row}")
        return None

    def detect_footer_boundaries(self, file_path: str, sheet_name: str) -> Optional[Tuple[int, int]]:
        """
        Detect table boundaries (header row to footer row) using SUM formula detection.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to analyze
            
        Returns:
            Tuple of (header_row, footer_row) if found, None otherwise
        """
        try:
            workbook = load_workbook(file_path, data_only=False)
            
            if sheet_name not in workbook.sheetnames:
                return None
            
            worksheet = workbook[sheet_name]
            
            # Find potential header rows (rows with table-like content)
            header_candidates = self._find_header_candidates(worksheet)
            
            for header_row in header_candidates:
                footer_info = self._detect_footer_in_worksheet(worksheet, header_row)
                if footer_info:
                    workbook.close()
                    return (header_row, footer_info.row)
            
            workbook.close()
            return None
            
        except Exception as e:
            print(f"[FOOTER_DETECTOR] Error detecting boundaries: {e}")
            return None
    
    def _find_header_candidates(self, worksheet: Worksheet) -> List[int]:
        """
        Find potential header rows by looking for rows with multiple text values.
        
        Args:
            worksheet: The worksheet to search
            
        Returns:
            List of potential header row numbers
        """
        header_candidates = []
        
        for row in range(1, min(50, worksheet.max_row + 1)):  # Check first 50 rows
            text_count = 0
            
            for col in range(1, min(20, worksheet.max_column + 1)):  # Check first 20 columns
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    # Look for descriptive text (not just numbers)
                    if len(cell_text) > 2 and not cell_text.replace('.', '').replace(',', '').isdigit():
                        text_count += 1
            
            # If row has multiple text values, it might be a header
            if text_count >= 3:
                header_candidates.append(row)
        
        return header_candidates


def main():
    # --- existing argument parsing code ---
    args = parser.parse_args()

    # --- Create output directory ---
    excel_file_path = Path(args.excel_file)
    output_dir = BASE_DIR / "result" / excel_file_path.stem
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"[ORCHESTRATOR] Output will be saved in: {output_dir}")

    # Step 1: Analyze the Excel File
    temp_analysis_file = tempfile.NamedTemporaryFile(
        mode='w',
        delete=False,
        suffix=".json",
        prefix="analysis_",
        dir=str(output_dir)  # <-- Change this line
    ).name

    # Step 1.5: Extract and log headers
    output_base_name = output_dir / excel_file_path.stem
    header_log_path = extract_and_log_headers(analysis_output_path, str(output_base_name), args.interactive)

    # Step 2: Generate the Configuration File
    if args.output:
        final_output_path = output_dir / args.output
    else:
        final_output_path = output_dir / f"{excel_file_path.stem}_config.json"

    # Step 3: Generate XLSX file if requested
    if args.generate_xlsx and XLSX_GENERATOR_AVAILABLE:
        xlsx_output_path = args.xlsx_output
        if not xlsx_output_path:
            xlsx_output_path = output_dir / f"{excel_file_path.stem}_processed.xlsx"

        xlsx_output = generator.generate_processed_xlsx(
            args.excel_file,
            str(xlsx_output_path), # <-- Make sure to use the new path
            #...
        )