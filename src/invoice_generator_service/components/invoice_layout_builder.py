#!/usr/bin/env python3
"""
Service-Oriented Invoice Generator
Builder Pattern: Invoice Layout Builder
"""

import logging
from typing import Dict, Any, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from ..models import InvoiceData, CompanyConfig
from ..operations import (
    CellOperations, RowOperations, MergeOperations, HeaderOperations, StyleOperations, FooterOperations
)

class InvoiceLayoutBuilder:
    """
    Builder pattern for constructing invoice layouts step by step.
    Provides a fluent interface for building complex invoice structures.
    """

    def __init__(self, workbook: Workbook, data: InvoiceData, config: CompanyConfig):
        """
        Initialize the builder with the core components.

        Args:
            workbook: The Excel workbook to build upon
            data: The invoice data to populate
            config: The company configuration
        """
        self.workbook = workbook
        self.data = data
        self.config = config
        self.sheet = workbook.active  # Default to active sheet

        # Initialize operation handlers
        self.cell_ops = CellOperations()
        self.row_ops = RowOperations()
        self.merge_ops = MergeOperations()
        self.header_ops = HeaderOperations()
        self.style_ops = StyleOperations()
        self.footer_ops = FooterOperations()

        # Builder state
        self.current_row = 1
        self.header_info = {}
        self._built = False

    def select_sheet(self, sheet_name: str):
        """Select which sheet to work on."""
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.create_sheet(sheet_name)
        return self

    def add_static_fields(self):
        """Add static placeholder field replacements."""
        logging.debug("Builder: Adding static field replacements")
        replacements = {f"{{{{{field}}}}}": getattr(self.data, field)
                       for field in self.data.model_fields
                       if getattr(self.data, field) is not None}
        self.cell_ops.find_and_replace_in_sheet(self.sheet, replacements)
        return self

    def add_header(self):
        """Add the table header and update builder state."""
        logging.debug("Builder: Adding table header")
        if not self.config.header_layout:
            logging.warning("Builder: No header layout defined in config. Skipping header.")
            return self

        self.header_info = self.header_ops.write_header(
            worksheet=self.sheet,
            start_row=self.config.table_start_row,
            header_layout=self.config.header_layout,
            sheet_styling_config=self.config.styling
        )
        return self

    def add_table_data(self):
        """Add the dynamic line item table data."""
        logging.debug("Builder: Adding table data")
        start_row = self.config.table_start_row + self.header_info.get('header_rows', 0)
        columns = self.header_info.get('column_map', {})
        num_items = len(self.data.items)

        if num_items == 0:
            logging.warning("Builder: No items to add to table.")
            self.current_row = start_row
            return self

        # Insert rows if needed
        rows_to_insert = num_items - 1
        if rows_to_insert > 0:
            logging.debug(f"Builder: Inserting {rows_to_insert} rows for table data")
            self.row_ops.insert_rows_and_copy_styles(self.sheet, start_row, rows_to_insert)

        # Populate data
        for i, item in enumerate(self.data.items):
            current_row = start_row + i
            for field, col_letter in columns.items():
                if hasattr(item, field):
                    value = getattr(item, field)
                    if value is not None:
                        cell = self.sheet[f"{col_letter}{current_row}"]
                        cell.value = value

        self.current_row = start_row + num_items
        return self

    def add_footer(self):
        """Add footer information below the table."""
        logging.debug(f"Builder: Adding footer starting at row {self.current_row}")
        self.current_row = self.footer_ops.write_footer(
            worksheet=self.sheet,
            start_row=self.current_row,
            data=self.data,
            config=self.config,
            header_info=self.header_info
        )
        return self

    def apply_merges(self):
        """Apply cell merge rules from configuration."""
        if self.config.merge_rules:
            logging.debug("Builder: Applying merge rules")
            self.merge_ops.merge_cells_by_config(self.sheet, self.config.merge_rules)
        return self

    def apply_styling(self):
        """Apply column widths and other styling."""
        if self.config.styling and 'column_widths' in self.config.styling:
            logging.debug("Builder: Applying column widths")
            column_map = self.header_info.get('column_map', {})
            for field, width in self.config.styling['column_widths'].items():
                if field in column_map:
                    col_letter = column_map[field]
                    self.sheet.column_dimensions[col_letter].width = width
        return self

    def build(self) -> Workbook:
        """
        Finalize the invoice construction and return the completed workbook.

        Returns:
            The completed Excel workbook
        """
        if self._built:
            raise RuntimeError("Builder has already been used. Create a new builder instance.")

        logging.info(f"Builder: Finalizing invoice construction for '{self.data.invoice_id}'")
        self._built = True
        return self.workbook

    # Convenience method for complete build
    @classmethod
    def build_complete_invoice(cls, workbook: Workbook, data: InvoiceData, config: CompanyConfig) -> Workbook:
        """
        Convenience method to build a complete invoice in one call.
        Follows the standard invoice construction pattern.
        """
        return (cls(workbook, data, config)
                .add_static_fields()
                .add_header()
                .add_table_data()
                .add_footer()
                .apply_merges()
                .apply_styling()
                .build())
