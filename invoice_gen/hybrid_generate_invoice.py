import argparse
import json
import shutil
import openpyxl
import sys
import re
from pathlib import Path
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

# --- Import Reusable and New Utilities ---
import text_replace_utils
import invoice_utils
import packing_list_utils
import merge_utils

# Helper function to copy sheets (remains unchanged)
def copy_sheet_between_workbooks(source_sheet: Worksheet, target_workbook: Workbook) -> Worksheet:
    """
    Manually copies a worksheet from a source workbook to a target workbook.
    This replaces the built-in copy_worksheet which only works within the same workbook.
    """
    target_sheet = target_workbook.create_sheet(title=source_sheet.title)
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for merge_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merge_range))
    for col_letter, dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
              target_sheet.column_dimensions[col_letter].hidden = True
    for row_idx, dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = dim.height
        if dim.hidden:
            target_sheet.row_dimensions[row_idx].hidden = True
    return target_sheet

# Other helper functions (calculate_and_inject_totals, preprocess_data_for_numerics, derive_paths, load_json_file) remain unchanged.

def calculate_and_inject_totals(data: dict) -> dict:
    """
    Calculates summary values from raw data (like total pallets)
    and injects them into the main data dictionary for easy access.
    """
    print("Calculating and injecting summary totals...")
    if 'raw_data' not in data:
        print("  -> No 'raw_data' found, skipping calculation.")
        return data
    grand_total_pallets = sum(len(table_data.get('pallet_count', [])) for table_data in data.get('raw_data', {}).values())
    print(f"  -> Calculated Grand Total Pallets: {grand_total_pallets}")
    data.setdefault('aggregated_summary', {})['total_pallets'] = grand_total_pallets
    return data

def preprocess_data_for_numerics(data: any, keys_to_convert: set) -> any:
    """
    Recursively traverses a data structure (dict or list) and converts
    string values of specified keys into floats.
    """
    if isinstance(data, dict):
        return {key: preprocess_data_for_numerics(value, keys_to_convert) if key not in keys_to_convert or not isinstance(value, list) else [float(item.replace(',', '')) if isinstance(item, str) else item for item in value] for key, value in data.items()}
    elif isinstance(data, list):
        return [preprocess_data_for_numerics(item, keys_to_convert) for item in data]
    return data

def derive_paths(input_data_path_str: str, template_dir_str: str, config_dir_str: str) -> dict | None:
    """
    Derives template and config file paths based on the input data filename.
    """
    print(f"Deriving paths from input: {input_data_path_str}")
    try:
        input_data_path, template_dir, config_dir = Path(input_data_path_str).resolve(), Path(template_dir_str).resolve(), Path(config_dir_str).resolve()
        
        # --- START OF DEBUG BLOCK 1 ---
        print("\n--- WEB APP PATH DIAGNOSTICS ---")
        print(f"DEBUG: Input data file path resolves to: {input_data_path}")
        print(f"DEBUG: Template directory resolves to: {template_dir}")
        print(f"DEBUG: Config directory resolves to: {config_dir}")
        print(f"DEBUG: Checking initial existence...")
        print(f"DEBUG:   Input exists? -> {input_data_path.exists()}")
        print(f"DEBUG:   Template dir exists? -> {template_dir.exists()}")
        print(f"DEBUG:   Config dir exists? -> {config_dir.exists()}")
        print("--- END OF DEBUG BLOCK 1 ---\n")
        # --- END OF DEBUG BLOCK 1 ---

        if not all([p.exists() for p in [input_data_path, template_dir, config_dir]]):
            print("Error: One or more paths (input file, template dir, config dir) not found.")
            return None
        
        template_name_part = re.sub(r'(_data|_input|_pkl)$', '', input_data_path.stem, flags=re.IGNORECASE)
        print(f"Derived template name part: '{template_name_part}'")
        
        for prefix in [template_name_part, (re.match(r'^([a-zA-Z]+)', template_name_part) or [''])[0]]:
            if not prefix: continue
            template_path, config_path = template_dir / f"{prefix}.xlsx", config_dir / f"{prefix}_config.json"
            
            # --- START OF DEBUG BLOCK 2 ---
            print("-" * 20)
            print(f"DEBUG: Attempting with prefix: '{prefix}'")
            print(f"DEBUG: Checking for template -> {template_path.resolve()}")
            print(f"DEBUG: Template exists? -> {template_path.is_file()}")
            print(f"DEBUG: Checking for config -> {config_path.resolve()}")
            print(f"DEBUG: Config exists? -> {config_path.is_file()}")
            print("-" * 20)
            # --- END OF DEBUG BLOCK 2 ---

            if template_path.is_file() and config_path.is_file():
                print(f"Found match for template and config using prefix: '{prefix}'")
                return {"data": input_data_path, "template": template_path, "config": config_path}
                
        print("Error: Could not find matching template/config files.")
        return None
    except Exception as e:
        print(f"Error deriving file paths: {e}"); return None

def load_json_file(file_path: Path, file_type: str) -> dict:
    """Loads and parses a JSON file (data or config)."""
    print(f"Loading {file_type} from: {file_path}")
    try:
        with open(file_path, 'r', encoding='utf-8') as f: return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"FATAL ERROR: Could not load or parse {file_type} file {file_path}. Error: {e}"); sys.exit(1)


def main():
    """Main function to orchestrate hybrid invoice generation."""
    parser = argparse.ArgumentParser(description="Generate invoice documents from a JSON data file.")
    parser.add_argument("input_data_file", help="Path to the input JSON data file. Filename determines template/config.")
    parser.add_argument("-o", "--outputdir", default=".", help="Output directory for the generated Excel files.")
    parser.add_argument("-t", "--templatedir", default="./TEMPLATE", help="Directory for template files.")
    parser.add_argument("-c", "--configdir", default="./config", help="Directory for config files.")
    args = parser.parse_args()

    print("--- Starting Hybrid Invoice Generation ---")
    
    paths = derive_paths(args.input_data_file, args.templatedir, args.configdir)
    if not paths: sys.exit(1)

    po_number = paths['data'].stem
    output_dir = Path(args.outputdir)
    output_dir.mkdir(parents=True, exist_ok=True)

    invoice_data = load_json_file(paths['data'], "data")
    config = load_json_file(paths['config'], "config")

    keys_to_convert = {'net', 'amount', 'price', 'unit', 'cbm'}
    invoice_data = preprocess_data_for_numerics(invoice_data, keys_to_convert)
    invoice_data = calculate_and_inject_totals(invoice_data)
    
    template_workbook = None
    try:
        print(f"Loading template from '{paths['template']}'...")
        template_workbook = openpyxl.load_workbook(paths['template'])
        sheets_to_process_config = config.get("sheets_to_process", {})

        for sheet_name, sheet_config in sheets_to_process_config.items():
            if sheet_name not in template_workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' from config not found in template. Skipping.")
                continue

            print(f"\n--- Preparing new file for sheet: '{sheet_name}' ---")
            output_workbook = openpyxl.Workbook()
            worksheet = copy_sheet_between_workbooks(template_workbook[sheet_name], output_workbook)
            if 'Sheet' in output_workbook.sheetnames: output_workbook.remove(output_workbook['Sheet'])
            
            print(f"--- Processing Content for: '{sheet_name}' ---")
            process_type = sheet_config.get("type")

            if process_type == "summary":
                print(f"Processing '{sheet_name}' as summary (text replacement).")
                text_replace_utils.find_and_replace(output_workbook, sheet_config.get("replacements", []), 50, 20, invoice_data)

            elif process_type == "packing_list":
                print(f"Processing '{sheet_name}' as a packing list.")
                
                # --- REVISION ---
                # First, perform the standard text replacement for any placeholders on the sheet.
                print(" -> Step 1: Performing text replacement for placeholders...")
                text_replace_utils.find_and_replace(output_workbook, sheet_config.get("replacements", []), 50, 20, invoice_data)

                # Second, continue with the detailed packing list table generation.
                print(" -> Step 2: Generating detailed packing list table...")
                start_row = sheet_config.get("start_row", 1)
                merges_to_restore = merge_utils.store_original_merges(output_workbook, [sheet_name])
                rows_to_add = packing_list_utils.calculate_rows_to_generate(invoice_data, sheet_config)
                if rows_to_add > 0:
                    print(f"    -> Inserting {rows_to_add} rows at row {start_row}...")
                    merge_utils.force_unmerge_from_row_down(worksheet, start_row)
                    worksheet.insert_rows(start_row, amount=rows_to_add)
                
                packing_list_utils.generate_full_packing_list(worksheet, start_row, invoice_data, sheet_config)
                merge_utils.find_and_restore_merges_heuristic(output_workbook, merges_to_restore, [sheet_name])
            
            else:
                print(f"Warning: Unknown process type '{process_type}' for sheet '{sheet_name}'. Skipping.")

            # --- THIS IS THE KEY LINE FOR THE FILENAME ---
            # It creates the filename as "{Sheet Name} {PO Number}.xlsx"
            sheet_output_path = output_dir / f"{sheet_name} {po_number}.xlsx"
            
            print(f"\n--- Saving final workbook to '{sheet_output_path}' ---")
            output_workbook.save(sheet_output_path)
            output_workbook.close()
            print(f"Processing complete for sheet '{sheet_name}'.")

    except Exception as e:
        print(f"\n--- A CRITICAL ERROR occurred: {e} ---")
        import traceback
        traceback.print_exc()
    finally:
        if template_workbook:
            template_workbook.close()
            print("\nTemplate workbook closed.")

if __name__ == "__main__":
    main()