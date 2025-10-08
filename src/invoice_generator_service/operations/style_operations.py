#!/usr/bin/env python3
"""
Style Operations - Clean Cell Styling Logic
Single responsibility: Cell and range styling operations
"""

from typing import Dict, Any, Optional, List, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class StyleOperations:
    """Clean, focused styling operations without scattered utility functions"""
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self._style_cache: Dict[str, Any] = {}
    
    def apply_cell_style(
        self,
        cell: Cell,
        style_config: Dict[str, Any],
        context: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Apply styling to a single cell
        
        Args:
            cell: Target cell
            style_config: Style configuration dictionary
            context: Additional context for conditional styling
        """
        try:
            # Apply font styling
            if 'font' in style_config:
                self._apply_font_style(cell, style_config['font'])
            
            # Apply alignment
            if 'alignment' in style_config:
                self._apply_alignment_style(cell, style_config['alignment'])
            
            # Apply border
            if 'border' in style_config:
                self._apply_border_style(cell, style_config['border'])
            
            # Apply fill/background
            if 'fill' in style_config:
                self._apply_fill_style(cell, style_config['fill'])
            
            # Apply number format
            if 'number_format' in style_config:
                cell.number_format = style_config['number_format']
            
            # Apply conditional styles based on context
            if context and 'conditional' in style_config:
                self._apply_conditional_styles(cell, style_config['conditional'], context)
                
        except Exception as e:
            if self.verbose:
                print(f"Error applying cell style: {e}")
    
    def apply_row_style(
        self,
        worksheet: Worksheet,
        row_num: int,
        style_config: Dict[str, Any],
        start_col: int = 1,
        end_col: Optional[int] = None
    ) -> None:
        """
        Apply styling to an entire row
        
        Args:
            worksheet: Target worksheet
            row_num: Row number to style
            style_config: Style configuration
            start_col: Starting column (default 1)
            end_col: Ending column (auto-detect if None)
        """
        if end_col is None:
            end_col = worksheet.max_column
        
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_num, column=col)
            self.apply_cell_style(cell, style_config)
    
    def apply_range_style(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        style_config: Dict[str, Any]
    ) -> None:
        """
        Apply styling to a range of cells
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row
            end_row: Ending row
            start_col: Starting column
            end_col: Ending column
            style_config: Style configuration
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                self.apply_cell_style(cell, style_config)
    
    def apply_alternating_row_styles(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        even_style: Dict[str, Any],
        odd_style: Dict[str, Any],
        start_col: int = 1,
        end_col: Optional[int] = None
    ) -> None:
        """
        Apply alternating styles to rows
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row
            end_row: Ending row
            even_style: Style for even rows
            odd_style: Style for odd rows
            start_col: Starting column
            end_col: Ending column
        """
        if end_col is None:
            end_col = worksheet.max_column
        
        for row in range(start_row, end_row + 1):
            style_config = even_style if row % 2 == 0 else odd_style
            self.apply_row_style(worksheet, row, style_config, start_col, end_col)
    
    def set_row_height(self, worksheet: Worksheet, row_num: int, height: float) -> None:
        """Set height for a specific row"""
        worksheet.row_dimensions[row_num].height = height
    
    def set_column_width(self, worksheet: Worksheet, col_num: int, width: float) -> None:
        """Set width for a specific column"""
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = width
    
    def apply_header_style(
        self,
        worksheet: Worksheet,
        header_rows: List[int],
        style_config: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Apply consistent header styling
        
        Args:
            worksheet: Target worksheet
            header_rows: List of header row numbers
            style_config: Custom style config, or use default header style
        """
        if style_config is None:
            style_config = self._get_default_header_style()
        
        for row_num in header_rows:
            self.apply_row_style(worksheet, row_num, style_config)
    
    def apply_footer_style(
        self,
        worksheet: Worksheet,
        footer_rows: List[int],
        style_config: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Apply consistent footer styling
        
        Args:
            worksheet: Target worksheet
            footer_rows: List of footer row numbers
            style_config: Custom style config, or use default footer style
        """
        if style_config is None:
            style_config = self._get_default_footer_style()
        
        for row_num in footer_rows:
            self.apply_row_style(worksheet, row_num, style_config)
    
    def _apply_font_style(self, cell: Cell, font_config: Dict[str, Any]) -> None:
        """Apply font styling to cell"""
        font_kwargs = {}
        
        if 'name' in font_config:
            font_kwargs['name'] = font_config['name']
        if 'size' in font_config:
            font_kwargs['size'] = font_config['size']
        if 'bold' in font_config:
            font_kwargs['bold'] = font_config['bold']
        if 'italic' in font_config:
            font_kwargs['italic'] = font_config['italic']
        if 'color' in font_config:
            font_kwargs['color'] = font_config['color']
        
        if font_kwargs:
            cell.font = Font(**font_kwargs)
    
    def _apply_alignment_style(self, cell: Cell, alignment_config: Dict[str, Any]) -> None:
        """Apply alignment styling to cell"""
        alignment_kwargs = {}
        
        if 'horizontal' in alignment_config:
            alignment_kwargs['horizontal'] = alignment_config['horizontal']
        if 'vertical' in alignment_config:
            alignment_kwargs['vertical'] = alignment_config['vertical']
        if 'wrap_text' in alignment_config:
            alignment_kwargs['wrap_text'] = alignment_config['wrap_text']
        
        if alignment_kwargs:
            cell.alignment = Alignment(**alignment_kwargs)
    
    def _apply_border_style(self, cell: Cell, border_config: Dict[str, Any]) -> None:
        """Apply border styling to cell"""
        border_kwargs = {}
        
        # Create Side objects for each border side
        for side in ['left', 'right', 'top', 'bottom']:
            if side in border_config:
                side_config = border_config[side]
                if isinstance(side_config, dict):
                    border_kwargs[side] = Side(
                        style=side_config.get('style', 'thin'),
                        color=side_config.get('color', '000000')
                    )
                elif isinstance(side_config, str):
                    border_kwargs[side] = Side(style=side_config)
        
        if border_kwargs:
            cell.border = Border(**border_kwargs)
    
    def _apply_fill_style(self, cell: Cell, fill_config: Dict[str, Any]) -> None:
        """Apply fill/background styling to cell"""
        if 'pattern_type' in fill_config and 'color' in fill_config:
            cell.fill = PatternFill(
                patternType=fill_config['pattern_type'],
                fgColor=fill_config['color']
            )
    
    def _apply_conditional_styles(
        self,
        cell: Cell,
        conditional_config: Dict[str, Any],
        context: Dict[str, Any]
    ) -> None:
        """Apply conditional styles based on context"""
        for condition, style_config in conditional_config.items():
            if self._evaluate_condition(condition, context):
                self.apply_cell_style(cell, style_config)
                break  # Apply only the first matching condition
    
    def _evaluate_condition(self, condition: str, context: Dict[str, Any]) -> bool:
        """Evaluate a style condition against context"""
        # Simple condition evaluation - can be expanded
        if condition == 'is_header':
            return context.get('is_header', False)
        elif condition == 'is_footer':
            return context.get('is_footer', False)
        elif condition == 'is_total':
            return context.get('is_total', False)
        elif condition == 'is_even_row':
            return context.get('row_number', 0) % 2 == 0
        else:
            return False
    
    def _get_default_header_style(self) -> Dict[str, Any]:
        """Get default header styling"""
        return {
            'font': {
                'bold': True,
                'size': 11
            },
            'alignment': {
                'horizontal': 'center',
                'vertical': 'middle'
            },
            'fill': {
                'pattern_type': 'solid',
                'color': 'E0E0E0'
            }
        }
    
    def _get_default_footer_style(self) -> Dict[str, Any]:
        """Get default footer styling"""
        return {
            'font': {
                'bold': True,
                'size': 10
            },
            'alignment': {
                'horizontal': 'right',
                'vertical': 'middle'
            }
        }