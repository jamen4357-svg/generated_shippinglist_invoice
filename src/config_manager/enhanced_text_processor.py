#!/usr/bin/env python3
"""
Enhanced Text Processor - Advanced pattern-based text replacement with circular pattern checking.

This module provides intelligent text replacement functionality that:
1. Finds label patterns in cells
2. Uses circular pattern checking on adjacent cells to find the correct target
3. Handles unusual label positioning and foreign language labels
4. Provides flexible pattern matching for various document layouts
"""

import re
import logging
from typing import Dict, List, Optional, Tuple, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EnhancedTextProcessor:
    """
    Enhanced text processor with circular pattern checking for adjacent cells.
    
    This processor finds label patterns and then checks adjacent cells in a circular
    pattern to determine the best target cell for replacement.
    """
    
    def __init__(self):
        """Initialize the enhanced text processor with default patterns."""
        self.replacement_patterns = {
            'date': {
                'patterns': [
                    r'date[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',     # "date" + any non-letters + dd/mm/yyyy
                    r'date[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',     # "date" + any non-letters + yyyy/mm/dd
                    r'date[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}',     # "date" + any non-letters + dd/mm/yy
                    r'dated[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',    # "dated" + any non-letters + dd/mm/yyyy
                    r'dated[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',    # "dated" + any non-letters + yyyy/mm/dd
                    r'dated[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}'     # "dated" + any non-letters + dd/mm/yy
                ],
                'replacement': 'JFTIME'
            },
            'date_label': {
                'patterns': [
                    r'date[^a-zA-Z]',                            # "date" + any non-letter character
                    r'dated[^a-zA-Z]',                           # "dated" + any non-letter character
                    r'invoice\s+date[^a-zA-Z]',                  # "invoice date" + any non-letter
                    r'contract\s+date[^a-zA-Z]'                  # "contract date" + any non-letter
                ],
                'replacement': 'JFTIME'
            },
            'invoice_label': {
                'patterns': [
                    r'invoice[^a-zA-Z]*no[^a-zA-Z]',             # "invoice" + any non-letters + "no" + any non-letter
                    r'inv[^a-zA-Z]*no[^a-zA-Z]',                 # "inv" + any non-letters + "no" + any non-letter
                    r'bill[^a-zA-Z]*no[^a-zA-Z]'                 # "bill" + any non-letters + "no" + any non-letter
                ],
                'replacement': 'JFINV'
            },
            'ref_label': {
                'patterns': [
                    r'ref[^a-zA-Z]*no[^a-zA-Z]',                 # "ref" + any non-letters + "no" + any non-letter
                    r'reference[^a-zA-Z]*no[^a-zA-Z]',           # "reference" + any non-letters + "no" + any non-letter
                    r'our[^a-zA-Z]*ref[^a-zA-Z]'                 # "our" + any non-letters + "ref" + any non-letter
                ],
                'replacement': 'JFREF'
            },
            'contract_no': {
                'patterns': [
                    r'contract[^a-zA-Z]*no[^a-zA-Z]*[\w\-/]+',   # "contract" + any non-letters + "no" + any non-letters + code
                    r'contract[^a-zA-Z]*[\w\-/]+',               # "contract" + any non-letters + code
                    r'cont[^a-zA-Z]*no[^a-zA-Z]*[\w\-/]+'        # "cont" + any non-letters + "no" + any non-letters + code
                ],
                'replacement': 'JFINV'
            },
            'invoice_no': {
                'patterns': [
                    r'invoice[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*',   # "invoice" + any non-letters + "no" + any non-letters + numbers
                    r'inv[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*',       # "inv" + any non-letters + "no" + any non-letters + numbers
                    r'bill[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*'       # "bill" + any non-letters + "no" + any non-letters + numbers
                ],
                'replacement': 'JFINV'
            },
            'ref_no': {
                'patterns': [
                    r'ref[^a-zA-Z]*no[^a-zA-Z]*[A-Z]+\d+-\d+',   # "ref" + any non-letters + "no" + any non-letters + code
                    r'reference[^a-zA-Z]*no[^a-zA-Z]*[A-Z]+\d+-\d+',  # "reference" + any non-letters + "no" + any non-letters + code
                    r'our[^a-zA-Z]*ref[^a-zA-Z]*[A-Z]+\d+-\d+'   # "our" + any non-letters + "ref" + any non-letters + code
                ],
                'replacement': 'JFREF'
            },
            'etd': {
                'patterns': [
                    r'etd[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',      # "etd" + any non-letters + dd/mm/yyyy
                    r'etd[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',      # "etd" + any non-letters + yyyy/mm/dd
                    r'etd[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}',      # "etd" + any non-letters + dd/mm/yy
                    r'estimated[^a-zA-Z]*time[^a-zA-Z]*of[^a-zA-Z]*departure[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',  # full ETD flexible
                    r'departure[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}', # "departure" + any non-letters + dd/mm/yyyy
                    r'departure[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}'  # "departure" + any non-letters + yyyy/mm/dd
                ],
                'replacement': 'JFTIME'
            }
        }
    
    def process_worksheet_with_circular_pattern(self, worksheet: Worksheet) -> Dict[str, int]:
        """
        Process a worksheet using circular pattern checking for text replacement.
        
        Args:
            worksheet: The worksheet to process
            
        Returns:
            Dictionary with replacement statistics
        """
        replacement_stats = {category: 0 for category in self.replacement_patterns.keys()}
        
        logger.info(f"Processing worksheet '{worksheet.title}' with circular pattern checking")
        
        # Create a set to track cells we've already processed to avoid infinite loops
        processed_cells = set()
        
        # Process each cell in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cell_coord = cell.coordinate
                    
                    # Skip if we've already processed this cell
                    if cell_coord in processed_cells:
                        continue
                    
                    # Check if this cell contains a label we want to process
                    label_match = self._find_label_match(cell.value)
                    if label_match:
                        logger.debug(f"Found label pattern '{cell.value}' at {cell.coordinate}")
                        
                        # Safety check: don't process cells that are clearly in data tables
                        if self._is_likely_table_data_area(worksheet, cell):
                            logger.debug(f"Skipping {cell.coordinate} - appears to be in table data area")
                            continue
                        
                        # Use circular pattern to find the best target cell
                        target_cell = self._find_target_cell_circular(worksheet, cell, label_match)
                        
                        if target_cell and target_cell.coordinate not in processed_cells:
                            original_value = str(target_cell.value) if target_cell.value else "None"
                            target_cell.value = label_match['replacement']
                            replacement_stats[label_match['category']] += 1
                            
                            # Mark both cells as processed
                            processed_cells.add(cell_coord)
                            processed_cells.add(target_cell.coordinate)
                            
                            logger.info(f"Replaced '{original_value}' -> '{label_match['replacement']}' "
                                      f"at {target_cell.coordinate} (label at {cell.coordinate})")
        
        return replacement_stats
    
    def _find_label_match(self, text: str) -> Optional[Dict[str, str]]:
        """
        Check if text matches any of our label patterns.
        
        Args:
            text: The text to check
            
        Returns:
            Dictionary with replacement info if match found, None otherwise
        """
        for category, config in self.replacement_patterns.items():
            for pattern in config['patterns']:
                regex = re.compile(pattern, re.IGNORECASE)
                if regex.search(text):
                    return {
                        'category': category,
                        'replacement': config['replacement']
                    }
        return None
    
    def _find_target_cell_circular(self, worksheet: Worksheet, label_cell: Cell, 
                                 label_match: Dict[str, str]) -> Optional[Cell]:
        """
        Find the target cell using circular pattern checking around the label cell.
        
        Args:
            worksheet: The worksheet containing the cells
            label_cell: The cell containing the label
            label_match: Information about the matched label
            
        Returns:
            The best target cell for replacement, or None if not found
        """
        row = label_cell.row
        col = label_cell.column
        
        # Define the circular pattern: adjacent cells in priority order
        # Format: (row_offset, col_offset, priority_weight, description)
        circular_pattern = [
            (0, 1, 10, "right"),           # D20 -> E20 (most common)
            (0, 2, 8, "right+2"),          # D20 -> F20
            (1, 0, 7, "below"),            # D20 -> D21 
            (-1, 0, 6, "above"),           # D20 -> D19
            (1, 1, 5, "below-right"),      # D20 -> E21
            (-1, 1, 5, "above-right"),     # D20 -> E19
            (0, -1, 4, "left"),            # D20 -> C20
            (1, -1, 3, "below-left"),      # D20 -> C21
            (-1, -1, 3, "above-left"),     # D20 -> C19
            (0, 3, 2, "right+3"),          # D20 -> G20
            (2, 0, 2, "below+2"),          # D20 -> D22
            (-2, 0, 2, "above+2"),         # D20 -> D18
        ]
        
        candidates = []
        
        # Check each position in the circular pattern
        for row_offset, col_offset, priority, description in circular_pattern:
            target_row = row + row_offset
            target_col = col + col_offset
            
            # Check bounds
            if target_row < 1 or target_col < 1:
                continue
                
            try:
                target_cell = worksheet.cell(row=target_row, column=target_col)
                
                # Evaluate this cell as a potential target
                score = self._evaluate_target_cell(target_cell, label_match, priority)
                
                if score > 0:
                    candidates.append({
                        'cell': target_cell,
                        'score': score,
                        'position': description,
                        'coordinate': target_cell.coordinate
                    })
                    
                    logger.debug(f"Target candidate at {target_cell.coordinate} ({description}): "
                               f"score={score}, value='{target_cell.value}'")
                
            except Exception as e:
                logger.debug(f"Error checking cell at ({target_row}, {target_col}): {e}")
                continue
        
        # Select the best candidate
        if candidates:
            best_candidate = max(candidates, key=lambda x: x['score'])
            logger.debug(f"Selected best target: {best_candidate['coordinate']} "
                        f"({best_candidate['position']}) with score {best_candidate['score']}")
            return best_candidate['cell']
        
        logger.warning(f"No suitable target cell found for label at {label_cell.coordinate}")
        return None
    
    def _evaluate_target_cell(self, cell: Cell, label_match: Dict[str, str], 
                            position_priority: int) -> int:
        """
        Evaluate how suitable a cell is as a replacement target.
        
        Args:
            cell: The cell to evaluate
            label_match: Information about the matched label
            position_priority: Priority weight based on position (higher = better)
            
        Returns:
            Score for this cell (higher = better, 0 = unsuitable)
        """
        if cell.value is None:
            return 0  # Empty cells are not good targets
        
        cell_value = str(cell.value).strip()
        if not cell_value:
            return 0  # Empty string values are not good targets
        
        # Start with a lower base score to be more restrictive
        score = max(1, position_priority // 2)  # Reduced base score
        category = label_match['category']
        
        # Category-specific scoring - hunt for all date formats (slash AND datetime)
        if category in ['date', 'etd']:
            # Hunt for various date formats including Excel datetime formats
            if re.search(r'^\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{1,2}:\d{1,2}$', cell_value):
                score += 35  # Excel datetime format (2025-07-09 00:00:00)
            elif re.search(r'^\d{4}-\d{1,2}-\d{1,2}$', cell_value):
                score += 35  # ISO date format (2025-07-09)
            elif re.search(r'^\d{1,2}/\d{1,2}/\d{4}$', cell_value):
                score += 30  # dd/mm/yyyy OR mm/dd/yyyy format (25/07/2025 or 07/25/2025)
            elif re.search(r'^\d{4}/\d{1,2}/\d{1,2}$', cell_value):
                score += 30  # yyyy/mm/dd format (2025/07/25)
            elif re.search(r'^\d{1,2}/\d{1,2}/\d{2}$', cell_value):
                score += 25  # Short year format (25/07/25 or 07/25/25)
            elif re.search(r'^\d{4}/\d{1,2}$', cell_value):
                score += 20  # Year/month only (2025/07)
            elif re.search(r'^\d{1,2}/\d{1,2}$', cell_value):
                score += 20  # Month/day or day/month only (25/07 or 07/25)
            elif re.search(r'^\d{1,4}/\d{1,4}$', cell_value):
                score += 15  # Any two-part date with slashes
            else:
                return 0  # Reject anything that doesn't look like a date
                
        elif category == 'ref_no':
            # For reference numbers: MUST have the exact pattern
            if re.search(r'^[A-Z]{2,}\d+-\d+$', cell_value):
                score += 25  # Perfect match for ref format
            elif re.search(r'^[A-Z]{2,}\d+$', cell_value):
                score += 15  # Has letters and numbers but missing dash
            else:
                return 0  # Reject anything that doesn't look like a reference
                
        elif category == 'invoice_no':
            # For invoice numbers: flexible alphanumeric patterns (JF25003, KB-20244, MOTO123123E, MT-21233)
            if re.search(r'^[A-Z]{2,}[-]?\d+[A-Z]?$', cell_value):
                score += 30  # Perfect match: letters + optional dash + numbers + optional letter (KB-20244, JF25003)
            elif re.search(r'^[A-Z]+\d+[A-Z]+$', cell_value):
                score += 30  # Perfect match: letters + numbers + letters (MOTO123123E)
            elif re.search(r'^[A-Z]{2,}\d+$', cell_value):
                score += 25  # Good match: letters + numbers (JF25003)
            elif re.search(r'^[A-Z]{2,}-\d+$', cell_value):
                score += 25  # Good match: letters + dash + numbers (MT-21233)
            elif re.search(r'^\d+\-?\d*$', cell_value):
                score += 15  # Numeric only (legacy support)
            elif re.search(r'^[A-Z0-9\-]+$', cell_value) and len(cell_value) >= 4:
                score += 20  # General alphanumeric with dashes, minimum length
            else:
                return 0  # Reject anything that doesn't look like an invoice number
                
        elif category == 'contract_no':
            # For contract numbers: flexible alphanumeric patterns (same as invoice patterns)
            if re.search(r'^[A-Z]{2,}[-]?\d+[A-Z]?$', cell_value):
                score += 30  # Perfect match: letters + optional dash + numbers + optional letter
            elif re.search(r'^[A-Z]+\d+[A-Z]+$', cell_value):
                score += 30  # Perfect match: letters + numbers + letters
            elif re.search(r'^[A-Z]{2,}\d+$', cell_value):
                score += 25  # Good match: letters + numbers
            elif re.search(r'^[A-Z]{2,}-\d+$', cell_value):
                score += 25  # Good match: letters + dash + numbers
            elif re.search(r'^[A-Z0-9\-/]+$', cell_value) and len(cell_value) >= 4:
                score += 20  # General alphanumeric with separators, minimum length
            else:
                return 0  # Reject anything that doesn't look like a contract code
                
        elif category == 'date_label':
            # For date labels, look for date values in adjacent cells (including Excel datetime formats)
            if re.search(r'^\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{1,2}:\d{1,2}$', cell_value):
                score += 35  # Excel datetime format (2025-07-09 00:00:00)
            elif re.search(r'^\d{4}-\d{1,2}-\d{1,2}$', cell_value):
                score += 35  # ISO date format (2025-07-09)
            elif re.search(r'^\d{1,2}/\d{1,2}/\d{4}$', cell_value):
                score += 30  # dd/mm/yyyy OR mm/dd/yyyy format (25/07/2025 or 07/25/2025)
            elif re.search(r'^\d{4}/\d{1,2}/\d{1,2}$', cell_value):
                score += 30  # yyyy/mm/dd format (2025/07/25)
            elif re.search(r'^\d{1,2}/\d{1,2}/\d{2}$', cell_value):
                score += 25  # Short year format (25/07/25 or 07/25/25)
            elif re.search(r'^\d{4}/\d{1,2}$', cell_value):
                score += 20  # Year/month only (2025/07)
            elif re.search(r'^\d{1,2}/\d{1,2}$', cell_value):
                score += 20  # Month/day or day/month only (25/07 or 07/25)
            elif re.search(r'^\d{1,4}/\d{1,4}$', cell_value):
                score += 15  # Any two-part date with slashes
            else:
                return 0  # Reject anything that doesn't look like a date
                
        elif category == 'invoice_label':
            # For invoice labels, look for flexible invoice number patterns in adjacent cells
            if re.search(r'^[A-Z]{2,}[-]?\d+[A-Z]?$', cell_value):
                score += 30  # Perfect match: letters + optional dash + numbers + optional letter (KB-20244, JF25003)
            elif re.search(r'^[A-Z]+\d+[A-Z]+$', cell_value):
                score += 30  # Perfect match: letters + numbers + letters (MOTO123123E)
            elif re.search(r'^[A-Z]{2,}\d+$', cell_value):
                score += 25  # Good match: letters + numbers (JF25003)
            elif re.search(r'^[A-Z]{2,}-\d+$', cell_value):
                score += 25  # Good match: letters + dash + numbers (MT-21233)
            elif re.search(r'^\d+\-?\d*$', cell_value):
                score += 15  # Numeric only (legacy support)
            elif re.search(r'^[A-Z0-9\-]+$', cell_value) and len(cell_value) >= 4:
                score += 20  # General alphanumeric with dashes, minimum length
            else:
                return 0  # Only replace cells that look like invoice numbers
                
        elif category == 'ref_label':
            # For ref labels, look for reference codes in adjacent cells  
            if re.search(r'^[A-Z]{2,}\d+-\d+$', cell_value):
                score += 25  # Perfect match for ref format
            elif re.search(r'^[A-Z]{2,}\d+$', cell_value):
                score += 15  # Has letters and numbers but missing dash
            else:
                return 0  # Only replace cells that look like reference codes
        
        # Additional safety checks - reject cells that look like data content
        # Reject cells that are clearly table data (very common in spreadsheets)
        if re.search(r'^\d+\.\d+$', cell_value):  # Decimal numbers like prices
            return 0
        if re.search(r'^\d{1,3}(,\d{3})*(\.\d+)?$', cell_value):  # Formatted numbers
            return 0
        if len(cell_value) > 30:  # Very long text is likely descriptive content
            return 0
        
        # Penalize cells in areas that might be data tables
        if re.search(r'^(item|qty|quantity|price|amount|total|sum)', cell_value, re.IGNORECASE):
            return 0  # Don't replace cells that look like table headers
        
        logger.debug(f"Cell {cell.coordinate} evaluation: '{cell_value}' -> score {score}")
        return max(0, score)  # Ensure non-negative score
    
    def _is_likely_table_data_area(self, worksheet: Worksheet, cell: Cell) -> bool:
        """
        Check if a cell is likely in a table data area and should be avoided.
        
        Args:
            worksheet: The worksheet containing the cell
            cell: The cell to check
            
        Returns:
            True if the cell appears to be in a table data area
        """
        row = cell.row
        col = cell.column
        
        # Check surrounding cells for table-like patterns
        table_indicators = 0
        
        # Check for numeric data in nearby cells (common in tables)
        for r_offset in [-1, 0, 1]:
            for c_offset in [-2, -1, 0, 1, 2]:
                if r_offset == 0 and c_offset == 0:
                    continue
                try:
                    check_cell = worksheet.cell(row=row + r_offset, column=col + c_offset)
                    if check_cell.value is not None:
                        cell_str = str(check_cell.value).strip()
                        
                        # Look for typical table content
                        if re.search(r'^\d+\.\d+$', cell_str):  # Decimal numbers
                            table_indicators += 3
                        elif re.search(r'^\d{1,3}(,\d{3})*(\.\d+)?$', cell_str):  # Formatted numbers
                            table_indicators += 3
                        elif re.search(r'^\d+$', cell_str):  # Plain integers
                            table_indicators += 2
                        elif re.search(r'^(qty|quantity|price|amount|total|item|no\.|pcs|unit)', cell_str, re.IGNORECASE):
                            table_indicators += 4  # Table headers
                        elif len(cell_str) > 20 and not re.search(r'[:：]', cell_str):  # Long text without colons
                            table_indicators += 1
                            
                except Exception:
                    continue
        
        # If we find multiple table indicators, it's likely a table area
        return table_indicators >= 6
    
    def get_replacement_patterns(self) -> Dict[str, Any]:
        """
        Get the current replacement patterns.
        
        Returns:
            Dictionary of replacement patterns
        """
        return self.replacement_patterns.copy()
    
    def update_replacement_patterns(self, new_patterns: Dict[str, Any]) -> None:
        """
        Update the replacement patterns.
        
        Args:
            new_patterns: New patterns to merge with existing ones
        """
        self.replacement_patterns.update(new_patterns)
        logger.info(f"Updated replacement patterns: {list(new_patterns.keys())}")
    
    def add_custom_pattern(self, category: str, patterns: List[str], replacement: str) -> None:
        """
        Add a custom replacement pattern.
        
        Args:
            category: Category name for the pattern
            patterns: List of regex patterns to match
            replacement: Replacement value
        """
        self.replacement_patterns[category] = {
            'patterns': patterns,
            'replacement': replacement
        }
        logger.info(f"Added custom pattern category: {category}")


# Example usage and testing
if __name__ == "__main__":
    # This would be used for testing the enhanced processor
    print("Enhanced Text Processor with Circular Pattern Checking")
    print("Features:")
    print("- Intelligent label detection")
    print("- Circular pattern checking for target cells")
    print("- Flexible positioning for unusual document layouts")
    print("- Priority-based cell selection") 