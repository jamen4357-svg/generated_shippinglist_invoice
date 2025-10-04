# Invoice Generation Strategies and Utilities
# Refactored from pages/0_Generate_Invoice.py to eliminate code duplication

import streamlit as st
import os
import sys
import subprocess
import openpyxl
import json
import datetime
import tempfile
import zipfile
import io
import re
from pathlib import Path
from abc import ABC, abstractmethod
from typing import Dict, List, Any, Optional, Tuple
from zoneinfo import ZoneInfo
import logging

# Get the directory where this script is located
SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
# Add project subdirectories to the Python path to ensure correct module resolution
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(SCRIPT_DIR / "create_json"))
sys.path.insert(0, str(SCRIPT_DIR / "invoice_gen"))

# Import shared modules - will be imported when needed to avoid circular imports
# from invoice_gen.print_area_config import PrintAreaConfig


class InvoiceGenerationStrategy(ABC):
    """Abstract base class for invoice generation strategies"""

    def __init__(self, name: str, description: str):
        self.name = name
        self.description = description

    @abstractmethod
    def get_required_fields(self) -> List[str]:
        """Return list of required fields for validation"""
        pass

    @abstractmethod
    def validate_excel_data(self, excel_path: Path) -> Tuple[bool, List[str]]:
        """Validate Excel data structure and return (is_valid, warnings_list)"""
        pass

    @abstractmethod
    def validate_json_data(self, json_path: Path) -> List[str]:
        """Validate JSON data and return list of missing fields"""
        pass

    @abstractmethod
    def process_excel_to_json(self, excel_path: Path, json_output_dir: Path, **kwargs) -> Tuple[Path, str]:
        """Process Excel file to JSON and return (json_path, identifier)"""
        pass

    @abstractmethod
    def get_override_ui_config(self) -> Dict[str, Any]:
        """Return UI configuration for manual overrides"""
        pass

    @abstractmethod
    def apply_overrides(self, json_path: Path, overrides: Dict[str, Any]) -> bool:
        """Apply user overrides to JSON data"""
        pass

    @abstractmethod
    def get_generation_options(self) -> List[Dict[str, Any]]:
        """Return available generation options (Normal, DAF, Combine, etc.)"""
        pass

    @abstractmethod
    def generate_documents(self, json_path: Path, output_dir: Path, options: List[str], **kwargs) -> List[Path]:
        """Generate final documents and return list of generated files"""
        pass

    def _run_subprocess(self, command: List[str], cwd: Path, identifier_for_error: str) -> None:
        """A shared helper to run a subprocess and handle common errors."""
        sub_env = os.environ.copy()
        sub_env['PYTHONIOENCODING'] = 'utf-8'

        try:
            result = subprocess.run(
                command,
                check=True,
                capture_output=True,
                text=True,
                cwd=cwd,
                encoding='utf-8',
                errors='replace',
                env=sub_env
            )
            # Optional: Log success if needed
            # st.info(f"Subprocess executed successfully: {result.stdout}")

        except subprocess.CalledProcessError as e:
            error_msg = ((e.stdout or '') + (e.stderr or '')).lower()
            if any(keyword in error_msg for keyword in ['config', 'template', 'not found', 'missing', 'no such file']):
                self._show_config_error(identifier_for_error)
            else:
                st.error(f"A process failed to execute. Error: {e.stderr or e.stdout or 'Unknown error'}")
            raise # Re-raise the exception to halt execution

    def _show_config_error(self, po_number: str):
        """Displays a consistent, formatted error message when a PO config is missing."""
        st.error(f"**Configuration Error:** No company configuration found for PO **{po_number}**.")
        st.warning(
            "Please ensure a company is assigned to this PO in the **Company Setup** page "
            "before generating documents."
        )
        # Append po_number to the key to ensure uniqueness when called multiple times
        if st.button("🏢 Go to Company Setup", key=f"setup_{self.name.replace(' ', '_')}_{po_number}", use_container_width=True):
            st.switch_page("pages/3_SHIPPING_HEADER.py")


class HighQualityLeatherStrategy(InvoiceGenerationStrategy):
    """Strategy for High-Quality Leather invoice generation"""

    def __init__(self):
        super().__init__(
            name="High-Quality Leather",
            description="Standard invoice generation with Normal/DAF/Combine options"
        )

    def get_required_fields(self) -> List[str]:
        return ['po', 'item', 'pcs', 'sqft', 'pallet_count', 'unit', 'amount', 'net', 'gross', 'cbm', 'production_order_no']

    def validate_excel_data(self, excel_path: Path) -> Tuple[bool, List[str]]:
        """Validate Excel data structure for high-quality leather"""
        warnings = []

        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # Check if we have worksheets
            if len(workbook.worksheets) == 0:
                warnings.append("❌ Excel file has no worksheets")
                return False, warnings

            # Check each worksheet for required data structure
            valid_sheets = 0
            for sheet in workbook.worksheets:
                sheet_name = sheet.title

                # Get all values from the sheet
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell for cell in row):  # Skip empty rows
                        data.append(row)

                if len(data) < 2:  # Need at least header + 1 data row
                    warnings.append(f"⚠️ Sheet '{sheet_name}' has insufficient data (less than 2 rows)")
                    continue

                # Check for required columns in header
                header = [str(cell).lower().strip() if cell else "" for cell in data[0]]
                required_cols = ['po', 'item', 'pcs', 'sqft', 'pallet_count', 'unit', 'amount', 'net', 'gross', 'cbm']

                missing_cols = []
                for col in required_cols:
                    if not any(col in h or h in col for h in header):
                        missing_cols.append(col)

                if missing_cols:
                    warnings.append(f"⚠️ Sheet '{sheet_name}' missing columns: {', '.join(missing_cols)}")
                else:
                    valid_sheets += 1

                    # Check data quality
                    data_rows = data[1:]  # Skip header
                    if len(data_rows) == 0:
                        warnings.append(f"⚠️ Sheet '{sheet_name}' has header but no data rows")
                    else:
                        # Check for empty values in required columns
                        empty_count = 0
                        for row_idx, row in enumerate(data_rows, 2):  # Start from row 2 (1-indexed)
                            for col_idx, cell in enumerate(row):
                                if col_idx < len(header) and any(req in header[col_idx] for req in required_cols):
                                    if cell is None or str(cell).strip() == "":
                                        empty_count += 1

                        if empty_count > 0:
                            warnings.append(f"⚠️ Sheet '{sheet_name}' has {empty_count} empty cells in required columns")

            if valid_sheets == 0:
                warnings.append("❌ No worksheets contain the required data structure for High-Quality Leather invoices")
                return False, warnings

            # If we have warnings but at least one valid sheet, allow continuation
            if warnings:
                warnings.insert(0, f"✅ Found {valid_sheets} valid worksheet(s), but there are some issues to review:")
            else:
                warnings.append(f"✅ Excel validation passed! Found {valid_sheets} valid worksheet(s)")

            return True, warnings

        except Exception as e:
            warnings.append(f"❌ Error reading Excel file: {str(e)}")
            return False, warnings

    def validate_json_data(self, json_path: Path) -> List[str]:
        """Validate JSON data for high-quality leather format"""
        if not json_path.exists():
            st.error(f"Validation failed: JSON file '{json_path.name}' not found.")
            return self.get_required_fields()

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            missing_or_empty_keys = set(self.get_required_fields())

            if 'processed_tables_data' in data and isinstance(data['processed_tables_data'], dict):
                all_tables_data = {k: v for table in data['processed_tables_data'].values()
                                 for k, v in table.items()}

                for key in self.get_required_fields():
                    if key in all_tables_data and isinstance(all_tables_data[key], list) and \
                       any(item is not None and str(item).strip() for item in all_tables_data[key]):
                        missing_or_empty_keys.discard(key)

            return sorted(list(missing_or_empty_keys))

        except (json.JSONDecodeError, Exception) as e:
            st.error(f"Validation failed due to invalid JSON: {e}")
            return self.get_required_fields()

    def process_excel_to_json(self, excel_path: Path, json_output_dir: Path, **kwargs) -> Tuple[Path, str]:
        """Process Excel using create_json automation script"""
        import subprocess
        import sys

        identifier = Path(excel_path).stem
        json_path = json_output_dir / f"{identifier}.json"

        # Get the path to create_json/main.py
        create_json_script = SCRIPT_DIR / "create_json" / "main.py"

        with st.spinner(f"Processing '{identifier}' to generate JSON..."):
            # Use the correct CLI arguments that main.py actually supports
            command = [
                sys.executable,
                str(SCRIPT_DIR / "create_json" / "main.py"),
                "--input-excel", str(excel_path),
                "--output-dir", str(json_output_dir)
            ]
            try:
                # We use the generic subprocess runner.
                # Pass the root directory as cwd.
                self._run_subprocess(command, cwd=SCRIPT_DIR, identifier_for_error=identifier)
                st.success("Excel processing completed.")
            except subprocess.CalledProcessError as e:
                # The _run_subprocess method already logs the error to streamlit.
                # We just need to raise a runtime error to stop execution.
                raise RuntimeError("Excel to JSON processing script failed.") from e

        # Verify that the JSON file was created and is not empty
        if not json_path.exists() or json_path.stat().st_size == 0:
            st.error("The processing script ran but did not create a valid JSON file.")
            raise RuntimeError("Excel processing failed to create a JSON file.")

        return json_path, identifier

    def get_override_ui_config(self) -> Dict[str, Any]:
        """Return UI config for high-quality leather overrides"""
        return {
            "inv_no": {"type": "text_input", "label": "Invoice No", "default": ""},
            "inv_ref": {"type": "text_input", "label": "Invoice Ref", "default": "auto"},
            "inv_date": {"type": "date_input", "label": "Invoice Date", "default": "tomorrow"},
            "containers": {"type": "text_area", "label": "Container / Truck (One per line)", "default": ""}
        }

    def apply_overrides(self, json_path: Path, overrides: Dict[str, Any]) -> bool:
        """Apply overrides to high-quality leather JSON"""
        try:
            with open(json_path, 'r+', encoding='utf-8') as f:
                data = json.load(f)
                was_modified = False

                cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
                creating_date_str = datetime.datetime.now(cambodia_tz).strftime("%Y-%m-%d %H:%M:%S")

                if 'processed_tables_data' in data:
                    for table_data in data['processed_tables_data'].values():
                        num_rows = len(table_data.get('amount', []))
                        if num_rows == 0:
                            continue

                        # Only add creating_date if it doesn't already exist
                        if 'creating_date' not in table_data or not table_data['creating_date']:
                            table_data['creating_date'] = [creating_date_str] * num_rows
                            was_modified = True

                        # Apply user overrides
                        if overrides.get('inv_no'):
                            table_data['inv_no'] = [overrides['inv_no'].strip()] * num_rows
                            was_modified = True
                        if overrides.get('inv_ref'):
                            table_data['inv_ref'] = [overrides['inv_ref'].strip()] * num_rows
                            was_modified = True
                        if overrides.get('inv_date'):
                            # Convert date object to string format DD/MM/YYYY
                            if isinstance(overrides['inv_date'], datetime.date):
                                date_str = overrides['inv_date'].strftime("%d/%m/%Y")
                            else:
                                date_str = str(overrides['inv_date'])
                            table_data['inv_date'] = [date_str] * num_rows
                            was_modified = True
                        if overrides.get('containers'):
                            container_list = [line.strip() for line in overrides['containers'].split('\n') if line.strip()]
                            table_data['container_type'] = [', '.join(container_list)] * num_rows
                            was_modified = True

                if was_modified:
                    f.seek(0)
                    json.dump(data, f, indent=4)
                    f.truncate()

            return True

        except Exception as e:
            st.error(f"Error during JSON Override: {e}")
            return False

    def get_generation_options(self) -> List[Dict[str, Any]]:
        """Return generation options for high-quality leather"""
        return [
            {"name": "Normal Invoice", "key": "normal", "flags": []},
            {"name": "DAF Version", "key": "daf", "flags": ["--DAF"]},
            {"name": "Combine Version", "key": "combine", "flags": ["--custom"]}
        ]

    def generate_documents(self, json_path: Path, output_dir: Path, options: List[str], **kwargs) -> List[Path]:
        """Generate documents for high-quality leather"""
        generated_files = []
        identifier = kwargs.get('identifier', json_path.stem)

        # Import here to avoid circular imports
        INVOICE_GEN_DIR = kwargs.get('invoice_gen_dir')
        TEMPLATE_DIR = kwargs.get('template_dir')
        CONFIG_DIR = kwargs.get('config_dir')

        # Find incoterm from template
        detected_term = self._find_incoterm_from_template(identifier, TEMPLATE_DIR)

        for option in options:
            mode_name = option.upper()
            if option == 'combine':
                mode_name = f"{(detected_term or '').upper()} COMBINE".strip()

            output_filename = f"CT&INV&PL {identifier} {mode_name}.xlsx"
            output_path = output_dir / output_filename

            flags = []
            if option == 'daf':
                flags = ["--DAF"]
            elif option == 'combine':
                flags = ["--custom"]

            command = [
                sys.executable,
                str(INVOICE_GEN_DIR / "generate_invoice.py"),
                str(json_path),
                "--output", str(output_path),
                "--templatedir", str(TEMPLATE_DIR),
                "--configdir", str(CONFIG_DIR)
            ] + flags

            self._run_subprocess(command, cwd=INVOICE_GEN_DIR, identifier_for_error=identifier)

            import time
            time.sleep(0.5)  # Allow file to be fully written

            generated_files.append(output_path)

        return generated_files

    def _find_incoterm_from_template(self, identifier: str, template_dir: Path) -> Optional[str]:
        """Find incoterm from template file"""
        terms_to_find = ["DAP", "FCA", "CIP"]
        if not identifier:
            return None

        match = re.match(r'([A-Za-z]+)', identifier)
        if not match:
            return None

        template_file_path = template_dir / f"{match.group(1)}.xlsx"
        if not template_file_path.exists():
            return None

        try:
            workbook = openpyxl.load_workbook(template_file_path, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for term in terms_to_find:
                            if term in cell.value:
                                workbook.close()
                                return term
            workbook.close()
        except Exception:
            pass
        return None

    def _get_company_config_for_po(self, po_number: str) -> Optional[Dict[str, Any]]:
        """Retrieve company configuration based on PO number"""
        # This function's implementation is not shown in the original code
        # Assuming it queries some config or database to find the company associated with the PO
        pass


class SecondLayerLeatherStrategy(InvoiceGenerationStrategy):
    """Strategy for 2nd Layer Leather invoice generation"""

    def __init__(self):
        super().__init__(
            name="2nd Layer Leather",
            description="Aggregated invoice generation with unit pricing"
        )

    def get_required_fields(self) -> List[str]:
        # 2nd layer leather required fields for mapping
        return ['description', 'item', 'po', 'pcs', 'net', 'gross', 'cbm']

    def validate_excel_data(self, excel_path: Path) -> Tuple[bool, List[str]]:
        """Validate Excel data structure for 2nd layer leather"""
        warnings = []

        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # Check if we have worksheets
            if len(workbook.worksheets) == 0:
                warnings.append("❌ Excel file has no worksheets")
                return False, warnings

            # For 2nd layer, we expect aggregated data with unit pricing
            # Check each worksheet for required data structure
            valid_sheets = 0
            for sheet in workbook.worksheets:
                sheet_name = sheet.title

                # Get all values from the sheet
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell for cell in row):  # Skip empty rows
                        data.append(row)

                if len(data) < 1:
                    warnings.append(f"⚠️ Sheet '{sheet_name}' has no data at all")
                    continue

                # Check for required columns in header using the same mapping as processing script
                header = [str(cell).lower().strip() if cell else "" for cell in data[0]]

                # Import the header mapping from create_json config
                try:
                    import sys
                    create_json_path = SCRIPT_DIR / "create_json"
                    if str(create_json_path) not in sys.path:
                        sys.path.insert(0, str(create_json_path))
                    from config import TARGET_HEADERS_MAP
                except ImportError:
                    # Fallback to hardcoded mapping if config import fails
                    TARGET_HEADERS_MAP = {
                        "po": ["PO NO.", "po", "PO", "Po", "订单号", "order number", "order no"],
                        "item": ["物料代码", "item no", "ITEM NO.", "item", "Item No"],
                        "pcs": ["pcs", "总张数", "张数", "PCS"],
                        "sqft": ["sqft", "出货数量 (sf)", "尺数", "SF"],
                        "pallet_count": ["pallet count", "拖数", "PALLET", "件数"],
                        "unit": ["unit price", "单价", "price", "unit"],
                        "amount": ["金额 USD", "金额USD", "金额", "USD", "amount", "总价"],
                        "net": ["NW", "net weight", "净重kg", "净重"],
                        "gross": ["GW", "gross weight", "毛重", "gross"],
                        "cbm": ["cbm", "材积", "CBM"],
                        "production_order_no": ["production order number", "生产单号", "po", "入库单号"]
                    }

                # Check which required fields are present - FIXED: Don't break early, check all headers for all fields
                found_fields = set()
                for header_cell in header:
                    header_lower = header_cell.lower().strip()
                    for canonical_name, variations in TARGET_HEADERS_MAP.items():
                        for variation in variations:
                            variation_lower = variation.lower().strip()
                            if variation_lower in header_lower or header_lower in variation_lower:
                                found_fields.add(canonical_name)
                                # Don't break - one header might match multiple canonical names

                # Check for required fields (2nd layer has different requirements than high-quality)
                required_fields = ['description', 'item', 'po', 'pcs', 'net', 'gross', 'cbm']  # Core fields for 2nd layer mapping
                missing_fields = [field for field in required_fields if field not in found_fields]

                if missing_fields:
                    # For 2nd layer processing, warn about missing headers but don't fail validation
                    # The processing script can work with partial headers and will aggregate what it finds
                    warnings.append(f"⚠️ Sheet '{sheet_name}' missing columns: {', '.join(missing_fields)} (processing will use defaults)")
                    valid_sheets += 1
                else:
                    # For 2nd layer processing, we don't require all fields to have data
                    # The processing script can handle partial data and will aggregate what it finds
                    valid_sheets += 1
                    warnings.append(f"✅ Sheet '{sheet_name}' has required headers - processing will handle data aggregation")

                    # Optional: Check data quality but don't fail validation
                    data_rows = data[1:] if len(data) > 1 else []  # Skip header if it exists

                    if not data_rows:
                        warnings.append(f"⚠️ Sheet '{sheet_name}' has headers but no data rows")
                    else:
                        # Count how many required fields have data (informational only)
                        fields_with_data = set()
                        for row in data_rows:
                            for col_idx, cell in enumerate(row):
                                if col_idx < len(header):
                                    header_text = header[col_idx].lower().strip()
                                    for field_name, variations in TARGET_HEADERS_MAP.items():
                                        if any(var.lower().strip() in header_text or header_text in var.lower().strip() for var in variations):
                                            if field_name in required_fields:
                                                if cell is not None and str(cell).strip():
                                                    fields_with_data.add(field_name)

                        fields_without_data = [field for field in required_fields if field not in fields_with_data]
                        if fields_without_data:
                            warnings.append(f"ℹ️ Sheet '{sheet_name}' has no data in fields: {', '.join(fields_without_data)} (processing will use defaults/aggregation)")

                        # Check data quality for 2nd layer (different from high-quality)
                        data_rows = data[1:] if len(data) > 1 else []

                        if not data_rows:
                            warnings.append(f"⚠️ Sheet '{sheet_name}' has headers but no data rows - processing will create data")
                        else:
                            # Check for unit pricing (important for 2nd layer amount calculation)
                            has_unit_pricing = any(any(unit_var in h.lower() for unit_var in ['unit', '单价', 'price']) for h in header)
                            if not has_unit_pricing:
                                warnings.append(f"⚠️ Sheet '{sheet_name}' may be missing unit pricing data")

                            # Check for empty values in key fields
                            empty_count = 0
                            for row_idx, row in enumerate(data_rows, 2):
                                for col_idx, cell in enumerate(row):
                                    if col_idx < len(header):
                                        header_text = header[col_idx].lower().strip()
                                        for field_name, variations in TARGET_HEADERS_MAP.items():
                                            if any(var.lower().strip() in header_text or header_text in var.lower().strip() for var in variations):
                                                if field_name in required_fields:
                                                    if cell is None or str(cell).strip() == "":
                                                        empty_count += 1
                                                break  # Found a match for this header, no need to check other fields

                            if empty_count > 0:
                                warnings.append(f"⚠️ Sheet '{sheet_name}' has {empty_count} empty cells in key fields (Description, Item, PO, PCS, Net, Gross)")

            if valid_sheets == 0:
                warnings.append("❌ No worksheets contain the required data structure for 2nd Layer Leather invoices")
                return False, warnings

            # If we have warnings but at least one valid sheet, allow continuation
            if warnings:
                warnings.insert(0, f"✅ Found {valid_sheets} valid worksheet(s), but there are some issues to review:")
            else:
                warnings.append(f"✅ Excel validation passed! Found {valid_sheets} valid worksheet(s)")

            return True, warnings

        except Exception as e:
            warnings.append(f"❌ Error reading Excel file: {str(e)}")
            return False, warnings

    def validate_json_data(self, json_path: Path) -> List[str]:
        """Validate JSON data for 2nd layer format"""
        # 2nd layer validation is simpler - just check if file exists and is valid JSON
        if not json_path.exists():
            st.error(f"Validation failed: JSON file '{json_path.name}' not found.")
            return ["json_file"]

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                json.load(f)
            return []  # No missing fields for 2nd layer
        except (json.JSONDecodeError, Exception) as e:
            st.error(f"Validation failed due to invalid JSON: {e}")
            return ["json_file"]

    def process_excel_to_json(self, excel_path: Path, json_output_dir: Path, **kwargs) -> Tuple[Path, str]:
        """Process Excel using Second_Layer script - Create initial JSON, invoice details added later"""
        CREATE_JSON_DIR = kwargs.get('create_json_dir')

        identifier = Path(excel_path).stem
        json_path = json_output_dir / f"{identifier}.json"

        # Call the original script to create initial JSON (without invoice details)
        with st.spinner("Processing Excel file..."):
            cmd = [
                sys.executable,
                str(CREATE_JSON_DIR / "second_layer_main.py"),
                str(excel_path),
                "-o", str(json_path)
            ]

            try:
                self._run_subprocess(cmd, cwd=CREATE_JSON_DIR, identifier_for_error=identifier)
                # Check if the script actually succeeded (JSON file was created)
                if json_path.exists() and json_path.stat().st_size > 0:
                    st.success(f"Excel processing complete: '{json_path.name}' created.")
                else:
                    st.error("Excel processing FAILED - no JSON file created.")
                    raise RuntimeError("Excel processing failed to create a JSON file.")

            except subprocess.CalledProcessError as e:
                # The _run_subprocess method logs the error to streamlit, but we re-raise
                # a more informative error to halt execution and provide context.
                error_message = f"2nd Layer Excel to JSON script failed for '{identifier}'. STDERR: {e.stderr}"
                raise RuntimeError(error_message) from e

        return json_path, identifier

    def get_override_ui_config(self) -> Dict[str, Any]:
        """Return UI config for 2nd layer overrides"""
        return {
            "inv_ref": {"type": "text_input", "label": "Invoice Reference", "default": "auto"},
            "inv_date": {"type": "date_input", "label": "Invoice Date", "default": "today"},
            "unit_price": {"type": "number_input", "label": "Unit Price", "default": 0.61, "min": 0.0, "step": 0.01}
        }

    def apply_overrides(self, json_path: Path, overrides: Dict[str, Any]) -> bool:
        """Apply overrides to 2nd layer JSON - handled during processing"""
        # For 2nd layer, overrides are applied during the JSON creation process
        return True

    def get_generation_options(self) -> List[Dict[str, Any]]:
        """Return generation options for 2nd layer"""
        return [
            {"name": "Standard Invoice", "key": "standard", "flags": []}
        ]

    def generate_documents(self, json_path: Path, output_dir: Path, options: List[str], **kwargs) -> List[Path]:
        """Generate documents for 2nd layer - SIMPLIFIED"""
        generated_files = []
        po_number = kwargs.get('identifier', json_path.stem)

        INVOICE_GEN_DIR = kwargs.get('invoice_gen_dir')
        TEMPLATE_DIR = kwargs.get('template_dir')
        CONFIG_DIR = kwargs.get('config_dir')

        # Generate documents directly
        with st.spinner("Generating final documents..."):
            cmd = [
                sys.executable,
                str(INVOICE_GEN_DIR / "hybrid_generate_invoice.py"),
                str(json_path),
                "--outputdir", str(output_dir),
                "--templatedir", str(TEMPLATE_DIR),
                "--configdir", str(CONFIG_DIR)
            ]

            self._run_subprocess(cmd, cwd=INVOICE_GEN_DIR, identifier_for_error=po_number)

            import time
            time.sleep(1.0)  # Allow files to be fully written

            st.success("Documents generated successfully.")

            # Find generated files
            generated_files = list(output_dir.glob(f"* {po_number}.xlsx"))

        return generated_files

    def apply_overrides(self, json_path: Path, overrides: Dict[str, Any]) -> bool:
        """Apply overrides to 2nd layer leather JSON using TARGET_HEADERS_MAP"""
        try:
            with open(json_path, 'r+', encoding='utf-8') as f:
                data = json.load(f)
                was_modified = False

                cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
                creating_date_str = datetime.datetime.now(cambodia_tz).strftime("%Y-%m-%d %H:%M:%S")

                # Handle invoice metadata overrides
                if 'metadata' not in data:
                    data['metadata'] = {}

                if overrides.get('inv_no'):
                    data['metadata']['inv_no'] = overrides['inv_no']
                    was_modified = True

                if overrides.get('inv_ref') and overrides['inv_ref'] != 'auto':
                    data['metadata']['inv_ref'] = overrides['inv_ref']
                    was_modified = True

                if overrides.get('unit_price') is not None:
                    try:
                        data['metadata']['unit_price'] = float(overrides['unit_price'])
                        was_modified = True
                    except (ValueError, TypeError):
                        # If conversion fails, skip this override
                        pass

                if overrides.get('inv_date'):
                    if overrides['inv_date'] == 'tomorrow':
                        tomorrow = datetime.datetime.now(cambodia_tz) + datetime.timedelta(days=1)
                        data['metadata']['inv_date'] = tomorrow.strftime("%Y-%m-%d")
                    elif overrides['inv_date'] == 'today':
                        today = datetime.datetime.now(cambodia_tz)
                        data['metadata']['inv_date'] = today.strftime("%Y-%m-%d")
                    else:
                        # Handle date objects from Streamlit date_input
                        if hasattr(overrides['inv_date'], 'strftime'):
                            # It's a date/datetime object
                            data['metadata']['inv_date'] = overrides['inv_date'].strftime("%Y-%m-%d")
                        else:
                            # It's already a string
                            data['metadata']['inv_date'] = str(overrides['inv_date'])
                    was_modified = True

                # Handle field-specific overrides in aggregated_summary
                if 'aggregated_summary' in data:
                    summary = data['aggregated_summary']

                    # Apply field overrides using TARGET_HEADERS_MAP knowledge
                    field_overrides = {k: v for k, v in overrides.items() if k.startswith('override_')}
                    for override_key, override_value in field_overrides.items():
                        if override_value.strip():  # Only apply non-empty overrides
                            field_name = override_key.replace('override_', '')
                            if field_name in summary:
                                summary[field_name] = override_value.strip()
                                was_modified = True

                # Update creating_date if any modifications were made
                if was_modified:
                    data['metadata']['creating_date'] = creating_date_str

                    # Write back to file
                    f.seek(0)
                    json.dump(data, f, indent=4, ensure_ascii=False)
                    f.truncate()

                return was_modified

        except Exception as e:
            st.error(f"Failed to apply overrides: {e}")
            return False

    def generate_documents(self, json_path: Path, output_dir: Path, options: List[str], **kwargs) -> List[Path]:
        """Generate final documents for 2nd layer leather"""
        INVOICE_GEN_DIR = kwargs.get('invoice_gen_dir', SCRIPT_DIR / "invoice_gen")
        TEMPLATE_DIR = kwargs.get('template_dir', INVOICE_GEN_DIR / "TEMPLATE")
        CONFIG_DIR = kwargs.get('config_dir', INVOICE_GEN_DIR / "config")

        po_number = self._get_po_from_json(json_path) or "UNKNOWN"

        # JSON should already be updated with invoice details from the processing step
        # No need to update again here

        # Step 2: Generate documents
        with st.spinner("Generating final documents..."):
            cmd = [
                sys.executable,
                str(INVOICE_GEN_DIR / "hybrid_generate_invoice.py"),
                str(json_path),
                "--outputdir", str(output_dir),
                "--templatedir", str(TEMPLATE_DIR),
                "--configdir", str(CONFIG_DIR)
            ]

            # Add generation options
            for option in options:
                if option == "daf":
                    cmd.append("--daf")
                elif option == "combine":
                    cmd.append("--combine")

            self._run_subprocess(cmd, cwd=INVOICE_GEN_DIR, identifier_for_error=po_number)

            import time
            time.sleep(1.0)  # Allow files to be fully written

            st.success("Documents generated successfully!")

            # Find generated files
            generated_files = list(output_dir.glob(f"* {po_number}.xlsx"))

            if not generated_files:
                st.warning(f"No files found matching pattern '* {po_number}.xlsx' in {output_dir}")
                # Try broader search
                all_files = list(output_dir.glob("*.xlsx"))
                if all_files:
                    st.info(f"Found {len(all_files)} XLSX files: {[f.name for f in all_files]}")
                    generated_files = all_files
                else:
                    st.error("No XLSX files found in output directory")
        return generated_files

    def _get_po_from_json(self, json_path: Path) -> Optional[str]:
        """Extract PO number from JSON file"""
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Try to get PO from aggregated_summary first
            if 'aggregated_summary' in data and 'po' in data['aggregated_summary']:
                po_value = data['aggregated_summary']['po']
                if po_value and str(po_value).strip():
                    return str(po_value).strip()

            # Fallback to raw_data
            if 'raw_data' in data:
                for table_data in data['raw_data'].values():
                    if 'po' in table_data and table_data['po']:
                        first_po = table_data['po'][0] if isinstance(table_data['po'], list) and table_data['po'] else table_data['po']
                        if first_po and str(first_po).strip():
                            return str(first_po).strip()

        except Exception:
            pass

        return None

    def _update_and_aggregate_json(self, json_path: Path, po_number: str, **kwargs) -> Optional[Dict[str, Any]]:
        """Update and aggregate JSON data for 2nd layer processing - matches old working implementation"""
        try:
            with open(json_path, 'r+', encoding='utf-8') as f:
                data = json.load(f)
                raw_data = data.get("raw_data", {})
                summary = data.get("aggregated_summary", {})

                # Get parameters from kwargs
                inv_ref = kwargs.get('inv_ref', f"INV{po_number}")
                inv_date = kwargs.get('inv_date', datetime.date.today().strftime("%d/%m/%Y"))
                unit_price = kwargs.get('unit_price', 0.61)

                cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
                creating_date_str = datetime.datetime.now(cambodia_tz).strftime("%Y-%m-%d %H:%M:%S")

                # Calculate totals
                net_value = float(summary.get("net", 0))
                total_pcs = sum(sum(t.get("pcs", [])) for t in raw_data.values())
                total_pallets = sum(len(t.get("pallet_count", [])) for t in raw_data.values())
                total_amount = unit_price * net_value
                date_str = inv_date

                # Get first item and description
                first_item = next((item[0] for table in raw_data.values() if table.get("item") for item in [table["item"]] if item), "N/A")
                first_desc = next((desc[0] for table in raw_data.values() if table.get("description") for desc in [table["description"]] if desc), "N/A")

                # Update raw data with invoice info
                for table in raw_data.values():
                    entries = len(table.get("po", []))
                    table.update({"inv_no": [po_number] * entries, "inv_ref": [inv_ref] * entries, "inv_date": [date_str] * entries, "unit": [unit_price] * entries})

                # Update summary
                summary.update({
                    "inv_no": po_number, "inv_ref": inv_ref, "inv_date": date_str, "unit": unit_price, "amount": total_amount,
                    "pcs": total_pcs, "pallet_count": total_pallets, "net": net_value, "creating_date": creating_date_str,
                    "gross": summary.get("gross", 0.0), "cbm": summary.get("cbm", 0.0), "item": first_item, "description": first_desc
                })

                data["aggregated_summary"] = summary

                # Write back to file
                f.seek(0)
                json.dump(data, f, indent=4, ensure_ascii=False)
                f.truncate()

                return {
                    "po_number": po_number, "amount": total_amount, "pcs": total_pcs, "pallet_count": total_pallets,
                    "net": net_value, "gross": summary.get("gross", 0.0), "cbm": summary.get("cbm", 0.0),
                    "item": first_item, "description": first_desc
                }

        except Exception as e:
            st.error(f"Failed to update JSON: {e}")
            return None


# Shared utility functions
def apply_print_settings_to_files(file_paths: List[Path], invoice_gen_dir: Path = None) -> Tuple[int, int]:
    """
    Apply print area configuration to multiple Excel files
    Returns (files_processed, total_sheets_processed)
    """
    if not file_paths:
        return 0, 0

    # Import here to avoid circular imports
    if invoice_gen_dir and str(invoice_gen_dir) not in sys.path:
        sys.path.insert(0, str(invoice_gen_dir))

    try:
        from invoice_gen.print_area_config import PrintAreaConfig
    except ImportError:
        st.error("Could not import PrintAreaConfig. Please check invoice_gen directory.")
        return 0, 0

    config = PrintAreaConfig()
    total_files_processed = 0
    total_sheets_processed = 0

    for excel_file in file_paths:
        if not excel_file.exists():
            continue

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=False)

            # Apply print area configuration to ALL worksheets
            sheets_processed = []

            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    # Skip hidden sheets
                    if ws.sheet_state != 'visible':
                        continue

                    config.configure_print_settings(ws)
                    sheets_processed.append(sheet_name)
                except Exception as sheet_error:
                    continue

            wb.save(excel_file)
            wb.close()

            # Additional delay to ensure save is complete
            import time
            time.sleep(0.2)

            total_files_processed += 1
            total_sheets_processed += len(sheets_processed)

            st.info(f"✅ Print settings applied to {excel_file.name} ({len(sheets_processed)} sheets)")

        except Exception as print_error:
            st.error(f"❌ Failed to apply print settings to {excel_file.name}: {print_error}")

    return total_files_processed, total_sheets_processed


def create_download_zip(files_to_zip: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Create a ZIP file from the provided files
    files_to_zip format: [{"name": filename, "data": file_bytes}, ...]
    """
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_info in files_to_zip:
            zf.writestr(file_info["name"], file_info["data"])
    return zip_buffer


def get_suggested_inv_ref() -> str:
    """Generate suggested invoice reference"""
    # This is a simplified version - you may want to implement more sophisticated logic
    today = datetime.date.today()
    # Generate a reference like HQ001, HQ002, etc. for the day
    return f"HQ{today.strftime('%d%m%y')}001"  # Simplified suggestion


def check_existing_identifiers(inv_no: str = None, inv_ref: str = None) -> Dict[str, bool]:
    """Check if invoice identifiers already exist"""
    # This would need access to the database
    # For now, return empty checks
    return {"inv_no": False, "inv_ref": False}


# Strategy registry
STRATEGIES = {
    "high_quality": HighQualityLeatherStrategy(),
    "second_layer": SecondLayerLeatherStrategy()
}