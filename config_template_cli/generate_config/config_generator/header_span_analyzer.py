"""
HeaderSpanAnalyzer component for detecting colspan and rowspan in Excel headers.

This module provides functionality to analyze Excel files and detect merged cells
in header sections, then map them to column IDs using existing header text mapping.
"""

from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from .header_text_updater import HeaderTextUpdater
from .models import QuantityAnalysisData


class HeaderSpanAnalyzerError(Exception):
    """Custom exception for HeaderSpanAnalyzer errors."""
    pass


class HeaderSpanAnalyzer:
    """Analyzes Excel headers for colspan and rowspan detection."""
    
    def __init__(self, excel_file_path: Optional[str] = None, mapping_config_path: str = "mapping_config.json"):
        """
        Initialize HeaderSpanAnalyzer.
        
        Args:
            excel_file_path: Path to Excel file to analyze (can be set later)
            mapping_config_path: Path to mapping configuration file
        """
        self.excel_file_path = excel_file_path
        self.header_text_updater = HeaderTextUpdater(mapping_config_path)
        
    def analyze_header_spans(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Analyze Excel file for header spans across all sheets.
        
        Returns:
            Dictionary with sheet names and their header span data
            Format: {
                "Contract": [
                    {
                        "text": "Product Information",
                        "col_id": "col_desc",
                        "rowspan": 1,
                        "colspan": 3
                    }
                ]
            }
        """
        try:
            if not self.excel_file_path:
                raise HeaderSpanAnalyzerError("Excel file path not set. Call update_header_spans with QuantityAnalysisData.")
                
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)
            header_spans = {}
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_spans = self._analyze_sheet_header_spans(worksheet, sheet_name)
                if sheet_spans:
                    header_spans[sheet_name] = sheet_spans
            
            workbook.close()
            return header_spans
            
        except Exception as e:
            raise HeaderSpanAnalyzerError(f"Failed to analyze header spans: {str(e)}") from e
    
    def _analyze_sheet_header_spans(self, worksheet: Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Analyze a single worksheet for header spans.
        
        Args:
            worksheet: Excel worksheet to analyze
            sheet_name: Name of the sheet for logging
            
        Returns:
            List of header span data for this sheet
        """
        header_spans = []
        
        # Special handling for Contract sheet based on actual Excel structure
        if sheet_name.lower() == 'contract':
            # Based on MCP Excel analysis, Contract sheet headers are in row 11:
            # A11:F11 merged = "Description of Goods."
            # G11 = "Unit Price" 
            # I11:J11 merged = "Quantity(SF)"
            
            # Check A11 (Description of Goods)
            cell_a11 = worksheet.cell(row=11, column=1)
            if cell_a11.value:
                header_text = str(cell_a11.value).strip()
                col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
                if col_id and not col_id.startswith('unmapped_'):
                    # Check if it's merged (A11:F11)
                    merged_ranges = list(worksheet.merged_cells.ranges)
                    merged_range = self._find_merged_range_for_cell(11, 1, merged_ranges)
                    if merged_range:
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                    else:
                        rowspan = 1
                        colspan = 1
                    
                    header_spans.append({
                        "text": header_text,
                        "col_id": col_id,
                        "rowspan": rowspan,
                        "colspan": colspan
                    })
            
            # Check G11 (Unit Price)
            cell_g11 = worksheet.cell(row=11, column=7)
            if cell_g11.value:
                header_text = str(cell_g11.value).strip()
                col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
                if col_id and not col_id.startswith('unmapped_'):
                    header_spans.append({
                        "text": header_text,
                        "col_id": col_id,
                        "rowspan": 1,
                        "colspan": 1
                    })
            
            # Check I11 (Quantity)
            cell_i11 = worksheet.cell(row=11, column=9)
            if cell_i11.value:
                header_text = str(cell_i11.value).strip()
                col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
                if col_id and not col_id.startswith('unmapped_'):
                    # Check if it's merged (I11:J11)
                    merged_ranges = list(worksheet.merged_cells.ranges)
                    merged_range = self._find_merged_range_for_cell(11, 9, merged_ranges)
                    if merged_range:
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                    else:
                        rowspan = 1
                        colspan = 1
                    
                    header_spans.append({
                        "text": header_text,
                        "col_id": col_id,
                        "rowspan": rowspan,
                        "colspan": colspan
                    })
            
            return header_spans
        
        # For other sheets, use the general approach
        # Get all merged cell ranges in the worksheet
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        # STEP 1: Find ALL table headers (merged and non-merged) in table area
        # Check rows 10-25 where table headers can be (expanded range)
        for row in range(10, 26):
            for col in range(1, 20):  # Columns A-S
                cell = worksheet.cell(row, col)
                
                # Skip empty cells
                if not cell.value:
                    continue
                
                header_text = str(cell.value).strip()
                
                # Skip if this looks like data, not a header
                if self._is_likely_data_cell(header_text):
                    continue
                
                # Check if this maps to existing col_id (actual table header)
                col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
                
                if col_id and not col_id.startswith('unmapped_'):
                    # This is a valid table header!
                    
                    # Check if it's part of a merged range
                    merged_range = self._find_merged_range_for_cell(row, col, merged_ranges)
                    
                    if merged_range:
                        # It's merged - calculate spans
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                    else:
                        # It's NOT merged - single cell
                        rowspan = 1
                        colspan = 1
                    
                    # Add to results (avoid duplicates)
                    span_data = {
                        "text": header_text,
                        "col_id": col_id,
                        "rowspan": rowspan,
                        "colspan": colspan
                    }
                    
                    # Check if we already have this header
                    if not any(h['text'] == header_text and h['col_id'] == col_id for h in header_spans):
                        header_spans.append(span_data)
        
        return header_spans
    
    def _is_likely_data_cell(self, text):
        """Check if this text looks like data rather than a header."""
        if not text:
            return True
            
        text = str(text).strip()
        
        # Check for pure numbers
        if text.replace('.', '').replace(',', '').replace('-', '').isdigit():
            return True
            
        # Check for currency/amount patterns
        if any(char in text for char in ['$', '€', '¥', '£']) and any(c.isdigit() for c in text):
            return True
            
        # Check for date patterns (MM/DD/YYYY, DD-MM-YYYY, etc.)
        import re
        date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{2,4}[/-]\d{1,2}[/-]\d{1,2}',
        ]
        if any(re.match(pattern, text) for pattern in date_patterns):
            return True
            
        # Check for very specific data indicators (removed "No." as it can be a valid header)
        data_indicators = [
            'total', 'subtotal', 'sum',
            'balance', 'amount due'
        ]
        text_lower = text.lower()
        if any(indicator in text_lower for indicator in data_indicators):
            return True
            
        return False
    
    def _extract_span_data(self, worksheet: Worksheet, merged_range, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Extract span data from a merged cell range.
        
        Args:
            worksheet: Excel worksheet
            merged_range: Merged cell range object
            sheet_name: Sheet name for logging
            
        Returns:
            Header span data or None if invalid
        """
        try:
            # Get the top-left cell (contains the text)
            top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            header_text = str(top_left_cell.value).strip() if top_left_cell.value else ""
            
            # Skip empty cells
            if not header_text or header_text.lower() in ['none', 'null', '']:
                return None
            
            # Only include headers that map to existing col_ids (table headers)
            col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
            
            if not col_id or col_id.startswith('unmapped_'):
                # Skip headers that don't map to existing col_ids
                return None
            
            # Calculate span dimensions
            rowspan = merged_range.max_row - merged_range.min_row + 1
            colspan = merged_range.max_col - merged_range.min_col + 1
            
            return {
                "text": header_text,
                "col_id": col_id,
                "rowspan": rowspan,
                "colspan": colspan
            }
            
        except Exception as e:
            print(f"[SPAN_ANALYZER] Error extracting span data: {e}")
            return None
    
    def _find_non_merged_table_headers(self, worksheet: Worksheet, merged_ranges: List) -> List[Dict[str, Any]]:
        """
        Find table headers that are not merged (single cell headers) in table area.
        
        Args:
            worksheet: Excel worksheet
            merged_ranges: List of merged cell ranges to avoid
            
        Returns:
            List of non-merged header span data
        """
        non_merged_headers = []
        
        # Check table header area (rows 15-25)
        for row in range(15, 26):  # Check rows 15-25
            for col in range(1, 20):  # Check columns A-S
                cell = worksheet.cell(row, col)
                
                # Skip empty cells
                if not cell.value:
                    continue
                
                # Skip if this cell is part of a merged range
                if self._is_cell_merged(row, col, merged_ranges):
                    continue
                
                header_text = str(cell.value).strip()
                
                # Only include if it maps to existing col_id (table headers)
                col_id = self.header_text_updater.map_header_to_column_id(header_text, strict_mode=True)
                
                if col_id and not col_id.startswith('unmapped_'):
                    non_merged_headers.append({
                        "text": header_text,
                        "col_id": col_id,
                        "rowspan": 1,
                        "colspan": 1
                    })
        
        return non_merged_headers
    
    def _find_merged_range_for_cell(self, row: int, col: int, merged_ranges: List):
        """Find the merged range that contains this cell, if any."""
        for merged_range in merged_ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return merged_range
        return None
    
    def _is_cell_merged(self, row: int, col: int, merged_ranges: List) -> bool:
        """Check if a cell is part of any merged range."""
        for merged_range in merged_ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return True
        return False
    
    def _looks_like_header(self, text: str) -> bool:
        """
        Determine if text looks like a header.
        
        Args:
            text: Cell text to check
            
        Returns:
            True if looks like a header
        """
        if len(text) < 2:
            return False
        
        # Common header indicators
        header_keywords = [
            'no', 'number', 'item', 'description', 'quantity', 'price', 'amount',
            'unit', 'total', 'po', 'mark', 'cbm', 'weight', 'pallet', 'remarks',
            'cargo', 'commodity', 'fca', 'usd', 'kgs', 'pcs', 'sf'
        ]
        
        text_lower = text.lower()
        return any(keyword in text_lower for keyword in header_keywords)
    
    def get_spans_for_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Get header spans for a specific sheet.
        
        Args:
            sheet_name: Name of the sheet to analyze
            
        Returns:
            List of header span data for the specified sheet
        """
        all_spans = self.analyze_header_spans()
        return all_spans.get(sheet_name, [])

    def update_header_spans(self, config: Dict[str, Any], quantity_data: QuantityAnalysisData) -> Dict[str, Any]:
        """
        Update configuration with header span information.
        
        This method follows the same pattern as other updaters in the ConfigGenerator workflow.
        It sets the Excel file path from quantity_data and then analyzes headers for spans.
        
        Args:
            config: Configuration to update
            quantity_data: Quantity analysis data containing file path and sheet info
            
        Returns:
            Updated configuration with span data added to header_to_write sections
            
        Raises:
            HeaderSpanAnalyzerError: If span analysis fails
        """
        try:
            # Set the Excel file path from quantity data
            self.excel_file_path = quantity_data.file_path
            
            # Analyze header spans
            span_data = self.analyze_header_spans()
            
            # Update configuration with span data
            updated_config = config.copy()
            
            # Process each sheet in the configuration
            for sheet_key, sheet_config in updated_config.items():
                if not isinstance(sheet_config, dict):
                    continue
                    
                # Look for header_to_write sections
                if 'header_to_write' in sheet_config:
                    sheet_name = sheet_key  # Assume sheet key matches sheet name
                    
                    # Get spans for this sheet
                    sheet_spans = span_data.get(sheet_name, [])
                    
                    # Update each header entry with span data
                    header_to_write = sheet_config['header_to_write']
                    if isinstance(header_to_write, dict):
                        for col_id, header_info in header_to_write.items():
                            if isinstance(header_info, dict):
                                # Find matching span data for this col_id
                                for span in sheet_spans:
                                    if span.get('col_id') == col_id:
                                        # Add span information
                                        header_info['rowspan'] = span.get('rowspan', 1)
                                        header_info['colspan'] = span.get('colspan', 1)
                                        break
            
            return updated_config
            
        except Exception as e:
            raise HeaderSpanAnalyzerError(f"Failed to update header spans: {str(e)}") from e
