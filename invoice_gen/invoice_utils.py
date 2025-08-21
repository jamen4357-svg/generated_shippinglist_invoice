from pickle import NONE
import openpyxl
import re
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from typing import List, Dict, Any, Optional, Tuple, Union
from decimal import Decimal
from decimal import Decimal, InvalidOperation
import merge_utils

# --- Constants for Styling ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) # Full grid border
no_border = Border(left=None, right=None, top=None, bottom=None)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold_font = Font(bold=True)

# --- Constants for Number Formats ---
FORMAT_GENERAL = 'General'
FORMAT_TEXT = '@'
FORMAT_NUMBER_COMMA_SEPARATED1 = '#,##0'
FORMAT_NUMBER_COMMA_SEPARATED2 = '#,##0.00'

# --- Utility Functions ---

def unmerge_row(worksheet: Worksheet, row_num: int, num_cols: int):
    """
    Unmerges any merged cells that overlap with the specified row within the given column range.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to unmerge.
        num_cols: The number of columns to check for merges.
    """
    if row_num <= 0:
        return
    merged_ranges_copy = list(worksheet.merged_cells.ranges) # Copy ranges before modification
    merged_ranges_to_remove = []

    # Identify ranges that overlap with the target row
    for merged_range in merged_ranges_copy:
        # Check if the range's row span includes the target row_num
        # And if the range's column span overlaps with columns 1 to num_cols
        overlap = (merged_range.min_row <= row_num <= merged_range.max_row and
                   max(merged_range.min_col, 1) <= min(merged_range.max_col, num_cols))
        if overlap:
            merged_ranges_to_remove.append(str(merged_range))

    if merged_ranges_to_remove:
        for range_str in merged_ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except KeyError:
                # Range might have been removed by unmerging an overlapping one
                pass
            except Exception as unmerge_err:
                # Log or handle other potential errors if needed
                pass
    else:
        # No overlapping merges found for this row
        pass


def unmerge_block(worksheet: Worksheet, start_row: int, end_row: int, num_cols: int):
    """
    Unmerges any merged cells that overlap with the specified row range and column range.
    Args:
        worksheet: The openpyxl Worksheet object.
        start_row: The 1-based starting row index of the block.
        end_row: The 1-based ending row index of the block.
        num_cols: The number of columns to check for merges.
    """
    if start_row <= 0 or end_row < start_row:
        return
    merged_ranges_copy = list(worksheet.merged_cells.ranges) # Copy ranges before modification
    merged_ranges_to_remove = []

    # Identify ranges that overlap with the target block
    for merged_range in merged_ranges_copy:
        mr_min_row, mr_min_col, mr_max_row, mr_max_col = merged_range.bounds
        row_overlap = max(mr_min_row, start_row) <= min(mr_max_row, end_row)
        col_overlap = max(mr_min_col, 1) <= min(mr_max_col, num_cols)

        if row_overlap and col_overlap:
            range_str = str(merged_range)
            if range_str not in merged_ranges_to_remove: # Avoid duplicates
                merged_ranges_to_remove.append(range_str)

    if merged_ranges_to_remove:
        for range_str in merged_ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except KeyError:
                # Range might have been removed by unmerging an overlapping one
                pass
            except Exception as unmerge_err:
                # Log or handle other potential errors if needed
                pass
    else:
        # No overlapping merges found in this block
        pass


def safe_unmerge_block(worksheet: Worksheet, start_row: int, end_row: int, num_cols: int):
    """
    Safely unmerges only cells within the specific target range, preventing unintended unmerging
    of cells completely outside the block.
    """
    if start_row <= 0 or end_row < start_row:
        return

    # Only process merges that actually intersect with our target range
    for merged_range in list(worksheet.merged_cells.ranges):
        # Check if this merge intersects our target range
        if (merged_range.min_row <= end_row and
            merged_range.max_row >= start_row and
            merged_range.min_col <= num_cols and
            merged_range.max_col >= 1):
            try:
                worksheet.unmerge_cells(merged_range.coord)
            except (KeyError, ValueError, AttributeError):
                # Ignore errors if the range is somehow invalid or already unmerged
                continue

    return True


def fill_static_row(worksheet: Worksheet, row_num: int, num_cols: int, static_content_dict: Dict[str, Any]):
    """
    Fills a specific row with static content defined in a dictionary.
    Applies default alignment and no border.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to fill.
        num_cols: The total number of columns in the table context (for bounds checking).
        static_content_dict: Dictionary where keys are column indices (as strings or ints)
                             and values are the static content to write.
    """
    if not static_content_dict:
        return # Nothing to do
    if row_num <= 0:
        return

    for col_key, value in static_content_dict.items():
        target_col_index = None
        try:
            # Attempt to convert key to integer column index
            target_col_index = int(col_key)
            # Check if the column index is within the valid range
            if 1 <= target_col_index <= num_cols:
                cell = worksheet.cell(row=row_num, column=target_col_index)
                cell.value = value
                # Apply default styling for static rows
                cell.alignment = center_alignment # Default alignment
                cell.border = no_border # Default: no border for static rows
                # Apply basic number formatting
                if isinstance(value, (int, float)):
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2 if isinstance(value, float) else FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    cell.number_format = FORMAT_TEXT # Treat as text otherwise
            else:
                # Column index out of range, log warning?
                pass
        except (ValueError, TypeError) as e:
            # Invalid column key, log warning?
            pass
        except Exception as cell_err:
            # Error accessing cell, log warning?
            pass


def apply_row_merges(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to apply merges to.
        num_cols: The total number of columns in the table context.
        merge_rules: Dictionary where keys are starting column indices (as strings or ints)
                     and values are the number of columns to span (colspan).
    """
    if not merge_rules or row_num <= 0:
        return # No rules or invalid row

    try:
        # Convert string keys to integers and sort for predictable application order
        rules_with_int_keys = {int(k): v for k, v in merge_rules.items()}
        sorted_keys = sorted(rules_with_int_keys.keys())
    except (ValueError, TypeError) as e:
        # Invalid key format in merge_rules
        return

    for start_col in sorted_keys:
        colspan_val = rules_with_int_keys[start_col]
        try:
            # Ensure colspan is an integer
            colspan = int(colspan_val)
        except (ValueError, TypeError):
            # Invalid colspan value
            continue

        # Basic validation for start column and colspan
        if not isinstance(start_col, int) or not isinstance(colspan, int) or start_col < 1 or colspan < 1:
            continue

        # Calculate end column, ensuring it doesn't exceed the table width
        end_col = start_col + colspan - 1
        if end_col > num_cols:
            end_col = num_cols
            # Check if clamping made the range invalid (start > end)
            if start_col > end_col:
                continue

        merge_range_str = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
        try:
            # --- Pre-Unmerge Overlapping Cells ---
            merges_to_clear = []
            current_merged_ranges = list(worksheet.merged_cells.ranges) # Work on a copy
            for merged_range in current_merged_ranges:
                # Check if the existing merge overlaps with the target row and column range
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    if max(merged_range.min_col, start_col) <= min(merged_range.max_col, end_col):
                        range_to_remove_str = str(merged_range)
                        if range_to_remove_str not in merges_to_clear:
                            merges_to_clear.append(range_to_remove_str)
            if merges_to_clear:
                for r_str in merges_to_clear:
                    try: worksheet.unmerge_cells(r_str)
                    except KeyError: pass
                    except Exception as unmerge_err_inner: pass # Log?
            # --- End Pre-Unmerge ---

            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            # Apply alignment to the top-left cell of the merged range
            top_left_cell = worksheet.cell(row=row_num, column=start_col)
            if not top_left_cell.alignment or top_left_cell.alignment.horizontal is None:
                top_left_cell.alignment = center_alignment # Apply center alignment if none exists
        except ValueError as ve:
            # This can happen if trying to merge over an existing merged cell that wasn't properly unmerged
            pass
        except Exception as merge_err:
            # Log or handle other merge errors
            pass

def _apply_cell_style(cell, column_id: Optional[str], sheet_styling_config: Optional[Dict[str, Any]] = None, fob_mode: Optional[bool] = False):
    """
    Applies font, alignment, and number format to a cell based on a column ID.
    """
    if not sheet_styling_config or not cell or not column_id:
        return

    try:
        # Get styling configurations using ID-based keys
        default_font_cfg = sheet_styling_config.get("default_font", {})
        default_align_cfg = sheet_styling_config.get("default_alignment", {})
        column_styles = sheet_styling_config.get("column_id_styles", {}) # <-- Uses "column_id_styles"

        # Find column-specific style rules if the ID matches
        col_specific_style = column_styles.get(column_id, {})

        # --- Apply Font ---
        final_font_cfg = default_font_cfg.copy()
        final_font_cfg.update(col_specific_style.get("font", {}))
        if final_font_cfg:
            cell.font = Font(**{k: v for k, v in final_font_cfg.items() if v is not None})

        # --- Apply Alignment ---
        final_align_cfg = default_align_cfg.copy()
        final_align_cfg.update(col_specific_style.get("alignment", {}))
        if final_align_cfg:
            cell.alignment = Alignment(**{k: v for k, v in final_align_cfg.items() if v is not None})
            
        # --- Apply Number Format ---
        number_format = col_specific_style.get("number_format")
        
        # PCS always uses config format, never forced format
        if column_id in ['col_pcs', 'col_qty_pcs']:
            if number_format and cell.number_format != FORMAT_TEXT:
                cell.number_format = number_format
        else:
            # Non-PCS columns follow FOB mode logic
            if number_format and cell.number_format != FORMAT_TEXT and not fob_mode:
                cell.number_format = number_format
            elif number_format and cell.number_format != FORMAT_TEXT and fob_mode:
                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2
            elif cell.number_format != FORMAT_TEXT and (cell.number_format == FORMAT_GENERAL or cell.number_format is None):
                if isinstance(cell.value, float): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2
                elif isinstance(cell.value, int): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    except Exception as style_err:
        print(f"Error applying cell style for ID {column_id}: {style_err}")



def write_grand_total_weight_summary(
    worksheet: Worksheet,
    start_row: int,
    header_info: Dict[str, Any],
    processed_tables_data: Dict[str, Dict[str, List[Any]]],
    weight_config: Dict[str, Any],
    styling_config: Optional[Dict[str, Any]] = None
) -> int:
    """
    Calculates GRAND TOTAL of Net/Gross weights, inserts two new rows,
    and writes a styled two-row summary using the main footer's style.

    Args:
        worksheet: The openpyxl worksheet to modify.
        start_row: The row index to start writing from.
        header_info: The header dictionary containing 'column_id_map' and 'num_columns'.
        processed_tables_data: The dictionary containing all table data.
        weight_config: The configuration object for the weight summary.
        footer_config: The main footer configuration for the sheet, used for styling.
    """
    footer_row_height = styling_config.get("styling", {}).get("row_heights", {}).get("footer", None)
    footer_config = styling_config.get("footer_configurations", {})

    if not weight_config.get("enabled"):
        return start_row

    print(f"--- Calculating and writing GRAND TOTAL Net/Gross Weight summary ---")

    # --- Calculation Logic (no changes here) ---
    grand_total_net = Decimal('0')
    grand_total_gross = Decimal('0')

    for table_data in processed_tables_data.values():
        net_weights = table_data.get("net", [])
        gross_weights = table_data.get("gross", [])
        for weight in net_weights:
            try:
                grand_total_net += Decimal(str(weight))
            except (InvalidOperation, TypeError, ValueError):
                continue
        for weight in gross_weights:
            try:
                grand_total_gross += Decimal(str(weight))
            except (InvalidOperation, TypeError, ValueError):
                continue

    # --- Get Column Indices and Dimensions (no changes here) ---
    col_id_map = header_info.get("column_id_map", {})
    num_columns = header_info.get("num_columns", 1)
    label_col_idx = col_id_map.get(weight_config.get("label_col_id"))
    value_col_idx = col_id_map.get(weight_config.get("value_col_id"))

    if not all([label_col_idx, value_col_idx]):
        print("Warning: Could not write grand total weight summary. Label/Value column ID not found.")
        return start_row

    # --- MODIFICATION: Parse Styling from the main footer_config ---
    # It now uses the new 'footer_config' parameter
    style_config = footer_config.get('style', {})
    font_to_apply = Font(**style_config.get('font', {'bold': True}))
    align_to_apply = Alignment(**style_config.get('alignment', {'horizontal': 'right', 'vertical': 'center'}))

    # --- Insert and unmerge rows (no changes here) ---
    try:
        worksheet.insert_rows(start_row, amount=2)
        unmerge_row(worksheet, start_row, num_columns)
        unmerge_row(worksheet, start_row + 1, num_columns)
    except Exception as insert_err:
        print(f"Error inserting/unmerging rows for weight summary: {insert_err}")
        return start_row

    # --- Write the final rows and apply styles (no changes here) ---
    net_weight_row = start_row
    gross_weight_row = start_row + 1

    try:
        cell_net_label = worksheet.cell(row=net_weight_row, column=label_col_idx, value="NW:")
        cell_net_value = worksheet.cell(row=net_weight_row, column=value_col_idx, value=float(grand_total_net))
        cell_net_value.number_format = FORMAT_NUMBER_COMMA_SEPARATED2

        cell_gross_label = worksheet.cell(row=gross_weight_row, column=label_col_idx, value="GW:")
        cell_gross_value = worksheet.cell(row=gross_weight_row, column=value_col_idx, value=float(grand_total_gross))
        cell_gross_value.number_format = FORMAT_NUMBER_COMMA_SEPARATED2

        for cell in [cell_net_label, cell_net_value, cell_gross_label, cell_gross_value]:
            cell.font = font_to_apply
            cell.alignment = align_to_apply
        
            # --- Get Column Indices and Dimensions (no changes here) ---
        col_id_map = header_info.get("column_id_map", {})
        num_columns = header_info.get("num_columns", 1)
        label_col_idx = col_id_map.get(weight_config.get("label_col_id"))
        value_col_idx = col_id_map.get(weight_config.get("value_col_id"))
        
        # ADD THIS LINE
        last_mapped_col_idx = max(col_id_map.values()) if col_id_map else 1
        for cell in [cell_net_label, cell_net_value, cell_gross_label, cell_gross_value]:
            cell.font = font_to_apply
            cell.alignment = align_to_apply

        # --- ADD THIS BLOCK TO APPLY BORDERS ---
        border_to_apply = thin_border
        for row_idx in [net_weight_row, gross_weight_row]:
            for col_idx in range(1, last_mapped_col_idx + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = border_to_apply
        # --- END OF BLOCK ---

        if footer_row_height:
            worksheet.row_dimensions[net_weight_row].height = footer_row_height
            worksheet.row_dimensions[gross_weight_row].height = footer_row_height
        

        print("--- Finished writing grand total weight summary. ---")
        return start_row + 2

    except Exception as e:
        print(f"Error writing grand total weight summary content: {e}")
        return start_row

def write_header(worksheet: Worksheet, start_row: int, header_layout_config: List[Dict[str, Any]],
                 sheet_styling_config: Optional[Dict[str, Any]] = None
                 ) -> Optional[Dict[str, Any]]:
    
    if not header_layout_config or start_row <= 0:
        return None
    
    merge_utils.force_unmerge_from_row_down(worksheet, start_row)


    # Determine header dimensions from the layout config
    num_header_rows = max(cell.get('row', 0) for cell in header_layout_config) + 1
    num_columns = max(cell.get('col', 0) + cell.get('colspan', 1) for cell in header_layout_config)
    end_row = start_row + num_header_rows - 1

    # Get header styling from config
    header_font_to_apply = bold_font
    header_alignment_to_apply = center_alignment
    header_border_to_apply = thin_border
    header_background_fill_to_apply = None # Default is no fill

    # --- NEW: Code to parse header styling from the config ---
    if sheet_styling_config:
        # Get font from config
        header_font_cfg = sheet_styling_config.get("header_font")
        if header_font_cfg and isinstance(header_font_cfg, dict):
            try:
                header_font_to_apply = Font(**header_font_cfg)
            except TypeError:
                pass # Keep default on error

        # Get alignment from config
        header_align_cfg = sheet_styling_config.get("header_alignment")
        if header_align_cfg and isinstance(header_align_cfg, dict):
            try:
                header_alignment_to_apply = Alignment(**header_align_cfg)
            except TypeError:
                pass # Keep default on error
        
        # Get background fill from config
        header_fill_cfg = sheet_styling_config.get("header_pattern_fill")
        if header_fill_cfg and isinstance(header_fill_cfg, dict):
            try:
                # Create a PatternFill object from the config dictionary
                header_background_fill_to_apply = PatternFill(**header_fill_cfg)
            except TypeError:
                # This could happen if config keys don't match PatternFill arguments
                print(f"Warning: Invalid parameters in header_pattern_fill config: {header_fill_cfg}")
                pass # Keep fill as None on error
    # --- END NEW CODE ---

    try:
        # 1. Unmerge the entire target area first
        unmerge_block(worksheet, start_row, end_row, num_columns)

        column_map_by_text = {}
        column_map_by_id = {}

        # 2. Loop through the explicit layout configuration
        for cell_config in header_layout_config:
            # Get cell properties from the config object
            relative_row = cell_config.get('row', 0)
            relative_col = cell_config.get('col', 0)
            text_to_write = cell_config.get('text')
            cell_id = cell_config.get('id')
            rowspan = cell_config.get('rowspan', 1)
            colspan = cell_config.get('colspan', 1)

            # Calculate absolute position on the worksheet
            abs_row = start_row + relative_row
            abs_col = 1 + relative_col # openpyxl is 1-based

            # 3. Write value and apply style to the top-left cell
            cell = worksheet.cell(row=abs_row, column=abs_col)
            cell.value = text_to_write
            cell.font = header_font_to_apply
            cell.alignment = header_alignment_to_apply
            cell.border = header_border_to_apply
            
            # This line now applies the fill object we created from the config
            if header_background_fill_to_apply:
                cell.fill = header_background_fill_to_apply

            # 4. Populate the ID and Text maps
            if cell_id:
                column_map_by_id[cell_id] = abs_col
            if text_to_write:
                column_map_by_text[str(text_to_write).strip()] = abs_col

            # 5. Apply merges if needed
            if rowspan > 1 or colspan > 1:
                end_merge_row = abs_row + rowspan - 1
                end_merge_col = abs_col + colspan - 1
                worksheet.merge_cells(
                    start_row=abs_row,
                    start_column=abs_col,
                    end_row=end_merge_row,
                    end_column=end_merge_col
                )

        return {
            'first_row_index': start_row,
            'second_row_index': end_row, # The last row of the entire header block
            'column_map': column_map_by_text,
            'column_id_map': column_map_by_id,
            'num_columns': num_columns
        }

    except Exception as e:
        print(f"Error in write_header during layout processing: {e}")
        traceback.print_exc()
        return None

    except Exception as e:
        print(f"Error in write_header during layout processing: {e}")
        traceback.print_exc()
        return None
def merge_contiguous_cells_by_id(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    col_id_to_merge: str,
    column_id_map: Dict[str, int]
):
    """
    Finds and merges contiguous vertical cells within a column that have the same value.
    This is called AFTER all data has been written to the sheet.
    """
    col_idx = column_id_map.get(col_id_to_merge)
    if not col_idx or start_row >= end_row:
        return

    current_merge_start_row = start_row
    value_to_match = worksheet.cell(row=start_row, column=col_idx).value

    for row_idx in range(start_row + 1, end_row + 2):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value if row_idx <= end_row else object()
        if cell_value != value_to_match:
            if row_idx - 1 > current_merge_start_row:
                if value_to_match is not None and str(value_to_match).strip():
                    try:
                        worksheet.merge_cells(
                            start_row=current_merge_start_row,
                            start_column=col_idx,
                            end_row=row_idx - 1,
                            end_column=col_idx
                        )
                    except Exception as e:
                        print(f"Could not merge cells for ID {col_id_to_merge} from row {current_merge_start_row} to {row_idx - 1}. Error: {e}")
            
            current_merge_start_row = row_idx
            if row_idx <= end_row:
                value_to_match = cell_value


def find_footer(worksheet: Worksheet, footer_rules: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Finds the footer row based on marker text and rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        footer_rules: Dictionary defining how to find the footer (marker text, columns, direction, etc.).

    Returns:
        A dictionary containing the footer start row {'start_row': index} or None if not found.
    """
    marker_text = footer_rules.get('marker_text'); search_type = footer_rules.get('search_type', 'exact'); case_sensitive = footer_rules.get('case_sensitive', True)
    search_columns = footer_rules.get('search_columns', [1]); search_direction = footer_rules.get('search_direction', 'down').lower()
    min_row_offset = footer_rules.get('min_row_offset', 1); max_row_search = footer_rules.get('max_row_to_search', worksheet.max_row)
    max_row_search = min(max_row_search, worksheet.max_row) # Ensure max_row_search doesn't exceed actual max row
    if not marker_text: return None
    if not isinstance(search_columns, list) or not search_columns: search_columns = [1]
    if min_row_offset <= 0: min_row_offset = 1

    try:
        # Determine Row Iteration Order
        row_iterator = None
        if search_direction == 'up': row_iterator = range(max_row_search, min_row_offset - 1, -1)
        else: row_iterator = range(min_row_offset, max_row_search + 1)
        marker_text_str = str(marker_text)

        # Search for Marker
        for r_idx in row_iterator:
            for c_idx in search_columns:
                if not (1 <= c_idx <= worksheet.max_column): continue # Skip invalid column index
                try:
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    # If it's a merged cell, only check the top-left origin cell of the merge range
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        is_origin = False
                        for merged_range in worksheet.merged_cells.ranges:
                            if merged_range.min_row == r_idx and merged_range.min_col == c_idx:
                                is_origin = True; break
                        if not is_origin: continue # Skip if not the top-left cell
                    cell_value_str = str(cell.value) if cell.value is not None else ""
                except IndexError: continue # Should not happen with max_column check, but safety first
                found = False
                if search_type == 'substring':
                    pattern = re.escape(marker_text_str); flags = 0 if case_sensitive else re.IGNORECASE
                    if re.search(pattern, cell_value_str, flags): found = True
                elif case_sensitive and cell_value_str == marker_text_str: found = True
                elif not case_sensitive and cell_value_str.lower() == marker_text_str.lower(): found = True
                if found: return {'start_row': r_idx} # Return immediately when found

        return None # Marker not found
    except Exception as e: return None # Error during search


# invoice_utils.py

# ... (imports, constants, other functions) ...

def write_configured_rows(
    worksheet: Worksheet,
    start_row_index: int,
    num_columns: int,
    rows_config_list: List[Dict[str, Any]], # Primary configuration for each row
    calculated_totals: Dict[str, Any],     # Data values to be inserted
    default_style_config: Optional[Dict[str, Any]] = None # Default styles from sheet config
):
    """
    Writes one or more rows with specified content (labels + dynamic values),
    styling, and merges based on configuration.
    Assumes the rows have already been inserted by the caller.
    Number formats are applied ONLY if specified in the cell's config,
    or set to Text ('@') for labels/non-numeric values.
    """
    if not rows_config_list or start_row_index <= 0:
        return

    print(f"--- Writing {len(rows_config_list)} configured rows starting at row {start_row_index} ---")
    calculated_totals = calculated_totals or {} # Ensure it's a dict

    # --- Get overall default styles from the sheet's styling configuration ---
    # These will be used if a row doesn't specify its own font/alignment.
    overall_default_font = Font() # Basic Openpyxl default
    overall_default_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False) # Basic Openpyxl default

    if default_style_config:
        # Use 'default_font' and 'default_alignment' from the sheet's styling config if available
        sheet_default_font_cfg = default_style_config.get("default_font")
        if sheet_default_font_cfg and isinstance(sheet_default_font_cfg, dict):
            try:
                overall_default_font = Font(**{k: v for k, v in sheet_default_font_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_font config. Using basic default font.")
        
        sheet_default_align_cfg = default_style_config.get("default_alignment")
        if sheet_default_align_cfg and isinstance(sheet_default_align_cfg, dict):
            try:
                overall_default_alignment = Alignment(**{k: v for k, v in sheet_default_align_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_alignment config. Using basic default alignment.")

    # Iterate through each row's configuration object
    for i, row_config_item in enumerate(rows_config_list):
        current_row_idx = start_row_index + i
        print(f"  Processing configured row {i+1} (Sheet Row: {current_row_idx})")

        # --- Get ROW-LEVEL configurations from the current row_config_item ---
        row_cell_definitions = row_config_item.get("content", []) # List of cell configs for this row
        
        row_specific_height = row_config_item.get("height")
        row_specific_font_config = row_config_item.get("font")      # Font for the whole row
        row_specific_align_config = row_config_item.get("alignment") # Alignment for the whole row
        row_specific_merge_rules = row_config_item.get("merge_rules") # Merges for this specific row
        row_specific_apply_border = row_config_item.get("apply_default_border", True) # Border for the whole row

        # --- Determine effective font and alignment FOR THIS ENTIRE ROW ---
        # Start with the overall defaults, then apply row-level overrides if they exist.
        effective_row_font = overall_default_font
        if row_specific_font_config and isinstance(row_specific_font_config, dict):
            font_params = {k: v for k, v in row_specific_font_config.items() if v is not None}
            if font_params:
                try:
                    effective_row_font = Font(**font_params)
                except TypeError:
                    print(f"Warning: Invalid font config for row {current_row_idx}. Using sheet/basic default.")

        effective_row_alignment = overall_default_alignment
        if row_specific_align_config and isinstance(row_specific_align_config, dict):
            align_params = {k: v for k, v in row_specific_align_config.items() if v is not None}
            if align_params:
                try:
                    effective_row_alignment = Alignment(**align_params)
                except TypeError:
                    print(f"Warning: Invalid alignment config for row {current_row_idx}. Using sheet/basic default.")

        # --- Write Content Items (Cells) for the current row and Apply Styles ---
        written_columns_in_row = set() # Keep track of columns explicitly written to in this row
        
        if isinstance(row_cell_definitions, list):
            for cell_config_item in row_cell_definitions: # Each item in 'content' array from your JSON
                if not isinstance(cell_config_item, dict):
                    print(f"Warning: Invalid cell config item in row {current_row_idx}: {cell_config_item}")
                    continue

                try:
                    target_col_idx = int(cell_config_item.get("col"))
                    if not (1 <= target_col_idx <= num_columns):
                        print(f"Warning: Column index {target_col_idx} out of range for row {current_row_idx}.")
                        continue

                    cell = worksheet.cell(row=current_row_idx, column=target_col_idx)
                    written_columns_in_row.add(target_col_idx)
                    
                    value_to_write = None
                    # Cell-specific number format, font, and alignment from its own config
                    cell_specific_number_format = cell_config_item.get("number_format")
                    # Note: Cell-specific font/alignment could also be added to JSON if needed,
                    # otherwise, the effective_row_font/alignment will be used.

                    if "label" in cell_config_item:
                        value_to_write = cell_config_item["label"]
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT # Use provided or default to Text
                    elif "value_key" in cell_config_item:
                        value_key = cell_config_item["value_key"]
                        raw_value = calculated_totals.get(value_key)
                        suffix = cell_config_item.get("suffix", "")
                        
                        numeric_value = None
                        if isinstance(raw_value, (int, float)):
                            numeric_value = float(raw_value)
                        elif isinstance(raw_value, str):
                            try:
                                # Attempt to convert if it looks like a number, handling commas
                                cleaned_raw_value = raw_value.replace(',', '')
                                if cleaned_raw_value.strip(): # Avoid empty strings
                                    numeric_value = float(cleaned_raw_value)
                            except (ValueError, TypeError):
                                pass # Keep as None if conversion fails

                        if numeric_value is not None:
                            # If there's a suffix, the value becomes a string.
                            # If no suffix, keep it as a number for Excel to handle.
                            value_to_write = f"{numeric_value}{suffix}" if suffix else numeric_value
                            
                            if cell_specific_number_format:
                                cell.number_format = cell_specific_number_format
                            elif suffix: # If suffix is present, it's text
                                cell.number_format = FORMAT_TEXT
                            # Else (numeric, no suffix, no specific format): Let Excel use default number format
                            
                        else: # Value is not numeric or not found, treat as text
                            value_to_write = f"{str(raw_value or '')}{suffix}" # Use empty string if raw_value is None
                            cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    
                    elif "value" in cell_config_item: # Direct static value
                        value_to_write = cell_config_item.get("value")
                        # Assume direct static values are text unless a number_format is given
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    else:
                        # No label, value_key, or value; cell might be intended to be blank but styled
                        pass


                    cell.value = value_to_write
                    cell.font = effective_row_font # Apply the determined row font
                    cell.alignment = effective_row_alignment # Apply the determined row alignment

                    # Apply border based on row-level setting
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        cell.border = no_border

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid data in cell config for row {current_row_idx}: {cell_config_item}. Error: {e}")
                except Exception as cell_err:
                    print(f"Warning: Error writing cell (Row: {current_row_idx}, Col: {cell_config_item.get('col', 'N/A')}): {cell_err}")

        # --- Ensure remaining (unwritten) cells in the row get default row styling (border) ---
        for c_idx_fill in range(1, num_columns + 1):
            if c_idx_fill not in written_columns_in_row: # Only touch columns not explicitly defined
                try:
                    cell = worksheet.cell(row=current_row_idx, column=c_idx_fill)
                    # Apply row's effective font and alignment to blank cells if desired (optional)
                    # cell.font = effective_row_font
                    # cell.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        # Only remove border if cell is truly blank and no border is intended for the row
                        if cell.value is None: # Check if cell is actually empty
                            cell.border = no_border
                except Exception as blank_cell_err:
                    print(f"Warning: Error styling blank cell ({current_row_idx},{c_idx_fill}): {blank_cell_err}")


        # --- Apply Merges for this entire row (using row-level merge rules) ---
        if row_specific_merge_rules and isinstance(row_specific_merge_rules, dict):
            apply_row_merges(worksheet, current_row_idx, num_columns, row_specific_merge_rules)
            # Re-apply style/border to the top-left cell of any merged ranges
            # to ensure consistent appearance, as merging can sometimes affect the primary cell's style.
            for start_col_str_merge in row_specific_merge_rules.keys():
                try:
                    start_col_idx_merge = int(start_col_str_merge)
                    merged_cell_anchor = worksheet.cell(row=current_row_idx, column=start_col_idx_merge)
                    merged_cell_anchor.font = effective_row_font
                    merged_cell_anchor.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        merged_cell_anchor.border = thin_border
                    else:
                        merged_cell_anchor.border = no_border
                except (ValueError, TypeError):
                    print(f"Warning: Invalid start column for merge rule on row {current_row_idx}: {start_col_str_merge}")
                except Exception as merge_style_err:
                    print(f"Warning: Error re-styling merged cell anchor at ({current_row_idx},{start_col_str_merge}): {merge_style_err}")

        # --- Apply Height for this entire row (using row-level height) ---
        if row_specific_height is not None:
            try:
                h_val = float(row_specific_height)
                if h_val > 0:
                    worksheet.row_dimensions[current_row_idx].height = h_val
            except (ValueError, TypeError):
                print(f"Warning: Invalid height value '{row_specific_height}' for row {current_row_idx}.")
            except Exception as height_err:
                print(f"Warning: Error setting height for row {current_row_idx}: {height_err}")

    print(f"--- Finished writing configured rows ---")

def apply_explicit_data_cell_merges_by_id(
    worksheet: Worksheet,
    row_num: int,
    column_id_map: Dict[str, int],  # Maps column ID to its 1-based column index
    num_total_columns: int,
    merge_rules_data_cells: Dict[str, Dict[str, Any]], # e.g., {'col_item': {'rowspan': 2}}
    sheet_styling_config: Optional[Dict[str, Any]] = None,
    fob_mode: Optional[bool] = False
):
    """
    Applies horizontal merges to data cells in a specific row based on column IDs.
    """
    if not merge_rules_data_cells or row_num <= 0:
        return

    thin_side = Side(border_style="thin", color="000000")
    full_thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Loop through rules where the key is now the column ID
    for col_id, rule_details in merge_rules_data_cells.items():
        colspan_to_apply = rule_details.get("rowspan")

        if not isinstance(colspan_to_apply, int) or colspan_to_apply <= 1:
            continue
        
        # Get column index from the ID map
        start_col_idx = column_id_map.get(col_id)
        if not start_col_idx:
            print(f"Warning: Could not find column for merge rule with ID '{col_id}'.")
            continue
            
        end_col_idx = start_col_idx + colspan_to_apply - 1
        end_col_idx = min(end_col_idx, num_total_columns)

        if start_col_idx >= end_col_idx:
            continue

        try:
            # Unmerge any existing ranges in the target area
            for mc_range in list(worksheet.merged_cells.ranges):
                if mc_range.min_row == row_num and mc_range.max_row == row_num:
                    if mc_range.min_col <= end_col_idx and mc_range.max_col >= start_col_idx:
                        worksheet.unmerge_cells(str(mc_range))
            
            # Apply the new merge
            worksheet.merge_cells(start_row=row_num, start_column=start_col_idx,
                                  end_row=row_num, end_column=end_col_idx)
            
            # Style the anchor cell of the new merged range
            anchor_cell = worksheet.cell(row=row_num, column=start_col_idx)
            
            # Apply base styling for the column ID
            _apply_cell_style(anchor_cell, col_id, sheet_styling_config, fob_mode)
            
            # Ensure the merged cell has the desired border and alignment
            anchor_cell.border = full_thin_border
            anchor_cell.alignment = center_alignment

        except Exception as e:
            print(f"Error applying explicit data cell merge for ID '{col_id}' on row {row_num}: {e}")

def _to_numeric(value: Any) -> Union[int, float, None, Any]:
    """
    Safely attempts to convert a value to a float or int.
    Handles strings with commas and returns the original value on failure.
    """
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        try:
            cleaned_val = value.replace(',', '').strip()
            if not cleaned_val:
                return None
            return float(cleaned_val) if '.' in cleaned_val else int(cleaned_val)
        except (ValueError, TypeError):
            return value # Return original string if conversion fails
    if isinstance(value, Decimal):
        return float(value)
    return value # Return original value for other types

def _apply_fallback(
    row_dict: Dict[int, Any],
    target_col_idx: int,
    mapping_rule: Dict[str, Any],
    fob_mode: bool
):
    """
    Applies a fallback value to the row_dict based on the fob_mode.
    """
    if fob_mode:
        fallback_key = "fallback_on_fob"
    else:
        fallback_key = "fallback_on_none"
        
    row_dict[target_col_idx] = mapping_rule.get(fallback_key) or mapping_rule.get("fallback_on_none")


def prepare_data_rows(
    data_source_type: str,
    data_source: Union[Dict, List],
    dynamic_mapping_rules: Dict[str, Any],
    column_id_map: Dict[str, int],
    idx_to_header_map: Dict[int, str],
    desc_col_idx: int,
    num_static_labels: int,
    static_value_map: Dict[int, Any],
    fob_mode: bool,
) -> Tuple[List[Dict[int, Any]], List[int], bool, int]:
    """
    Corrected version with typo fix and improved fallback flexibility.
    """
    data_rows_prepared = []
    pallet_counts_for_rows = []
    num_data_rows_from_source = 0
    dynamic_desc_used = False
    
    NUMERIC_IDS = {"col_qty_pcs", "col_qty_sf", "col_unit_price", "col_amount", "col_net", "col_gross", "col_cbm"}

    # --- Handler for FOB Aggregation (Uses new fallback logic) ---
    if data_source_type == 'fob_aggregation':
        fob_data = data_source or {}
        num_data_rows_from_source = len(fob_data)
        id_to_data_key_map = {"col_po": "combined_po", "col_item": "combined_item", "col_desc": "combined_description", "col_qty_sf": "total_sqft", "col_amount": "total_amount"}
        price_col_idx = column_id_map.get("col_unit_price")
        
        for row_key in sorted(fob_data.keys()):
            row_value_dict = fob_data.get(row_key, {})
            row_dict = {}
            for col_id, data_key in id_to_data_key_map.items():
                target_col_idx = column_id_map.get(col_id)
                if not target_col_idx: continue

                data_value = row_value_dict.get(data_key)
                is_empty = data_value is None or (isinstance(data_value, str) and not data_value.strip())

                if not is_empty:
                    row_dict[target_col_idx] = _to_numeric(data_value)
                    if col_id == "col_desc":
                        dynamic_desc_used = True
                else:
                    mapping_rule_for_id = {}
                    for rule in dynamic_mapping_rules.values():
                        if rule.get("id") == col_id:
                            mapping_rule_for_id = rule
                            break
                    _apply_fallback(row_dict, target_col_idx, mapping_rule_for_id, fob_mode)

            if price_col_idx:
                row_dict[price_col_idx] = {"type": "formula", "template": "{col_ref_1}{row}/{col_ref_0}{row}", "inputs": ["col_qty_sf", "col_amount"]}
            data_rows_prepared.append(row_dict)

    # --- Handler for Custom Aggregation (FOB Check and Fallback Added) ---
    elif data_source_type == 'custom_aggregation':
        custom_data = data_source or {}
        num_data_rows_from_source = len(custom_data)
        price_col_idx = column_id_map.get("col_unit_price")
        desc_col_idx_local = column_id_map.get("col_desc")

        for key_tuple, value_dict in custom_data.items():
            if not isinstance(key_tuple, tuple) or len(key_tuple) < 4: continue
            
            row_dict = {}
            # Directly map known values first
            row_dict[column_id_map.get("col_po")] = key_tuple[0]
            row_dict[column_id_map.get("col_item")] = key_tuple[1]
            row_dict[column_id_map.get("col_qty_sf")] = _to_numeric(value_dict.get("sqft_sum"))
            row_dict[column_id_map.get("col_amount")] = _to_numeric(value_dict.get("amount_sum"))

            if desc_col_idx_local:
                desc_value = key_tuple[3]
                if desc_value:
                    row_dict[desc_col_idx_local] = desc_value
                    dynamic_desc_used = True
            
            # Apply fallbacks for any unmapped columns based on fob_mode
            for header, mapping_rule in dynamic_mapping_rules.items():
                target_id = mapping_rule.get("id")
                target_col_idx = column_id_map.get(target_id)
                if not target_col_idx or target_col_idx in row_dict:
                    continue

                # If the column is not already populated, apply the appropriate fallback.
                # The _apply_fallback function should handle the logic for
                # 'fallback_on_non' vs 'fallback_fob' based on the fob_mode flag.
                _apply_fallback(row_dict, target_col_idx, mapping_rule, fob_mode)

            if price_col_idx:
                row_dict[price_col_idx] = {"type": "formula", "template": "{col_ref_1}{row}/{col_ref_0}{row}", "inputs": ["col_qty_sf", "col_amount"]}
            
            data_rows_prepared.append({k: v for k, v in row_dict.items() if k is not None})

    # --- Unified Handler for Aggregation & Processed Tables (TYPO FIXED) ---
    else:
        normalized_data = []
        if data_source_type == 'aggregation':
            aggregation_data = data_source or {}
            num_data_rows_from_source = len(aggregation_data)
            for key_tuple, value_dict in aggregation_data.items():
                normalized_data.append({'key_tuple': key_tuple, 'value_dict': value_dict})

        elif data_source_type == 'processed_tables':
            table_data = data_source or {}
            if isinstance(table_data, dict):
                max_len = max((len(v) for v in table_data.values() if isinstance(v, list)), default=0)
                num_data_rows_from_source = max_len
                raw_pallet_counts = table_data.get("pallet_count", [])
                pallet_counts_for_rows = raw_pallet_counts[:max_len] + [0] * (max_len - len(raw_pallet_counts)) if isinstance(raw_pallet_counts, list) else [0] * max_len
                for i in range(max_len):
                    normalized_data.append({'table_row_index': i, 'table_data': table_data})
        
        for item in normalized_data:
            row_dict = {}
            for header, mapping_rule in dynamic_mapping_rules.items():
                target_id = mapping_rule.get("id")
                target_col_idx = column_id_map.get(target_id)
                if not target_col_idx: continue

                data_value = None
                if 'key_tuple' in item:
                    key_tuple, value_dict = item['key_tuple'], item['value_dict']
                    if 'key_index' in mapping_rule and mapping_rule['key_index'] < len(key_tuple):
                        data_value = key_tuple[mapping_rule['key_index']]
                    elif 'value_key' in mapping_rule:
                        data_value = value_dict.get(mapping_rule['value_key'])
                elif 'table_row_index' in item:
                    i, table_data = item['table_row_index'], item['table_data']
                    source_list = table_data.get(header, [])
                    if i < len(source_list):
                        data_value = source_list[i]

                is_empty = data_value is None or (isinstance(data_value, str) and not data_value.strip())
                
                if not is_empty:
                    if target_id in NUMERIC_IDS: data_value = _to_numeric(data_value)
                    row_dict[target_col_idx] = data_value
                    if target_id == 'col_desc':
                        dynamic_desc_used = True
                else:
                    _apply_fallback(row_dict, target_col_idx, mapping_rule, fob_mode)
            
            if data_source_type == 'aggregation':
                amount_col_idx = column_id_map.get("col_amount")
                if amount_col_idx:
                    row_dict[amount_col_idx] = {"type": "formula", "template": "{col_ref_1}{row}*{col_ref_0}{row}", "inputs": ["col_qty_sf", "col_unit_price"]}
            
            data_rows_prepared.append(row_dict)

    # --- Final Processing Steps (Unchanged) ---
    if static_value_map:
        for row_data in data_rows_prepared:
            for col_idx, static_val in static_value_map.items():
                if col_idx not in row_data:
                    row_data[col_idx] = static_val
    
    if num_static_labels > len(data_rows_prepared):
        data_rows_prepared.extend([{}] * (num_static_labels - len(data_rows_prepared)))
    
    return data_rows_prepared, pallet_counts_for_rows, dynamic_desc_used, num_data_rows_from_source

# Helper functions _to_numeric and _apply_fallback would be defined elsewhere in your code.
# For example:
def _to_numeric(value: Any) -> Any:
    """Converts a value to a numeric type if possible."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        try:
            return int(value)
        except (ValueError, TypeError):
            try:
                return float(value)
            except (ValueError, TypeError):
                pass
    return value



def parse_mapping_rules(
    mapping_rules: Dict[str, Any],
    column_id_map: Dict[str, int],
    idx_to_header_map: Dict[int, str]
) -> Dict[str, Any]:
    """
    Parses the mapping rules from a standardized, ID-based configuration.

    This function is refined to handle different mapping structures, such as a
    flat structure for aggregation sheets and a nested 'data_map' for table-based sheets.

    Args:
        mapping_rules: The raw mapping rules dictionary from the sheet's configuration.
        column_id_map: A dictionary mapping column IDs to their 1-based column index.
        idx_to_header_map: A dictionary mapping a column index back to its header text.

    Returns:
        A dictionary containing all the parsed information required for data filling.
    """
    # --- Initialize all return values ---
    parsed_result = {
        "static_value_map": {},
        "initial_static_col1_values": [],
        "dynamic_mapping_rules": {},
        "formula_rules": {},
        "col1_index": -1,
        "num_static_labels": 0,
        "static_column_header_name": None,
        "apply_special_border_rule": False
    }

    # --- Process all rules in a single, intelligent pass ---
    for rule_key, rule_value in mapping_rules.items():
        if not isinstance(rule_value, dict):
            continue # Skip non-dictionary rules

        # --- Handler for nested 'data_map' (used by 'processed_tables_multi') ---
        if rule_key == "data_map":
            # The entire dictionary under "data_map" is our set of dynamic rules.
            parsed_result["dynamic_mapping_rules"].update(rule_value)
            continue

        rule_type = rule_value.get("type")

        # --- Handler for Initial Static Rows ---
        if rule_type == "initial_static_rows":
            static_column_id = rule_value.get("column_header_id")
            target_col_idx = column_id_map.get(static_column_id)

            if target_col_idx:
                parsed_result["static_column_header_name"] = idx_to_header_map.get(target_col_idx)
                parsed_result["col1_index"] = target_col_idx
                parsed_result["initial_static_col1_values"] = rule_value.get("values", [])
                parsed_result["num_static_labels"] = len(parsed_result["initial_static_col1_values"])
                
                header_text = parsed_result["static_column_header_name"]
                parsed_result["apply_special_border_rule"] = header_text and header_text.strip() in ["Mark & Nº", "Mark & N °"]
            else:
                print(f"Warning: Initial static rows column with ID '{static_column_id}' not found.")
            continue

        # For all other rules, get the target column index using the RELIABLE ID
        target_id = rule_value.get("id")
        target_col_idx = column_id_map.get(target_id)

        # --- Handler for Formulas ---
        if rule_type == "formula":
            if target_col_idx:
                parsed_result["formula_rules"][target_col_idx] = {
                    "template": rule_value.get("formula_template"),
                    "input_ids": rule_value.get("inputs", [])
                }
            else:
                print(f"Warning: Could not find target column for formula rule with id '{target_id}'.")

        # --- Handler for Static Values ---
        elif "static_value" in rule_value:
            if target_col_idx:
                parsed_result["static_value_map"][target_col_idx] = rule_value["static_value"]
            else:
                print(f"Warning: Could not find target column for static_value rule with id '{target_id}'.")
        
        # --- Handler for top-level Dynamic Rules (used by 'aggregation') ---
        else:
            # If it's not a special rule, it's a dynamic mapping rule for the aggregation data type.
            parsed_result["dynamic_mapping_rules"][rule_key] = rule_value
            
    return parsed_result


def write_summary_rows(
    worksheet: Worksheet,
    start_row: int,
    header_info: Dict[str, Any],
    all_tables_data: Dict[str, Any],
    table_keys: List[str],
    footer_config: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    styling_config: Optional[Dict[str, Any]] = None,
    fob_mode: Optional[bool] = False
) -> int:
    """
    Calculates and writes ID-driven summary rows, ensuring text cells are
    formatted as text and the final bold font style is applied correctly.
    """
    buffalo_summary_row = start_row
    leather_summary_row = start_row + 1
    next_available_row = start_row + 2

    try:
        # --- Get Styles from Footer Config ---
        style_config = footer_config.get('style', {})
        font_to_apply = Font(**style_config.get('font', {'bold': True}))
        align_to_apply = Alignment(**style_config.get('alignment', {'horizontal': 'center', 'vertical': 'center'}))

        # --- Calculation and Writing Logic (remains the same) ---
        # ... (all the calculation and cell writing logic is unchanged) ...
        column_id_map = header_info.get('column_id_map', {})
        idx_to_id_map = {v: k for k, v in column_id_map.items()}
        data_map = mapping_rules.get('data_map', {})
        numeric_ids_to_sum = ["col_qty_pcs", "col_qty_sf", "col_net", "col_gross", "col_cbm"]
        id_to_data_key_map = {v['id']: k for k, v in data_map.items() if v.get('id') in numeric_ids_to_sum}
        ids_to_sum = list(id_to_data_key_map.keys())
        buffalo_totals = {col_id: 0 for col_id in ids_to_sum}
        cow_totals = {col_id: 0 for col_id in ids_to_sum}
        buffalo_pallet_total = 0
        cow_pallet_total = 0
        for table_key in table_keys:
            table_data = all_tables_data.get(str(table_key), {})
            descriptions = table_data.get("description", [])
            pallet_counts = table_data.get("pallet_count", [])
            for i in range(len(descriptions)):
                raw_val = descriptions[i]
                desc_val = raw_val
                if isinstance(raw_val, list) and raw_val:
                    desc_val = raw_val[0]
                is_buffalo = desc_val and "BUFFALO" in str(desc_val).upper()
                target_dict = buffalo_totals if is_buffalo else cow_totals
                try:
                    pallet_val = int(pallet_counts[i]) if i < len(pallet_counts) else 0
                    if is_buffalo:
                        buffalo_pallet_total += pallet_val
                    else:
                        cow_pallet_total += pallet_val
                except (ValueError, TypeError): pass
                for col_id in ids_to_sum:
                    data_key = id_to_data_key_map.get(col_id)
                    if not data_key: continue
                    data_list = table_data.get(data_key, [])
                    if i < len(data_list):
                        try:
                            value_to_add = data_list[i]
                            if isinstance(value_to_add, (int, float)):
                                target_dict[col_id] += float(value_to_add)
                            elif isinstance(value_to_add, str) and value_to_add.strip():
                                target_dict[col_id] += float(value_to_add.replace(',', ''))
                        except (ValueError, TypeError, IndexError): pass
        num_columns = header_info['num_columns']
        desc_col_idx = column_id_map.get("col_desc")
        label_col_idx = column_id_map.get("col_pallet") or 2
        unmerge_row(worksheet, buffalo_summary_row, num_columns)
        worksheet.cell(row=buffalo_summary_row, column=label_col_idx, value="TOTAL OF:").number_format = FORMAT_TEXT
        worksheet.cell(row=buffalo_summary_row, column=label_col_idx + 1, value="BUFFALO LEATHER").number_format = FORMAT_TEXT
        if desc_col_idx:
            worksheet.cell(row=buffalo_summary_row, column=desc_col_idx, value=f"{buffalo_pallet_total} PALLETS").number_format = FORMAT_TEXT
        for col_id, total_value in buffalo_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                worksheet.cell(row=buffalo_summary_row, column=col_idx, value=total_value)
        unmerge_row(worksheet, leather_summary_row, num_columns)
        worksheet.cell(row=leather_summary_row, column=label_col_idx, value="TOTAL OF:").number_format = FORMAT_TEXT
        worksheet.cell(row=leather_summary_row, column=label_col_idx + 1, value="COW LEATHER").number_format = FORMAT_TEXT
        if desc_col_idx:
            worksheet.cell(row=leather_summary_row, column=desc_col_idx, value=f"{cow_pallet_total} PALLETS").number_format = FORMAT_TEXT
        for col_id, total_value in cow_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                worksheet.cell(row=leather_summary_row, column=col_idx, value=total_value)


        # --- Apply Styles to Both Rows with Correct Order ---
        for row_num in [buffalo_summary_row, leather_summary_row]:
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=row_num, column=c_idx)
                current_col_id = idx_to_id_map.get(c_idx)

                # Step 1: Apply column-specific styles first (like number formats).
                _apply_cell_style(cell, current_col_id, styling_config, fob_mode)

                # Step 2: Enforce the general footer style as the final rule.
                cell.font = font_to_apply
                cell.alignment = align_to_apply
                cell.border = no_border
        
        # --- Apply Row Height (remains the same) ---
        footer_height = None
        if styling_config:
            row_heights_cfg = styling_config.get("row_heights", {})
            footer_height = row_heights_cfg.get("footer", row_heights_cfg.get("header"))
        if footer_height is not None:
            try:
                h_val = float(footer_height)
                worksheet.row_dimensions[buffalo_summary_row].height = h_val
                worksheet.row_dimensions[leather_summary_row].height = h_val
            except (ValueError, TypeError): pass

        return next_available_row

    except Exception as summary_err:
        print(f"Warning: Failed processing summary rows: {summary_err}")
        traceback.print_exc()
        return start_row + 2

def write_footer_row(
    worksheet: Worksheet,
    footer_row_num: int,
    header_info: Dict[str, Any],
    sum_ranges: List[Tuple[int, int]],
    footer_config: Dict[str, Any],
    pallet_count: int,
    override_total_text: Optional[str] = None,
    fob_mode: bool = False,
    grand_total_flag: bool = False
) -> int:
    """
    Writes a fully configured footer row, including styling, borders, merges,
    summed totals with number formatting, and a pallet count summary.

    This function is driven entirely by a footer configuration object and can sum
    over multiple, non-contiguous data ranges.

    Args:
        worksheet: The openpyxl worksheet to modify.
        footer_row_num: The 1-based row index for the footer.
        header_info: The header dictionary containing 'column_id_map' and 'num_columns'.
        sum_ranges: A list of tuples, where each tuple is a (start_row, end_row) of data to sum.
        footer_config: The footer configuration object from the JSON.
        pallet_count: The total number of pallets to display in the footer.
        override_total_text: Optional text to use instead of the one in the config.

    Returns:
        The row index (footer_row_num) on success, or -1 on failure.
    """
    if not footer_config or footer_row_num <= 0:
        return -1

    try:
        # --- 1. Parse Configs and Prepare Style Objects ---
        num_columns = header_info.get('num_columns', 1)
        column_map_by_id = header_info.get('column_id_map', {})

        # Get style configurations with sensible defaults
        style_config = footer_config.get('style', {})
        font_config = style_config.get('font', {'bold': True})
        align_config = style_config.get('alignment', {'horizontal': 'center', 'vertical': 'center'})
        border_config = style_config.get('border', {'apply': True})
        
        # Get number format configuration
        number_format_config = footer_config.get("number_formats", {})

        # Create openpyxl style objects
        font_to_apply = Font(**font_config)
        align_to_apply = Alignment(**align_config)
        border_to_apply = None
        if border_config.get('apply'):
            side = Side(border_style=border_config.get('style', 'thin'), color=border_config.get('color', '000000'))
            border_to_apply = Border(left=side, right=side, top=side, bottom=side)

        unmerge_row(worksheet, footer_row_num, num_columns)

        # --- 2. Write Content (Labels, Formulas, and Pallet Count) ---
        total_text = override_total_text if override_total_text is not None else footer_config.get("total_text", "TOTAL:")
        total_text_col_id = footer_config.get("total_text_column_id")
        if total_text_col_id and column_map_by_id.get(total_text_col_id):
            cell = worksheet.cell(row=footer_row_num, column=column_map_by_id[total_text_col_id], value=total_text)
            cell.font = font_to_apply
            cell.alignment = align_to_apply

        # Write Pallet Count Text
        pallet_col_id = footer_config.get("pallet_count_column_id")
        if pallet_col_id and pallet_count > 0:
            pallet_col_idx = column_map_by_id.get(pallet_col_id)
            if pallet_col_idx:
                pallet_text = f"{pallet_count} PALLET{'S' if pallet_count != 1 else ''}"
                cell = worksheet.cell(row=footer_row_num, column=pallet_col_idx, value=pallet_text)
                cell.font = font_to_apply
                cell.alignment = align_to_apply

        sum_column_ids = footer_config.get("sum_column_ids", [])
        if sum_ranges:
            for col_id in sum_column_ids:
                col_idx = column_map_by_id.get(col_id)
                if col_idx:
                    col_letter = get_column_letter(col_idx)
                    sum_parts = [f"{col_letter}{start}:{col_letter}{end}" for start, end in sum_ranges]
                    formula = f"=SUM({','.join(sum_parts)})"
                    cell = worksheet.cell(row=footer_row_num, column=col_idx, value=formula)
                    cell.font = font_to_apply
                    cell.alignment = align_to_apply
                    
                    # Apply Number Formatting from Config if fob
                    number_format_str = number_format_config.get(col_id)
                    if number_format_str and fob_mode and col_id not in ['col_pcs', 'col_qty_pcs']:
                        cell.number_format = "##,00.00"
                    elif number_format_str:
                        cell.number_format = number_format_str["number_format"]
        # --- 3. Apply Border and Final Styling to the Whole Row ---
        if grand_total_flag != True:
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=footer_row_num, column=c_idx)
                if cell.font != font_to_apply: cell.font = font_to_apply
                if cell.alignment != align_to_apply: cell.alignment = align_to_apply
                if border_to_apply:
                    cell.border = border_to_apply

        # --- 4. Apply Merges ---
        merge_rules = footer_config.get("merge_rules", [])
        for rule in merge_rules:
            start_col_id = rule.get("start_column_id")
            colspan = rule.get("colspan")
            start_col = column_map_by_id.get(start_col_id)
            if start_col and colspan:
                end_col = min(start_col + colspan - 1, num_columns)
                worksheet.merge_cells(start_row=footer_row_num, start_column=start_col, end_row=footer_row_num, end_column=end_col)
                worksheet.cell(row=footer_row_num, column=start_col).alignment = align_to_apply

        return footer_row_num

    except Exception as e:
        print(f"ERROR: An error occurred during footer generation on row {footer_row_num}: {e}")
        return -1

    except Exception as e:
        print(f"ERROR: An error occurred during footer generation on row {footer_row_num}: {e}")
        # On failure, return -1
        return -1


def _style_row_before_footer(
    worksheet: Worksheet,
    row_num: int,
    num_columns: int,
    sheet_styling_config: Optional[Dict[str, Any]],
    idx_to_id_map: Dict[int, str],
    col1_index: int, # The index of the first column to receive special border handling
    fob_mode: bool
):
    """
    Applies column-specific styles, a full border, and a specific height
    to the static row before the footer. The first column will only have
    side borders.
    """
    if not sheet_styling_config or row_num <= 0:
        return

    # Set the row height using the 'header' value from the styling config.
    try:
        row_heights = sheet_styling_config.get("row_heights", {})
        header_height = row_heights.get("header")

        if header_height:
            worksheet.row_dimensions[row_num].height = header_height
    except Exception as e:
        print(f"Warning: Could not set row height for row {row_num}. Error: {e}")

    # --- START: Refactored Logic ---
    # Define the two border styles needed for this row
    thin_side = Side(border_style="thin", color="000000")
    
    # Style 1: Full border for all columns except the first
    full_thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Style 2: Side-only border for the first column
    side_only_border = Border(left=thin_side, right=thin_side)
    # --- END: Refactored Logic ---

    # Iterate through each column of the row to apply cell-level styles
    for c_idx in range(1, num_columns + 1):
        try:
            cell = worksheet.cell(row=row_num, column=c_idx)
            current_col_id = idx_to_id_map.get(c_idx)

            # 1. Apply font, alignment, and number formats based on the column ID.
            _apply_cell_style(cell, current_col_id, sheet_styling_config, fob_mode)

            # --- START: Refactored Logic ---
            # 2. Apply a conditional border based on the column index.
            if c_idx == col1_index:
                # First column gets a border on the sides only
                cell.border = side_only_border
            else:
                # All other columns get a full border
                cell.border = full_thin_border
            # --- END: Refactored Logic ---

        except Exception as e:
            print(f"Warning: Could not style cell at ({row_num}, {c_idx}). Error: {e}")



def fill_invoice_data(
    worksheet: Worksheet,
    sheet_name: str,
    sheet_config: Dict[str, Any], # Keep current sheet config param
    all_sheet_configs: Dict[str, Any], # <--- Add param for all sheet configs
    data_source: Union[Dict[str, List[Any]], Dict[Tuple, Dict[str, Any]]],
    data_source_type: str,
    header_info: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    sheet_styling_config: Optional[Dict[str, Any]] = None,
    add_blank_after_header: bool = False,
    static_content_after_header: Optional[Dict[str, Any]] = None,
    add_blank_before_footer: bool = False,
    static_content_before_footer: Optional[Dict[str, Any]] = None,
    merge_rules_after_header: Optional[Dict[str, int]] = None,
    merge_rules_before_footer: Optional[Dict[str, int]] = None,
    merge_rules_footer: Optional[Dict[str, int]] = None, # Added footer merge rules
    footer_info: Optional[Dict[str, Any]] = None, # Currently unused
    max_rows_to_fill: Optional[int] = None,
    grand_total_pallets: int = 0, # RE-ADDED parameter
    custom_flag: bool = False, # Added custom flag parameter
    data_cell_merging_rules: Optional[Dict[str, Any]] = None, # Added data cell merging rules 29/05/2025
    fob_mode: Optional[bool] = False,
    ) -> Tuple[bool, int, int, int, int]: # Still 5 return values
    """
    REVISED LOGIC V13: Added merge_rules_footer parameter.
    Footer pallet count uses local_chunk_pallets for processed_tables,
    and grand_total_pallets for aggregation/fob_aggregation.
    """

    # --- Initialize Variables --- (Keep existing initializations)
    actual_rows_to_process = 0; data_rows_prepared = []; col1_index = 1; num_static_labels = 0
    static_column_header_name = None; data_row_indices_written = [];
    columns_to_grid = []
    desc_col_idx = None
    local_chunk_pallets = 0
    dynamic_desc_used = False



    # get data source pallet count and hanle null
    for pallet_count in data_source.get("pallet_count", []):
        if pallet_count is not None:
            local_chunk_pallets += pallet_count

    # --- Row Index Tracking --- (Keep existing)
    row_after_header_idx = -1
    data_start_row = -1
    data_end_row = -1
    row_before_footer_idx = -1
    footer_row_final = -1

    # Ensure dictionaries/lists are initialized (Keep existing)
    static_content_after_header = static_content_after_header or {}
    static_content_before_footer = static_content_before_footer or {}
    merge_rules_after_header = merge_rules_after_header or {}
    merge_rules_before_footer = merge_rules_before_footer or {}
    merge_rules_footer = merge_rules_footer or {} # Initialize footer merge rules
    mapping_rules = mapping_rules or {}
    col_id_map = header_info.get('column_id_map', {})
    column_map = header_info.get('column_map', {})
    idx_to_header_map = {v: k for k, v in column_map.items()}


    try:
        data_cell_merging_rules = data_cell_merging_rules or {}
        # --- Validate Header Info ---
        if not header_info or 'second_row_index' not in header_info or 'column_map' not in header_info or 'num_columns' not in header_info:
            print("Error: Invalid header_info provided.")
            return False, -1, -1, -1, 0

        # --- FIX: Extract num_columns and other values from header_info ---
        num_columns = header_info['num_columns']
        data_writing_start_row = header_info['second_row_index'] + 1
 
        # --- Find Description & Pallet Info Column Indices --- (Keep existing)
        desc_col_idx = col_id_map.get("col_desc")
        pallet_info_col_idx = col_id_map.get("col_pallet")
        if pallet_info_col_idx is None: print("Warning: Header 'Pallet Info' not found.")

        # --- ADD/MODIFY THIS PART FOR PALLET INFO INDEX ---
        if pallet_info_col_idx is None:
            print("Warning: Could not find a 'Pallet Info' (e.g., 'Pallet\\nNo') column header.")
        # --- END OF ADDITION/MODIFICATION FOR PALLET INFO INDEX ---

        # --- Get Styling Config --- (Keep existing)
        force_text_headers = []
        effective_header_font = bold_font # Start with default
        effective_header_align = center_alignment # Start with default

        if sheet_styling_config:
            columns_to_grid = sheet_styling_config.get("column_ids_with_full_grid", [])
            if not isinstance(columns_to_grid, list): columns_to_grid = []

            force_text_headers = sheet_styling_config.get("force_text_format_ids", [])
            if not isinstance(force_text_headers, list): force_text_headers = []

            header_font_cfg = sheet_styling_config.get("header_font")
            if header_font_cfg and isinstance(header_font_cfg, dict):
                 font_params = {k: v for k, v in header_font_cfg.items() if v is not None}
                 if font_params:
                     try: # Expanded try block
                         effective_header_font = Font(**font_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_font config: {font_params}. Using default.")
                         pass # Keep default font on error
                     except Exception as font_err: # Catch other potential errors
                         print(f"Warning: Error applying header_font config: {font_err}. Using default.")
                         pass # Keep default font on error

            header_align_cfg = sheet_styling_config.get("header_alignment")
            if header_align_cfg and isinstance(header_align_cfg, dict):
                 align_params = {k: v for k, v in header_align_cfg.items() if v is not None}
                 if align_params:
                     try: # Expanded try block
                         effective_header_align = Alignment(**align_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_alignment config: {align_params}. Using default.")
                         pass # Keep default alignment on error
                     except Exception as align_err: # Catch other potential errors
                          print(f"Warning: Error applying header_alignment config: {align_err}. Using default.")
                          pass # Keep default alignment on error
        parsed_rules = parse_mapping_rules(
            mapping_rules=mapping_rules,
            column_id_map=col_id_map,
            idx_to_header_map=idx_to_header_map
        )

        # Unpack the results into local variables for the rest of the function to use
        static_value_map = parsed_rules["static_value_map"]
        initial_static_col1_values = parsed_rules["initial_static_col1_values"]
        dynamic_mapping_rules = parsed_rules["dynamic_mapping_rules"]
        formula_rules = parsed_rules["formula_rules"]
        col1_index = parsed_rules["col1_index"]
        num_static_labels = parsed_rules["num_static_labels"]
        static_column_header_name = parsed_rules["static_column_header_name"]
        apply_special_border_rule = parsed_rules["apply_special_border_rule"]
        fallback_on_none = parsed_rules.get("dynamic_mapping_rules", {}).get("description", {}).get("fallback_on_none")

        # --- Prepare Data Rows for Writing (Determine number of rows needed from source) ---
        # This section remains largely the same, preparing the `data_rows_prepared` list
        # which holds the *input* data, not the calculated formulas.
        desc_col_idx = col_id_map.get("col_desc") # Get the description column index
        data_rows_prepared, pallet_counts_for_rows, dynamic_desc_used, num_data_rows_from_source = prepare_data_rows(
            data_source_type=data_source_type,
            data_source=data_source,
            dynamic_mapping_rules=dynamic_mapping_rules,
            column_id_map=col_id_map,
            idx_to_header_map=idx_to_header_map,
            desc_col_idx=desc_col_idx,
            num_static_labels=num_static_labels,
            static_value_map=static_value_map,
            fob_mode=fob_mode,
        )
# --- Determine Final Number of Data Rows ---
# The number of rows to process is the greater of the number of data rows or static labels.
        actual_rows_to_process = max(len(data_rows_prepared), num_static_labels)

        # Optional: Apply max_rows_to_fill constraint if it exists
        if max_rows_to_fill is not None and max_rows_to_fill >= 0:
            actual_rows_to_process = min(actual_rows_to_process, max_rows_to_fill)

        # Ensure pallet counts list matches the number of rows we intend to process
        if len(pallet_counts_for_rows) < actual_rows_to_process: pallet_counts_for_rows.extend([0] * (actual_rows_to_process - len(pallet_counts_for_rows)))
        elif len(pallet_counts_for_rows) > actual_rows_to_process: pallet_counts_for_rows = pallet_counts_for_rows[:actual_rows_to_process]

        # --- Calculate Total Rows to Insert and Row Indices ---
        total_rows_to_insert = 0
        current_row_offset = 0

        # Row after header (static/blank)
        if add_blank_after_header:
            row_after_header_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_after_header_idx = -1 # Indicate no blank row

        # Data rows
        data_start_row = data_writing_start_row + current_row_offset
        if actual_rows_to_process > 0:
            data_end_row = data_start_row + actual_rows_to_process - 1
            total_rows_to_insert += actual_rows_to_process
            current_row_offset += actual_rows_to_process
        else:
            # No data rows to process (can happen if source is empty)
            data_end_row = data_start_row - 1 # Indicate no data rows

        # Row before footer (static/blank)
        if add_blank_before_footer:
            row_before_footer_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_before_footer_idx = -1 # Indicate no blank row

        # Calculate final footer row index relative to where this chunk starts
        footer_row_final = data_writing_start_row + total_rows_to_insert
        total_rows_to_insert += 1 # Add 1 for the footer itself

        # --- Bulk Insert Rows --- # V11: Only insert if NOT pre-inserted by caller (i.e., for single-table modes)
        if data_source_type in ['aggregation', 'fob_aggregation', "custom_aggregation"]:
            if total_rows_to_insert > 0:
                try:
                    worksheet.insert_rows(data_writing_start_row, amount=total_rows_to_insert)
                    # Unmerge the block covering the inserted rows *before* the footer starts
                    safe_unmerge_block(worksheet, data_writing_start_row, footer_row_final - 1, num_columns)
                    print("Rows inserted and unmerged successfully.")
                except Exception as bulk_insert_err:
                    print(f"Error during single-table bulk row insert/unmerge: {bulk_insert_err}")
                    # Adjust fallback row calculation
                    fallback_row = max(header_info.get('second_row_index', 0) + 1, footer_row_final)
                    return False, fallback_row, -1, -1, 0

        # --- Fill Row After Header (if applicable) --- 

        # --- Prepare FOB Data Dictionary (inside loop now, safer) ---
        # Removed the premature preparation block here.
        # FOB data dict will be prepared inside the loop if data_source_type is fob_aggregation.

        # --- Fill Data Rows Loop ---
        if actual_rows_to_process > 0:
            print(f"--- DEBUG START LOOP (Sheet: {worksheet.title}) ---")
            print(f"  data_start_row: {data_start_row}")
            print(f"  actual_rows_to_process: {actual_rows_to_process}")
            print(f"  num_static_labels: {num_static_labels}")
            print(f"  col1_index: {col1_index}")
            print(f"  initial_static_col1_values: {initial_static_col1_values}")
            print(f"  data_source_type: {data_source_type}")
            # --- END DEBUG START LOOP ---
        try:
            # --- Create a reverse map from index to ID for easy lookups inside the loop ---
            idx_to_id_map = {v: k for k, v in col_id_map.items()}

            # --- Get column indices directly using their stable IDs ---
            no_col_idx = col_id_map.get("col_no")
            pallet_info_col_idx = col_id_map.get("col_pallet")
            
            # Get the list of column IDs that need to be formatted as text
            force_text_format_ids = sheet_styling_config.get("force_text_format_ids", []) if sheet_styling_config else []
            
            # Get the list of column IDs that should have a full grid border
            grid_column_ids = sheet_styling_config.get("column_ids_with_full_grid", []) if sheet_styling_config else []

            row_pallet_index = 0
            
            # --- Main Data-Writing Loop ---
            for i in range(actual_rows_to_process):
                target_row = data_start_row + i
                data_row_indices_written.append(target_row)
                row_data_dict = data_rows_prepared[i] if i < len(data_rows_prepared) else {}
                is_last_data_row = (i == actual_rows_to_process - 1)

                # --- Pallet Count Logic (remains the same) ---
                current_row_pallet_count = pallet_counts_for_rows[i] if i < len(pallet_counts_for_rows) else 0
                if current_row_pallet_count is not None and current_row_pallet_count > 0:
                    row_pallet_index += 1
                display_pallet_order = row_pallet_index
                
                # --- Cell Filling and Styling Loop ---
                for c_idx in range(1, num_columns + 1):
                    cell = worksheet.cell(row=target_row, column=c_idx)
                    current_id = idx_to_id_map.get(c_idx)
                    value_to_write = None

                    # --- Priority 1: Handle Initial Static Label Column ---
                    if i < num_static_labels and c_idx == col1_index:
                        cell.value = initial_static_col1_values[i]
                    
                    # --- Priority 2: Handle Regular Data Rows ---
                    else:
                        # Get the value that was prepared by the "kitchen"
                        prepared_value = row_data_dict.get(c_idx)
                        value_to_write = None

                        # First, check if the prepared value is a formula hint
                        if isinstance(prepared_value, dict) and prepared_value.get("type") == "formula":
                            rule = prepared_value
                            formula_template = rule.get("template")
                            input_ids = rule.get("inputs", [])
                            formula_params = {'row': target_row}
                            valid_inputs = True
                            for idx, input_id in enumerate(input_ids):
                                input_col_idx = col_id_map.get(input_id)
                                if input_col_idx:
                                    formula_params[f'col_ref_{idx}'] = get_column_letter(input_col_idx)
                                else:
                                    valid_inputs = False; break
                            
                            if valid_inputs and formula_template:
                                value_to_write = f"={formula_template.format(**formula_params)}"
                            else:
                                value_to_write = "#REF!"
                        
                        # If not a hint, check for your original special columns
                        elif c_idx == no_col_idx:
                            value_to_write = i + 1
                        elif c_idx == pallet_info_col_idx:
                            value_to_write = f"{display_pallet_order}-{local_chunk_pallets}"

                        # If none of the above, it's just plain data
                        else:
                            value_to_write = prepared_value
                        
                        # Finally, write the chosen value to the cell once
                        cell.value = value_to_write
                    # --- Apply Cell Styling and Formatting ---
                    if current_id in force_text_format_ids:
                        cell.number_format = FORMAT_TEXT
                    
                    _apply_cell_style(cell, current_id, sheet_styling_config, fob_mode)

                # --- Apply Border Rules for the entire row ---
                for c_idx_border in range(1, num_columns + 1):
                    cell_to_border = worksheet.cell(row=target_row, column=c_idx_border)
                    border_id = idx_to_id_map.get(c_idx_border)
                    
                    apply_grid = border_id and border_id in grid_column_ids
                    
                    top_b = thin_side if i == 0 else (thin_side if apply_grid else None)
                    bottom_b = thin_side if is_last_data_row else (thin_side if apply_grid else None)
                    
                    if apply_special_border_rule and c_idx_border == col1_index:
                        cell_to_border.border = Border(left=thin_side, right=thin_side, top=(thin_side if i == 0 else None), bottom=None)
                    elif apply_grid:
                        cell_to_border.border = thin_border
                    else:
                        cell_to_border.border = Border(left=thin_side, right=thin_side, top=top_b, bottom=bottom_b)

                # --- Apply explicit cell merging for this row ---
                if data_cell_merging_rules:
                    apply_explicit_data_cell_merges_by_id(
                        worksheet=worksheet,
                        row_num=target_row,
                        column_id_map=col_id_map,
                        num_total_columns=num_columns,
                        merge_rules_data_cells=data_cell_merging_rules,
                        sheet_styling_config=sheet_styling_config,
                        fob_mode=fob_mode
                    )

        except Exception as fill_data_err:
            print(f"Error during data filling loop: {fill_data_err}\n{traceback.format_exc()}")
            return False, footer_row_final + 1, data_start_row, data_end_row, 0

    # Merge Description Column if the layout used fallback/static data
        if not dynamic_desc_used and data_start_row > 0 and data_end_row > data_start_row:
            desc_col_id = "col_desc" 
            if col_id_map.get(desc_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=desc_col_id,
                    column_id_map=col_id_map
                )

        # Always try to merge the Pallet Info Column if it exists
        if data_start_row > 0 and data_end_row > data_start_row:
            pallet_col_id = "col_pallet" 
            if col_id_map.get(pallet_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=pallet_col_id,
                    column_id_map=col_id_map
                )
        if data_start_row > 0 and data_end_row > data_start_row:
            pallet_col_id = "col_hs" 
            if col_id_map.get(pallet_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=pallet_col_id,
                    column_id_map=col_id_map
                )

# --- Fill Row Before Footer ---
        if add_blank_before_footer and row_before_footer_idx > 0:
            try:
                # Step 1: Fill the row with content (this applies default styles)
                fill_static_row(worksheet, row_before_footer_idx, num_columns, static_content_before_footer)
                
                # Step 2: Apply the special styling and borders for this specific row
                _style_row_before_footer(
                    worksheet=worksheet,
                    row_num=row_before_footer_idx,
                    num_columns=num_columns,
                    sheet_styling_config=sheet_styling_config,
                    idx_to_id_map=idx_to_id_map, # Pass the ID map here
                    col1_index=col1_index,
                    fob_mode=fob_mode)
            except Exception as fill_bf_err:
                print(f"Warning: Error filling/styling row before footer: {fill_bf_err}")
        

        # --- Fill Footer Row --- (Keep existing logic)
        # The SUM formulas here should correctly sum the results of the formulas
        # written in the data rows above.
        if footer_row_final > 0:
            # Get the footer configuration object from the main sheet config
            footer_config = sheet_config.get("footer_configurations", {})
            data_range_to_sum = [(data_start_row, data_end_row)]

            pallet_count = 0
            if data_source_type == "processed_tables":
                pallet_count = local_chunk_pallets
            else:
                pallet_count = grand_total_pallets

            write_footer_row(
                worksheet=worksheet,
                footer_row_num=footer_row_final,
                header_info=header_info,
                sum_ranges=data_range_to_sum,
                footer_config=footer_config,
                pallet_count=pallet_count,
                fob_mode=data_source_type == "fob_aggregation"
            )
    # No need to pass font, alignment, num_columns, etc. as the
    # function gets this info from header_info and footer_config.
        # --- Apply Merges ---
        # Apply merges to row after header (if applicable)
        if add_blank_after_header and row_after_header_idx > 0 and merge_rules_after_header:
            apply_row_merges(worksheet, row_after_header_idx, num_columns, merge_rules_after_header)

        # Apply merges to row before footer (if applicable)
        target_row_for_bf_merge = row_before_footer_idx if add_blank_before_footer and row_before_footer_idx > 0 else -1
        if target_row_for_bf_merge > 0 and merge_rules_before_footer:
            apply_row_merges(worksheet, target_row_for_bf_merge, num_columns, merge_rules_before_footer)

        # Apply merges to the footer row itself (if applicable)
        if footer_row_final > 0 and merge_rules_footer:
            print(f"Applying footer merges to row {footer_row_final} with rules: {merge_rules_footer}") # Optional Debug
            try:
                apply_row_merges(worksheet, footer_row_final, num_columns, merge_rules_footer)
            except Exception as footer_merge_err:
                 print(f"Warning: Error applying footer merges: {footer_merge_err}")

        # --- Apply Row Heights --- (Keep existing)
        apply_row_heights(worksheet=worksheet, sheet_styling_config=sheet_styling_config, header_info=header_info, data_row_indices=data_row_indices_written, footer_row_index=footer_row_final, row_after_header_idx=row_after_header_idx, row_before_footer_idx=row_before_footer_idx)

        # --- Finalization --- (Keep existing)
        next_available_row_final = footer_row_final + 1
        if actual_rows_to_process == 0: data_start_row = -1; data_end_row = -1
        return True, next_available_row_final, data_start_row, data_end_row, local_chunk_pallets

    except Exception as e:
        # --- Error Handling --- (Keep existing)
        print(f"Critical error in fill_invoice_data: {e}\n{traceback.format_exc()}")
        fallback_row = header_info.get('second_row_index', 0) + 1; frf_local = locals().get('footer_row_final', -1)
        if frf_local > 0: fallback_row = max(fallback_row, frf_local + 1)
        else: est_footer = locals().get('initial_insert_point', fallback_row) + locals().get('total_rows_to_insert', 0); fallback_row = max(fallback_row, est_footer)
        return False, fallback_row, -1, -1, 0

def apply_column_widths(worksheet: Worksheet, sheet_styling_config: Optional[Dict[str, Any]], header_map: Optional[Dict[str, int]]):
    """
    Sets column widths based on the configuration.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'column_widths' dictionary.
        header_map: Dictionary mapping header text to column index (1-based).
    """
    if not sheet_styling_config or not header_map: return
    column_widths_cfg = sheet_styling_config.get("column_widths")
    if not column_widths_cfg or not isinstance(column_widths_cfg, dict): return
    for header_text, width in column_widths_cfg.items():
        col_idx = header_map.get(header_text)
        if col_idx:
            col_letter = get_column_letter(col_idx)
            try:
                width_val = float(width)
                if width_val > 0: worksheet.column_dimensions[col_letter].width = width_val
                else: pass # Ignore non-positive widths
            except (ValueError, TypeError): pass # Ignore invalid width values
            except Exception as width_err: pass # Log other errors?
        else: pass # Header text not found in map


def apply_row_heights(worksheet: Worksheet, sheet_styling_config: Optional[Dict[str, Any]], header_info: Dict[str, Any], data_row_indices: List[int], footer_row_index: int, row_after_header_idx: int, row_before_footer_idx: int):
    """
    Sets row heights based on the configuration for header, data, footer, and specific rows.
    Footer height can now optionally match the header height.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'row_heights' dictionary.
        header_info: Dictionary with header row indices.
        data_row_indices: List of 1-based indices for the actual data rows written.
        footer_row_index: 1-based index of the footer row.
        row_after_header_idx: 1-based index of the static/blank row after the header (-1 if none).
        row_before_footer_idx: 1-based index of the static/blank row before the footer (-1 if none).
    """
    if not sheet_styling_config: return
    row_heights_cfg = sheet_styling_config.get("row_heights")
    if not row_heights_cfg or not isinstance(row_heights_cfg, dict): return

    actual_header_height = None # Store the applied header height

    def set_height(r_idx, height_val, desc): # Helper function
        nonlocal actual_header_height # Ensure actual_header_height is modified
        if r_idx <= 0: return
        try:
            h_val = float(height_val)
            if h_val > 0:
                worksheet.row_dimensions[r_idx].height = h_val
                if desc == "header": # Store the height applied to the header
                    actual_header_height = h_val
            else: pass # Ignore non-positive heights
        except (ValueError, TypeError): pass # Ignore invalid height values
        except Exception as height_err: pass # Log other errors?

    # Apply Heights Based on Config
    header_height = row_heights_cfg.get("header")
    if header_height is not None and header_info:
        h_start = header_info.get('first_row_index', -1); h_end = header_info.get('second_row_index', -1)
        if h_start > 0 and h_end >= h_start:
            for r in range(h_start, h_end + 1): set_height(r, header_height, "header")

    after_header_height = row_heights_cfg.get("after_header")
    if after_header_height is not None and row_after_header_idx > 0: set_height(row_after_header_idx, after_header_height, "after_header")
    data_default_height = row_heights_cfg.get("data_default")
    if data_default_height is not None and data_row_indices:
        for r in data_row_indices: set_height(r, data_default_height, "data_default")
    before_footer_height = row_heights_cfg.get("before_footer")
    if before_footer_height is not None and row_before_footer_idx > 0: set_height(row_before_footer_idx, before_footer_height, "before_footer")

    # --- Footer Height Logic ---
    footer_height_config = row_heights_cfg.get("footer")
    match_header_height_flag = row_heights_cfg.get("footer_matches_header_height", True) # Default to True

    final_footer_height = None
    if match_header_height_flag and actual_header_height is not None:
        final_footer_height = actual_header_height # Use header height if flag is true and header height was set
    elif footer_height_config is not None:
        final_footer_height = footer_height_config # Otherwise, use specific footer height if defined

    if final_footer_height is not None and footer_row_index > 0:
        set_height(footer_row_index, final_footer_height, "footer")
    # --- End Footer Height Logic ---

    specific_heights = row_heights_cfg.get("specific_rows")
    if isinstance(specific_heights, dict):
        for row_str, height_val in specific_heights.items():
            try: row_num = int(row_str); set_height(row_num, height_val, f"specific_row_{row_num}")
            except ValueError: pass # Ignore invalid row numbers


# --- Main Execution Guard --- (Keep existing)
if __name__ == "__main__":
    print("invoice_utils.py executed directly.")

   