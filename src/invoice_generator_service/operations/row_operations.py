#!/usr/bin/env python3
"""
Row Operations - Clean Row Manipulation Logic
"""

import logging
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet

class RowOperations:
    """A class to handle row-level operations like inserting and styling."""

    def insert_rows_and_copy_styles(self, worksheet: Worksheet, start_row: int, num_rows: int):
        """
        Inserts a specified number of rows and copies the style from the starting row.

        Args:
            worksheet: The openpyxl worksheet object.
            start_row: The row index where insertion should begin.
            num_rows: The number of rows to insert.
        """
        if num_rows <= 0:
            return

        logging.debug(f"Inserting {num_rows} rows at index {start_row + 1} in sheet '{worksheet.title}'")
        
        # The source row for styles
        source_row_idx = start_row
        
        for i in range(num_rows):
            new_row_idx = start_row + i + 1
            worksheet.insert_rows(new_row_idx)

            # Copy styles from the source row to the new row
            for col_idx in range(1, worksheet.max_column + 1):
                source_cell = worksheet.cell(row=source_row_idx, column=col_idx)
                new_cell = worksheet.cell(row=new_row_idx, column=col_idx)

                if source_cell.has_style:
                    new_cell.font = copy(source_cell.font)
                    new_cell.border = copy(source_cell.border)
                    new_cell.fill = copy(source_cell.fill)
                    new_cell.number_format = copy(source_cell.number_format)
                    new_cell.protection = copy(source_cell.protection)
                    new_cell.alignment = copy(source_cell.alignment)
        
        logging.debug("Finished inserting rows and copying styles.")
