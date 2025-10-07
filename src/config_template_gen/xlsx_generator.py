#!/usr/bin/env python3
"""
XLSX Generator - Comprehensive Excel file processing and generation.

This module combines text replacement and row removal functionality to:
1. Process Excel files with text replacements
2. Remove specified rows based on table structure
3. Generate modified XLSX files with proper naming
4. Preserve formatting during processing
5. Provide comprehensive reporting
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from excel_processor import ExcelProcessor
from row_processor import RowProcessor

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class XLSXGenerator:
    """Comprehensive XLSX file generator with text replacement and row removal."""
    
    def __init__(self):
        """Initialize the XLSX generator with processors."""
        self.text_processor = ExcelProcessor()
        self.row_processor = RowProcessor()
    
    def generate_processed_xlsx(self, 
                              input_file: str, 
                              output_file: Optional[str] = None,
                              enable_text_replacement: bool = True,
                              enable_row_removal: bool = True) -> str:
        """
        Generate a processed XLSX file with text replacement and row removal.
        
        Args:
            input_file: Path to the input Excel file
            output_file: Path for the output file (optional)
            enable_text_replacement: Whether to perform text replacements
            enable_row_removal: Whether to perform row removal
            
        Returns:
            Path to the generated XLSX file
        """
        try:
            logger.info(f"Starting XLSX generation for: {input_file}")
            
            # Determine output file path
            if output_file is None:
                input_path = Path(input_file)
                suffix = "_processed"
                if enable_text_replacement:
                    suffix += "_text"
                if enable_row_removal:
                    suffix += "_rows"
                output_file = str(input_path.parent / f"{input_path.stem}{suffix}{input_path.suffix}")
            
            # Load the workbook
            workbook = load_workbook(input_file)
            
            # --- Store merges before any destructive operations ---
            from merge_utils import (store_original_merges, find_and_restore_merges_heuristic,
                                   store_empty_merges_with_coordinates, restore_empty_merges_with_offset,
                                   MergeOffsetTracker)
            sheet_names = workbook.sheetnames
            
            # Store value-based merges (existing system)
            stored_merges = store_original_merges(workbook, sheet_names)
            
            # Store empty merges (new system)
            empty_merges = store_empty_merges_with_coordinates(workbook, sheet_names)
            
            # Initialize offset tracker
            offset_tracker = MergeOffsetTracker()
            
            # Step 1: Text Replacement
            if enable_text_replacement:
                logger.info("Performing text replacements...")
                self._apply_text_replacements(workbook)
            
            # Step 2: Row Removal
            if enable_row_removal:
                logger.info("Performing row removal...")
                self._apply_row_removal(workbook, offset_tracker)
            
            # --- Restore merges after all destructive operations ---
            # Restore value-based merges (existing system)
            find_and_restore_merges_heuristic(workbook, stored_merges, sheet_names)
            
            # Restore empty merges (new system)
            restore_empty_merges_with_offset(workbook, empty_merges, offset_tracker, sheet_names)
            
            # Step 3: Unhide all rows (preserve column visibility)
            logger.info("Unhiding all rows while preserving column visibility...")
            self._unhide_all_rows(workbook)
            
            # Save the processed workbook
            workbook.save(output_file)
            logger.info(f"Generated XLSX file: {output_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"Error generating XLSX file: {e}")
            raise
    
    def _apply_text_replacements(self, workbook: Workbook) -> None:
        """
        Apply text replacements to the workbook using enhanced circular pattern checking.
        
        Args:
            workbook: The workbook to process
        """
        # Import the enhanced text processor
        from enhanced_text_processor import EnhancedTextProcessor
        
        # Create enhanced processor instance
        enhanced_processor = EnhancedTextProcessor()
        
        # Initialize stats for all available categories dynamically
        total_replacement_stats = {category: 0 for category in enhanced_processor.get_replacement_patterns().keys()}
        
        # Process each worksheet with enhanced circular pattern checking
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            logger.info(f"Processing text replacements in sheet: {sheet_name} (Enhanced Mode)")
            
            # Use the enhanced processor with circular pattern checking
            sheet_stats = enhanced_processor.process_worksheet_with_circular_pattern(worksheet)
            
            # Accumulate statistics
            for category, count in sheet_stats.items():
                total_replacement_stats[category] += count
        
        # Log replacement statistics
        total_replacements = sum(total_replacement_stats.values())
        logger.info(f"Enhanced text replacements completed: {total_replacements} total")
        for category, count in total_replacement_stats.items():
            if count > 0:
                logger.info(f"  {category}: {count} replacements (circular pattern)")
    
    def _apply_row_removal(self, workbook: Workbook, offset_tracker) -> None:
        """
        Apply row removal to the workbook.
        
        Args:
            workbook: The workbook to process
            offset_tracker: MergeOffsetTracker instance for logging operations
        """
        removal_stats = {
            'total_header_rows': 0,
            'tables_found': 0,
            'rows_removed': 0
        }
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            logger.info(f"Processing row removal in sheet: {sheet_name}")
            
            # Get statistics before processing
            stats = self.row_processor.get_table_statistics(worksheet)
            removal_stats['total_header_rows'] += stats['total_header_rows']
            removal_stats['tables_found'] += stats['tables_found']
            
            # Process the worksheet WITHOUT internal merge handling
            # (xlsx_generator handles merges at the top level)
            self.row_processor._process_worksheet_rows_with_offset_tracking(worksheet, offset_tracker)
            
            # Calculate rows removed
            removal_stats['rows_removed'] += stats['rows_to_remove']
        
        # Log removal statistics
        logger.info(f"Row removal completed:")
        logger.info(f"  Total header rows found: {removal_stats['total_header_rows']}")
        logger.info(f"  Tables found: {removal_stats['tables_found']}")
        logger.info(f"  Rows removed: {removal_stats['rows_removed']}")
    
    def _unhide_all_rows(self, workbook: Workbook) -> None:
        """
        Unhide all rows in all worksheets while preserving column visibility.
        
        Args:
            workbook: The workbook to process
        """
        logger.info("Unhiding all rows in all worksheets...")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_unhidden = 0
            
            # Unhide all rows
            for row_idx in range(1, worksheet.max_row + 1):
                if worksheet.row_dimensions[row_idx].hidden:
                    worksheet.row_dimensions[row_idx].hidden = False
                    rows_unhidden += 1
            
            if rows_unhidden > 0:
                logger.info(f"  Sheet '{sheet_name}': {rows_unhidden} rows unhidden")
            else:
                logger.info(f"  Sheet '{sheet_name}': no hidden rows found")
    
    def generate_comprehensive_report(self, 
                                    input_file: str,
                                    output_file: str,
                                    text_replacements: Dict[str, int],
                                    row_removals: Dict[str, int]) -> str:
        """
        Generate a comprehensive report of the processing.
        
        Args:
            input_file: Path to the input file
            output_file: Path to the output file
            text_replacements: Statistics about text replacements
            row_removals: Statistics about row removals
            
        Returns:
            Path to the generated report file
        """
        report_path = Path(output_file).with_suffix('.txt')
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("XLSX PROCESSING REPORT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Input File: {input_file}\n")
            f.write(f"Output File: {output_file}\n")
            f.write(f"Report Generated: {Path(__file__).stat().st_mtime}\n")
            f.write("=" * 80 + "\n\n")
            
            # Text Replacement Summary
            f.write("TEXT REPLACEMENT SUMMARY\n")
            f.write("-" * 40 + "\n")
            total_text_replacements = sum(text_replacements.values())
            f.write(f"Total Replacements: {total_text_replacements}\n")
            for category, count in text_replacements.items():
                if count > 0:
                    f.write(f"  {category.replace('_', ' ').title()}: {count}\n")
            f.write("\n")
            
            # Row Removal Summary
            f.write("ROW REMOVAL SUMMARY\n")
            f.write("-" * 40 + "\n")
            f.write(f"Total Header Rows Found: {row_removals.get('total_header_rows', 0)}\n")
            f.write(f"Tables Found: {row_removals.get('tables_found', 0)}\n")
            f.write(f"Rows Removed: {row_removals.get('rows_removed', 0)}\n")
            f.write("\n")
            
            # Processing Details
            f.write("PROCESSING DETAILS\n")
            f.write("-" * 40 + "\n")
            f.write("Text Replacement Patterns:\n")
            f.write("  - Date patterns → JFTIME\n")
            f.write("  - Invoice No patterns → JFINV\n")
            f.write("  - Ref No patterns → JFREF\n")
            f.write("  - ETD patterns → JFTIME\n")
            f.write("\n")
            f.write("Row Removal Logic:\n")
            f.write("  - Detect header rows with table-like content\n")
            f.write("  - Find CBM columns in headers\n")
            f.write("  - Locate SUM formula rows\n")
            f.write("  - Remove rows from header to formula\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("END OF REPORT\n")
            f.write("=" * 80 + "\n")
        
        logger.info(f"Comprehensive report generated: {report_path}")
        return str(report_path)
    
    def batch_process_files(self, 
                           input_files: List[str],
                           output_directory: Optional[str] = None,
                           enable_text_replacement: bool = True,
                           enable_row_removal: bool = True) -> List[str]:
        """
        Process multiple Excel files in batch.
        
        Args:
            input_files: List of input file paths
            output_directory: Directory for output files (optional)
            enable_text_replacement: Whether to perform text replacements
            enable_row_removal: Whether to perform row removal
            
        Returns:
            List of generated output file paths
        """
        output_files = []
        
        for input_file in input_files:
            try:
                logger.info(f"Processing file: {input_file}")
                
                # Determine output path
                if output_directory:
                    input_path = Path(input_file)
                    output_path = Path(output_directory) / input_path.name
                    output_file = self.generate_processed_xlsx(
                        input_file, 
                        str(output_path),
                        enable_text_replacement,
                        enable_row_removal
                    )
                else:
                    output_file = self.generate_processed_xlsx(
                        input_file,
                        enable_text_replacement=enable_text_replacement,
                        enable_row_removal=enable_row_removal
                    )
                
                output_files.append(output_file)
                logger.info(f"Successfully processed: {output_file}")
                
            except Exception as e:
                logger.error(f"Error processing {input_file}: {e}")
                continue
        
        logger.info(f"Batch processing completed: {len(output_files)} files processed")
        return output_files


def main():
    """Main function for the XLSX generator."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate processed XLSX files with text replacement and row removal')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    parser.add_argument('--no-text-replacement', action='store_true', help='Skip text replacement')
    parser.add_argument('--no-row-removal', action='store_true', help='Skip row removal')
    parser.add_argument('--generate-report', action='store_true', help='Generate detailed report')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    generator = XLSXGenerator()
    
    try:
        # Generate the processed XLSX file
        output_file = generator.generate_processed_xlsx(
            args.input_file,
            args.output,
            enable_text_replacement=not args.no_text_replacement,
            enable_row_removal=not args.no_row_removal
        )
        
        print(f"✅ Successfully generated XLSX file: {output_file}")
        
        # Generate report if requested
        if args.generate_report:
            # This would require tracking statistics during processing
            # For now, we'll create a basic report
            report_path = Path(output_file).with_suffix('.txt')
            with open(report_path, 'w') as f:
                f.write(f"XLSX Processing Report\n")
                f.write(f"Input: {args.input_file}\n")
                f.write(f"Output: {output_file}\n")
                f.write(f"Text Replacement: {'Enabled' if not args.no_text_replacement else 'Disabled'}\n")
                f.write(f"Row Removal: {'Enabled' if not args.no_row_removal else 'Disabled'}\n")
            
            print(f"📋 Report generated: {report_path}")
        
    except Exception as e:
        print(f"❌ Error generating XLSX file: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main()) 