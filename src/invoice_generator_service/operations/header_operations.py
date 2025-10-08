#!/usr/bin/env python3
"""
Header Operations - Clean Header Writing Logic
"""

import logging
from typing import Dict, Any, List, Tuple, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from .style_operations import StyleOperations

class HeaderOperations:
    """A class to handle writing and styling headers in a worksheet."""

    def __init__(self, style_ops: Optional[StyleOperations] = None):
        self.style_ops = style_ops if style_ops else StyleOperations()

    def write_header(
        self,
        worksheet: Worksheet,
        start_row: int,
        header_layout: List[Dict[str, Any]],
        sheet_styling_config: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Writes a complex header to the worksheet based on a layout configuration.

        Args:
            worksheet: The openpyxl worksheet object.
            start_row: The starting row for the header.
            header_layout: A list of dictionaries defining each cell in the header.
            sheet_styling_config: The styling configuration for the sheet.

        Returns:
            A dictionary containing information about the written header, including
            a column map and the number of rows occupied.
        """
        if not header_layout:
            logging.warning("Header layout is empty. Nothing to write.")
            return {}

        logging.debug(f"Writing header at start_row {start_row} on sheet '{worksheet.title}'")
        
        column_map = {}
        max_row_offset = 0

        for cell_config in header_layout:
            label = cell_config.get('label', '')
            row_offset = cell_config.get('row', 1) - 1
            col_offset = cell_config.get('col', 1) - 1
            rowspan = cell_config.get('rowspan', 1)
            colspan = cell_config.get('colspan', 1)
            
            target_row = start_row + row_offset
            target_col = 1 + col_offset # Assuming start at column 'A'

            # Merge cells if needed
            if rowspan > 1 or colspan > 1:
                worksheet.merge_cells(
                    start_row=target_row,
                    start_column=target_col,
                    end_row=target_row + rowspan - 1,
                    end_column=target_col + colspan - 1
                )

            cell = worksheet.cell(row=target_row, column=target_col)
            cell.value = label
            
            # Apply styling
            style_key = cell_config.get('style_key')
            if sheet_styling_config and style_key and style_key in sheet_styling_config:
                style = sheet_styling_config[style_key]
                self.style_ops.apply_cell_style(cell, style)

            # Build column map from the 'field' property
            field = cell_config.get('field')
            if field:
                column_map[field] = get_column_letter(target_col)
            
            max_row_offset = max(max_row_offset, row_offset + rowspan)

        header_info = {
            'column_map': column_map,
            'header_rows': max_row_offset,
            'start_row': start_row,
            'end_row': start_row + max_row_offset - 1
        }
        
        logging.debug(f"Header written successfully. Info: {header_info}")
        return header_info
