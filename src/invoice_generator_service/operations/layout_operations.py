#!/usr/bin/env python3
"""
Layout Operations - Clean Layout Management Logic
Single responsibility: Worksheet layout, headers, footers, and structure
"""

from typing import Dict, Any, List, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet


class LayoutOperations:
    """Clean, focused layout operations without scattered utility functions"""
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
    
    def write_header_section(
        self,
        worksheet: Worksheet,
        start_row: int,
        header_config: List[Dict[str, Any]],
        data_source: Dict[str, Any]
    ) -> int:
        """
        Write header section to worksheet
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row for header
            header_config: Header configuration
            data_source: Data for header substitution
            
        Returns:
            Next available row after header
        """
        current_row = start_row
        
        for row_config in header_config:
            self._write_header_row(worksheet, current_row, row_config, data_source)
            current_row += 1
        
        return current_row
    
    def write_data_section(
        self,
        worksheet: Worksheet,
        start_row: int,
        data_rows: List[List[Any]],
        column_config: Dict[str, Any]
    ) -> int:
        """
        Write data section to worksheet
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row for data
            data_rows: List of data rows
            column_config: Column configuration
            
        Returns:
            Next available row after data
        """
        current_row = start_row
        
        for row_data in data_rows:
            self._write_data_row(worksheet, current_row, row_data, column_config)
            current_row += 1
        
        return current_row
    
    def write_footer_section(
        self,
        worksheet: Worksheet,
        start_row: int,
        footer_config: Dict[str, Any],
        summary_data: Optional[Dict[str, Any]] = None
    ) -> int:
        """
        Write footer section to worksheet
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row for footer
            footer_config: Footer configuration
            summary_data: Summary data for footer
            
        Returns:
            Next available row after footer
        """
        current_row = start_row
        
        # Write summary rows if configured
        if 'summary_rows' in footer_config and summary_data:
            for summary_config in footer_config['summary_rows']:
                self._write_summary_row(worksheet, current_row, summary_config, summary_data)
                current_row += 1
        
        # Write footer rows
        if 'footer_rows' in footer_config:
            for footer_row_config in footer_config['footer_rows']:
                self._write_footer_row(worksheet, current_row, footer_row_config)
                current_row += 1
        
        return current_row
    
    def find_footer_position(
        self,
        worksheet: Worksheet,
        footer_rules: Dict[str, Any]
    ) -> Optional[Tuple[int, Dict[str, Any]]]:
        """
        Find footer position based on rules
        
        Args:
            worksheet: Target worksheet
            footer_rules: Rules for finding footer
            
        Returns:
            Tuple of (row_number, footer_info) or None if not found
        """
        search_method = footer_rules.get('method', 'scan_from_bottom')
        
        if search_method == 'scan_from_bottom':
            return self._find_footer_by_bottom_scan(worksheet, footer_rules)
        elif search_method == 'search_keyword':
            return self._find_footer_by_keyword(worksheet, footer_rules)
        elif search_method == 'fixed_position':
            return self._find_footer_by_position(worksheet, footer_rules)
        else:
            return None
    
    def adjust_layout_spacing(
        self,
        worksheet: Worksheet,
        layout_config: Dict[str, Any]
    ) -> None:
        """
        Adjust spacing between layout sections
        
        Args:
            worksheet: Target worksheet
            layout_config: Layout configuration with spacing rules
        """
        # Set row heights
        if 'row_heights' in layout_config:
            for row_num, height in layout_config['row_heights'].items():
                worksheet.row_dimensions[int(row_num)].height = height
        
        # Set column widths
        if 'column_widths' in layout_config:
            from openpyxl.utils import get_column_letter
            for col_num, width in layout_config['column_widths'].items():
                col_letter = get_column_letter(int(col_num))
                worksheet.column_dimensions[col_letter].width = width
        
        # Insert blank rows for spacing
        if 'spacing_rows' in layout_config:
            for spacing_config in layout_config['spacing_rows']:
                row_num = spacing_config['after_row']
                blank_rows = spacing_config.get('blank_rows', 1)
                
                # Insert blank rows
                worksheet.insert_rows(row_num + 1, blank_rows)
    
    def calculate_section_bounds(
        self,
        worksheet: Worksheet,
        section_type: str,
        start_row: int,
        data_count: int
    ) -> Tuple[int, int]:
        """
        Calculate section boundaries
        
        Args:
            worksheet: Target worksheet
            section_type: Type of section ('header', 'data', 'footer')
            start_row: Starting row
            data_count: Number of data items
            
        Returns:
            Tuple of (start_row, end_row)
        """
        if section_type == 'header':
            # Headers typically have fixed height
            return (start_row, start_row + 5)
        elif section_type == 'data':
            # Data section grows with data
            return (start_row, start_row + data_count - 1)
        elif section_type == 'footer':
            # Footers typically have fixed height
            return (start_row, start_row + 3)
        else:
            return (start_row, start_row)
    
    def _write_header_row(
        self,
        worksheet: Worksheet,
        row_num: int,
        row_config: Dict[str, Any],
        data_source: Dict[str, Any]
    ) -> None:
        """Write a single header row"""
        if 'columns' in row_config:
            for col_config in row_config['columns']:
                col_num = col_config.get('column', 1)
                
                # Get value from data source or use static value
                if 'data_key' in col_config:
                    value = data_source.get(col_config['data_key'], '')
                else:
                    value = col_config.get('value', '')
                
                # Set cell value
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
    
    def _write_data_row(
        self,
        worksheet: Worksheet,
        row_num: int,
        row_data: List[Any],
        column_config: Dict[str, Any]
    ) -> None:
        """Write a single data row"""
        for col_index, value in enumerate(row_data):
            col_num = col_index + 1
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
    
    def _write_summary_row(
        self,
        worksheet: Worksheet,
        row_num: int,
        summary_config: Dict[str, Any],
        summary_data: Dict[str, Any]
    ) -> None:
        """Write a summary row"""
        label_col = summary_config.get('label_column', 1)
        value_col = summary_config.get('value_column', 2)
        
        # Write label
        label = summary_config.get('label', '')
        worksheet.cell(row=row_num, column=label_col).value = label
        
        # Write value
        data_key = summary_config.get('data_key', '')
        value = summary_data.get(data_key, '')
        worksheet.cell(row=row_num, column=value_col).value = value
    
    def _write_footer_row(
        self,
        worksheet: Worksheet,
        row_num: int,
        footer_config: Dict[str, Any]
    ) -> None:
        """Write a footer row"""
        if 'columns' in footer_config:
            for col_config in footer_config['columns']:
                col_num = col_config.get('column', 1)
                value = col_config.get('value', '')
                
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
    
    def _find_footer_by_bottom_scan(
        self,
        worksheet: Worksheet,
        footer_rules: Dict[str, Any]
    ) -> Optional[Tuple[int, Dict[str, Any]]]:
        """Find footer by scanning from bottom"""
        max_row = worksheet.max_row
        scan_range = footer_rules.get('scan_range', 10)
        
        # Scan from bottom up
        for row in range(max_row, max(max_row - scan_range, 1) - 1, -1):
            # Check if row has content that looks like footer
            has_content = False
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and str(cell_value).strip():
                    has_content = True
                    break
            
            if has_content:
                return (row, {'method': 'bottom_scan', 'found_row': row})
        
        return None
    
    def _find_footer_by_keyword(
        self,
        worksheet: Worksheet,
        footer_rules: Dict[str, Any]
    ) -> Optional[Tuple[int, Dict[str, Any]]]:
        """Find footer by searching for keyword"""
        keyword = footer_rules.get('keyword', 'Total')
        search_column = footer_rules.get('search_column', 1)
        
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=search_column).value
            if cell_value and keyword.lower() in str(cell_value).lower():
                return (row, {'method': 'keyword', 'keyword': keyword, 'found_row': row})
        
        return None
    
    def _find_footer_by_position(
        self,
        worksheet: Worksheet,
        footer_rules: Dict[str, Any]
    ) -> Optional[Tuple[int, Dict[str, Any]]]:
        """Find footer by fixed position"""
        fixed_row = footer_rules.get('row', worksheet.max_row)
        return (fixed_row, {'method': 'fixed_position', 'row': fixed_row})