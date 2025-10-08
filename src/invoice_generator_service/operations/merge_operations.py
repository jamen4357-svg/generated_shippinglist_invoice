#!/usr/bin/env python3
"""
Merge Operations - Clean Cell Merging Logic
Single responsibility: Cell merging and unmerging operations
"""

from typing import Dict, List, Optional, Tuple, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook


class MergeOperations:
    """Clean, focused merge operations without scattered utility functions"""
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self._merge_registry: Dict[str, List[Tuple[int, Any, Optional[float]]]] = {}
    
    def store_original_merges(self, workbook: Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple]]:
        """
        Store original merge information before processing
        
        Args:
            workbook: Target workbook
            sheet_names: List of sheet names to process
            
        Returns:
            Dictionary mapping sheet names to merge information
        """
        merge_info = {}
        
        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_merges = []
                
                # Store merge ranges
                for merge_range in worksheet.merged_cells.ranges:
                    min_row = merge_range.min_row
                    max_row = merge_range.max_row
                    min_col = merge_range.min_col
                    max_col = merge_range.max_col
                    
                    sheet_merges.append((min_row, max_row, min_col, max_col))
                
                merge_info[sheet_name] = sheet_merges
                
        self._merge_registry.update(merge_info)
        return merge_info
    
    def unmerge_row(self, worksheet: Worksheet, row_num: int, num_cols: int) -> None:
        """
        Unmerge all cells in a specific row
        
        Args:
            worksheet: Target worksheet
            row_num: Row number to unmerge
            num_cols: Number of columns to check
        """
        ranges_to_unmerge = []
        
        for merge_range in worksheet.merged_cells.ranges:
            if merge_range.min_row <= row_num <= merge_range.max_row:
                ranges_to_unmerge.append(merge_range)
        
        for merge_range in ranges_to_unmerge:
            worksheet.unmerge_cells(str(merge_range))
    
    def unmerge_block(
        self, 
        worksheet: Worksheet, 
        start_row: int, 
        end_row: int, 
        num_cols: int
    ) -> None:
        """
        Unmerge all cells in a block range
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row number
            end_row: Ending row number  
            num_cols: Number of columns to check
        """
        ranges_to_unmerge = []
        
        for merge_range in worksheet.merged_cells.ranges:
            # Check if merge range overlaps with our block
            if (merge_range.min_row <= end_row and merge_range.max_row >= start_row):
                ranges_to_unmerge.append(merge_range)
        
        for merge_range in ranges_to_unmerge:
            try:
                worksheet.unmerge_cells(str(merge_range))
            except Exception as e:
                if self.verbose:
                    print(f"Could not unmerge {merge_range}: {e}")
    
    def merge_cells_in_row(
        self,
        worksheet: Worksheet,
        row_num: int,
        merge_rules: Dict[str, int]
    ) -> None:
        """
        Merge cells in a row according to rules
        
        Args:
            worksheet: Target worksheet
            row_num: Row number to merge
            merge_rules: Dictionary mapping column identifiers to merge spans
        """
        for column_id, merge_span in merge_rules.items():
            if merge_span > 1:
                try:
                    # Convert column ID to column number
                    start_col = self._column_id_to_number(column_id)
                    end_col = start_col + merge_span - 1
                    
                    # Merge the range
                    worksheet.merge_cells(
                        start_row=row_num,
                        start_column=start_col,
                        end_row=row_num,
                        end_column=end_col
                    )
                except Exception as e:
                    if self.verbose:
                        print(f"Could not merge {column_id} in row {row_num}: {e}")
    
    def merge_vertical_cells(
        self,
        worksheet: Worksheet,
        col: int,
        start_row: int,
        end_row: int
    ) -> None:
        """
        Merge cells vertically in a column
        
        Args:
            worksheet: Target worksheet
            col: Column number
            start_row: Starting row
            end_row: Ending row
        """
        if start_row < end_row:
            try:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col
                )
            except Exception as e:
                if self.verbose:
                    print(f"Could not merge vertical cells col {col}, rows {start_row}-{end_row}: {e}")
    
    def merge_contiguous_cells(
        self,
        worksheet: Worksheet,
        row_num: int,
        column_data: List[Tuple[str, Any]]
    ) -> None:
        """
        Merge contiguous cells with the same value in a row
        
        Args:
            worksheet: Target worksheet
            row_num: Row number to process
            column_data: List of (column_id, value) tuples
        """
        if not column_data:
            return
        
        current_value = column_data[0][1]
        start_col = self._column_id_to_number(column_data[0][0])
        current_col = start_col
        
        for i in range(1, len(column_data)):
            col_id, value = column_data[i]
            col_num = self._column_id_to_number(col_id)
            
            if value == current_value and col_num == current_col + 1:
                current_col = col_num
            else:
                # Merge the previous contiguous range
                if current_col > start_col:
                    try:
                        worksheet.merge_cells(
                            start_row=row_num,
                            start_column=start_col,
                            end_row=row_num,
                            end_column=current_col
                        )
                    except Exception as e:
                        if self.verbose:
                            print(f"Could not merge contiguous cells: {e}")
                
                # Start new range
                current_value = value
                start_col = col_num
                current_col = col_num
        
        # Merge the final range if needed
        if current_col > start_col:
            try:
                worksheet.merge_cells(
                    start_row=row_num,
                    start_column=start_col,
                    end_row=row_num,
                    end_column=current_col
                )
            except Exception as e:
                if self.verbose:
                    print(f"Could not merge final contiguous cells: {e}")
    
    def restore_original_merges(
        self,
        workbook: Workbook,
        sheet_name: str,
        preserve_data_area: bool = True
    ) -> None:
        """
        Restore original merges from registry
        
        Args:
            workbook: Target workbook
            sheet_name: Sheet name to restore
            preserve_data_area: Whether to avoid restoring merges in data areas
        """
        if sheet_name not in self._merge_registry:
            return
        
        worksheet = workbook[sheet_name]
        original_merges = self._merge_registry[sheet_name]
        
        for min_row, max_row, min_col, max_col in original_merges:
            try:
                # Skip if in data area and preserve_data_area is True
                if preserve_data_area and self._is_in_data_area(min_row, max_row):
                    continue
                
                worksheet.merge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col
                )
            except Exception as e:
                if self.verbose:
                    print(f"Could not restore merge {min_row}-{max_row}, {min_col}-{max_col}: {e}")
    
    def _column_id_to_number(self, column_id: str) -> int:
        """Convert column ID (like 'A', 'B', 'AA') to column number"""
        try:
            from openpyxl.utils import column_index_from_string
            return column_index_from_string(column_id)
        except Exception:
            # Fallback for numeric column IDs
            try:
                return int(column_id)
            except Exception:
                return 1
    
    def _is_in_data_area(self, min_row: int, max_row: int) -> bool:
        """Check if merge range is in typical data area (rows 18+)"""
        return min_row >= 18