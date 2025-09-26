#!/usr/bin/env python3
"""
Row Processor - Handles row removal and table processing in Excel files.

This module provides functionality to:
1. Detect header rows in Excel sheets
2. Find CBM columns and formula rows
3. Remove rows from header to formula row
4. Handle multiple tables per sheet
5. Process circular patterns for table detection
"""

import re
import json
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
# --- Add merge utility imports ---
from merge_utils import store_original_merges, find_and_restore_merges_heuristic

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class RowProcessor:
    """Handles row removal and table processing in Excel files."""
    
    def __init__(self, mapping_config_path: str = "mapping_config.json"):
        """Initialize the row processor."""
        self.mapping_config_path = mapping_config_path
        self.header_mappings = self._load_header_mappings()
        
        # Extract header keywords from mapping config
        self.header_keywords = list(self.header_mappings.keys()) if self.header_mappings else [
            "P.O", "ITEM", "Description", "Quantity", "Amount", "Unit Price", 
            "Mark", "No", "Pallet", "N.W", "G.W", "CBM", "PCS", "SF"
        ]
        
        self.formula_patterns = [
            r'=sum\(',
            r'=SUM\(',
            r'=Sum\('
        ]
    
    def _load_header_mappings(self) -> Dict[str, str]:
        """Load header mappings from mapping_config.json."""
        try:
            if Path(self.mapping_config_path).exists():
                with open(self.mapping_config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('header_text_mappings', {}).get('mappings', {})
            else:
                logger.warning(f"Mapping config file not found: {self.mapping_config_path}")
                return {}
        except Exception as e:
            logger.warning(f"Error loading mapping config: {e}")
            return {}
    
    def process_excel_file(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Process an Excel file and remove specified rows.
        
        Args:
            input_file: Path to the input Excel file
            output_file: Path for the output file (optional)
            
        Returns:
            Path to the processed Excel file
        """
        try:
            logger.info(f"Processing Excel file for row removal: {input_file}")
            
            # Load the workbook
            workbook = load_workbook(input_file)
            # --- Store merges before any row operations ---
            sheet_names = workbook.sheetnames
            stored_merges = store_original_merges(workbook, sheet_names)
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                self._process_worksheet_rows(worksheet)
            
            # --- Restore merges after all row operations, before saving ---
            find_and_restore_merges_heuristic(workbook, stored_merges, sheet_names)
            
            # Determine output file path
            if output_file is None:
                input_path = Path(input_file)
                output_file = str(input_path.parent / f"{input_path.stem}_rows_processed{input_path.suffix}")
            
            # Save the processed workbook
            workbook.save(output_file)
            logger.info(f"Row-processed file saved to: {output_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"Error processing Excel file for row removal: {e}")
            raise
    
    def _process_worksheet_rows(self, worksheet: Worksheet) -> None:
        """
        Process rows in a single worksheet with comprehensive merge handling.
        
        Args:
            worksheet: The worksheet to process
        """
        logger.info(f"Starting row processing for sheet '{worksheet.title}'")
        
        # Find all header rows
        header_rows = self._find_header_rows(worksheet)
        logger.info(f"Found {len(header_rows)} header rows in sheet '{worksheet.title}'")
        
        # Note: We don't need to clean up ALL data area merges upfront
        # The targeted unmerging in _remove_row_range handles conflicts properly
        
        # Process each header row and its associated table
        for header_row in header_rows:
            self._process_table_from_header(worksheet, header_row)
        
        # Clean up any empty rows at the end after deletions
        self._cleanup_empty_rows(worksheet)
        
        logger.info(f"Completed row processing for sheet '{worksheet.title}'")
    
    def _process_worksheet_rows_no_merge_handling(self, worksheet: Worksheet) -> None:
        """
        Process rows in a single worksheet WITHOUT internal merge handling.
        This is used when the caller (like xlsx_generator) handles merges at a higher level.
        
        Args:
            worksheet: The worksheet to process
        """
        logger.info(f"Starting row processing for sheet '{worksheet.title}' (no merge handling)")
        
        # Find all header rows
        header_rows = self._find_header_rows(worksheet)
        logger.info(f"Found {len(header_rows)} header rows in sheet '{worksheet.title}'")
        
        # Process each header row and its associated table
        for header_row in header_rows:
            self._process_table_from_header(worksheet, header_row)
        
        # Clean up any empty rows at the end after deletions
        self._cleanup_empty_rows(worksheet)
        
        logger.info(f"Completed row processing for sheet '{worksheet.title}' (no merge handling)")
    
    def _process_worksheet_rows_with_offset_tracking(self, worksheet: Worksheet, offset_tracker) -> None:
        """
        Process rows in a single worksheet WITH offset tracking for empty merges.
        
        Args:
            worksheet: The worksheet to process
            offset_tracker: MergeOffsetTracker instance to log operations
        """
        logger.info(f"Starting row processing for sheet '{worksheet.title}' (with offset tracking)")
        
        # Store the offset tracker for use in row operations
        self.offset_tracker = offset_tracker
        
        # Find all header rows
        header_rows = self._find_header_rows(worksheet)
        logger.info(f"Found {len(header_rows)} header rows in sheet '{worksheet.title}'")
        
        # Process each header row and its associated table
        for header_row in header_rows:
            self._process_table_from_header_with_tracking(worksheet, header_row)
        
        # Clean up any empty rows at the end after deletions
        self._cleanup_empty_rows_with_tracking(worksheet)
        
        logger.info(f"Completed row processing for sheet '{worksheet.title}' (with offset tracking)")
    
    def _cleanup_data_area_merges(self, worksheet: Worksheet) -> None:
        """
        Clean up merges in the data area (row 16 and below) before processing.
        This ensures clean row deletion without merge corruption.
        
        Args:
            worksheet: The worksheet to clean
        """
        # Import the utility function
        from merge_utils import force_unmerge_from_row_down
        
        # Unmerge everything from row 16 down (preserves header area)
        # This is exactly what we want - clean slate for row processing
        force_unmerge_from_row_down(worksheet, 16)
        
        logger.info("Cleaned up data area merges (row 16+) for clean row processing")
    
    def _find_header_rows(self, worksheet: Worksheet) -> List[int]:
        """
        Find all header rows in the worksheet.
        
        Args:
            worksheet: The worksheet to search
            
        Returns:
            List of row numbers that contain headers
        """
        header_rows = []
        
        for row in range(1, worksheet.max_row + 1):
            if self._is_header_row(worksheet, row):
                header_rows.append(row)
                logger.debug(f"Found header row: {row}")
        
        return header_rows
    
    def _is_header_row(self, worksheet: Worksheet, row: int) -> bool:
        """
        Check if a row is a header row using the same logic as header detection system.
        
        Args:
            worksheet: The worksheet to check
            row: Row number to check
            
        Returns:
            True if the row is a header row
        """
        header_cells = 0
        total_cells = 0
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                total_cells += 1
                cell_text = str(cell.value).strip()
                
                # Check if cell contains any of our header keywords
                for keyword in self.header_keywords:
                    if self._matches_keyword(cell_text, keyword):
                        header_cells += 1
                        break
        
        # Consider it a header row if at least 50% of non-empty cells contain header text
        if total_cells > 0:
            header_ratio = header_cells / total_cells
            return header_ratio >= 0.5
        
        return False
    
    def _matches_keyword(self, cell_value: str, keyword: str) -> bool:
        """
        Check if a cell value matches a header keyword.
        Uses the same logic as the header detection system.
        
        Args:
            cell_value: The cell value to check
            keyword: The keyword to match against
            
        Returns:
            True if the cell value matches the keyword
        """
        cell_lower = cell_value.lower().strip()
        keyword_lower = keyword.lower()
        
        # Exact match
        if cell_lower == keyword_lower:
            return True
            
        # Allow some common variations but keep it strict
        # Remove common punctuation and extra spaces
        import re
        cell_clean = re.sub(r'[^\w\s]', ' ', cell_lower)
        cell_clean = ' '.join(cell_clean.split())  # normalize whitespace
        
        keyword_clean = re.sub(r'[^\w\s]', ' ', keyword_lower)
        keyword_clean = ' '.join(keyword_clean.split())
        
        # Check if cleaned versions match
        if cell_clean == keyword_clean:
            return True
            
        # For very short cells (likely headers), allow if keyword is majority of content
        if len(cell_lower) <= 20 and keyword_lower in cell_lower:
            # Calculate similarity - keyword should be significant portion
            similarity = len(keyword_lower) / len(cell_lower)
            return similarity >= 0.6
        
        return False
    
    def _process_table_from_header(self, worksheet: Worksheet, header_row: int) -> None:
        """
        Process a table starting from a header row with comprehensive merge handling.
        
        Args:
            worksheet: The worksheet to process
            header_row: The header row number
        """
        logger.info(f"Processing table starting from header row {header_row}")
        
        # Find any column with a formula (SUM formula)
        formula_col = self._find_formula_column(worksheet, header_row)
        if formula_col is None:
            logger.warning(f"No formula column found for header row {header_row}")
            return
        
        # Find formula row (SUM formula) - now looks for rows with 2+ SUM formulas
        formula_row = self._find_formula_row(worksheet, header_row)
        if formula_row is None:
            logger.warning(f"No formula row found for header row {header_row}")
            return
        
        logger.info(f"Table identified: header at row {header_row}, formula at row {formula_row}")
        
        # Remove rows from header to formula row (this includes proper merge cleanup)
        self._remove_row_range(worksheet, header_row, formula_row)
        
        # After deletion, insert 2 empty rows at the original header position
        # The header position is still valid because we deleted from header_row down
        self._insert_empty_rows(worksheet, header_row, 2)
        
        logger.info(f"Successfully processed table: removed rows {header_row}-{formula_row}, inserted 2 empty rows at position {header_row}")
    
    def _find_formula_column(self, worksheet: Worksheet, header_row: int) -> Optional[int]:
        """
        Find any column that has a SUM formula below the header row.
        
        Args:
            worksheet: The worksheet to search
            header_row: The header row number
            
        Returns:
            Column number if found, None otherwise
        """
        # Look for SUM formulas in the next 50 rows after the header
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell_text = str(cell.value)
                    
                    # Check if cell contains a SUM formula
                    for pattern in self.formula_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            logger.debug(f"Found formula column: {get_column_letter(col)} (col {col}) in row {row}")
                            return col
        
        return None
    
    def _find_formula_row(self, worksheet: Worksheet, header_row: int, formula_col: int = None) -> Optional[int]:
        """
        Find the formula row (SUM formula) in the worksheet.
        Now looks for rows with 2 or more SUM formulas to identify true table footers.
        The formula_col parameter is kept for backward compatibility but is no longer used.
        
        Args:
            worksheet: The worksheet to search
            header_row: The header row number
            formula_col: (Deprecated) Column number to search - no longer used
            
        Returns:
            Row number of formula row, or None if not found
        """
        # Find all rows with SUM formulas in any column
        sum_rows = []
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            sum_count = 0
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell_text = str(cell.value)
                    # Check if cell contains a SUM formula
                    for pattern in self.formula_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            sum_count += 1
                            break
            
            if sum_count >= 2:  # Row has 2 or more SUM formulas
                sum_rows.append((row, sum_count))
                logger.debug(f"Found footer row candidate: {row} with {sum_count} SUM formulas")
        
        if sum_rows:
            # Return the first row with 2+ SUM formulas (closest to header)
            best_row = sum_rows[0][0]
            logger.debug(f"Selected footer row: {best_row} (first row with 2+ SUM formulas)")
            return best_row
        
        # Fallback: if no row has 2+ SUM formulas, use the old logic
        logger.warning("No rows found with 2+ SUM formulas, falling back to single SUM formula detection")
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell_text = str(cell.value)
                    # Check if cell contains a SUM formula
                    for pattern in self.formula_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            logger.debug(f"Found formula row (fallback): {row} with formula: {cell_text}")
                            return row
        
        return None
    
    def _remove_row_range(self, worksheet: Worksheet, start_row: int, end_row: int) -> None:
        """
        Remove a range of rows from the worksheet with proper merge cleanup.
        This completely removes rows and their associated merges (as intended).
        
        Args:
            worksheet: The worksheet to modify
            start_row: Starting row number (inclusive)
            end_row: Ending row number (inclusive)
        """
        logger.info(f"Preparing to delete rows {start_row} to {end_row} from sheet '{worksheet.title}'")
        
        # CRITICAL: Unmerge all cells in the deletion range BEFORE deleting rows
        # This prevents merge corruption and ensures clean deletion
        self._unmerge_cells_in_range(worksheet, start_row, end_row)
        
        # Calculate how many rows to delete
        rows_to_delete = end_row - start_row + 1
        
        # Delete rows from bottom to top to avoid shifting issues
        for i in range(rows_to_delete):
            # Delete the start_row (it will shift up after each deletion)
            worksheet.delete_rows(start_row)
        
        logger.info(f"Successfully deleted {rows_to_delete} rows from {start_row} to {end_row} (merges properly cleaned)")
    
    def _unmerge_cells_in_range(self, worksheet: Worksheet, start_row: int, end_row: int) -> None:
        """
        Unmerge ALL cells that could be affected by the row deletion operation.
        This includes:
        1. Merges that intersect with the deletion range
        2. Merges that start at or below the deletion range (these get shifted up)
        
        When rows are deleted, cells below the deletion point get shifted up.
        Any merge that was below the deletion range will have invalid coordinates
        after the shift, so we must unmerge them all.

        Args:
            worksheet: The worksheet to modify
            start_row: Starting row number (inclusive)
            end_row: Ending row number (inclusive)
        """
        # Create a copy of merged ranges to avoid modification during iteration
        merged_ranges_copy = list(worksheet.merged_cells.ranges)
        unmerged_count = 0
        preserved_count = 0

        logger.debug(f"Checking {len(merged_ranges_copy)} merged ranges for conflicts with deletion rows {start_row}-{end_row}")

        for merged_range in merged_ranges_copy:
            # CRITICAL FIX: Unmerge ALL merges that start at or below the deletion START row
            # When rows are deleted, ALL merges from the deletion start downward get shifted up
            # and their coordinates become invalid, so we must unmerge them all
            if merged_range.min_row >= start_row:
                # This merge starts at or below the deletion range and will be shifted up
                try:
                    logger.debug(f"Unmerging merge that will be shifted up: {merged_range.coord} (starts at row {merged_range.min_row})")
                    worksheet.unmerge_cells(str(merged_range))
                    unmerged_count += 1
                except Exception as e:
                    logger.warning(f"Failed to unmerge range {merged_range.coord}: {e}")
            else:
                # Merge starts above the deletion range and won't be affected
                logger.debug(f"Preserving merge above deletion range: {merged_range.coord}")
                preserved_count += 1

        if unmerged_count > 0:
            logger.info(f"Unmerged {unmerged_count} merges that could be affected by row deletion")
        if preserved_count > 0:
            logger.info(f"Preserved {preserved_count} merges completely below deletion range")
        if unmerged_count == 0 and preserved_count == 0:
            logger.debug("No merges found in worksheet")
    
    def _merge_completely_within_deletion_range(self, merge_start: int, merge_end: int, delete_start: int, delete_end: int) -> bool:
        """
        Check if a merge is completely contained within the deletion range.
        
        Args:
            merge_start: Start row of merged range
            merge_end: End row of merged range  
            delete_start: Start row of deletion range
            delete_end: End row of deletion range
            
        Returns:
            True if merge is completely within deletion range
        """
        # Merge is completely within deletion range if both start and end are within bounds
        return merge_start >= delete_start and merge_end <= delete_end
    
    def _ranges_intersect(self, merge_start: int, merge_end: int, delete_start: int, delete_end: int) -> bool:
        """
        Check if two row ranges intersect.
        
        Args:
            merge_start: Start row of merged range
            merge_end: End row of merged range  
            delete_start: Start row of deletion range
            delete_end: End row of deletion range
            
        Returns:
            True if ranges intersect
        """
        # Ranges intersect if one starts before the other ends
        return not (merge_end < delete_start or merge_start > delete_end)
    
    def _insert_empty_rows(self, worksheet: Worksheet, position: int, num_rows: int) -> None:
        """
        Insert empty rows at the specified position.
        
        Args:
            worksheet: The worksheet to modify
            position: The row position where to insert (1-based)
            num_rows: Number of empty rows to insert
        """
        for i in range(num_rows):
            worksheet.insert_rows(position)
        
        logger.debug(f"Inserted {num_rows} empty rows at position {position}")
    
    def _cleanup_empty_rows(self, worksheet: Worksheet) -> None:
        """
        Remove empty rows at the end of the worksheet after row deletions.
        But preserve the 2 inserted rows at header positions.
        
        Args:
            worksheet: The worksheet to clean up
        """
        max_row = worksheet.max_row
        
        # Start from the bottom and work up to find the last non-empty row
        last_non_empty_row = 0
        for row in range(max_row, 0, -1):
            row_has_content = False
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    row_has_content = True
                    break
            
            if row_has_content:
                last_non_empty_row = row
                break
        
        # If we found empty rows at the end, delete them
        # But leave at least 2 rows after the last content (for inserted rows)
        if last_non_empty_row < max_row:
            rows_to_delete = max_row - last_non_empty_row
            # Don't delete if we only have 2 or fewer empty rows (might be our inserted rows)
            if rows_to_delete > 2:
                actual_delete = rows_to_delete - 2
                for i in range(actual_delete):
                    worksheet.delete_rows(last_non_empty_row + 3)  # Delete from 3rd empty row onwards
                logger.info(f"Cleaned up {actual_delete} empty rows at the end of sheet '{worksheet.title}' (preserved 2 inserted rows)")
            else:
                logger.info(f"Preserved {rows_to_delete} empty rows at the end of sheet '{worksheet.title}' (likely inserted rows)")
    
    def circular_search_for_tables(self, worksheet: Worksheet) -> List[Dict[str, int]]:
        """
        Perform a circular search for tables in the worksheet.
        
        Args:
            worksheet: The worksheet to search
            
        Returns:
            List of dictionaries with table information
        """
        tables = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 0 or max_col == 0:
            return tables
        
        # Calculate center
        center_row = max_row // 2
        center_col = max_col // 2
        
        # Spiral search pattern
        directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]  # Right, Down, Left, Up
        current_dir = 0
        steps = 1
        step_count = 0
        
        row, col = center_row, center_col
        visited_cells = set()
        
        for _ in range(max_row * max_col):
            # Check if cell is within bounds and not visited
            if (1 <= row <= max_row and 1 <= col <= max_col and 
                (row, col) not in visited_cells):
                
                visited_cells.add((row, col))
                
                # Check if this cell might be part of a table
                if self._is_header_row(worksheet, row):
                    formula_col = self._find_formula_column(worksheet, row)
                    if formula_col:
                        formula_row = self._find_formula_row(worksheet, row)
                        if formula_row:
                            tables.append({
                                'header_row': row,
                                'formula_col': formula_col,
                                'formula_row': formula_row
                            })
            
            # Move in current direction
            row += directions[current_dir][0]
            col += directions[current_dir][1]
            
            step_count += 1
            
            # Change direction when step count reaches current step limit
            if step_count >= steps:
                step_count = 0
                current_dir = (current_dir + 1) % 4
                
                # Increase step limit every 2 direction changes
                if current_dir % 2 == 0:
                    steps += 1
        
        return tables
    
    def get_table_statistics(self, worksheet: Worksheet) -> Dict[str, int]:
        """
        Get statistics about tables in the worksheet.
        
        Args:
            worksheet: The worksheet to analyze
            
        Returns:
            Dictionary with table statistics
        """
        header_rows = self._find_header_rows(worksheet)
        tables_found = 0
        rows_to_remove = 0
        
        for header_row in header_rows:
            formula_col = self._find_formula_column(worksheet, header_row)
            if formula_col:
                formula_row = self._find_formula_row(worksheet, header_row)
                if formula_row:
                    tables_found += 1
                    rows_to_remove += (formula_row - header_row + 1)
        
        return {
            'total_header_rows': len(header_rows),
            'tables_found': tables_found,
            'rows_to_remove': rows_to_remove
        }
    
    def _process_table_from_header_with_tracking(self, worksheet: Worksheet, header_row: int) -> None:
        """
        Process a table starting from a header row WITH offset tracking.
        
        Args:
            worksheet: The worksheet to process
            header_row: The header row number
        """
        logger.info(f"Processing table starting from header row {header_row} (with tracking)")
        
        # Find any column with a formula (SUM formula)
        formula_col = self._find_formula_column(worksheet, header_row)
        if formula_col is None:
            logger.warning(f"No formula column found for header row {header_row}")
            return
        
        # Find formula row (SUM formula) - now looks for rows with 2+ SUM formulas
        formula_row = self._find_formula_row(worksheet, header_row)
        if formula_row is None:
            logger.warning(f"No formula row found for header row {header_row}")
            return
        
        logger.info(f"Table identified: header at row {header_row}, formula at row {formula_row}")
        
        # Remove rows from header to formula row (with tracking)
        self._remove_row_range_with_tracking(worksheet, header_row, formula_row)
        
        # After deletion, insert 2 empty rows at the original header position (with tracking)
        self._insert_empty_rows_with_tracking(worksheet, header_row, 2)
        
        logger.info(f"Successfully processed table: removed rows {header_row}-{formula_row}, inserted 2 empty rows at position {header_row}")
    
    def _remove_row_range_with_tracking(self, worksheet: Worksheet, start_row: int, end_row: int) -> None:
        """
        Remove a range of rows from the worksheet with offset tracking.
        
        Args:
            worksheet: The worksheet to modify
            start_row: Starting row number (inclusive)
            end_row: Ending row number (inclusive)
        """
        logger.info(f"Preparing to delete rows {start_row} to {end_row} from sheet '{worksheet.title}' (with tracking)")
        
        # CRITICAL: Unmerge all cells in the deletion range BEFORE deleting rows
        self._unmerge_cells_in_range(worksheet, start_row, end_row)
        
        # Calculate how many rows to delete
        rows_to_delete = end_row - start_row + 1
        
        # Log the deletion operation for offset tracking
        if hasattr(self, 'offset_tracker') and self.offset_tracker:
            self.offset_tracker.log_delete_rows(start_row, rows_to_delete, worksheet.title)
        
        # Delete rows from bottom to top to avoid shifting issues
        for i in range(rows_to_delete):
            # Delete the start_row (it will shift up after each deletion)
            worksheet.delete_rows(start_row)
        
        logger.info(f"Successfully deleted {rows_to_delete} rows from {start_row} to {end_row} (tracked)")
    
    def _insert_empty_rows_with_tracking(self, worksheet: Worksheet, position: int, num_rows: int) -> None:
        """
        Insert empty rows at the specified position with offset tracking.
        
        Args:
            worksheet: The worksheet to modify
            position: The row position where to insert (1-based)
            num_rows: Number of empty rows to insert
        """
        # Log the insertion operation for offset tracking
        if hasattr(self, 'offset_tracker') and self.offset_tracker:
            self.offset_tracker.log_insert_rows(position, num_rows, worksheet.title)
        
        for i in range(num_rows):
            worksheet.insert_rows(position)
        
        logger.debug(f"Inserted {num_rows} empty rows at position {position} (tracked)")
    
    def _cleanup_empty_rows_with_tracking(self, worksheet: Worksheet) -> None:
        """
        Remove empty rows at the end of the worksheet with offset tracking.
        
        Args:
            worksheet: The worksheet to clean up
        """
        max_row = worksheet.max_row
        
        # Start from the bottom and work up to find the last non-empty row
        last_non_empty_row = 0
        for row in range(max_row, 0, -1):
            row_has_content = False
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    row_has_content = True
                    break
            
            if row_has_content:
                last_non_empty_row = row
                break
        
        # If we found empty rows at the end, delete them
        if last_non_empty_row < max_row:
            rows_to_delete = max_row - last_non_empty_row
            # Don't delete if we only have 2 or fewer empty rows (might be our inserted rows)
            if rows_to_delete > 2:
                actual_delete = rows_to_delete - 2
                delete_start = last_non_empty_row + 3
                
                # Log the cleanup deletion for offset tracking
                if hasattr(self, 'offset_tracker') and self.offset_tracker:
                    self.offset_tracker.log_delete_rows(delete_start, actual_delete, worksheet.title)
                
                for i in range(actual_delete):
                    worksheet.delete_rows(delete_start)  # Delete from 3rd empty row onwards
                logger.info(f"Cleaned up {actual_delete} empty rows at the end of sheet '{worksheet.title}' (tracked)")
            else:
                logger.info(f"Preserved {rows_to_delete} empty rows at the end of sheet '{worksheet.title}' (likely inserted rows)")


def main():
    """Main function for testing the row processor."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Excel files for row removal')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
    parser.add_argument('--analyze-only', action='store_true', help='Only analyze, don\'t process')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    processor = RowProcessor()
    
    try:
        if args.analyze_only:
            # Just analyze the file
            workbook = load_workbook(args.input_file)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                stats = processor.get_table_statistics(worksheet)
                print(f"\n📊 Sheet: {sheet_name}")
                print(f"  Header rows found: {stats['total_header_rows']}")
                print(f"  Tables found: {stats['tables_found']}")
                print(f"  Rows to remove: {stats['rows_to_remove']}")
        else:
            # Process the file
            output_file = processor.process_excel_file(args.input_file, args.output)
            print(f"✅ Successfully processed file: {output_file}")
        
    except Exception as e:
        print(f"❌ Error processing file: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())