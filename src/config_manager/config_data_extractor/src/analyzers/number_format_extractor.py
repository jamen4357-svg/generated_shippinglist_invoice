"""
Number Format Extractor for Excel Analysis Tool.

This module provides functionality to extract number format information from Excel worksheets,
focusing on data rows while avoiding headers and footers.
"""

from typing import List, Dict, Optional, Tuple, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import numbers

from models.data_models import NumberFormatInfo, HeaderMatch


class NumberFormatExtractor:
    """Extracts number format information from Excel worksheets."""
    
    def __init__(self):
        """Initialize the NumberFormatExtractor."""
        # Common number format mappings for better descriptions
        self.format_descriptions = {
            'General': 'General format',
            '#,##0': 'Whole numbers with comma separators',
            '#,##0.00': 'Decimal numbers (2 places) with comma separators',
            '0': 'Whole numbers without separators',
            '0.00': 'Decimal numbers (2 places) without separators',
            '0.0': 'Decimal numbers (1 place) without separators',
            '0.000': 'Decimal numbers (3 places) without separators',
            '$#,##0.00': 'Currency format (USD)',
            '£#,##0.00': 'Currency format (GBP)',
            '€#,##0.00': 'Currency format (EUR)',
            '0%': 'Percentage (whole numbers)',
            '0.00%': 'Percentage (2 decimal places)',
            'm/d/yyyy': 'Date format (US)',
            'd/m/yyyy': 'Date format (European)',
            'yyyy-mm-dd': 'Date format (ISO)',
            '#,##0_);(#,##0)': 'Accounting format (negative in parentheses)',
            '#,##0_);[Red](#,##0)': 'Accounting format (negative in red)',
        }
    
    def extract_number_formats(self, worksheet: Worksheet, header_positions: List[HeaderMatch], 
                              start_row: int, max_sample_rows: int = 10, mapping_config: Optional[Dict[str, Any]] = None) -> List[NumberFormatInfo]:
        """
        Extract number formats from data rows in the worksheet.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_positions: List of detected header positions
            start_row: The row where data starts (after headers)
            max_sample_rows: Maximum number of data rows to sample for formats
            
        Returns:
            List of NumberFormatInfo objects for each column with number formats
        """
        if not header_positions:
            return []
        
        # Determine data column range based on headers
        min_col, max_col = self._get_data_column_range(header_positions)
        
        # Sample data rows to extract number formats
        number_formats = []
        
        for col in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            column_formats = self._extract_column_formats(
                worksheet, col, start_row, max_sample_rows
            )
            
            if column_formats:
                # Use the most common format for this column
                most_common_format = self._get_most_common_format(column_formats)
                if most_common_format:
                    column_id = self._generate_column_id(col, header_positions, mapping_config)
                    description = self._get_format_description(most_common_format)
                    
                    number_formats.append(NumberFormatInfo(
                        column_id=column_id,
                        excel_format=most_common_format,
                        description=description
                    ))
        
        return number_formats
    
    def _get_data_column_range(self, header_positions: List[HeaderMatch]) -> Tuple[int, int]:
        """Get the column range that contains data based on header positions."""
        if not header_positions:
            return (1, 1)
        
        min_col = min(pos.column for pos in header_positions)
        max_col = max(pos.column for pos in header_positions)
        
        return (min_col, max_col)
    
    def _extract_column_formats(self, worksheet: Worksheet, col: int, start_row: int, 
                               max_sample_rows: int) -> List[str]:
        """Extract number formats from a specific column in data rows."""
        formats = []
        
        for row in range(start_row, min(start_row + max_sample_rows, worksheet.max_row + 1)):
            cell = worksheet.cell(row=row, column=col)
            
            # Skip empty cells
            if cell.value is None or cell.value == '':
                continue
            
            # Only process cells that contain numbers
            try:
                float(cell.value)
                # Get the number format
                format_code = cell.number_format
                if format_code and format_code != 'General':
                    formats.append(format_code)
            except (ValueError, TypeError):
                # Not a number, skip
                continue
        
        return formats
    
    def _get_most_common_format(self, formats: List[str]) -> Optional[str]:
        """Get the most common format from a list of formats."""
        if not formats:
            return None
        
        # Count occurrences of each format
        format_counts = {}
        for fmt in formats:
            format_counts[fmt] = format_counts.get(fmt, 0) + 1
        
        # Return the most common format
        most_common = max(format_counts.items(), key=lambda x: x[1])
        return most_common[0]
    
    def _generate_column_id(self, col: int, header_positions: List[HeaderMatch], mapping_config: Optional[Dict[str, Any]] = None) -> str:
        """Generate a column ID based on column position and header information."""
        # Debug: print what we're looking for
        print(f"[DEBUG] Looking for column {col} in header positions")
        
        # First try to use the mapping config if available
        if mapping_config:
            header_mappings = mapping_config.get('header_text_mappings', {}).get('mappings', {})
            
            # Try to find a matching header for this column
            for header_pos in header_positions:
                if header_pos.column == col:
                    header_text = header_pos.keyword.strip()
                    
                    # Try exact match first
                    if header_text in header_mappings:
                        print(f"[DEBUG] Found exact mapping: '{header_text}' -> {header_mappings[header_text]}")
                        return header_mappings[header_text]
                    
                    # Try case-insensitive match
                    for mapped_header, column_id in header_mappings.items():
                        if mapped_header.lower() == header_text.lower():
                            print(f"[DEBUG] Found case-insensitive mapping: '{header_text}' -> {column_id}")
                            return column_id
        
        # Fallback to the original logic if no mapping config or no match found
        for header_pos in header_positions:
            print(f"[DEBUG] Checking header: col={header_pos.column}, keyword='{header_pos.keyword}'")
            if header_pos.column == col:
                keyword = header_pos.keyword.lower()
                print(f"[DEBUG] Found matching column {col} with keyword: '{keyword}'")
                
                # Map common keywords to column IDs (more specific matches first)
                if 'sqft' in keyword or 'sf' in keyword:
                    print(f"[DEBUG] Matched 'sqft' or 'sf' -> col_qty_sf")
                    return 'col_qty_sf'
                elif 'pcs' in keyword:
                    print(f"[DEBUG] Matched 'pcs' -> col_qty_pcs")
                    return 'col_qty_pcs'
                elif 'quantity' in keyword:
                    print(f"[DEBUG] Matched 'quantity' -> col_qty_pcs")
                    return 'col_qty_pcs'
                elif 'amount' in keyword or 'total' in keyword:
                    print(f"[DEBUG] Matched 'amount' or 'total' -> col_amount")
                    return 'col_amount'
                elif 'price' in keyword or 'unit' in keyword:
                    print(f"[DEBUG] Matched 'price' or 'unit' -> col_unit_price")
                    return 'col_unit_price'
                elif 'gross' in keyword:
                    return 'col_gross'
                elif 'net' in keyword:
                    return 'col_net'
                elif 'cbm' in keyword:
                    return 'col_cbm'
                elif 'description' in keyword or 'desc' in keyword:
                    return 'col_desc'
                elif 'po' in keyword:
                    return 'col_po'
                elif 'item' in keyword:
                    return 'col_item'
        
        # Final fallback to generic column ID
        col_letter = openpyxl.utils.get_column_letter(col).lower()
        print(f"[DEBUG] No match found for column {col}, using fallback: col_{col_letter}")
        return f'col_{col_letter}'
    
    def _get_format_description(self, format_code: str) -> str:
        """Get a human-readable description for a number format code."""
        return self.format_descriptions.get(format_code, f'Custom format: {format_code}')
