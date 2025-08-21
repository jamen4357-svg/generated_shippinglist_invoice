import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
# from openpyxl.worksheet.dimensions import RowDimension # Not strictly needed for access
from typing import Dict, List, Optional, Tuple, Any

center_alignment = Alignment(horizontal='center', vertical='center')# --- store_original_merges FILTERED to ignore merges ABOVE row 16 ---
def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    WARNING: Does NOT store starting coordinates... (rest of docstring unchanged)

    Args: (args unchanged)

    Returns:
        A dictionary where keys are sheet names and values are lists of
        tuples: (col_span, top_left_cell_value, row_height).
        row_height will be None if the original row had default height.
    """
    original_merges = {}
    print("\n" + "="*60)
    print("DEBUG: STARTING MERGE STORAGE PROCESS")
    print("="*60)
    print("Storing original merge horizontal spans, top-left values, and row heights (NO coordinates)...")
    print("  (Ignoring merges that start above row 10)") # Updated filter info
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name] # Type hint for clarity
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_above_16_count = 0 # Counter for this filter

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                # --- Check 1: Skip if multi-row ---
                if max_row != min_row:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - it spans multiple rows ({min_row} to {max_row}).")
                    continue

                # ***** MODIFIED CHECK 2: Only skip merges in the very top area (rows 1-10) *****
                # This preserves signature area merges while still filtering out header merges
                if min_row < 10:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - starts at row {min_row} (above row 10).") # Keep commented unless needed
                    skipped_above_16_count += 1
                    continue
                # ***** END MODIFIED CHECK *****

                # --- If not skipped, proceed to get span, height, value ---
                col_span = max_col - min_col + 1
                row_height = None # Default to None
                try:
                    # Get Row Height
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                    # print(f"    DEBUG Store: Sheet='{sheet_name}', MergeCoord='{merged_range.coord}', StartRow={min_row}, Storing Height={row_height} (Type: {type(row_height)})")

                    # Get Value
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value

                    # Store Data (span, value, height)
                    merges_data.append((col_span, top_left_value, row_height))

                except KeyError:
                     print(f"    Warning: Could not find row dimension for row {min_row} on sheet '{sheet_name}' while getting height. Storing height as None.")
                     try:
                         top_left_value = worksheet.cell(row=min_row, column=min_col).value
                     except Exception as val_e:
                         print(f"    Warning: Also failed to get value for merge at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value as None. Error: {val_e}")
                         top_left_value = None
                     merges_data.append((col_span, top_left_value, None))

                except Exception as e:
                    print(f"    Warning: Could not get value/height for merge starting at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value/height as None. Error: {e}")
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
            print(f"  DEBUG: Stored {len(original_merges[sheet_name])} horizontal merge span/value/height entries for sheet '{sheet_name}'.")
            # Report skipped count for this filter
            if skipped_above_16_count > 0:
                print(f"    DEBUG: (Skipped {skipped_above_16_count} merges starting above row 16)")
            
            # DEBUG: Show what was actually stored
            if merges_data:
                print(f"    DEBUG: Sample stored merges for '{sheet_name}':")
                for i, (col_span, stored_value, stored_height) in enumerate(merges_data[:3]):  # Show first 3
                    print(f"      {i+1}. Span: {col_span}, Value: '{stored_value}', Height: {stored_height}")
                if len(merges_data) > 3:
                    print(f"      ... and {len(merges_data) - 3} more")
            else:
                print(f"    DEBUG: No merges found in valid range for '{sheet_name}'")
        else:
             print(f"  DEBUG: Warning: Sheet '{sheet_name}' specified but not found during merge storage.")
             original_merges[sheet_name] = []
    
    total_merges = sum(len(merges) for merges in original_merges.values())
    print("="*60)
    print(f"DEBUG: MERGE STORAGE COMPLETE - Total merges stored: {total_merges}")
    print("="*60)
    return original_merges

# --- find_and_restore_merges_heuristic remains unchanged (still searches bottom-up, applies stored value/height) ---
def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A10:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights
    by searching for the value within a specified range (default A16:H200).
    This version is silent, with no detailed logging.

    WARNING: This is a HEURISTIC approach... (rest of docstring unchanged)

    Args: (args unchanged)
    """
    print("\n" + "="*60)
    print("DEBUG: STARTING MERGE RESTORATION PROCESS")
    print("="*60)
    print(f"DEBUG: Search range: {search_range_str}")
    print(f"DEBUG: Sheets to process: {processed_sheet_names}")
    
    # Debug: Show what we're trying to restore
    total_merges_to_restore = sum(len(stored_merges.get(sheet, [])) for sheet in processed_sheet_names)
    print(f"DEBUG: Total merges to restore: {total_merges_to_restore}")
    
    for sheet_name in processed_sheet_names:
        if sheet_name in stored_merges and stored_merges[sheet_name]:
            print(f"DEBUG: Sheet '{sheet_name}' has {len(stored_merges[sheet_name])} merges to restore")
            for i, (col_span, stored_value, stored_height) in enumerate(stored_merges[sheet_name][:2]):  # Show first 2
                print(f"  {i+1}. Looking for value: '{stored_value}' with span: {col_span}")
        else:
            print(f"DEBUG: Sheet '{sheet_name}' has no merges to restore")

    # These counters are still used by the logic but are no longer printed.
    restored_count = 0
    failed_count = 0
    skipped_count = 0
    skipped_duplicate_value_count = 0

    # --- Define search boundaries (critical errors are still reported) ---
    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
    except TypeError as te:
        print(f"Error processing search range '{search_range_str}'. Check openpyxl version compatibility or range format. Internal error: {te}")
        traceback.print_exc()
        return
    except Exception as e:
        print(f"Error: Invalid search range string '{search_range_str}'. Cannot proceed with restoration. Error: {e}")
        return

    # --- Loop through sheets ---
    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            successfully_restored_values_on_sheet = set()
            
            print(f"\nDEBUG: Processing sheet '{sheet_name}' with {len(original_merges_data)} merges to restore")

            # --- Loop through stored merge info ---
            for merge_idx, (col_span, stored_value, stored_height) in enumerate(original_merges_data):
                print(f"  DEBUG: Merge {merge_idx+1}/{len(original_merges_data)} - Looking for value: '{stored_value}', span: {col_span}")

                if col_span <= 1:
                    print(f"    DEBUG: Skipping merge - span too small ({col_span})")
                    skipped_count += 1
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    print(f"    DEBUG: Skipping merge - value already restored on this sheet")
                    skipped_duplicate_value_count += 1
                    continue

                found = False
                # --- Search range loop - ROW SEARCH REVERSED ---
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        if current_val == stored_value:
                            print(f"    DEBUG: Found matching value at row {r}, col {c}")
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1
                            print(f"    DEBUG: Will merge from ({start_row},{start_col}) to ({end_row},{end_col})")

                            # --- Proactively unmerge any conflicting ranges ---
                            merged_ranges_copy = list(worksheet.merged_cells.ranges)
                            for existing_merge in merged_ranges_copy:
                                rows_overlap = (existing_merge.min_row <= end_row) and (existing_merge.max_row >= start_row)
                                cols_overlap = (existing_merge.min_col <= end_col) and (existing_merge.max_col >= start_col)

                                if rows_overlap and cols_overlap:
                                    try:
                                        worksheet.unmerge_cells(str(existing_merge))
                                    except Exception:
                                        # Fails silently as requested
                                        pass

                            # --- Apply the new merge, Row Height, AND Value ---
                            try:
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                                print(f"    DEBUG: Successfully merged cells")

                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                        print(f"    DEBUG: Set row height to {stored_height}")
                                    except Exception as e:
                                        print(f"    DEBUG: Failed to set row height: {e}")

                                top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                top_left_cell_to_set.value = stored_value
                                print(f"    DEBUG: Set cell value to '{stored_value}'")

                                successfully_restored_values_on_sheet.add(stored_value)
                                restored_count += 1
                                found = True
                                print(f"    DEBUG: ✅ Merge restored successfully!")
                                break

                            except Exception as e:
                                print(f"    DEBUG: ❌ Failed to apply merge: {e}")
                                failed_count += 1
                                found = True
                                break

                    if found:
                        break

                if not found:
                    print(f"    DEBUG: ❌ Could not find value '{stored_value}' in search range")
                    if stored_value not in successfully_restored_values_on_sheet:
                        failed_count += 1

    print("\n" + "="*60)
    print("DEBUG: MERGE RESTORATION COMPLETE")
    print("="*60)
    print(f"DEBUG: Results summary:")
    print(f"  ✅ Restored: {restored_count}")
    print(f"  ❌ Failed: {failed_count}")
    print(f"  ⏭️  Skipped (span <= 1): {skipped_count}")
    print(f"  ⏭️  Skipped (duplicate value): {skipped_duplicate_value_count}")
    print(f"  📊 Total processed: {restored_count + failed_count + skipped_count + skipped_duplicate_value_count}")
    print("="*60)


def force_unmerge_from_row_down(worksheet: Worksheet, start_row: int):
    """
    Forcefully unmerges all cells that start on or after a specific row.

    This is the ideal way to clean a 'data area' while leaving a
    'header area' completely untouched.

    Args:
        worksheet: The openpyxl worksheet object to modify.
        start_row: The row number from which to start unmerging. All merges
                   at this row or any row below it will be removed.
    """
    print(f"--- Selectively unmerging cells from row {start_row} downwards on sheet '{worksheet.title}' ---")
    
    # Create a copy of the list to avoid issues while modifying it
    all_merged_ranges = list(worksheet.merged_cells.ranges)
    unmerged_count = 0
    
    for merged_range in all_merged_ranges:
        # The key condition: only unmerge if the merge starts in the target zone.
        if merged_range.min_row >= start_row:
            try:
                worksheet.unmerge_cells(str(merged_range))
                unmerged_count += 1
            except Exception:
                pass # Ignore errors, as the goal is a clean slate anyway
    
    if unmerged_count > 0:
        print(f"--- Removed {unmerged_count} merges from the data area (row {start_row}+) ---")
    else:
        print(f"--- No merges found in the data area (row {start_row}+) to remove ---")

def apply_row_merges(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on a dictionary of rules.
    This is the only function needed for your request.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to apply merges to.
        num_cols: The total number of columns in the table for validation.
        merge_rules: Dictionary where keys are the starting column index (as a string)
                     and values are the number of columns to span (colspan).
    """
    # Exit if there are no rules to apply
    if not merge_rules:
        return

    print(f"  Applying custom merge rules for row {row_num}...")
    for start_col_str, colspan_val in merge_rules.items():
        try:
            start_col = int(start_col_str)
            colspan = int(colspan_val)

            # Skip if the rule is invalid (e.g., merging 1 or fewer columns)
            if start_col < 1 or colspan <= 1:
                continue

            # Calculate the end column and ensure it doesn't exceed the table's width
            end_col = start_col + colspan - 1
            if end_col > num_cols:
                end_col = num_cols

            # Perform the merge and apply center alignment
            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            cell = worksheet.cell(row=row_num, column=start_col)
            cell.alignment = center_alignment
            print(f"    - Merged row {row_num} from column {start_col} to {end_col}.")

        except (ValueError, TypeError):
            # Ignore if the rule is badly formatted in the JSON (e.g., "A": 5)
            continue


def merge_vertical_cells_in_range(worksheet: Worksheet, scan_col: int, start_row: int, end_row: int):
    """
    Scans a single column and merges adjacent cells that have the same value.

    Args:
        worksheet: The openpyxl Worksheet object.
        scan_col: The 1-based column index to scan and merge.
        start_row: The 1-based starting row index.
        end_row: The 1-based ending row index.
    """
    if not all(isinstance(i, int) and i > 0 for i in [scan_col, start_row, end_row]) or start_row >= end_row:
        return

    row_idx = start_row
    while row_idx < end_row:
        start_of_merge_row = row_idx
        cell_to_match = worksheet.cell(row=start_of_merge_row, column=scan_col)
        value_to_match = cell_to_match.value

        # Skip merging for empty cells
        if value_to_match is None:
            row_idx += 1
            continue

        # Scan downwards to find how many cells match
        end_of_merge_row = start_of_merge_row
        for next_row_idx in range(start_of_merge_row + 1, end_row + 1):
            next_cell = worksheet.cell(row=next_row_idx, column=scan_col)
            if next_cell.value == value_to_match:
                end_of_merge_row = next_row_idx
            else:
                break  # Stop when a different value is found

        # If a sequence of 2 or more was found, perform the merge
        if end_of_merge_row > start_of_merge_row:
            try:
                worksheet.merge_cells(
                    start_row=start_of_merge_row,
                    start_column=scan_col,
                    end_row=end_of_merge_row,
                    end_column=scan_col
                )
                # Apply center alignment to the merged cell
                cell_to_match.alignment = center_alignment
            except Exception:
                pass  # Fails silently

        # Move the main index past the just-scanned range
        row_idx = end_of_merge_row + 1


# ============================================================================
# EMPTY MERGE OFFSET TRACKING SYSTEM
# ============================================================================

class MergeOffsetTracker:
    """
    Tracks row operations to calculate position offsets for empty merge restoration.
    """
    
    def __init__(self):
        self.operations = []  # List of (operation_type, position, count, sheet_name)
        self.debug = True
    
    def log_delete_rows(self, start_row: int, count: int, sheet_name: str):
        """Log a row deletion operation."""
        self.operations.append(('delete', start_row, count, sheet_name))
        if self.debug:
            print(f"[OFFSET] Logged deletion: {count} rows from {start_row} in '{sheet_name}'")
    
    def log_insert_rows(self, position: int, count: int, sheet_name: str):
        """Log a row insertion operation."""
        self.operations.append(('insert', position, count, sheet_name))
        if self.debug:
            print(f"[OFFSET] Logged insertion: {count} rows at {position} in '{sheet_name}'")
    
    def calculate_new_position(self, original_row: int, sheet_name: str) -> int:
        """Calculate the new position of a row after all operations."""
        current_row = original_row
        
        if self.debug:
            print(f"[OFFSET] Calculating position for original row {original_row} in '{sheet_name}'")
        
        for op_type, position, count, op_sheet in self.operations:
            if op_sheet != sheet_name:
                continue  # Skip operations on other sheets
            
            old_row = current_row
            
            if op_type == 'delete':
                # Delete operation: rows from 'position' to 'position + count - 1' are deleted
                delete_start = position
                delete_end = position + count - 1
                
                if current_row < delete_start:
                    # Row is BEFORE deletion range - no effect
                    pass
                elif current_row <= delete_end:
                    # Row is WITHIN deletion range - it gets deleted (invalid position)
                    current_row = -1  # Mark as invalid
                    if self.debug:
                        print(f"[OFFSET] Row {old_row} is WITHIN deletion range [{delete_start}-{delete_end}] - DELETED")
                    break  # No point in continuing if row is deleted
                else:
                    # Row is AFTER deletion range - shift up by count
                    current_row -= count
                    if self.debug:
                        print(f"[OFFSET] Delete [{delete_start}-{delete_end}] shifts row {old_row} → {current_row}")
            
            elif op_type == 'insert':
                # Insert operation: 'count' rows inserted at 'position'
                if current_row >= position:
                    # Row is at or after insertion point - shift down
                    current_row += count
                    if self.debug:
                        print(f"[OFFSET] Insert {count} at {position} shifts row {old_row} → {current_row}")
                # If current_row < position, no effect
        
        if self.debug:
            print(f"[OFFSET] Final position: {original_row} → {current_row}")
        
        return current_row


def store_empty_merges_with_coordinates(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Store empty merges (merges with no value) along with their coordinates for offset-based restoration.
    
    Args:
        workbook: The openpyxl workbook
        sheet_names: List of sheet names to process
    
    Returns:
        Dictionary mapping sheet names to lists of empty merge data
    """
    empty_merges = {}
    
    print("\n" + "="*60)
    print("DEBUG: STORING EMPTY MERGES WITH COORDINATES")
    print("="*60)
    
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_empty_merges = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            
            print(f"Processing sheet '{sheet_name}' for empty merges...")
            
            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds
                
                # Only process single-row merges (same as existing system)
                if max_row != min_row:
                    continue
                
                # Only process merges in our target area (row 10+)
                if min_row < 10:
                    continue
                
                # Get the top-left cell value
                top_left_cell = worksheet.cell(row=min_row, column=min_col)
                cell_value = top_left_cell.value
                
                # Only store if the merge is EMPTY (no value)
                if cell_value is None or str(cell_value).strip() == "":
                    col_span = max_col - min_col + 1
                    
                    # Get row height
                    row_height = None
                    try:
                        row_dim = worksheet.row_dimensions[min_row]
                        row_height = row_dim.height
                    except KeyError:
                        pass
                    
                    empty_merge_data = {
                        'original_row': min_row,
                        'col': min_col,
                        'span': col_span,
                        'height': row_height,
                        'coord': merged_range.coord  # For debugging
                    }
                    
                    sheet_empty_merges.append(empty_merge_data)
                    print(f"  Found empty merge: {merged_range.coord} (row={min_row}, col={min_col}, span={col_span})")
            
            empty_merges[sheet_name] = sheet_empty_merges
            print(f"  Stored {len(sheet_empty_merges)} empty merges for sheet '{sheet_name}'")
        else:
            empty_merges[sheet_name] = []
    
    total_empty_merges = sum(len(merges) for merges in empty_merges.values())
    print("="*60)
    print(f"DEBUG: EMPTY MERGE STORAGE COMPLETE - Total: {total_empty_merges}")
    print("="*60)
    
    return empty_merges


def restore_empty_merges_with_offset(workbook: openpyxl.Workbook, 
                                   empty_merges: Dict[str, List[Dict[str, Any]]], 
                                   offset_tracker: MergeOffsetTracker,
                                   sheet_names: List[str]):
    """
    Restore empty merges using offset calculations.
    
    Args:
        workbook: The openpyxl workbook
        empty_merges: Dictionary of empty merge data from storage
        offset_tracker: The offset tracker with logged operations
        sheet_names: List of sheet names to process
    """
    print("\n" + "="*60)
    print("DEBUG: RESTORING EMPTY MERGES WITH OFFSET CALCULATION")
    print("="*60)
    
    restored_count = 0
    failed_count = 0
    
    for sheet_name in sheet_names:
        if sheet_name not in empty_merges or not empty_merges[sheet_name]:
            continue
        
        worksheet = workbook[sheet_name]
        sheet_merges = empty_merges[sheet_name]
        
        print(f"Restoring {len(sheet_merges)} empty merges in sheet '{sheet_name}'...")
        
        for i, merge_data in enumerate(sheet_merges):
            original_row = merge_data['original_row']
            col = merge_data['col']
            span = merge_data['span']
            height = merge_data['height']
            
            # Calculate new position using offset tracker
            new_row = offset_tracker.calculate_new_position(original_row, sheet_name)
            
            print(f"  Empty merge {i+1}: {original_row} → {new_row} (col={col}, span={span})")
            
            # Validate the new position
            if new_row <= 0:
                if new_row == -1:
                    print(f"    SKIP: Merge was in deleted range (row {original_row})")
                else:
                    print(f"    SKIP: Invalid position {new_row}")
                failed_count += 1
                continue
            
            try:
                # Restore the merge
                end_col = col + span - 1
                worksheet.merge_cells(start_row=new_row, start_column=col, 
                                    end_row=new_row, end_column=end_col)
                
                # Restore row height if available
                if height is not None:
                    worksheet.row_dimensions[new_row].height = height
                
                print(f"    SUCCESS: Restored merge at ({new_row}, {col}) span={span}")
                restored_count += 1
                
            except Exception as e:
                print(f"    FAILED: {e}")
                failed_count += 1
    
    print("="*60)
    print(f"DEBUG: EMPTY MERGE RESTORATION COMPLETE")
    print(f"  Restored: {restored_count}")
    print(f"  Failed: {failed_count}")
    print("="*60)
