"""
Main Excel analysis functionality.

This module provides the ExcelAnalyzer class that coordinates header detection
and font extraction to analyze Excel files and generate analysis results.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Optional
import json
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.data_models import AnalysisResult, SheetAnalysis, FontInfo, HeaderMatch
from analyzers.header_detector import HeaderDetector
from analyzers.font_extractor import FontExtractor
from analyzers.number_format_extractor import NumberFormatExtractor
from analyzers.alignment_extractor import AlignmentExtractor
from analyzers.description_fallback_extractor import DescriptionFallbackExtractor


class ExcelAnalyzer:
    """Main class that coordinates Excel file analysis."""
    
    def __init__(self, quantity_mode: bool = False, mapping_config: Optional[dict] = None):
        """Initialize the ExcelAnalyzer with detector and extractor components.
        
        Args:
            quantity_mode: If True, adds PCS and SQFT columns for packing list sheets
            mapping_config: Optional mapping configuration for column ID mapping
        """
        self.header_detector = HeaderDetector(quantity_mode=quantity_mode, mapping_config=mapping_config)
        self.font_extractor = FontExtractor()
        self.number_format_extractor = NumberFormatExtractor()
        self.alignment_extractor = AlignmentExtractor()
        self.description_fallback_extractor = DescriptionFallbackExtractor()
        self.mapping_config = mapping_config
    
    def analyze_file(self, file_path: str) -> AnalysisResult:
        """
        Analyze an Excel file and extract font information and start row data.
        
        Args:
            file_path: Path to the Excel file to analyze
            
        Returns:
            AnalysisResult containing analysis data for all sheets
            
        Raises:
            FileNotFoundError: If the Excel file doesn't exist
            Exception: If the file cannot be opened or analyzed
        """
        # Validate file exists
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            # Load Excel file using openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Process each worksheet separately
            sheet_analyses = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_analysis = self._analyze_worksheet(worksheet, sheet_name)
                if sheet_analysis:
                    sheet_analyses.append(sheet_analysis)
            
            # Generate AnalysisResult with text output
            return AnalysisResult(
                file_path=str(path.absolute()),
                sheets=sheet_analyses,
                timestamp=datetime.now()
            )
            
        except Exception as e:
            raise Exception(f"Failed to analyze Excel file '{file_path}': {str(e)}")
    
    def _analyze_worksheet(self, worksheet: Worksheet, sheet_name: str) -> Optional[SheetAnalysis]:
        """
        Analyze a single worksheet for headers and font information.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            sheet_name: Name of the worksheet
            
        Returns:
            SheetAnalysis object or None if analysis fails
        """
        try:
            # Find headers using HeaderDetector
            header_positions = self.header_detector.find_headers(worksheet)
            
            # Calculate start row
            start_row = self.header_detector.calculate_start_row(header_positions)
            
            # Extract font information using FontExtractor
            header_font, data_font = self.font_extractor.extract_header_and_data_fonts(
                worksheet, header_positions
            )
            
            # Extract number formats from data rows
            number_formats = self.number_format_extractor.extract_number_formats(
                worksheet, header_positions, start_row, mapping_config=self.mapping_config
            )
            
            # Extract alignments from data rows
            alignments = self.alignment_extractor.extract_alignments(
                worksheet, header_positions, start_row, self.mapping_config
            )
            
            # Extract description fallbacks from merged cells (only for packing list sheets)
            fallbacks = self.description_fallback_extractor.extract_description_fallbacks(
                worksheet, sheet_name, header_positions, start_row
            )
            
            # Detect FOB summary conditions (only for packing list sheets)
            fob_summary_description = self._detect_fob_summary_description(worksheet, sheet_name, header_positions, start_row)
            
            # Detect weight summary conditions (only for invoice sheets)
            weight_summary_enabled = self._detect_weight_summary_enabled(worksheet, sheet_name, header_positions, start_row)
            
            # Create SheetAnalysis object
            return SheetAnalysis(
                sheet_name=sheet_name,
                header_font=header_font,
                data_font=data_font,
                start_row=start_row,
                header_positions=header_positions,
                number_formats=number_formats,
                alignments=alignments,
                fallbacks=fallbacks,
                fob_summary_description=fob_summary_description,
                weight_summary_enabled=weight_summary_enabled
            )
            
        except Exception as e:
            # Log warning but continue with other sheets
            print(f"Warning: Failed to analyze sheet '{sheet_name}': {str(e)}")
            return None
    
    def _detect_fob_summary_description(self, worksheet: Worksheet, sheet_name: str, header_positions: List[HeaderMatch], start_row: int) -> bool:
        """
        Detect FOB summary description conditions in packing list sheets.
        
        Searches for cells containing 'buffalo' text (case insensitive).
        Only applies to sheets named "Packing list".
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            sheet_name: Name of the worksheet
            header_positions: List of detected header positions
            start_row: The start row for data
            
        Returns:
            True if cells with 'buffalo' text are found, False otherwise
        """
        # Only check for packing list sheets
        if sheet_name.lower() != "packing list":
            return False
        
        try:
            # Scan all cells for 'buffalo' text
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            for row in range(start_row, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    # Check if cell contains 'buffalo' (case insensitive)
                    if cell_value and isinstance(cell_value, str) and 'buffalo' in cell_value.lower():
                        print(f"[FOB_SUMMARY_DESC] Found 'buffalo' text in {sheet_name}, cell {chr(64 + col)}{row}: '{cell_value}'")
                        return True
            
            return False
            
        except Exception as e:
            print(f"Warning: Failed to detect FOB summary description in sheet '{sheet_name}': {str(e)}")
            return False
    
    def _detect_weight_summary_enabled(self, worksheet: Worksheet, sheet_name: str, header_positions: List[HeaderMatch], start_row: int) -> bool:
        """
        Detect weight summary conditions in invoice sheets.
        
        Searches for cells containing 'NW(KGS):' text (case insensitive).
        Only applies to sheets named "Invoice".
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            sheet_name: Name of the worksheet
            header_positions: List of detected header positions
            start_row: The start row for data
            
        Returns:
            True if cells with 'NW(KGS):' text are found, False otherwise
        """
        # Only check for invoice sheets
        if sheet_name.lower() != "invoice":
            return False
        
        try:
            # Scan all cells for 'NW(KGS):' text
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            for row in range(1, max_row + 1):  # Start from row 1 to check headers too
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    # Check if cell contains 'NW(KGS):' (case insensitive)
                    if cell_value and isinstance(cell_value, str) and 'nw(kgs):' in cell_value.lower():
                        print(f"[WEIGHT_SUMMARY] Found 'NW(KGS):' text in {sheet_name}, cell {chr(64 + col)}{row}: '{cell_value}'")
                        return True
            
            return False
            
        except Exception as e:
            print(f"Warning: Failed to detect weight summary in sheet '{sheet_name}': {str(e)}")
        return False
    
    def analyze_and_output_text(self, file_path: str) -> str:
        """
        Analyze an Excel file and return results as formatted text.
        
        Args:
            file_path: Path to the Excel file to analyze
            
        Returns:
            Formatted text string with analysis results
        """
        try:
            analysis_result = self.analyze_file(file_path)
            return analysis_result.to_text()
        except Exception as e:
            return f"Error analyzing file: {str(e)}"
    
    def analyze_and_output_json(self, file_path: str, indent: int = 2) -> str:
        """
        Analyze an Excel file and return results as JSON.
        
        Args:
            file_path: Path to the Excel file to analyze
            indent: JSON indentation level
            
        Returns:
            JSON string with analysis results
        """
        try:
            analysis_result = self.analyze_file(file_path)
            return analysis_result.to_json(indent=indent)
        except Exception as e:
            error_result = {
                "error": f"Error analyzing file: {str(e)}",
                "file_path": file_path,
                "timestamp": datetime.now().isoformat()
            }
            return json.dumps(error_result, indent=indent)