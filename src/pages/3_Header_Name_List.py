import streamlit as st
import json
import os
import pandas as pd
from difflib import SequenceMatcher
import openpyxl
import sys
sys.path.append('create_json')
from config import TARGET_HEADERS_MAP

# Load existing headers from JSON (adjust path as needed)
HEADERS_FILE = "config_template_cli/mapping_config.json"
if os.path.exists(HEADERS_FILE):
    with open(HEADERS_FILE, "r", encoding="utf-8") as f:
        config = json.load(f)
    headers = config.get("shipping_list_header_map", {}).get("mappings", {})
else:
    headers = {}

def score_header_match(header_text, candidate_name):
    """Score how well header_text matches candidate_name."""
    if not header_text or not candidate_name:
        return 0
    header_lower = str(header_text).lower().strip()
    candidate_lower = candidate_name.lower()
    if header_lower == candidate_lower:
        return 100
    if header_lower in candidate_lower or candidate_lower in header_lower:
        return 80
    # Partial match
    score = SequenceMatcher(None, header_lower, candidate_lower).ratio() * 60
    return score

def find_best_mapping(header, mappings, threshold=0.7):
    """Find best mapping for header using case-insensitive and partial matching."""
    # Exact match
    if header in mappings:
        return mappings[header]
    # Case insensitive
    header_lower = header.lower()
    for h, m in mappings.items():
        if h.lower() == header_lower:
            return m
    # Partial matching
    best_match = None
    best_score = 0
    for h, m in mappings.items():
        score = SequenceMatcher(None, header_lower, h.lower()).ratio()
        if score > threshold and score > best_score:
            best_match = m
            best_score = score
    return best_match

def find_best_header_row(sheet, max_rows=20):
    """Find the best header row using the algorithm from create_json."""
    best_row = None
    highest_score = 0
    best_mapping = {}
    
    for row_num in range(1, min(max_rows + 1, sheet.max_row + 1)):
        potential_mapping = {}
        row_score = 0
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value:
                best_candidate = None
                best_score = 0
                for canonical, variations in TARGET_HEADERS_MAP.items():
                    for variation in variations:
                        score = score_header_match(str(cell_value), variation)
                        if score > best_score:
                            best_candidate = {'name': canonical}
                            best_score = score
                if best_candidate and best_score > 50:  # Threshold
                    potential_mapping[canonical] = best_candidate['name']
                    row_score += 1
        if row_score > highest_score:
            highest_score = row_score
            best_row = row_num
            best_mapping = potential_mapping
    return best_row, best_mapping

st.title("Shipping List Header Config")

tab1, tab2 = st.tabs(["Process Shipping List", "Edit Mappings"])

with tab1:
    st.header("Process Shipping List")
    uploaded_file = st.file_uploader("Upload Shipping List Excel", type=["xlsx"])
    if uploaded_file is not None:
        # Load workbook
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        
        # Find best header row
        header_row, _ = find_best_header_row(sheet)
        if header_row:
            st.write(f"Detected header row: {header_row}")
            # Extract headers and sample data
            headers_in_file = []
            sample_data = {}
            for col_num in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=header_row, column=col_num).value
                if header_value:
                    header_str = str(header_value)
                    headers_in_file.append(header_str)
                    # Get sample values from next few rows
                    samples = []
                    for row_offset in range(1, 4):  # Next 3 rows
                        sample_row = header_row + row_offset
                        if sample_row <= sheet.max_row:
                            val = sheet.cell(row=sample_row, column=col_num).value
                            samples.append(str(val) if val is not None else "")
                    sample_data[header_str] = samples
        else:
            st.error("No suitable header row found.")
            headers_in_file = []
            sample_data = {}
        
        if headers_in_file:
            # Check mappings with smart algorithm
            mapped = {}
            missing = []
            for header in headers_in_file:
                mapping = find_best_mapping(header, headers)
                if mapping:
                    mapped[header] = mapping
                else:
                    missing.append(header)
            
            st.write("Mapped headers:")
            if mapped:
                st.table(pd.DataFrame(list(mapped.items()), columns=["Header", "Mapping"]))
            else:
                st.write("None mapped.")
            
            if missing:
                st.write("Missing mappings for the following headers:")
                
                # Get possible mapping names
                possible_mappings = list(set(headers.values())) if headers else []
                
                for header in missing:
                    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                    with col1:
                        st.write(f"**{header}**")
                        if header in sample_data:
                            st.write("Samples:", ", ".join(sample_data[header][:3]))
                    with col2:
                        suggested = find_best_mapping(header, headers, threshold=0.5)
                        default_index = possible_mappings.index(suggested) if suggested and suggested in possible_mappings else 0
                        selected_mapping = st.selectbox(
                            "Select Mapping",
                            options=possible_mappings,
                            index=default_index if possible_mappings else 0,
                            key=f"select_{header}"
                        )
                    with col3:
                        st.write("")  # Spacer
                    with col4:
                        if st.button("Add", key=f"add_{header}"):
                            headers[header] = selected_mapping
                            config["shipping_list_header_map"]["mappings"] = headers
                            with open(HEADERS_FILE, "w", encoding="utf-8") as f:
                                json.dump(config, f, indent=4, ensure_ascii=False)
                            st.success(f"Mapping for '{header}' added!")
                            st.rerun()
            else:
                st.success("All headers in the file are mapped!")

with tab2:
    st.header("Edit Mappings")
    # Display and edit current mappings using data_editor
    st.subheader("Current Header Mappings")
    if headers:
        df = pd.DataFrame(list(headers.items()), columns=["Header", "Mapping"])
        edited_df = st.data_editor(df, num_rows="dynamic", key="edit_mappings_editor")
        if st.button("Save Changes"):
            new_headers = dict(zip(edited_df["Header"], edited_df["Mapping"]))
            config["shipping_list_header_map"]["mappings"] = new_headers
            with open(HEADERS_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            st.success("Mappings saved!")
            st.rerun()
    else:
        st.write("No mappings yet.")
        if st.button("Add First Mapping"):
            st.rerun()  # To refresh and show editor
