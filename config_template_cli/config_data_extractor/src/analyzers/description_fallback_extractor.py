"""
Description Fallback Extractor for Excel Analysis Tool.

This module provides functionality to extract fallback description text from
merged cells in packing list sheets.
"""

from typing import List, Optional, Dict, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from models.data_models import HeaderMatch, FallbackInfo


class DescriptionFallbackExtractor:
    """Extracts fallback description text from merged cells in packing list sheets."""

    def __init__(self):
        """Initialize the DescriptionFallbackExtractor."""
        pass

"""
Description Fallback Extractor for Excel Analysis Tool.

This module provides functionality to extract fallback description text from
merged cells in packing list sheets.
"""

from typing import List, Optional, Dict, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from models.data_models import HeaderMatch, FallbackInfo
import re


class DescriptionFallbackExtractor:
    """Extracts fallback description text from merged cells in packing list sheets."""

    def __init__(self):
        """Initialize the DescriptionFallbackExtractor."""
        pass

    def extract_description_fallbacks(self, worksheet: Worksheet, sheet_name: str,
                                    header_positions: List[HeaderMatch],
                                    start_row: int) -> Optional[FallbackInfo]:
        """
        Extract fallback descriptions from the row immediately after headers in the description column.

        Only processes packing list sheets. The logic is:
        1. Find description column
        2. Identify header rows (usually 2 rows)
        3. Skip header rows
        4. Extract text from the next row in description column

        Args:
            worksheet: The openpyxl worksheet to analyze
            sheet_name: Name of the worksheet
            header_positions: List of detected header positions
            start_row: The row where data starts

        Returns:
            FallbackInfo with fallback information or None if no fallbacks found
        """
        # Only process packing list sheets
        if "packing list" not in sheet_name.lower():
            return None

        # Find the description column
        desc_col = self._find_description_column(header_positions)
        if not desc_col:
            return None

        # Find the header rows for the description column
        header_rows = self._find_header_rows_for_column(header_positions, desc_col)
        if not header_rows:
            return None

        # The headers span 2 rows, so the fallback row is start_row + 2
        fallback_row = start_row + 2

        # Extract the fallback text from the description column at the fallback row
        fallback_text = self._extract_cell_text(worksheet, fallback_row, desc_col)
        if not fallback_text or not fallback_text.strip():
            return None

        # Create both versions: original and DAF-filtered
        original_text = fallback_text.strip()
        daf_text = self._filter_english_only(original_text)

        return FallbackInfo(
            column_id='col_desc',
            fallback_texts=[original_text],
            fallback_DAF_texts=[daf_text] if daf_text else [original_text]  # fallback to original if filtering removes everything
        )

    def _filter_english_only(self, text: str) -> str:
        """Filter text to keep only English letters, numbers, and spaces (remove punctuation and non-ASCII)."""
        # Keep only ASCII letters, numbers, and spaces
        filtered = re.sub(r'[^a-zA-Z0-9\s]', '', text)
        # Clean up extra spaces
        filtered = ' '.join(filtered.split())
        return filtered.strip()

    def _find_description_column(self, header_positions: List[HeaderMatch]) -> Optional[int]:
        """Find the column number for the description header."""
        for header_pos in header_positions:
            keyword = header_pos.keyword.lower()
            if 'description' in keyword or 'desc' in keyword:
                return header_pos.column
        return None

    def _find_header_rows_for_column(self, header_positions: List[HeaderMatch], target_col: int) -> List[int]:
        """Find all header rows that contain headers in the target column."""
        header_rows = []
        for header_pos in header_positions:
            if header_pos.column == target_col:
                header_rows.append(header_pos.row)
        return sorted(header_rows)

    def _extract_cell_text(self, worksheet: Worksheet, row: int, col: int) -> str:
        """Extract text from a specific cell."""
        try:
            cell = worksheet.cell(row=row, column=col)
            return str(cell.value) if cell.value is not None else ""
        except Exception:
            return ""