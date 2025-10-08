#!/usr/bin/env python3
"""
Invoice Generator IO - Excel Writer
Clean Excel file operations and management
"""

from pathlib import Path
from typing import Dict, Any, Optional
import shutil
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:
    """Handles Excel file operations for invoice generation"""
    
    def __init__(self, template_file: Path, output_file: Path):
        self.template_file = template_file
        self.output_file = output_file
        self.workbook: Optional[Workbook] = None
        
        self._load_template()
    
    def _load_template(self):
        """Load template and prepare for editing"""
        if not self.template_file.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_file}")
        
        # Copy template to output location
        shutil.copy2(self.template_file, self.output_file)
        
        # Load workbook for editing
        try:
            self.workbook = openpyxl.load_workbook(self.output_file)
        except Exception as e:
            raise Exception(f"Failed to load template {self.template_file}: {e}")
    
    def get_worksheet(self, sheet_name: str) -> Optional[Worksheet]:
        """Get worksheet by name"""
        if not self.workbook:
            return None
        
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        
        return None
    
    def save(self):
        """Save the workbook"""
        if not self.workbook:
            raise RuntimeError("No workbook loaded")
        
        try:
            self.workbook.save(self.output_file)
        except Exception as e:
            raise Exception(f"Failed to save workbook to {self.output_file}: {e}")
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        """Context manager entry"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()
    
    def list_sheets(self) -> list[str]:
        """List all sheet names"""
        if not self.workbook:
            return []
        
        return self.workbook.sheetnames
    
    def sheet_exists(self, sheet_name: str) -> bool:
        """Check if sheet exists"""
        if not self.workbook:
            return False
        
        return sheet_name in self.workbook.sheetnames