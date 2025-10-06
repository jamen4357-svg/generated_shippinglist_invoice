# --- START OF FULL FILE: excel_handler.py ---

import openpyxl
import os
import logging # Using logging is better than print for info/errors

# Basic config moved to main.py, logger will inherit settings
# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelHandler:
    """Handles loading and accessing data from Excel files using openpyxl."""
    def __init__(self, file_path):
        if not os.path.exists(file_path):
            logging.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"The file '{file_path}' was not found.")
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        logging.info(f"Initialized ExcelHandler for: {file_path}")

    def load_sheet(self, sheet_name=None, data_only=True):
        """
        Loads the workbook and a specific sheet.

        Args:
            sheet_name (str, optional): Name of the sheet. Defaults to None (active sheet).
            data_only (bool, optional): Get cell values (True) or formulas (False). Defaults to True.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The loaded sheet object, or None on failure.
        """
        try:
            logging.info(f"Attempting to load workbook '{self.file_path}' with data_only={data_only}")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=data_only)
            active_sheet_title = self.workbook.active.title # Get active sheet title early

            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.sheet = self.workbook[sheet_name]
                    logging.info(f"Successfully loaded specified sheet: '{self.sheet.title}'")
                else:
                    logging.warning(f"Sheet '{sheet_name}' not found in '{self.file_path}'. Loading active sheet: '{active_sheet_title}'")
                    self.sheet = self.workbook.active
            else:
                self.sheet = self.workbook.active
                logging.info(f"No sheet name specified. Successfully loaded active sheet: '{self.sheet.title}'")

            logging.info(f"Sheet dimensions: Max Row={self.sheet.max_row}, Max Col={self.sheet.max_column}")
            
            # Add diagnostic info for MOTO files
            if "MOTO" in str(self.file_path):
                logging.warning(f"MOTO FILE DETECTED - Large max_row might cause performance issues: {self.sheet.max_row}")
            
            return self.sheet
        except FileNotFoundError: # Already handled in __init__, but belt-and-suspenders
             logging.error(f"File not found exception during load: {self.file_path}")
             raise # Re-raise the specific error
        except Exception as e:
            logging.error(f"Failed to load workbook/sheet from '{self.file_path}': {e}", exc_info=True)
            self.workbook = None
            self.sheet = None
            return None

    def get_sheet(self):
        """Returns the currently loaded sheet object."""
        if not self.sheet:
            logging.warning("Sheet not loaded. Call load_sheet() first.")
        return self.sheet

    def close(self):
        """Closes the workbook if it's open."""
        # openpyxl doesn't require explicit closing for reading,
        # but closing the workbook object might release resources sooner.
        if self.workbook:
            try:
                # Although load_workbook doesn't keep the file open,
                # calling close releases the workbook object from memory sooner.
                self.workbook.close()
                logging.info(f"Closed workbook object reference for: {self.file_path}")
            except Exception as e:
                # This shouldn't typically happen on read-only workbooks
                logging.warning(f"Exception while closing workbook reference (this is usually okay): {e}")
            finally:
                 self.workbook = None
                 self.sheet = None


# --- END OF FULL FILE: excel_handler.py ---