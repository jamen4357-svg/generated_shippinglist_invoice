# integration_example.py
"""
Example of how to integrate print area configuration with existing invoice generation.
"""

from pathlib import Path
import sys

# Add current directory to path for imports
sys.path.append(str(Path(__file__).parent))

from openpyxl import load_workbook
from print_area_config import PrintAreaConfig

def integrate_print_config_with_existing_invoice(invoice_path: str) -> None:
    """
    Example of integrating print area configuration with an existing invoice file.

    Args:
        invoice_path: Path to the Excel invoice file
    """
    try:
        # Load existing invoice
        wb = load_workbook(invoice_path)

        # Configure print settings for each worksheet
        config = PrintAreaConfig()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Configuring print settings for sheet: {sheet_name}")

            # Apply print configuration
            config.configure_print_settings(ws)

            print(f"  Print area: {ws.print_area}")
            print(f"  Paper size: A4")
            print(f"  Margins: L={ws.page_margins.left}, R={ws.page_margins.right}")

        # Save the updated invoice
        output_path = Path(invoice_path).stem + "_with_print_config.xlsx"
        wb.save(output_path)

        print(f"\n✅ Invoice saved with print configuration: {output_path}")

    except Exception as e:
        print(f"❌ Error integrating print config: {e}")
        raise

def demo_integration():
    """Demonstrate the integration with a sample invoice."""
    # Create a sample invoice for demonstration
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Add sample invoice data
    ws['A1'] = 'INVOICE'
    ws['A3'] = 'PO Number:'
    ws['B3'] = 'PO001'
    ws['A4'] = 'Date:'
    ws['B4'] = '2025-01-15'

    # Headers
    ws['A6'] = 'Item'
    ws['B6'] = 'Description'
    ws['C6'] = 'Quantity'
    ws['D6'] = 'Unit Price'
    ws['E6'] = 'Amount'

    # Data
    data = [
        ['001', 'Product A', 10, 15.50, 155.00],
        ['002', 'Product B', 5, 25.00, 125.00],
        ['003', 'Product C', 8, 12.75, 102.00],
    ]

    for i, row in enumerate(data, 7):
        ws[f'A{i}'] = row[0]
        ws[f'B{i}'] = row[1]
        ws[f'C{i}'] = row[2]
        ws[f'D{i}'] = row[3]
        ws[f'E{i}'] = row[4]

    # Save sample invoice
    sample_path = "sample_invoice.xlsx"
    wb.save(sample_path)
    print(f"Sample invoice created: {sample_path}")

    # Now integrate print configuration
    integrate_print_config_with_existing_invoice(sample_path)

if __name__ == "__main__":
    print("Print Area Configuration Integration Example")
    print("=" * 50)
    demo_integration()
