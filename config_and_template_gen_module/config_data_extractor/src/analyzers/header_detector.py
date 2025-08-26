"""
Header detection functionality for Excel analysis tool.

This module provides the HeaderDetector class that searches for specific header keywords
and determines the start row for data insertion.
"""

from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from models.data_models import HeaderMatch


class HeaderDetector:
    """Detects header keywords and calculates start row positions."""
    
    # Core header keywords to search for
    HEADER_KEYWORDS = [
        "P.O", "ITEM", "Description", "Quantity", "Amount"
    ]
    
    def __init__(self, quantity_mode: bool = False):
        """Initialize the HeaderDetector.
        
        Args:
            quantity_mode: If True, adds PCS and SQFT columns for packing list sheets
        """
        self.quantity_mode = quantity_mode
    
    def find_headers(self, worksheet: Worksheet) -> List[HeaderMatch]:
        """
        Search for header keywords in the worksheet and record their positions.
        Once a header row is found, extract all headers from that entire row.
        Handles both single-row and two-row headers by checking for merged cells.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            
        Returns:
            List of HeaderMatch objects containing keyword, row, and column positions
        """
        header_matches = []
        header_row_found = None
        
        # First pass: Find any header keyword to identify the header row
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    
                    # Check if cell contains any of our header keywords
                    for keyword in self.HEADER_KEYWORDS:
                        if self._matches_keyword(cell_value, keyword):
                            header_row_found = cell.row
                            break
                    
                    if header_row_found:
                        break
            
            if header_row_found:
                break
        
        # If we found a header row, determine if it's a single or double header
        if header_row_found:
            is_double_header = self._is_double_header(worksheet, header_row_found)
            
            if is_double_header:
                # Extract headers from both rows for double header
                header_matches = self._extract_double_header(worksheet, header_row_found)
            else:
                # Extract headers from single row
                header_matches = self._extract_all_headers_from_row(worksheet, header_row_found)
            
            # Apply quantity mode enhancement if enabled
            if self.quantity_mode:
                header_matches = self._apply_quantity_mode_enhancement(header_matches, worksheet)
        
        return header_matches
    
    def calculate_start_row(self, header_positions: List[HeaderMatch]) -> int:
        """
        Calculate the start row where headers begin.
        
        Args:
            header_positions: List of HeaderMatch objects
            
        Returns:
            The row number where headers start (min_header_row)
        """
        if not header_positions:
            return 1  # Default to row 1 if no headers found
        
        # Find the minimum header row (where headers start)
        min_header_row = min(match.row for match in header_positions)
        return min_header_row
    
    def _extract_all_headers_from_row(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract all non-empty headers from the specified row.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The row number containing headers
            
        Returns:
            List of HeaderMatch objects for all headers in the row
        """
        header_matches = []
        
        # Get the specific row and extract all non-empty cells
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:  # Only include non-empty values
                    header_match = HeaderMatch(
                        keyword=cell_value,  # Use the actual cell value as the keyword
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _apply_quantity_mode_enhancement(self, header_matches: List[HeaderMatch], worksheet: Worksheet) -> List[HeaderMatch]:
        """
        Apply quantity mode enhancement for packing list sheets.
        Adds PCS and SQFT columns after Quantity column.
        
        Args:
            header_matches: Original list of header matches
            worksheet: The worksheet being analyzed
            
        Returns:
            Enhanced list of header matches with PCS and SQFT columns
        """
        # Check if this is a packing list sheet
        sheet_name = worksheet.title.lower()
        if not any(keyword in sheet_name for keyword in ['packing', 'pkl', 'packing list']):
            return header_matches  # Not a packing list, return original
        
        # Find the Quantity column
        quantity_match = None
        for match in header_matches:
            if 'quantity' in match.keyword.lower():
                quantity_match = match
                break
        
        if not quantity_match:
            return header_matches  # No quantity column found
        
        # Create enhanced header list with original headers
        enhanced_headers = header_matches.copy()
        
        # Add PCS and SQFT in the row BELOW the Quantity header
        # PCS: same column as Quantity, but row + 1
        pcs_header = HeaderMatch(
            keyword="PCS",
            row=quantity_match.row + 1,
            column=quantity_match.column
        )
        enhanced_headers.append(pcs_header)
        
        # SQFT: same row as PCS, but next column
        sqft_header = HeaderMatch(
            keyword="SF", 
            row=quantity_match.row + 1,
            column=quantity_match.column + 1
        )
        enhanced_headers.append(sqft_header)
        
        return enhanced_headers
    
    def _is_double_header(self, worksheet: Worksheet, header_row: int) -> bool:
        """
        Check if the header has two rows by examining if the first column is merged.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The row number where header was found
            
        Returns:
            True if this is a double header (first column is merged), False otherwise
        """
        # Get the first column cell (column A) of the header row
        first_cell = worksheet.cell(row=header_row, column=1)
        
        # Check if this cell is part of a merged range
        for merged_range in worksheet.merged_cells.ranges:
            if first_cell.coordinate in merged_range:
                # If the merged range spans multiple rows, it's a double header
                if merged_range.max_row > merged_range.min_row:
                    return True
        
        return False
    
    def _extract_double_header(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract headers from a two-row header structure.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The first row of the header
            
        Returns:
            List of HeaderMatch objects for all headers in both rows
        """
        header_matches = []
        
        # Extract headers from the first row
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        # Extract headers from the second row (header_row + 1)
        second_row = header_row + 1
        for cell in worksheet[second_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _matches_keyword(self, cell_value: str, keyword: str) -> bool:
        """
        Check if a cell value matches a header keyword.
        Uses strict matching to avoid false positives - cell should be primarily the keyword.
        
        Args:
            cell_value: The cell value to check
            keyword: The keyword to match against
            
        Returns:
            True if the cell value is primarily the keyword (case-insensitive)
        """
        cell_lower = cell_value.lower().strip()
        keyword_lower = keyword.lower()
        
        # Exact match
        if cell_lower == keyword_lower:
            return True
            
        # Allow some common variations but keep it strict
        # Remove common punctuation and extra spaces
        import re
        cell_clean = re.sub(r'[^\w\s]', ' ', cell_lower)
        cell_clean = ' '.join(cell_clean.split())  # normalize whitespace
        
        keyword_clean = re.sub(r'[^\w\s]', ' ', keyword_lower)
        keyword_clean = ' '.join(keyword_clean.split())
        
        # Check if cleaned versions match
        if cell_clean == keyword_clean:
            return True
            
        # For very short cells (likely headers), allow if keyword is majority of content
        if len(cell_lower) <= 20 and keyword_lower in cell_lower:
            # Calculate similarity - keyword should be significant portion
            similarity = len(keyword_lower) / len(cell_lower)
            return similarity >= 0.6
        
        return False