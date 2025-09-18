"""
Test to see if HeaderSpanAnalyzer finds actual table headers vs company info
"""

import sys
import os
sys.path.append(os.path.dirname(__file__))

from config_generator.header_span_analyzer import HeaderSpanAnalyzer
import openpyxl


def check_actual_headers():
    """Check what's in the Excel file to see if we're missing table headers."""
    
    excel_file = "../config_data_extractor/CT&INV&PL MT2-25005E DAP.xlsx"
    
    if not os.path.exists(excel_file):
        print("❌ Excel file not found")
        return
    
    print(f"🔍 EXAMINING EXCEL FILE: {excel_file}")
    print("="*60)
    
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=False)
        
        for sheet_name in workbook.sheetnames:
            print(f"\n📋 SHEET: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            print(f"   📊 MERGED CELLS:")
            merged_ranges = list(worksheet.merged_cells.ranges)
            for i, merged_range in enumerate(merged_ranges[:10], 1):  # First 10 only
                top_left = worksheet.cell(merged_range.min_row, merged_range.min_col)
                text = str(top_left.value).strip() if top_left.value else ""
                rowspan = merged_range.max_row - merged_range.min_row + 1
                colspan = merged_range.max_col - merged_range.min_col + 1
                print(f"      [{i}] Row {merged_range.min_row}: '{text}' ({rowspan}×{colspan})")
            
            print(f"   📝 REGULAR CELLS IN HEADER AREA (Rows 1-10):")
            for row in range(1, 11):
                row_data = []
                for col in range(1, 15):  # Columns A-N
                    cell = worksheet.cell(row, col)
                    if cell.value:
                        text = str(cell.value).strip()[:20]  # First 20 chars
                        if text:
                            row_data.append(f"Col{col}:'{text}'")
                if row_data:
                    print(f"      Row {row}: {', '.join(row_data)}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")


def test_with_table_headers():
    """Test if we can detect actual table headers."""
    
    excel_file = "../config_data_extractor/CT&INV&PL MT2-25005E DAP.xlsx"
    
    print(f"\n🎯 TESTING HEADER SPAN ANALYZER")
    print("="*40)
    
    try:
        analyzer = HeaderSpanAnalyzer(excel_file)
        spans = analyzer.analyze_header_spans()
        
        print(f"Current detection focuses on:")
        for sheet_name, sheet_spans in spans.items():
            print(f"\n{sheet_name}:")
            for span in sheet_spans:
                print(f"  - '{span['text'][:30]}...' (spans {span['colspan']} cols)")
        
        print(f"\nℹ️  These are mostly company information headers.")
        print(f"    Table headers like 'Item', 'Description', 'Quantity' might be:")
        print(f"    1. Non-merged (single cells)")
        print(f"    2. In rows below company info")
        print(f"    3. Need different detection logic")
        
    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    check_actual_headers()
    test_with_table_headers()
