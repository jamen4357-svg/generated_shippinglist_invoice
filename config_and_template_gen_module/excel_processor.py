#!/usr/bin/env python3
"""
Excel Content Processor - Handles text replacement and content modification in Excel files.

This module provides functionality to:
1. Replace date patterns with JFTIME
2. Replace Invoice No with JFINV
3. Replace Ref No with JFREF
4. Handle case-insensitive searches
5. Perform circular searches for patterns
"""

import re
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Handles Excel file content processing and text replacement."""
    
    def __init__(self):
        """Initialize the Excel processor with replacement patterns."""
        self.replacement_patterns = {
            'date': {
                'patterns': [
                    r'\b\d{4}-\d{2}-\d{2}\b',  # YYYY-MM-DD
                    r'\b\d{2}/\d{2}/\d{4}\b',  # MM/DD/YYYY
                    r'\b\d{2}-\d{2}-\d{4}\b',  # MM-DD-YYYY
                    r'\b\d{4}/\d{2}/\d{2}\b',  # YYYY/MM/DD
                    r'\b2025\b',  # Specific year 2025
                ],
                'replacement': 'JFTIME'
            },
            'contract_no': {
                'patterns': [
                    r'\bCONTRACT\s*NO\.?\s*:?\b',  # CONTRACT NO.:
                    r'\bCONTRACT\s*#\s*:?\b',
                ],
                'replacement': 'JFINV'
            },
            'invoice_no': {
                'patterns': [
                    r'\bINVOICE\s*NO\s*:?\b',  # INVOICE NO :
                    r'\bINV\s*NO\.?\s*:?\b',
                    r'\bINVOICE\s*#\s*:?\b',
                ],
                'replacement': 'JFINV'
            },
            'ref_no': {
                'patterns': [
                    r'\bREF\s*NO\.?\s*:?\b',  # Ref No.:
                    r'\bREFERENCE\s*NO\.?\s*:?\b',
                    r'\bREF\s*#\s*:?\b',
                ],
                'replacement': 'JFREF'
            },
            'date_label': {
                'patterns': [
                    r'\bDATE\s*:?\b',  # Date:
                ],
                'replacement': 'JFTIME'
            },
            'etd': {
                'patterns': [
                    r'\bETD\s*:?\b',  # ETD:
                    r'\bESTIMATED\s+TIME\s+OF\s+DEPARTURE\s*:?\b',
                ],
                'replacement': 'JFTIME'
            }
        }
    
    def process_excel_file(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Process an Excel file and perform text replacements.
        
        Args:
            input_file: Path to the input Excel file
            output_file: Path for the output file (optional)
            
        Returns:
            Path to the processed Excel file
        """
        try:
            logger.info(f"Processing Excel file: {input_file}")
            
            # Load the workbook
            workbook = load_workbook(input_file)
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                self._process_worksheet(worksheet)
            
            # Determine output file path
            if output_file is None:
                input_path = Path(input_file)
                output_file = str(input_path.parent / f"{input_path.stem}_processed{input_path.suffix}")
            
            # Save the processed workbook
            workbook.save(output_file)
            logger.info(f"Processed file saved to: {output_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise
    
    def _process_worksheet(self, worksheet: Worksheet) -> None:
        """
        Process a single worksheet for text replacements.
        
        Args:
            worksheet: The worksheet to process
        """
        replacements_made = 0
        
        # Process each cell in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    # Check if this cell contains a label we want to replace
                    label_match = self._find_label_match(cell.value)
                    if label_match:
                        # Find the value cell - check next few columns
                        value_cell = None
                        original_value = None
                        
                        # Check the next 3 columns for a value
                        for offset in range(1, 4):
                            check_cell = worksheet.cell(row=cell.row, column=cell.column + offset)
                            if check_cell.value is not None:
                                value_cell = check_cell
                                original_value = str(check_cell.value)
                                break
                        
                        if value_cell is not None:
                            # Replace the value
                            value_cell.value = label_match['replacement']
                            replacements_made += 1
                            logger.debug(f"Found label '{cell.value}' in {cell.coordinate}, replaced value '{original_value}' -> '{label_match['replacement']}' in {value_cell.coordinate}")
        
        logger.info(f"Made {replacements_made} replacements in sheet '{worksheet.title}'")
    
    def _find_label_match(self, text: str) -> Optional[Dict[str, str]]:
        """
        Check if text matches any of our label patterns.
        
        Args:
            text: The text to check
            
        Returns:
            Dictionary with replacement info if match found, None otherwise
        """
        for category, config in self.replacement_patterns.items():
            for pattern in config['patterns']:
                regex = re.compile(pattern, re.IGNORECASE)
                if regex.search(text):
                    return {
                        'category': category,
                        'replacement': config['replacement']
                    }
        return None
    
    def _process_cell_text(self, text: str) -> str:
        """
        Process a single cell's text content for replacements.
        
        Args:
            text: The text content to process
            
        Returns:
            Processed text with replacements applied
        """
        processed_text = text
        
        # Apply each replacement pattern
        for category, config in self.replacement_patterns.items():
            for pattern in config['patterns']:
                # Case-insensitive search
                regex = re.compile(pattern, re.IGNORECASE)
                
                if category in ['date', 'etd', 'date_label']:
                    # For dates, ETD, and date labels, replace the entire match
                    processed_text = regex.sub(config['replacement'], processed_text)
                else:
                    # For invoice_no, ref_no, and contract_no, replace the entire pattern match
                    processed_text = regex.sub(config['replacement'], processed_text)
        
        return processed_text
    
    def circular_search_for_patterns(self, worksheet: Worksheet, pattern: str) -> List[Tuple[str, str]]:
        """
        Perform a circular search for patterns in a worksheet.
        
        Args:
            worksheet: The worksheet to search
            pattern: The pattern to search for
            
        Returns:
            List of tuples containing (cell_coordinate, cell_value)
        """
        matches = []
        regex = re.compile(pattern, re.IGNORECASE)
        
        # Search in a circular pattern (spiral from center)
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 0 or max_col == 0:
            return matches
        
        # Calculate center
        center_row = max_row // 2
        center_col = max_col // 2
        
        # Spiral search pattern
        directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]  # Right, Down, Left, Up
        current_dir = 0
        steps = 1
        step_count = 0
        
        row, col = center_row, center_col
        
        for _ in range(max_row * max_col):
            # Check if cell is within bounds
            if 1 <= row <= max_row and 1 <= col <= max_col:
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, str):
                    if regex.search(cell.value):
                        matches.append((cell.coordinate, cell.value))
            
            # Move in current direction
            row += directions[current_dir][0]
            col += directions[current_dir][1]
            
            step_count += 1
            
            # Change direction when step count reaches current step limit
            if step_count >= steps:
                step_count = 0
                current_dir = (current_dir + 1) % 4
                
                # Increase step limit every 2 direction changes
                if current_dir % 2 == 0:
                    steps += 1
        
        return matches
    
    def validate_replacements(self, original_file: str, processed_file: str) -> Dict[str, int]:
        """
        Validate that replacements were made correctly.
        
        Args:
            original_file: Path to original file
            processed_file: Path to processed file
            
        Returns:
            Dictionary with replacement counts by category
        """
        original_wb = load_workbook(original_file)
        processed_wb = load_workbook(processed_file)
        
        replacement_counts = {
            'date': 0,
            'date_label': 0,
            'contract_no': 0,
            'invoice_no': 0,
            'ref_no': 0,
            'etd': 0
        }
        
        for sheet_name in original_wb.sheetnames:
            if sheet_name in processed_wb.sheetnames:
                orig_sheet = original_wb[sheet_name]
                proc_sheet = processed_wb[sheet_name]
                
                for row in range(1, min(orig_sheet.max_row, proc_sheet.max_row) + 1):
                    for col in range(1, min(orig_sheet.max_column, proc_sheet.max_column) + 1):
                        orig_cell = orig_sheet.cell(row=row, column=col)
                        proc_cell = proc_sheet.cell(row=row, column=col)
                        
                        if (orig_cell.value != proc_cell.value and 
                            orig_cell.value is not None and 
                            proc_cell.value is not None):
                            
                            # Determine what type of replacement was made
                            for category, config in self.replacement_patterns.items():
                                for pattern in config['patterns']:
                                    if re.search(pattern, str(orig_cell.value), re.IGNORECASE):
                                        replacement_counts[category] += 1
                                        break
        
        return replacement_counts


def main():
    """Main function for testing the Excel processor."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Excel files for text replacement')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    processor = ExcelProcessor()
    
    try:
        output_file = processor.process_excel_file(args.input_file, args.output)
        print(f"✅ Successfully processed file: {output_file}")
        
        # Validate replacements
        validation = processor.validate_replacements(args.input_file, output_file)
        print("\n📊 Replacement Summary:")
        for category, count in validation.items():
            if count > 0:
                print(f"  {category}: {count} replacements")
        
    except Exception as e:
        print(f"❌ Error processing file: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main()) 