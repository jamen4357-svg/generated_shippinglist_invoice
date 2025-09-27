import streamlit as st
import os
import sys
from pathlib import Path
import subprocess
import openpyxl
import re
import io
import zipfile
import json
import datetime
import sqlite3
import time
import tempfile
from zoneinfo import ZoneInfo
from auth_wrapper import setup_page_auth

# --- Enhanced Authentication Setup ---
user_info = setup_page_auth(
    page_title="Invoice Generation", 
    page_name="Invoice Generation Suite",
    layout="wide"
)

st.title("Unified Invoice Generation Suite ⚙️📄")

# --- Project Path & Directory Configuration (Unified) ---
# Assuming this script is located in a subfolder of the project root
try:
    PROJECT_ROOT = Path(__file__).resolve().parent.parent
    CREATE_JSON_DIR = PROJECT_ROOT / "create_json"
    INVOICE_GEN_DIR = PROJECT_ROOT / "invoice_gen"
    DATA_DIR = PROJECT_ROOT / "data"
    JSON_OUTPUT_DIR = DATA_DIR / "invoices_to_process"
    TEMP_UPLOAD_DIR = DATA_DIR / "temp_uploads"
    TEMPLATE_DIR = INVOICE_GEN_DIR / "TEMPLATE"
    CONFIG_DIR = INVOICE_GEN_DIR / "config"
    DATA_DIRECTORY = DATA_DIR / 'Invoice Record'
    DATABASE_FILE = DATA_DIRECTORY / 'master_invoice_data.db'
    TABLE_NAME = 'invoices'

    # Create necessary directories
    for dir_path in [JSON_OUTPUT_DIR, TEMP_UPLOAD_DIR, DATA_DIRECTORY, CONFIG_DIR]:
        dir_path.mkdir(parents=True, exist_ok=True)

    # Add script directories to path for imports
    if str(CREATE_JSON_DIR) not in sys.path: sys.path.insert(0, str(CREATE_JSON_DIR))
    if str(INVOICE_GEN_DIR) not in sys.path: sys.path.insert(0, str(INVOICE_GEN_DIR))
    from main import run_invoice_automation # For High-Quality Leather
    from print_area_config import PrintAreaConfig  # Import print area configuration
except (ImportError, IndexError, NameError) as e:
    st.error(f"Error: Could not configure project paths or import necessary scripts. Please check your project's directory structure. Details: {e}")
    st.exception(e)
    st.stop()


# --- Database Initialization (Consolidated) ---
def initialize_database(db_file: Path):
    """
    Initializes the database by creating all required tables and performance indexes
    if they don't already exist. Merged from both original scripts.
    """
    try:
        with sqlite3.connect(db_file) as conn:
            cursor = conn.cursor()
            # Main invoices table
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT, inv_no TEXT, inv_date TEXT,
                inv_ref TEXT UNIQUE, po TEXT, item TEXT, description TEXT, pcs TEXT,
                sqft TEXT, pallet_count TEXT, unit TEXT, amount TEXT, net TEXT,
                gross TEXT, cbm TEXT, production_order_no TEXT, creating_date TEXT,
                status TEXT DEFAULT 'active'
            );
            """)
            # Containers table (from 2nd Layer script)
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS invoice_containers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                inv_ref TEXT NOT NULL,
                container_description TEXT NOT NULL,
                FOREIGN KEY (inv_ref) REFERENCES invoices (inv_ref)
            );
            """)
            # Function-based indexes for fast, case-insensitive lookups
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoices_lower_inv_ref ON invoices (LOWER(inv_ref));")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoices_lower_inv_no ON invoices (LOWER(inv_no));")
            # Index for container lookups
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_inv_ref ON invoice_containers (inv_ref);")
            conn.commit()
        return True
    except sqlite3.Error as e:
        st.error(f"Database Initialization Failed: {e}")
        return False

# --- Initialize Database ---
DB_ENABLED = initialize_database(DATABASE_FILE)
if not DB_ENABLED:
    st.error("Database connection could not be established. The app cannot function correctly.")
    st.stop()


# --- Shared Helper Functions ---
def cleanup_old_files(directories: list, max_age_seconds: int = 3600):
    """Deletes files older than a specified age in a list of directories."""
    cutoff_time = time.time() - max_age_seconds
    for directory in directories:
        if not directory.exists(): continue
        try:
            for filepath in directory.iterdir():
                if filepath.is_file() and filepath.stat().st_mtime < cutoff_time:
                    try:
                        filepath.unlink()
                    except OSError:
                        pass # Ignore if file is locked
        except Exception:
            pass # Ignore directory-level errors

def get_suggested_inv_ref():
    """
    Efficiently suggests the next invoice reference number for the current year
    by querying the database for the maximum existing number.
    """
    # Use Cambodia timezone for current year
    cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
    current_year = datetime.datetime.now(cambodia_tz).strftime('%Y')
    prefix = "INV"
    suggestion = f"{prefix}{current_year}-1"
    if not DB_ENABLED: return suggestion
    try:
        with sqlite3.connect(DATABASE_FILE) as conn:
            cursor = conn.cursor()
            query = f"""
                SELECT inv_ref FROM {TABLE_NAME}
                WHERE inv_ref LIKE ?
                ORDER BY CAST(SUBSTR(inv_ref, INSTR(inv_ref, '-') + 1) AS INTEGER) DESC
                LIMIT 1
            """
            pattern = f"_%{current_year}-%"
            cursor.execute(query, (pattern,))
            last_ref = cursor.fetchone()

            if not last_ref: return suggestion

            match = re.match(r"([a-zA-Z]+)(\d{4})-(\d+)", last_ref[0])
            if match:
                prefix, _, last_num = match.groups()
                return f"{prefix}{current_year}-{int(last_num) + 1}"
            return suggestion
    except sqlite3.Error as e:
        st.warning(f"DB error suggesting Invoice Ref: {e}")
        return suggestion

def check_existing_identifiers(inv_no: str = None, inv_ref: str = None) -> dict:
    """
    Checks for the existence of an invoice number and/or reference in a single,
    fast, indexed database query.
    """
    results = {}
    if not DB_ENABLED or (not inv_no and not inv_ref): return results
    try:
        with sqlite3.connect(DATABASE_FILE) as conn:
            cursor = conn.cursor()
            if inv_no:
                cursor.execute(f"SELECT 1 FROM {TABLE_NAME} WHERE LOWER(inv_no) = LOWER(?) LIMIT 1", (inv_no,))
                if cursor.fetchone(): results['inv_no'] = True
            if inv_ref:
                cursor.execute(f"SELECT 1 FROM {TABLE_NAME} WHERE LOWER(inv_ref) = LOWER(?) LIMIT 1", (inv_ref,))
                if cursor.fetchone(): results['inv_ref'] = True
    except sqlite3.Error as e:
        st.warning(f"DB error checking for existing values: {e}")
    return results

# --- Session State Initialization ---
def reset_hq_workflow_state():
    """Callback to reset the High-Quality tab's state."""
    st.session_state['hq_validation_done'] = False
    st.session_state['hq_json_path'] = None
    st.session_state['hq_missing_fields'] = []
    st.session_state['hq_identifier'] = None

if 'hq_validation_done' not in st.session_state:
    reset_hq_workflow_state()

# --- Create Tabs ---
tab1, tab2 = st.tabs(["For High-Quality Leather", "For 2nd Layer Leather"])


# ==============================================================================
# --- TAB 1: FOR HIGH-QUALITY LEATHER ---
# ==============================================================================
with tab1:
    st.header("High-Quality Leather Invoice Workflow")

    # --- Helper Functions Specific to High-Quality Workflow ---
    def find_incoterm_from_template(identifier: str):
        terms_to_find = ["DAP", "FCA", "CIP"]
        if not identifier: return None
        match = re.match(r'([A-Za-z]+)', identifier)
        if not match: return None
        template_file_path = TEMPLATE_DIR / f"{match.group(1)}.xlsx"
        if not template_file_path.exists(): return None
        try:
            workbook = openpyxl.load_workbook(template_file_path, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for term in terms_to_find:
                            if term in cell.value:
                                workbook.close(); return term
            workbook.close()
        except Exception: pass
        return None

    def validate_json_data(json_path: Path, required_keys: list) -> list:
        if not json_path.exists():
            st.error(f"Validation failed: JSON file '{json_path.name}' not found."); return required_keys
        try:
            with open(json_path, 'r', encoding='utf-8') as f: data = json.load(f)
            missing_or_empty_keys = set(required_keys)
            if 'processed_tables_data' in data and isinstance(data['processed_tables_data'], dict):
                all_tables_data = {k: v for table in data['processed_tables_data'].values() for k, v in table.items()}
                for key in required_keys:
                    if key in all_tables_data and isinstance(all_tables_data[key], list) and any(item is not None and str(item).strip() for item in all_tables_data[key]):
                        missing_or_empty_keys.discard(key)
            return sorted(list(missing_or_empty_keys))
        except (json.JSONDecodeError, Exception) as e:
            st.error(f"Validation failed due to invalid JSON: {e}"); return required_keys

    # --- UI Step 1: Upload ---
    st.subheader("1. Upload Excel File")
    hq_uploaded_file = st.file_uploader("Choose an XLSX file for High-Quality Leather", type="xlsx", key="hq_uploader", on_change=reset_hq_workflow_state)
    
    # --- Processing on Upload ---
    if hq_uploaded_file and not st.session_state.get('hq_validation_done'):
        cleanup_old_files([TEMP_UPLOAD_DIR, JSON_OUTPUT_DIR])
        st.session_state['hq_identifier'] = Path(hq_uploaded_file.name).stem
        temp_file_path = TEMP_UPLOAD_DIR / hq_uploaded_file.name
        
        try:
            with open(temp_file_path, "wb") as f: f.write(hq_uploaded_file.getbuffer())

            with st.spinner("Automatically processing and validating your file..."):
                run_invoice_automation(input_excel_override=str(temp_file_path), output_dir_override=str(JSON_OUTPUT_DIR))
                json_path = JSON_OUTPUT_DIR / f"{st.session_state['hq_identifier']}.json"

                if not json_path.exists():
                    st.error("Processing failed: The JSON data file was not created by the automation script.")
                    st.stop()
                
                required_columns = ['po', 'item', 'pcs', 'sqft', 'pallet_count', 'unit', 'amount', 'net', 'gross', 'cbm', 'production_order_no']
                st.session_state['hq_missing_fields'] = validate_json_data(json_path, required_columns)
                st.session_state['hq_json_path'] = str(json_path)
                st.session_state['hq_validation_done'] = True
                st.rerun()
        except Exception as e:
            st.error(f"An error occurred during initial processing: {e}")
            st.exception(e)
            st.stop()
        finally:
            if temp_file_path.exists(): temp_file_path.unlink()

    # --- UI Steps 2-5: Post-validation ---
    if st.session_state.get('hq_validation_done'):
        st.subheader("✔️ Automatic Validation Complete")
        missing_fields = st.session_state.get('hq_missing_fields', [])
        if missing_fields: st.warning(f"⚠️ **Validation Warning:** Missing fields: **{', '.join(missing_fields)}**.")
        else: st.success("✅ **Validation Complete:** All required data fields are present.")
        st.divider()

        st.subheader("2. Optional Invoice Overrides")
        col1, col2 = st.columns(2)
        suggested_ref = get_suggested_inv_ref()
        with col1:
            user_inv_no = st.text_input("Invoice No", key="hq_user_inv_no")
            user_inv_ref = st.text_input("Invoice Ref", value=suggested_ref, key="hq_user_inv_ref", help="This is automatically suggested. You can override it.")
            tomorrow = datetime.date.today() + datetime.timedelta(days=1)
            selected_date_obj = st.date_input("Invoice Date", value=tomorrow, format="DD/MM/YYYY", key="hq_inv_date")
            user_inv_date = selected_date_obj.strftime("%d/%m/%Y") if selected_date_obj else ""
            
            existing = check_existing_identifiers(inv_no=user_inv_no.strip(), inv_ref=user_inv_ref.strip() if user_inv_ref != suggested_ref else None)
            if existing.get('inv_no'): st.warning(f"⚠️ Invoice No `{user_inv_no.strip()}` already exists.")
            if existing.get('inv_ref'): st.warning(f"⚠️ Invoice Ref `{user_inv_ref.strip()}` already exists.")

        with col2:
            user_container_types = st.text_area("Container / Truck (One per line)", height=150, key="hq_containers")

        st.subheader("3. Select Final Invoice Versions")
        c1, c2, c3 = st.columns(3)
        with c1: gen_normal = st.checkbox("Normal Invoice", value=True, key="hq_normal")
        with c2: gen_DAF = st.checkbox("DAF Version", value=True, key="hq_DAF")
        with c3: gen_combine = st.checkbox("Combine Version", value=True, key="hq_combine")

        st.subheader("4. Generate Final Invoices")
        if st.button("Generate Final Invoices", use_container_width=True, type="primary", key="hq_generate"):
            if not (gen_normal or gen_DAF or gen_combine): st.error("Please select at least one invoice version."); st.stop()

            json_path = Path(st.session_state['hq_json_path'])
            final_user_inv_ref = user_inv_ref if user_inv_ref else suggested_ref
            container_list = [line.strip() for line in user_container_types.split('\n') if line.strip()]

            # Apply Overrides
            with st.spinner("Applying manual overrides..."):
                try:
                    with open(json_path, 'r+', encoding='utf-8') as f:
                        data = json.load(f); was_modified = False
                        
                        # **MODIFICATION START**
                        # Only set creating_date if it doesn't exist (preserve original creation time)
                        cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
                        creating_date_str = datetime.datetime.now(cambodia_tz).strftime("%Y-%m-%d %H:%M:%S")
                        
                        if 'processed_tables_data' in data:
                            for table_data in data['processed_tables_data'].values():
                                num_rows = len(table_data.get('amount', []))
                                if num_rows == 0: continue
                                
                                # Only add creating_date if it doesn't already exist
                                if 'creating_date' not in table_data or not table_data['creating_date']:
                                    table_data['creating_date'] = [creating_date_str] * num_rows; was_modified = True
                                # **MODIFICATION END**
                                
                                if user_inv_no: table_data['inv_no'] = [user_inv_no.strip()] * num_rows; was_modified = True
                                if final_user_inv_ref: table_data['inv_ref'] = [final_user_inv_ref.strip()] * num_rows; was_modified = True
                                if user_inv_date: table_data['inv_date'] = [user_inv_date] * num_rows; was_modified = True
                                if container_list: table_data['container_type'] = [', '.join(container_list)] * num_rows; was_modified = True
                        if was_modified:
                            f.seek(0); json.dump(data, f, indent=4); f.truncate()
                except Exception as e: st.error(f"Error during JSON Override: {e}"); st.stop()

            # Generate Files
            with st.spinner("Generating selected invoice files..."):
                identifier = st.session_state['hq_identifier']
                files_to_zip = [{"name": json_path.name, "data": json_path.read_bytes()}]
                detected_term = find_incoterm_from_template(identifier)
                modes_to_run = []
                if gen_normal: modes_to_run.append((detected_term if detected_term else "normal", []))
                if gen_DAF: modes_to_run.append(("DAF", ["--DAF"]))
                if gen_combine: modes_to_run.append(("combine", ["--custom"]))
                
                success_count = 0
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_dir_path = Path(temp_dir)
                    for mode_name, mode_flags in modes_to_run:
                        final_mode_name = mode_name.upper()
                        if mode_name == 'combine':
                            final_mode_name = f"{(detected_term or '').upper()} COMBINE".strip()
                        
                        output_filename = f"CT&INV&PL {identifier} {final_mode_name}.xlsx"
                        output_path = temp_dir_path / output_filename
                        command = [sys.executable, str(INVOICE_GEN_DIR / "generate_invoice.py"), str(json_path), "--output", str(output_path), "--templatedir", str(TEMPLATE_DIR), "--configdir", str(CONFIG_DIR)] + mode_flags
                        
                        # Set the environment for the subprocess to handle Unicode correctly
                        sub_env = os.environ.copy()
                        sub_env['PYTHONIOENCODING'] = 'utf-8'

                        try:
                            # Add the 'env=sub_env' parameter to the call
                            result = subprocess.run(command, check=True, capture_output=True, text=True, cwd=INVOICE_GEN_DIR, encoding='utf-8', errors='replace', env=sub_env)
                            
                            # Small delay to ensure file is fully written
                            import time
                            time.sleep(0.5)
                            
                            # Apply print area configuration to the generated Excel file
                            try:
                                wb = openpyxl.load_workbook(output_path, read_only=False)

                                # Apply print area configuration to ALL worksheets
                                config = PrintAreaConfig()
                                sheets_processed = []

                                for sheet_name in wb.sheetnames:
                                    try:
                                        ws = wb[sheet_name]
                                        # Skip hidden sheets
                                        if ws.sheet_state != 'visible':
                                            continue

                                        config.configure_print_settings(ws)
                                        sheets_processed.append(sheet_name)
                                    except Exception as sheet_error:
                                        continue

                                wb.save(output_path)
                                wb.close()

                                # Additional delay to ensure save is complete
                                time.sleep(0.2)

                                # Simple success message
                                st.success(f"✅ Print settings applied to {output_filename}")

                            except Exception as print_error:
                                st.error(f"❌ Failed to apply print settings to {output_filename}: {print_error}")
                            
                            # Additional delay before reading file for zipping
                            time.sleep(0.5)
                            
                            files_to_zip.append({"name": output_filename, "data": output_path.read_bytes()})
                            success_count += 1
                        except subprocess.CalledProcessError as e:
                            error_msg = e.stderr.lower() if e.stderr else ""

                            # Check for config-related errors
                            if any(keyword in error_msg for keyword in ['config', 'template', 'not found', 'missing', 'no such file']):
                                st.error("⚠️ **Company Configuration Missing**")
                                st.markdown(f"""
                                <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px; padding: 15px; margin: 10px 0;">
                                <h4 style="color: #856404; margin-top: 0;">🏢 This company's invoice template hasn't been set up yet</h4>
                                <p style="color: #856404; margin-bottom: 10px;">
                                To generate invoices for <strong>{identifier}</strong>, you'll need to:
                                </p>
                                <ol style="color: #856404;">
                                <li>Go to the <strong>🏢 Company Setup Assistant</strong> page</li>
                                <li>Upload the company's Excel invoice template</li>
                                <li>Complete the setup process (takes 5-10 minutes)</li>
                                <li>Return here to generate invoices</li>
                                </ol>
                                </div>
                                """, unsafe_allow_html=True)

                                col_help1, col_help2 = st.columns(2)
                                with col_help1:
                                    if st.button("🏢 Go to Company Setup", key=f"setup_{mode_name}", use_container_width=True):
                                        st.switch_page("pages/2_Config_Generator.py")
                                with col_help2:
                                    with st.expander("🔍 Technical Details"):
                                        st.code(f"Error: {e.stderr}")
                            else:
                                st.error(f"Failed to generate '{final_mode_name}' version. Error: {e.stderr}")
                        except Exception as general_error:
                            st.error(f"❌ Unexpected error for {identifier} - {mode_name}: {str(general_error)}")
            
            # Offer download
            if success_count > 0:
                st.success(f"Successfully created {success_count} invoice file(s) with print settings configured!")
                st.info("📄 **Print Settings Applied:** A4 paper, 0.1\" margins, horizontal centering, page break preview enabled")
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for file_info in files_to_zip: zf.writestr(file_info["name"], file_info["data"])
                st.subheader("5. Download Your Files")
                st.download_button(label=f"📥 Download All Files ({len(files_to_zip)}) as ZIP", data=zip_buffer.getvalue(), file_name=f"Invoices-{identifier}.zip", mime="application/zip", use_container_width=True)
            else:
                st.error("Processing finished, but no files were generated. Check errors above.")

# ==============================================================================
# --- TAB 2: FOR 2ND LAYER LEATHER ---
# ==============================================================================
with tab2:
    st.header("2nd Layer Leather Invoice Workflow")

    # --- Helper Functions Specific to 2nd Layer Workflow ---
    def update_and_aggregate_json(json_path: Path, inv_ref: str, inv_date: datetime.date, unit_price: float, po_number: str) -> dict | None:
        try:
            with open(json_path, 'r+', encoding='utf-8') as f:
                data = json.load(f)
                raw_data = data.get("raw_data", {})
                summary = data.get("aggregated_summary", {})
                
                cambodia_tz = ZoneInfo("Asia/Phnom_Penh")
                creating_date_str = datetime.datetime.now(cambodia_tz).strftime("%Y-%m-%d %H:%M:%S")

                net_value = float(summary.get("net", 0))
                total_pcs = sum(sum(t.get("pcs", [])) for t in raw_data.values())
                total_pallets = sum(len(t.get("pallet_count", [])) for t in raw_data.values())
                total_amount = unit_price * net_value
                date_str = inv_date.strftime("%d/%m/%Y")

                first_item = next((item[0] for table in raw_data.values() if table.get("item") for item in [table["item"]] if item), "N/A")
                first_desc = next((desc[0] for table in raw_data.values() if table.get("description") for desc in [table["description"]] if desc), "N/A")

                for table in raw_data.values():
                    entries = len(table.get("po", []))
                    table.update({"inv_no": [po_number] * entries, "inv_ref": [inv_ref] * entries, "inv_date": [date_str] * entries, "unit": [unit_price] * entries})

                summary.update({
                    "inv_no": po_number, "inv_ref": inv_ref, "inv_date": date_str, "unit": unit_price, "amount": total_amount, 
                    "pcs": total_pcs, "pallet_count": total_pallets, "net": net_value, "creating_date": creating_date_str
                })
                data["aggregated_summary"] = summary
                
                f.seek(0); json.dump(data, f, indent=4); f.truncate()

                return {"po_number": po_number, "amount": total_amount, "pcs": total_pcs, "pallet_count": total_pallets, "net": net_value,
                        "gross": summary.get("gross", 0.0), "cbm": summary.get("cbm", 0.0), "item": first_item, "description": first_desc}
        except Exception as e:
            st.error(f"Failed to update JSON. Details: {e}"); return None

    def get_po_from_json(json_path: Path) -> str | None:
        try:
            with open(json_path, 'r') as f: return str(json.load(f).get("aggregated_summary", {}).get("po", "")).strip() or None
        except Exception: return None

    # --- UI & Processing ---
    st.subheader("1. Upload Source Excel File")
    sl_uploaded_file = st.file_uploader("Choose an XLSX file for 2nd Layer Leather", type="xlsx", key="sl_uploader")

    if sl_uploaded_file:
        st.markdown("---")
        st.subheader("2. Enter Invoice Details")
        col1, col2, col3 = st.columns(3)
        with col1:
            sl_suggested_ref = get_suggested_inv_ref()
            sl_inv_ref = st.text_input("Invoice Reference", value=sl_suggested_ref, key="sl_inv_ref")
            if sl_inv_ref and sl_inv_ref != sl_suggested_ref and check_existing_identifiers(inv_ref=sl_inv_ref).get('inv_ref'): 
                st.warning("Ref already exists in the database.")
        with col2: sl_inv_date = st.date_input("Invoice Date", datetime.date.today(), key="sl_inv_date")
        with col3: sl_unit_price = st.number_input("Unit Price", min_value=0.0, value=0.61, step=0.01, key="sl_unit_price")

        st.markdown("---")
        if st.button(f"Process '{sl_uploaded_file.name}'", use_container_width=True, type="primary", key="sl_process"):
            temp_file_path = TEMP_UPLOAD_DIR / sl_uploaded_file.name
            try:
                with open(temp_file_path, "wb") as f: f.write(sl_uploaded_file.getbuffer())
            except Exception as e:
                st.error(f"Could not save uploaded file: {e}"); st.stop()

            final_json_path = None
            summary_data = None
            
            # Set the environment for all subprocesses
            sub_env = os.environ.copy()
            sub_env['PYTHONIOENCODING'] = 'utf-8'

            # Step 1: Create JSON from Excel
            with st.spinner("Step 1 of 2: Creating data file from Excel..."):
                try:
                    buffer_file = JSON_OUTPUT_DIR / "__buffer.json"
                    cmd = [sys.executable, str(CREATE_JSON_DIR / "Second_Layer(main).py"), str(temp_file_path), "-o", str(buffer_file)]
                    subprocess.run(cmd, check=True, capture_output=True, text=True, cwd=str(CREATE_JSON_DIR), encoding='utf-8', env=sub_env)
                    
                    po_number = get_po_from_json(buffer_file) or Path(sl_uploaded_file.name).stem
                    summary_data = update_and_aggregate_json(buffer_file, sl_inv_ref, sl_inv_date, sl_unit_price, po_number)
                    if not summary_data: st.stop()
                    
                    final_json_path = JSON_OUTPUT_DIR / f"{po_number}.json"
                    buffer_file.replace(final_json_path)
                    st.success(f"Step 1 complete: Data file created as '{final_json_path.name}'.")
                except subprocess.CalledProcessError as e:
                    st.error("Step 1 FAILED."); st.text_area("Full Error Log:", e.stdout + e.stderr, height=200); st.stop()
                finally:
                    if 'buffer_file' in locals() and buffer_file.exists(): buffer_file.unlink()

            # Step 2: Generate documents and ZIP them
            with tempfile.TemporaryDirectory() as temp_output_dir:
                try:
                    with st.spinner("Step 2 of 2: Generating final documents..."):
                        cmd = [sys.executable, str(INVOICE_GEN_DIR / "hybrid_generate_invoice.py"), str(final_json_path),
                                "--outputdir", str(temp_output_dir), "--templatedir", str(TEMPLATE_DIR), "--configdir", str(CONFIG_DIR)]
                        subprocess.run(cmd, check=True, capture_output=True, text=True, cwd=str(INVOICE_GEN_DIR), encoding='utf-8', env=sub_env)
                        
                        # Small delay to ensure all files are fully written
                        import time
                        time.sleep(1.0)
                        
                        st.success("Step 2 complete: Documents generated.")

                    if summary_data:
                        st.markdown("---")
                        st.subheader("Invoice Summary")
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("PO Number", summary_data.get("po_number", "N/A"))
                        c2.metric("Total Amount", f"${summary_data.get('amount', 0):,.2f}")
                        c3.metric("Net Weight (KG)", f"{summary_data.get('net', 0):,.2f}")
                        c4.metric("Gross Weight", f"{summary_data.get('gross', 0):,.2f}")

                        st.subheader("3. Download Generated Documents")
                        generated_files = list(Path(temp_output_dir).glob(f"* {summary_data['po_number']}.xlsx"))
                        
                        # Apply print area configuration to all generated Excel files
                        if generated_files:
                            st.info("📄 Applying print settings to generated Excel files...")
                            total_files_processed = 0
                            total_sheets_processed = 0

                            for excel_file in generated_files:
                                try:
                                    wb = openpyxl.load_workbook(excel_file, read_only=False)

                                    # Apply print area configuration to ALL worksheets
                                    config = PrintAreaConfig()
                                    sheets_processed = []

                                    for sheet_name in wb.sheetnames:
                                        try:
                                            ws = wb[sheet_name]
                                            # Skip hidden sheets
                                            if ws.sheet_state != 'visible':
                                                continue

                                            config.configure_print_settings(ws)
                                            sheets_processed.append(sheet_name)
                                        except Exception as sheet_error:
                                            continue

                                    wb.save(excel_file)
                                    wb.close()

                                    # Additional delay to ensure save is complete
                                    import time
                                    time.sleep(0.2)

                                    total_files_processed += 1
                                    total_sheets_processed += len(sheets_processed)

                                    st.info(f"✅ Print settings applied to {excel_file.name} ({len(sheets_processed)} sheets)")
                                except Exception as print_error:
                                    st.error(f"❌ Failed to apply print settings to {excel_file.name}: {print_error}")

                            st.success(f"📄 **Print Settings Applied:** A4 paper, 0.1\" margins, horizontal centering, page break preview enabled to {total_files_processed} files ({total_sheets_processed} total sheets)")
                        
                        # Additional delay before zipping to ensure all file operations are complete
                        import time
                        time.sleep(1.0)
                        
                        zip_filename = f"{summary_data['po_number']}.zip"
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            for file_path in generated_files: zipf.write(file_path, arcname=file_path.name)
                            if final_json_path and final_json_path.exists(): zipf.write(final_json_path, arcname=final_json_path.name)
                        
                        st.download_button(label=f"Download All Documents and Data (.zip)", data=zip_buffer.getvalue(), file_name=zip_filename, mime="application/zip", use_container_width=True)
                
                except subprocess.CalledProcessError as e:
                    error_msg = (e.stdout + e.stderr).lower() if (e.stdout or e.stderr) else ""
                    
                    # Check for config-related errors
                    if any(keyword in error_msg for keyword in ['config', 'template', 'not found', 'missing', 'no such file']):
                        st.error("⚠️ **Company Configuration Missing**")
                        st.markdown(f"""
                        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px; padding: 15px; margin: 10px 0;">
                        <h4 style="color: #856404; margin-top: 0;">🏢 This company's invoice template hasn't been set up yet</h4>
                        <p style="color: #856404; margin-bottom: 10px;">
                        To generate invoices for this company, you'll need to:
                        </p>
                        <ol style="color: #856404;">
                        <li>Go to the <strong>🏢 Company Setup Assistant</strong> page</li>
                        <li>Upload the company's Excel invoice template</li>
                        <li>Complete the setup process (takes 5-10 minutes)</li>
                        <li>Return here to generate invoices</li>
                        </ol>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        col_help1, col_help2 = st.columns(2)
                        with col_help1:
                            if st.button("🏢 Go to Company Setup", key="setup_2nd_layer", use_container_width=True):
                                st.switch_page("pages/2_Config_Generator.py")
                        with col_help2:
                            with st.expander("🔍 Technical Details"):
                                st.text_area("Full Error Log:", e.stdout + e.stderr, height=300)
                    else:
                        st.error("Step 2 FAILED.")
                        st.text_area("Full Error Log:", e.stdout + e.stderr, height=300)
                    st.stop()
                except Exception as e:
                    st.error(f"An unexpected error occurred: {e}")
                finally:
                    if temp_file_path and temp_file_path.exists(): temp_file_path.unlink()