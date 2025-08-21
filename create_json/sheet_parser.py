# --- START OF FULL REFACTORED FILE: sheet_parser.py ---

import re
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation

# Import config values, now including the new pattern-matching configs
from config import (
    TARGET_HEADERS_MAP,
    HEADER_SEARCH_ROW_RANGE,
    HEADER_SEARCH_COL_RANGE,
    HEADER_IDENTIFICATION_PATTERN,
    STOP_EXTRACTION_ON_EMPTY_COLUMN,
    MAX_DATA_ROWS_TO_SCAN,
    DISTRIBUTION_BASIS_COLUMN,
    COLUMNS_TO_DISTRIBUTE,
    # --- Dependencies for the new smart function ---
    EXPECTED_HEADER_DATA_TYPES,
    EXPECTED_HEADER_PATTERNS,
    HEADERLESS_COLUMN_PATTERNS,
    EXPECTED_HEADER_VALUES,
)


# --- NEW: Helper functions for smart validation ---

def _is_numeric(value: Any) -> bool:
    """Helper to check if a value is a number."""
    if value is None:
        return False
    return isinstance(value, (int, float, Decimal))

def _is_string_like(value: Any) -> bool:
    """Helper to check if a value is a non-empty string or a number."""
    if value is None:
        return False
    if isinstance(value, str) and value.strip():
        return True
    if _is_numeric(value):
        return True
    return False

def _matches_any_pattern(value: Any, patterns: Union[str, List[str]]) -> bool:
    """
    Helper to check if a value's string representation matches ANY of the regex patterns in a list.
    """
    # Convert value to a stripped string for reliable matching. Handles numbers, None, etc.
    value_str = str(value or '').strip()
    if not value_str:
        return False

    # Ensure patterns is always a list for iteration
    if isinstance(patterns, str):
        patterns_list = [patterns]
    else:
        patterns_list = patterns

    # Check against each pattern in the list
    for pattern in patterns_list:
        try:
            if re.match(pattern, value_str):
                # If any pattern matches, we return True immediately
                return True
        except re.error as e:
            logging.error(f"[Pattern Check] Invalid regex pattern provided in config '{pattern}': {e}")
            continue # Try the next pattern
            
    # If no patterns matched after checking all of them
    return False


# --- THE NEW SMART HEADER DETECTION FUNCTION ---
def find_and_map_smart_headers(sheet: Worksheet) -> Optional[Tuple[int, Dict[str, str]]]:
    """
    Finds and maps headers using a scoring system. It now evaluates all rows
    in the search range and selects the one with the highest cumulative score,
    making it robust against stray keywords outside the main table.
    """
    prefix = "[find_and_map_smart_headers_v12]" # Version increment
    logging.info(f"{prefix} Starting best-fit header search...")

    best_result: Optional[Tuple[int, Dict[str, str]]] = None
    highest_row_score = 0

    for row_num in range(HEADER_SEARCH_ROW_RANGE[0], HEADER_SEARCH_ROW_RANGE[1] + 1):
        if row_num + 1 > sheet.max_row: continue

        all_column_candidates: Dict[int, List[Dict]] = {}
        for col_num in range(HEADER_SEARCH_COL_RANGE[0], HEADER_SEARCH_COL_RANGE[1] + 1):
            header_cell = sheet.cell(row=row_num, column=col_num)
            header_value = str(header_cell.value or '').strip().upper()

            if header_value:
                candidate_canonicals = [
                    canonical for canonical, aliases in TARGET_HEADERS_MAP.items()
                    if header_value in [str(a).upper() for a in aliases]
                ]
                if not candidate_canonicals: continue

                col_scores = []
                for canonical_name in candidate_canonicals:
                    data_cell = sheet.cell(row=row_num + 1, column=col_num)
                    data_value = data_cell.value
                    score = 0
                    used_strict_value_check = False

                    # --- START: MODIFIED SCORING LOGIC ---
                    allowed_values = EXPECTED_HEADER_VALUES.get(canonical_name)
                    if allowed_values is not None:
                        used_strict_value_check = True
                        processed_data_value = int(data_value) if isinstance(data_value, str) and data_value.isdigit() else data_value
                        if processed_data_value in allowed_values:
                            score = 25 # Increased score for specific value matches
                    
                    if not used_strict_value_check:
                        patterns_to_check = EXPECTED_HEADER_PATTERNS.get(canonical_name)
                        if patterns_to_check:
                            if _matches_any_pattern(data_value, patterns_to_check):
                                score = 15 # Increased score for pattern matches
                        else:
                            allowed_types = EXPECTED_HEADER_DATA_TYPES.get(canonical_name, [])
                            if ('numeric' in allowed_types and _is_numeric(data_value)) or \
                               ('string' in allowed_types and _is_string_like(data_value)):
                                score = 5 # Kept as the baseline score
                    # --- END: MODIFIED SCORING LOGIC ---
                    
                    if score == 0:
                        # --- NEW MERCY RULE ---
                        # If all other checks fail but we did find a header match for a column
                        # that is allowed to be a string, give it a minimal score.
                        # This handles cases where the cell below the header is empty (None).
                        allowed_types = EXPECTED_HEADER_DATA_TYPES.get(canonical_name, [])
                        if 'string' in allowed_types:
                            score = 1 # A minimal "mercy" score to prevent it from being ignored

                    if score > 0:
                        col_scores.append({'score': score, 'name': canonical_name})
                
                if col_scores:
                    all_column_candidates[col_num] = col_scores
            else:
                data_cell = sheet.cell(row=row_num + 1, column=col_num)
                data_value = data_cell.value
                for canonical_name, patterns in HEADERLESS_COLUMN_PATTERNS.items():
                    if _matches_any_pattern(data_value, patterns):
                        all_column_candidates[col_num] = [{'score': 4, 'name': canonical_name}]
                        break

        potential_mapping: Dict[str, str] = {}
        current_row_score = 0
        processed_canonicals = set()
        
        unit_amt_tie_cols = [
            col for col, candidates in all_column_candidates.items()
            if {c['name'] for c in candidates} == {'unit', 'amount'} and all(c['score'] == 5 for c in candidates)
        ]
        if len(unit_amt_tie_cols) == 2:
            col1, col2 = sorted(unit_amt_tie_cols)
            potential_mapping['unit'] = get_column_letter(col1)
            potential_mapping['amount'] = get_column_letter(col2)
            processed_canonicals.update(['unit', 'amount'])
            current_row_score += 10
            del all_column_candidates[col1], all_column_candidates[col2]
        
        for col_num, candidates in sorted(all_column_candidates.items()):
            valid_candidates = [c for c in candidates if c['name'] not in processed_canonicals]
            if not valid_candidates: continue
            best_candidate = sorted(valid_candidates, key=lambda x: x['score'], reverse=True)[0]
            potential_mapping[best_candidate['name']] = get_column_letter(col_num)
            processed_canonicals.add(best_candidate['name'])
            current_row_score += best_candidate['score']

        # --- NEW LOGGING LINE FOR DIAGNOSTICS ---
        logging.info(f"{prefix} Evaluated Row {row_num} | Score: {current_row_score} | Mapping: {potential_mapping}")

        if len(potential_mapping) >= 3 and current_row_score > highest_row_score:
            highest_row_score = current_row_score
            best_result = (row_num, potential_mapping)
            logging.info(f"{prefix} Found new best candidate row at {row_num} with score {highest_row_score}.")

    if best_result:
        logging.info(f"{prefix} SUCCESS: Confirmed header row at {best_result[0]} with final score {highest_row_score}.")
        return best_result

    logging.error(f"{prefix} FAILED: Could not find any row that passed smart validation.")
    return None

def map_columns_to_headers(sheet, header_row: int, col_range: int) -> Dict[str, int]:
    """
    DEPRECATED in favor of find_and_map_smart_headers for primary mapping,
    but can be kept for diagnostics or simple cases.
    """
    pass


def extract_multiple_tables(sheet, header_rows: List[int], column_mapping: Dict[str, str]) -> Dict[int, Dict[str, List[Any]]]:
    """
    Extracts data for multiple tables defined by header_rows using the validated column_mapping.
    """
    if not header_rows or not column_mapping:
        logging.warning("[extract_multiple_tables] No header rows or column mapping provided.")
        return {}

    all_tables_data: Dict[int, Dict[str, List[Any]]] = {}
    stop_col_letter = column_mapping.get(STOP_EXTRACTION_ON_EMPTY_COLUMN)
    prefix = "[extract_multiple_tables]"

    logging.info(f"{prefix} Starting extraction for {len(header_rows)} tables: {header_rows}")

    for i, header_row in enumerate(header_rows):
        table_index = i + 1
        start_data_row = header_row + 1
        
        if i + 1 < len(header_rows):
            max_possible_end_row = header_rows[i + 1]
        else:
            max_possible_end_row = sheet.max_row + 1
            
        scan_limit_row = start_data_row + MAX_DATA_ROWS_TO_SCAN
        end_data_row = min(max_possible_end_row, scan_limit_row)

        if start_data_row >= end_data_row:
            all_tables_data[table_index] = {key: [] for key in column_mapping.keys()}
            continue

        logging.info(f"{prefix} Table {table_index}: Extracting Data Rows {start_data_row} to {end_data_row - 1}")
        current_table_data: Dict[str, List[Any]] = {key: [] for key in column_mapping.keys()}
        
        for current_row in range(start_data_row, end_data_row):
            if stop_col_letter:
                stop_cell_value = sheet[f"{stop_col_letter}{current_row}"].value
                if stop_cell_value is None or (isinstance(stop_cell_value, str) and not stop_cell_value.strip()):
                    logging.info(f"{prefix} Stopping extraction for Table {table_index} at row {current_row}: Empty cell in stop column '{STOP_EXTRACTION_ON_EMPTY_COLUMN}'.")
                    break

            col_letter_to_canonical = {v: k for k, v in column_mapping.items()}
            for col_letter, canonical_name in col_letter_to_canonical.items():
                cell_value = sheet[f"{col_letter}{current_row}"].value
                processed_value = cell_value.strip() if isinstance(cell_value, str) else cell_value
                current_table_data[canonical_name].append(processed_value)

        all_tables_data[table_index] = current_table_data
        logging.info(f"{prefix} Successfully stored {len(current_table_data.get('po',[]))} rows for Table Index {table_index}.")
        
    return all_tables_data


def find_all_header_rows(sheet, search_pattern, row_range, col_range, start_after_row: int = 0) -> List[int]:
    """
    Finds all 1-indexed row numbers containing a header based on a pattern,
    optionally starting the search after a specific row.
    """
    found_rows: set[int] = set()
    try:
        regex = re.compile(search_pattern, re.IGNORECASE)
        start_row = max(row_range[0], start_after_row + 1)
        max_row_to_search = min(row_range[1], sheet.max_row)
        max_col_to_search = min(col_range[1], sheet.max_column)

        if start_row > max_row_to_search:
             return []

        logging.info(
            f"[find_all_header_rows] Searching for additional headers using pattern '{search_pattern}' "
            f"in rows {start_row}-{max_row_to_search}"
        )

        for r_idx in range(start_row, max_row_to_search + 1):
            for c_idx in range(col_range[0], max_col_to_search + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                if cell.value is not None:
                    cell_value_str = str(cell.value).strip()
                    if regex.search(cell_value_str):
                        found_rows.add(r_idx)
                        break
        
        if not found_rows:
            return []

        header_rows = sorted(list(found_rows))
        logging.info(f"[find_all_header_rows] Found {len(header_rows)} additional header rows at: {header_rows}")
        return header_rows

    except Exception as e:
        logging.error(f"[find_all_header_rows] Error finding header rows: {e}", exc_info=True)
        return []

# --- NEWLY ADDED FUNCTION ---
def parse_and_calculate_cbm(cbm_value: Any) -> Optional[Decimal]:
    """
    Parses a CBM value which can be a direct number, a string representation of a number,
    or a string with dimensions separated by 'x' or '*'.
    
    Args:
        cbm_value: The value to parse, e.g., 1.23, "1.23", "1.2*0.8*0.5".

    Returns:
        The calculated CBM as a Decimal, or None if parsing fails.
    """
    if cbm_value is None:
        return None

    # If it's already a number, convert to Decimal and return
    if isinstance(cbm_value, (int, float, Decimal)):
        return Decimal(cbm_value)

    # If it's a string, process it
    if isinstance(cbm_value, str):
        cbm_str = cbm_value.strip().lower() # Standardize to lowercase
        if not cbm_str:
            return None
            
        # Try direct conversion first for simple cases like "1.23"
        try:
            return Decimal(cbm_str)
        except InvalidOperation:
            # If direct conversion fails, it might be a calculation string
            pass

        # Check for multiplication patterns (e.g., "1.2*0.8*0.5" or "1.2 x 0.8 x 0.5")
        try:
            # Split by 'x' or '*'
            parts = re.split(r'\s*[x*]\s*', cbm_str)
            if len(parts) > 1:
                total_cbm = Decimal('1')
                for part in parts:
                    total_cbm *= Decimal(part)
                return total_cbm
        except (InvalidOperation, TypeError):
            logging.warning(f"Could not parse CBM expression: '{cbm_value}'")
            return None

    # Log a warning if the type is unexpected or parsing failed
    logging.warning(f"Unexpected type or format for CBM value: '{cbm_value}'")
    return None

# --- END OF FULL REFACTORED FILE: sheet_parser.py ---