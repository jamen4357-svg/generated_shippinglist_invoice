"""
Header detection functionality for Excel analysis tool.

This module provides the HeaderDetector class that searches for specific header keywords
and determines the start row for data insertion.
"""

from typing import List, Optional, Dict
import json
import os
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from models.data_models import HeaderMatch


class HeaderDetector:
    """Detects header keywords and calculates start row positions."""
    
    def __init__(self, quantity_mode: bool = False, mapping_config: Optional[Dict] = None):
        """Initialize the HeaderDetector.
        
        Args:
            quantity_mode: If True, adds PCS and SQFT columns for packing list sheets
            mapping_config: Optional mapping configuration dictionary
        """
        self.quantity_mode = quantity_mode
        self.mapping_config = mapping_config
        self.header_keywords = self._load_header_keywords()
        # Load exact headers from mapping config if available
        if self.mapping_config:
            self.exact_headers = list(self.mapping_config.get('header_text_mappings', {}).get('mappings', {}).keys())
        else:
            self.exact_headers = []
    
    def _load_header_keywords(self) -> List[str]:
        """Load header keywords from mapping config or use defaults."""
        keywords = set()
        
        # Try to load from mapping config first
        if self.mapping_config:
            try:
                header_mappings = self.mapping_config.get('header_text_mappings', {}).get('mappings', {})
                for header in header_mappings.keys():
                    # Extract base keywords from headers
                    keywords.update(self._extract_keywords_from_header(header))
            except Exception as e:
                import logging
                logging.warning(f"Could not load keywords from mapping config: {e}")
                print(f"Warning: Could not load keywords from mapping config: {e}")
        
        # If no keywords loaded from config, use defaults
        if not keywords:
            keywords.update([
                "P.O", "ITEM", "Description", "Quantity", "Amount",
                "Mark", "Unit price", "Price", "Total", "Weight", "CBM", "Pallet",
                "Remarks", "HS CODE", "Name", "Commodity", "Goods", "Product",
                "PCS", "SF", "No.", "N.W", "G.W", "Net", "Gross", "FCA"
            ])
        
        return list(keywords)
    
    def _extract_keywords_from_header(self, header: str) -> List[str]:
        """Extract meaningful keywords from a header string."""
        import re
        
        keywords = []
        header_lower = header.lower()
        
        # Split by common separators and extract meaningful words
        filtered_words = [word for word in words if word not in generic_words]
        keywords.extend(filtered_words)
        
        # Add common combinations that might appear in headers
        if 'unit' in header_lower and 'price' in header_lower:
            keywords.append('unit price')
        
        # Add common combinations that might appear in headers
        if 'unit' in header_lower and 'price' in header_lower:
            keywords.append('unit price')
        if 'gross' in header_lower and 'weight' in header_lower:
            keywords.append('gross weight')
        if 'net' in header_lower and 'weight' in header_lower:
            keywords.append('net weight')
        if 'p.o' in header_lower:
            keywords.append('p.o')
        if 'hs' in header_lower and 'code' in header_lower:
            keywords.append('hs code')
        
        return keywords
    
    def find_headers(self, worksheet: Worksheet) -> List[HeaderMatch]:
        """
        Search for header keywords in the worksheet and record their positions.
        Once a header row is found, extract all headers from that entire row.
        Handles both single-row and two-row headers by checking for merged cells.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            
        Returns:
            List of HeaderMatch objects containing keyword, row, and column positions
        """
        header_matches = []
        header_matches = []
        # First pass: Find any header keyword to identify the header row
        max_rows_to_check = 30
        header_row_found = None  # Initialize before use
        for idx, row in enumerate(worksheet.iter_rows()):
            if idx >= max_rows_to_check:
                break
            for cell in row:
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    
                    # Check if cell contains any of our header keywords
                    for keyword in self.header_keywords:
                        if self._matches_keyword(cell_value, keyword):
                            header_row_found = cell.row
            if header_row_found:
                break
            if header_row_found:
                break
        if header_row_found:
            is_double_header = self._is_double_header(worksheet, header_row_found)
            
            if is_double_header:
                # Extract headers from both rows for double header
                header_matches = self._extract_double_header(worksheet, header_row_found)
            else:
                # Extract headers from single row
                header_matches = self._extract_all_headers_from_row(worksheet, header_row_found)
            
            # Apply quantity mode enhancement if enabled
            if self.quantity_mode:
                header_matches = self._apply_quantity_mode_enhancement(header_matches, worksheet)
        
        return header_matches
    
    def calculate_start_row(self, header_positions: List[HeaderMatch]) -> int:
        """
        Calculate the start row where headers begin.
        
        Args:
            header_positions: List of HeaderMatch objects
            
        Returns:
            The row number where headers start (min_header_row)
        """
        if not header_positions:
            return 1  # Default to row 1 if no headers found
        
        # Find the minimum header row (where headers start)
        min_header_row = min(match.row for match in header_positions)
        return min_header_row
    
    def _extract_all_headers_from_row(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract all non-empty headers from the specified row.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The row number containing headers
            
        Returns:
            List of HeaderMatch objects for all headers in the row
        """
        header_matches = []
        
        # Get the specific row and extract all non-empty cells
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:  # Only include non-empty values
                    header_match = HeaderMatch(
                        keyword=cell_value,  # Use the actual cell value as the keyword
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _apply_quantity_mode_enhancement(self, header_matches: List[HeaderMatch], worksheet: Worksheet) -> List[HeaderMatch]:
        """
        Apply quantity mode enhancement for packing list sheets.
        Adds PCS and SQFT columns after Quantity column.
        
        Args:
            header_matches: Original list of header matches
            worksheet: The worksheet being analyzed
            
        Returns:
            Enhanced list of header matches with PCS and SQFT columns
        """
        # Check if this is a packing list sheet
        sheet_name = worksheet.title.lower()
        if not any(keyword in sheet_name for keyword in ['packing', 'pkl', 'packing list']):
            return header_matches  # Not a packing list, return original
        
        # Find the Quantity column
        quantity_match = None
        for match in header_matches:
            if 'quantity' in match.keyword.lower():
                quantity_match = match
                break
        
        if not quantity_match:
            return header_matches  # No quantity column found
        
        # Create enhanced header list with original headers
        enhanced_headers = header_matches.copy()
        
        # Add PCS and SQFT in the row BELOW the Quantity header
        # PCS: same column as Quantity, but row + 1
        pcs_header = HeaderMatch(
            keyword="PCS",
            row=quantity_match.row + 1,
            column=quantity_match.column
        )
        enhanced_headers.append(pcs_header)
        
        # SQFT: same row as PCS, but next column
        sqft_header = HeaderMatch(
            keyword="SF", 
            row=quantity_match.row + 1,
            column=quantity_match.column + 1
        )
        enhanced_headers.append(sqft_header)
        
        return enhanced_headers
    
    def _is_double_header(self, worksheet: Worksheet, header_row: int) -> bool:
        """
        Check if the header has two rows by examining if the first column is merged.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The row number where header was found
            
        Returns:
            True if this is a double header (first column is merged), False otherwise
        """
        # Get the first column cell (column A) of the header row
        first_cell = worksheet.cell(row=header_row, column=1)
        
        # Check if this cell is part of a merged range
        for merged_range in worksheet.merged_cells.ranges:
            if first_cell.coordinate in merged_range:
                # If the merged range spans multiple rows, it's a double header
                if merged_range.max_row > merged_range.min_row:
                    return True
        
        return False
    
    def _extract_double_header(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract headers from a two-row header structure.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The first row of the header
            
        Returns:
            List of HeaderMatch objects for all headers in both rows
        """
        header_matches = []
        
        # Extract headers from the first row
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        # Extract headers from the second row (header_row + 1)
        second_row = header_row + 1
        for cell in worksheet[second_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _matches_keyword(self, cell_value: str, keyword: str) -> bool:
        """
        Check if a cell value matches a header keyword.
        Uses strict matching to avoid false positives - cell should be primarily the keyword.
        
        Args:
            cell_value: The cell value to check
            keyword: The keyword to match against
            
        Returns:
            True if the cell value is primarily the keyword (case-insensitive)
        """
        cell_lower = cell_value.lower().strip()
        keyword_lower = keyword.lower()
        
        # Exact match
        if cell_lower == keyword_lower:
            return True
            
        # Allow some common variations but keep it strict
        # Remove common punctuation and extra spaces
        import re
        cell_clean = re.sub(r'[^\w\s]', ' ', cell_lower)
        cell_clean = ' '.join(cell_clean.split())  # normalize whitespace
        
        keyword_clean = re.sub(r'[^\w\s]', ' ', keyword_lower)
        keyword_clean = ' '.join(keyword_clean.split())
        
        # Check if cleaned versions match
        if cell_clean == keyword_clean:
            return True
            
        # For very short cells (likely headers), allow if keyword is majority of content
        if len(cell_lower) <= 20 and keyword_lower in cell_lower:
            # Calculate similarity - keyword should be significant portion
            similarity = len(keyword_lower) / len(cell_lower)
            return similarity >= 0.6
        
        return False