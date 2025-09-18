"""
Excel Height Analyzer - Extract actual row heights from Excel files using openpyxl.

This module provides functionality to extract actual row heights from Excel files
instead of relying on font-based estimations.
"""

from typing import Dict, Optional, Any
from openpyxl import load_workbook
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ExcelHeightAnalyzer:
    """Analyzes and extracts actual row heights from Excel files."""
    
    def __init__(self, excel_file_path: str):
        """
        Initialize the analyzer with an Excel file.
        
        Args:
            excel_file_path: Path to the Excel file to analyze
        """
        self.excel_file_path = Path(excel_file_path)
        self.workbook = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Load the Excel workbook."""
        try:
            self.workbook = load_workbook(self.excel_file_path)
            logger.info(f"[EXCEL_HEIGHT] Loaded workbook: {self.excel_file_path}")
        except Exception as e:
            logger.error(f"[EXCEL_HEIGHT] Failed to load workbook: {e}")
            self.workbook = None
    
    def extract_sheet_row_heights(self, sheet_name: str, header_row: int, 
                                 data_start_row: int, footer_row: Optional[int] = None) -> Dict[str, float]:
        """
        Extract actual row heights from a specific sheet.
        
        Args:
            sheet_name: Name of the sheet to analyze
            header_row: Row number of the header (1-based)
            data_start_row: Row number where data starts (1-based)
            footer_row: Row number of the footer (1-based, optional)
            
        Returns:
            Dictionary with actual row heights: {'header': float, 'data_default': float, 'footer': float}
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            logger.warning(f"[EXCEL_HEIGHT] Sheet '{sheet_name}' not found, using fallback heights")
            return self._get_fallback_heights()
        
        sheet = self.workbook[sheet_name]
        heights = {}
        
        try:
            # Extract header height
            header_height = sheet.row_dimensions[header_row].height
            if header_height:
                heights['header'] = round(header_height, 1)
                logger.info(f"[EXCEL_HEIGHT] {sheet_name} header row {header_row}: {header_height}pt")
            else:
                heights['header'] = 30  # Default if no height set
                logger.warning(f"[EXCEL_HEIGHT] {sheet_name} header row {header_row}: no height set, using default 30pt")
            
            # Extract data row height (sample from first few data rows)
            data_heights = []
            for row in range(data_start_row, min(data_start_row + 3, sheet.max_row + 1)):
                row_height = sheet.row_dimensions[row].height
                if row_height:
                    data_heights.append(row_height)
            
            if data_heights:
                avg_data_height = sum(data_heights) / len(data_heights)
                heights['data_default'] = round(avg_data_height, 1)
                logger.info(f"[EXCEL_HEIGHT] {sheet_name} data rows avg: {avg_data_height}pt")
            else:
                heights['data_default'] = 25  # Default if no height set
                logger.warning(f"[EXCEL_HEIGHT] {sheet_name} data rows: no height set, using default 25pt")
            
            # Extract footer height if provided
            if footer_row:
                footer_height = sheet.row_dimensions[footer_row].height
                if footer_height:
                    heights['footer'] = round(footer_height, 1)
                    logger.info(f"[EXCEL_HEIGHT] {sheet_name} footer row {footer_row}: {footer_height}pt")
                else:
                    heights['footer'] = heights['header']  # Use header height as fallback
                    logger.warning(f"[EXCEL_HEIGHT] {sheet_name} footer row {footer_row}: no height set, using header height {heights['footer']}pt")
            else:
                heights['footer'] = heights['header']  # Use header height as fallback
                logger.info(f"[EXCEL_HEIGHT] {sheet_name} no footer specified, using header height {heights['footer']}pt")
                
        except Exception as e:
            logger.error(f"[EXCEL_HEIGHT] Error extracting heights for {sheet_name}: {e}")
            return self._get_fallback_heights()
        
        return heights
    
    def _get_fallback_heights(self) -> Dict[str, float]:
        """Get fallback height values when Excel extraction fails."""
        return {
            'header': 30.0,
            'data_default': 25.0,
            'footer': 30.0
        }
    
    def find_header_row(self, sheet_name: str, search_keywords: Optional[list] = None) -> Optional[int]:
        """
        Find the header row by searching for typical header keywords.
        
        Args:
            sheet_name: Name of the sheet to search
            search_keywords: List of keywords to search for (optional)
            
        Returns:
            Row number (1-based) of the header row, or None if not found
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return None
        
        sheet = self.workbook[sheet_name]
        
        if not search_keywords:
            search_keywords = ['ITEM', 'DESCRIPTION', 'QUANTITY', 'PRICE', 'AMOUNT', 'P.O', 'PALLET']
        
        for row in range(1, min(30, sheet.max_row + 1)):
            cell_values = []
            for col in range(1, min(15, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    cell_values.append(str(cell.value).strip().upper())
            
            row_text = ' '.join(cell_values)
            if any(keyword.upper() in row_text for keyword in search_keywords):
                logger.info(f"[EXCEL_HEIGHT] Found header row {row} in {sheet_name}: {cell_values[:5]}...")
                return row
        
        logger.warning(f"[EXCEL_HEIGHT] Header row not found in {sheet_name}")
        return None
    
    def find_footer_row(self, sheet_name: str, start_search_row: int = 1) -> Optional[int]:
        """
        Find the footer row by searching for SUM formulas.
        
        Args:
            sheet_name: Name of the sheet to search
            start_search_row: Row to start searching from (1-based)
            
        Returns:
            Row number (1-based) of the footer row, or None if not found
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return None
        
        sheet = self.workbook[sheet_name]
        
        for row in range(start_search_row, min(sheet.max_row + 1, start_search_row + 50)):
            for col in range(1, min(15, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and 'SUM(' in str(cell.value).upper():
                    logger.info(f"[EXCEL_HEIGHT] Found footer row {row} in {sheet_name} with SUM formula")
                    return row
        
        logger.warning(f"[EXCEL_HEIGHT] Footer row not found in {sheet_name}")
        return None
    
    def analyze_sheet_structure(self, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze the complete structure of a sheet including header, data, and footer positions.
        
        Args:
            sheet_name: Name of the sheet to analyze
            
        Returns:
            Dictionary with sheet structure information
        """
        structure = {
            'sheet_name': sheet_name,
            'header_row': None,
            'footer_row': None,
            'heights': None
        }
        
        # Find header row
        header_row = self.find_header_row(sheet_name)
        if header_row:
            structure['header_row'] = header_row
            
            # Find footer row (search after header)
            footer_row = self.find_footer_row(sheet_name, header_row + 1)
            structure['footer_row'] = footer_row
            
            # Extract heights
            data_start_row = header_row + 1
            heights = self.extract_sheet_row_heights(sheet_name, header_row, data_start_row, footer_row)
            structure['heights'] = heights
        
        return structure