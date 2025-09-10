# printer_area_config.py
"""
Print Area Configuration Module for Invoice Generation

This module provides functionality to configure print settings for Excel worksheets,
including dynamic print area detection, A4 paper size, and margin settings.
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple


class PrintAreaConfig:
    """
    Handles print area configuration for Excel worksheets used in invoice generation.
    """

    def __init__(self):
        """Initialize print area configuration with default settings."""
        self.paper_size = "A4"
        self.margin_left = 0.1
        self.margin_right = 0.1
        self.margin_top = 0.75  # Default top margin
        self.margin_bottom = 0.75  # Default bottom margin
        self.center_horizontally = True
        self.center_vertically = False
        self.show_page_breaks = True  # Show page breaks in worksheet view
        self.show_grid_lines = True   # Show grid lines
        self.show_row_col_headers = True  # Show row/column headers

    def configure_print_settings(self, worksheet: Worksheet) -> None:
        """
        Configure all print settings for the worksheet.

        Args:
            worksheet: The openpyxl worksheet to configure
        """
        try:
            # Skip hidden sheets
            if worksheet.sheet_state != 'visible':
                return

            # Set paper size to A4
            self._set_paper_size(worksheet)

            # Set margins
            self._set_margins(worksheet)

            # Set centering options
            self._set_centering(worksheet)

            # Set worksheet view options (including page breaks)
            self._set_worksheet_view(worksheet)

            # Set dynamic print area
            self._set_dynamic_print_area(worksheet)

        except Exception as e:
            raise

    def _set_paper_size(self, worksheet: Worksheet) -> None:
        """Set paper size to A4."""
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = 'portrait'  # Default to portrait

    def _set_margins(self, worksheet: Worksheet) -> None:
        """Set page margins."""
        worksheet.page_margins.left = self.margin_left
        worksheet.page_margins.right = self.margin_right
        worksheet.page_margins.top = self.margin_top
        worksheet.page_margins.bottom = self.margin_bottom

    def _set_centering(self, worksheet: Worksheet) -> None:
        """Set page centering options."""
        worksheet.print_options.horizontalCentered = self.center_horizontally
        worksheet.print_options.verticalCentered = self.center_vertically

    def _set_worksheet_view(self, worksheet: Worksheet) -> None:
        """Set worksheet view options including page breaks display."""
        try:
            # Set view to show page breaks
            if self.show_page_breaks:
                worksheet.sheet_view.view = 'pageBreakPreview'
            else:
                worksheet.sheet_view.view = 'normal'

            # Set grid lines visibility
            worksheet.sheet_view.showGridLines = self.show_grid_lines

            # Set row/column headers visibility
            worksheet.sheet_view.showRowColHeaders = self.show_row_col_headers

        except Exception as e:
            # Continue without failing - view settings are not critical
            pass

    def _set_dynamic_print_area(self, worksheet: Worksheet) -> None:
        """
        Dynamically determine and set the print area based on non-empty cells.

        The print area will include:
        - Columns from the first non-empty column to the last non-empty column
        - Rows from 1 to the last row with any non-null value
        """
        try:
            # Find the boundaries of non-empty data
            min_row, max_row, min_col, max_col = self._find_data_boundaries(worksheet)

            if max_row is None or max_col is None:
                return

            # Convert column numbers to letters
            start_col_letter = get_column_letter(min_col)  # Already 1-based from _find_data_boundaries
            end_col_letter = get_column_letter(max_col)

            # Create print area range (rows are 1-based, columns are 1-based)
            print_area = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"

            # Clear any existing print area first
            if hasattr(worksheet, 'print_area') and worksheet.print_area:
                worksheet.print_area = None

            # Set the print area
            worksheet.print_area = print_area

        except Exception as e:
            raise

    def _find_data_boundaries(self, worksheet: Worksheet) -> Tuple[int, int, int, int]:
        """
        Find the boundaries of non-empty data in the worksheet.

        Returns:
            Tuple of (min_row, max_row, min_col, max_col)
            All values are 1-based indices
        """
        min_row = None
        max_row = None
        min_col = None
        max_col = None

        # Iterate through all cells to find data boundaries
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    row_idx = cell.row
                    col_idx = cell.column
                    
                    # Update boundaries
                    if min_row is None or row_idx < min_row:
                        min_row = row_idx
                    if max_row is None or row_idx > max_row:
                        max_row = row_idx
                    if min_col is None or col_idx < min_col:
                        min_col = col_idx
                    if max_col is None or col_idx > max_col:
                        max_col = col_idx

        return min_row, max_row, min_col, max_col

    def set_custom_print_area(self, worksheet: Worksheet, start_cell: str, end_cell: str) -> None:
        """
        Set a custom print area range.

        Args:
            worksheet: The worksheet to configure
            start_cell: Starting cell (e.g., 'A1')
            end_cell: Ending cell (e.g., 'H50')
        """
        try:
            print_area = f"{start_cell}:{end_cell}"
            worksheet.print_area = print_area
        except Exception as e:
            raise

    def set_print_titles(self, worksheet: Worksheet, title_rows: Optional[str] = None,
                        title_cols: Optional[str] = None) -> None:
        """
        Set rows and/or columns to repeat on each printed page.

        Args:
            worksheet: The worksheet to configure
            title_rows: Rows to repeat (e.g., '1:2' for rows 1-2)
            title_cols: Columns to repeat (e.g., 'A:B' for columns A-B)
        """
        try:
            if title_rows:
                worksheet.print_title_rows = title_rows

            if title_cols:
                worksheet.print_title_cols = title_cols

        except Exception as e:
            raise

    def set_view_options(self, show_page_breaks: bool = True, show_grid_lines: bool = True,
                        show_headers: bool = True) -> None:
        """
        Configure worksheet view display options.

        Args:
            show_page_breaks: Show page breaks in worksheet view
            show_grid_lines: Show grid lines in worksheet
            show_headers: Show row and column headers
        """
        self.show_page_breaks = show_page_breaks
        self.show_grid_lines = show_grid_lines
        self.show_row_col_headers = show_headers


# Convenience function for quick configuration
def configure_print_area(worksheet: Worksheet) -> None:
    """
    Convenience function to quickly configure print settings for a worksheet.

    Args:
        worksheet: The worksheet to configure
    """
    config = PrintAreaConfig()
    config.configure_print_settings(worksheet)


# Example usage function
def example_usage():
    """
    Example of how to use the PrintAreaConfig class.
    """
    from openpyxl import Workbook

    # Create a sample workbook
    wb = Workbook()
    ws = wb.active

    # Add some sample data
    ws['A1'] = 'Header 1'
    ws['B1'] = 'Header 2'
    ws['C1'] = 'Header 3'
    ws['A2'] = 'Data 1'
    ws['B2'] = 'Data 2'
    ws['C2'] = 'Data 3'

    # Configure print settings with page breaks enabled
    config = PrintAreaConfig()
    config.configure_print_settings(ws)

    # Save the workbook
    wb.save('example_with_print_settings.xlsx')


if __name__ == "__main__":
    example_usage()
