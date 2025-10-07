"""
StyleUpdater component for extracting and updating number formats and cell styles.

This module provides functionality to extract number formats from Excel analysis data
and update configuration templates with the actual formatting patterns used in the source Excel files.
"""

from typing import Dict, List, Any, Optional
import copy
import os
from openpyxl import load_workbook
from .models import QuantityAnalysisData, SheetData
from .mapping_manager import MappingManager, MappingManagerError


class StyleUpdaterError(Exception):
    """Custom exception for StyleUpdater errors."""
    pass


class StyleUpdater:
    """Extracts and updates number formats and cell styles in configuration templates."""
    
    def __init__(self, mapping_config_path: str = None):
        """Initialize StyleUpdater with mapping manager."""
        # If no path provided, find it relative to this script
        if mapping_config_path is None:
            from pathlib import Path
            script_dir = Path(__file__).resolve().parent
            # Go up two levels: generate_config/config_generator -> generate_config -> config_template_cli
            base_dir = script_dir.parent.parent
            mapping_config_path = str(base_dir / "mapping_config.json")
        
        # Initialize mapping manager
        try:
            self.mapping_manager = MappingManager(mapping_config_path)
        except MappingManagerError as e:
            print(f"Warning: Could not load mapping config: {e}")
            self.mapping_manager = None
        
        # Store Excel file path for format extraction
        self.excel_file_path = None
    
    def update_alignments(self, template: Dict[str, Any], quantity_data: QuantityAnalysisData) -> Dict[str, Any]:
        """
        Update alignments in styling sections using actual Excel cell alignments.
        
        Args:
            template: Configuration template dictionary
            quantity_data: Quantity analysis data containing alignment information
            
        Returns:
            Updated configuration template
        """
        print("🎯 [ALIGNMENTS] Starting alignment update process...")
        
        for sheet_data in quantity_data.sheets:
            sheet_name = sheet_data.sheet_name
            print(f"📋 [ALIGNMENTS] Processing sheet: {sheet_name}")
            
            # Find the sheet configuration
            sheet_config = self._find_sheet_config(template, sheet_name)
            if not sheet_config:
                print(f"⚠️ [ALIGNMENTS] No configuration found for sheet: {sheet_name}")
                continue
                
            # Check if sheet has styling configuration
            if 'styling' not in sheet_config:
                print(f"⚠️ [ALIGNMENTS] No styling section found for sheet: {sheet_name}")
                continue
                
            styling_config = sheet_config['styling']
            
            # Extract alignments for this sheet
            alignments = self._extract_alignments_from_sheet(sheet_data, sheet_name)
            
            if alignments:
                # Integrate alignments into the new consolidated styles structure
                self._integrate_alignments_into_column_styles(styling_config, alignments, sheet_name)
            else:
                print(f"ℹ️ [ALIGNMENTS] No alignments found for sheet: {sheet_name}")
        
        print("✅ [ALIGNMENTS] Alignment update process completed")
        return template
    
    def update_number_formats(self, template: Dict[str, Any], quantity_data: QuantityAnalysisData) -> Dict[str, Any]:
        """
        Update number formats in styling sections using actual Excel cell formats.
        
        Args:
            template: Configuration template dictionary
            quantity_data: Quantity analysis data containing cell format information
            
        Returns:
            Updated template with number_formats extracted from Excel
            
        Raises:
            StyleUpdaterError: If template structure is invalid or update fails
        """
        print("🔍 [DEBUG] update_number_formats() method called!")
        
        try:
            if not isinstance(template, dict):
                raise StyleUpdaterError("Template must be a dictionary")
            
            if not isinstance(quantity_data, QuantityAnalysisData):
                raise StyleUpdaterError("Quantity data must be QuantityAnalysisData instance")
            
            print(f"🔍 [DEBUG] Processing {len(quantity_data.sheets)} sheets for number formats")
            
            # Create deep copy to avoid modifying original template
            updated_template = copy.deepcopy(template)
            
            # Process each sheet in the template
            data_mapping = updated_template.get('data_mapping', {})
            
            for sheet_data in quantity_data.sheets:
                quantity_sheet_name = sheet_data.sheet_name
                mapped_sheet_name = self._map_sheet_name(quantity_sheet_name)
                
                print(f"🔍 [DEBUG] Processing sheet: {quantity_sheet_name} -> {mapped_sheet_name}")
                
                if mapped_sheet_name not in data_mapping:
                    print(f"🔍 [DEBUG] Sheet {mapped_sheet_name} not found in data_mapping, skipping")
                    continue
                    
                sheet_config = data_mapping[mapped_sheet_name]
                
                print(f"🔍 [DEBUG] Extracting number formats for {mapped_sheet_name}")
                
                # Extract number formats from the sheet data
                number_formats = self._extract_number_formats_from_sheet(sheet_data, mapped_sheet_name)
                
                # Update the styling section with extracted number formats
                if 'styling' not in sheet_config:
                    sheet_config['styling'] = {}
                    print(f"🔍 [DEBUG] Created styling section for {mapped_sheet_name}")
                
                if number_formats:
                    # Instead of adding to styling.number_formats, integrate into column_id_styles
                    self._integrate_number_formats_into_column_styles(sheet_config['styling'], number_formats, mapped_sheet_name)
                    print(f"✅ [NUMBER_FORMATS] Integrated {len(number_formats)} formats into column_id_styles for {mapped_sheet_name}")
                else:
                    print(f"⚠️ [NUMBER_FORMATS] No formats found for {mapped_sheet_name}")
            
            return updated_template
            
        except Exception as e:
            if isinstance(e, StyleUpdaterError):
                raise
            raise StyleUpdaterError(f"Number format update failed: {str(e)}") from e
    
    def _extract_number_formats_from_sheet(self, sheet_data: SheetData, sheet_name: str) -> Dict[str, str]:
        """
        Extract number formats from sheet data or actual Excel file.
        
        First tries to use pre-extracted number format data from config_data_extractor,
        falls back to extracting from Excel file if not available.
        
        Args:
            sheet_data: Sheet data containing number format information
            sheet_name: Name of the sheet for logging
            
        Returns:
            Dictionary mapping column identifiers to number format patterns
        """
        # First, check if we have pre-extracted number format data
        if hasattr(sheet_data, 'number_formats') and sheet_data.number_formats:
            print(f"[NUMBER_FORMATS] Using pre-extracted formats for {sheet_name}: {len(sheet_data.number_formats)} formats")
            
            # Convert NumberFormatInfo objects to dictionary format
            number_formats = {}
            for fmt_info in sheet_data.number_formats:
                number_formats[fmt_info.column_id] = fmt_info.excel_format
                print(f"[NUMBER_FORMATS] {fmt_info.column_id} -> {fmt_info.excel_format} ({fmt_info.description})")
            
            return number_formats
        
        # Fallback: Try to extract from actual Excel file
        try:
            excel_file_path = self._get_excel_file_path()
            
            if excel_file_path and os.path.exists(excel_file_path):
                print(f"[NUMBER_FORMATS] Analyzing Excel file: {excel_file_path}")
                number_formats = self._extract_formats_from_excel_file(excel_file_path, sheet_name, sheet_data)
                
                if number_formats:
                    print(f"[NUMBER_FORMATS] Extracted {len(number_formats)} formats from Excel for {sheet_name}")
                    return number_formats
                else:
                    print(f"[NUMBER_FORMATS] No formats found in Excel for {sheet_name}")
            else:
                print(f"[NUMBER_FORMATS] Excel file not accessible: {excel_file_path}")
                
        except Exception as e:
            print(f"[NUMBER_FORMATS] Error extracting from Excel: {e}")
        
        # Final fallback: Return empty dict
        print(f"[NUMBER_FORMATS] No number formats available for {sheet_name}")
        return {}
    
    def _integrate_number_formats_into_column_styles(self, styling_config: Dict[str, Any],
                                                   number_formats: Dict[str, str],
                                                   sheet_name: str) -> None:
        """
        Integrate extracted number formats into the column_id_styles structure.
        
        Args:
            styling_config: The styling configuration section
            number_formats: Dictionary of column_id -> format_string
            sheet_name: Name of the sheet for logging
        """
        # Ensure the column_id_styles section exists
        if 'column_id_styles' not in styling_config:
            styling_config['column_id_styles'] = {}
            print(f"🔍 [DEBUG] Created column_id_styles section for {sheet_name}")
            
        column_id_styles = styling_config['column_id_styles']
        
        # Integrate each number format into the appropriate column style
        for column_id, format_string in number_formats.items():
            if column_id not in column_id_styles:
                column_id_styles[column_id] = {}
                
            # Add or update the number_format
            column_id_styles[column_id]['number_format'] = format_string
            print(f"✅ [COLUMN_STYLE] Added number_format '{format_string}' to {column_id} in {sheet_name}")
    
    def _extract_alignments_from_sheet(self, sheet_data: SheetData, sheet_name: str) -> Dict[str, Dict[str, str]]:
        """
        Extract alignment information from sheet data.
        
        Args:
            sheet_data: Sheet data containing alignment information
            sheet_name: Name of the sheet for logging
            
        Returns:
            Dictionary of column_id -> alignment_info
        """
        alignments = {}
        
        if not sheet_data.alignments:
            return alignments
            
        for alignment_info in sheet_data.alignments:
            alignments[alignment_info.column_id] = {
                'horizontal': alignment_info.horizontal,
                'vertical': alignment_info.vertical
            }
            
        print(f"📊 [ALIGNMENTS] Extracted {len(alignments)} alignments from {sheet_name}")
        return alignments
    
    def _integrate_alignments_into_column_styles(self, styling_config: Dict[str, Any],
                                                alignments: Dict[str, Dict[str, str]],
                                                sheet_name: str) -> None:
        """
        Integrate extracted alignments into the column_id_styles structure.
        
        Args:
            styling_config: The styling configuration section
            alignments: Dictionary of column_id -> alignment_info
            sheet_name: Name of the sheet for logging
        """
        # Ensure the column_id_styles section exists
        if 'column_id_styles' not in styling_config:
            styling_config['column_id_styles'] = {}
            print(f"🔍 [DEBUG] Created column_id_styles section for {sheet_name}")
            
        column_id_styles = styling_config['column_id_styles']
        
        # Integrate each alignment into the appropriate column style
        for column_id, alignment_info in alignments.items():
            if column_id not in column_id_styles:
                column_id_styles[column_id] = {}
                
            # Add alignment info to the column style
            column_id_styles[column_id]['alignment'] = alignment_info
            print(f"✅ [ALIGNMENT] Added alignment '{alignment_info.get('horizontal', 'unknown')}' to {column_id} in {sheet_name}")
    
    def _map_excel_format_pattern(self, excel_format: str) -> Optional[str]:
        """
        Map Excel number format patterns to standardized format strings.
        
        Args:
            excel_format: Excel number format pattern
            
        Returns:
            Standardized format pattern or None if not recognized
        """
        if not isinstance(excel_format, str):
            return None
        
        # Clean up the format string
        format_clean = excel_format.strip().lower()
        
        # Common Excel format mappings
        format_mappings = {
            # Currency and decimal formats
            '#,##0.00': '#,##0.00',
            '#,##0.0000000': '#,##0.0000000',
            '#,##0': '#,##0',
            '0.00': '#,##0.00',
            '0.0000000': '#,##0.0000000',
            
            # Percentage formats
            '0.00%': '0.00%',
            '0%': '0%',
            
            # Date formats
            'mm/dd/yyyy': 'MM/DD/YYYY',
            'dd/mm/yyyy': 'DD/MM/YYYY',
            'yyyy-mm-dd': 'YYYY-MM-DD',
            
            # Text format
            '@': '@',
            'general': 'General'
        }
        
        # Try exact match first
        for pattern, mapped in format_mappings.items():
            if pattern in format_clean:
                return mapped
        
        # Try to detect decimal places
        if '.' in format_clean and '#' in format_clean:
            # Count decimal places
            if '.0000000' in format_clean:
                return '#,##0.0000000'
            elif '.00' in format_clean:
                return '#,##0.00'
            elif '.' in format_clean:
                return '#,##0.00'
        
        # Try to detect integer format
        if '#,##0' in format_clean or '#,###' in format_clean:
            return '#,##0'
        
        return None
    
    def _get_default_number_formats(self, sheet_name: str) -> Dict[str, str]:
        """
        Get default number formats based on sheet type and common column patterns.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary with default number format patterns
        """
        # Default formats based on common column types
        default_formats = {}
        
        # Common amount/price columns
        amount_columns = ['col_amount', 'col_total', 'col_subtotal', 'col_price', 'col_value']
        for col in amount_columns:
            default_formats[col] = '#,##0.00'
        
        # Unit price columns (higher precision)
        unit_price_columns = ['col_unit_price', 'col_rate', 'col_unit_cost']
        for col in unit_price_columns:
            default_formats[col] = '#,##0.0000000'
        
        # Quantity columns (integers)
        quantity_columns = ['col_quantity', 'col_qty', 'col_count', 'col_number']
        for col in quantity_columns:
            default_formats[col] = '#,##0'
        
        # Percentage columns
        percentage_columns = ['col_percentage', 'col_percent', 'col_rate_percent']
        for col in percentage_columns:
            default_formats[col] = '0.00%'
        
        print(f"[NUMBER_FORMATS] Applied default formats: {default_formats}")
        return default_formats
    
    def _map_sheet_name(self, quantity_sheet_name: str) -> str:
        """
        Map quantity data sheet name to template config sheet name.
        
        Args:
            quantity_sheet_name: Sheet name from quantity data
            
        Returns:
            Mapped sheet name for template config, or original name if no mapping found
        """
        if self.mapping_manager:
            return self.mapping_manager.map_sheet_name(quantity_sheet_name)
        
        # Fallback to hardcoded mappings if mapping manager is not available
        fallback_mappings = {
            'INV': 'Invoice',
            'PAK': 'Packing list',
            'CON': 'Contract',
            'CONTRACT': 'Contract',
            'INVOICE': 'Invoice',
            'PACKING': 'Packing list',
            'PACKING LIST': 'Packing list'
        }
        
        return fallback_mappings.get(quantity_sheet_name.upper(), quantity_sheet_name)
    
    def _get_excel_file_path(self) -> Optional[str]:
        """
        Get the Excel file path from the current analysis context.
        
        Returns:
            Excel file path if available, None otherwise
        """
        return self.excel_file_path
    
    def set_excel_file_path(self, file_path: str) -> None:
        """
        Set the Excel file path for format extraction.
        
        Args:
            file_path: Path to the Excel file
        """
        self.excel_file_path = file_path
    
    def set_template_config(self, template_config: Dict[str, Any]) -> None:
        """
        Set the template configuration for column ID mapping.
        
        Args:
            template_config: Template configuration dictionary
        """
        self.template_config = template_config
    
    def _extract_formats_from_excel_file(self, excel_file_path: str, sheet_name: str, sheet_data: SheetData) -> Dict[str, str]:
        """
        Extract actual number formats from Excel file using proper table boundary detection.
        
        Args:
            excel_file_path: Path to the Excel file
            sheet_name: Name of the sheet to analyze
            sheet_data: Sheet data with header positions
            
        Returns:
            Dictionary mapping column identifiers to actual Excel number formats
        """
        try:
            # Load the workbook
            workbook = load_workbook(excel_file_path, data_only=False)
            
            if sheet_name not in workbook.sheetnames:
                print(f"[NUMBER_FORMATS] Sheet '{sheet_name}' not found in workbook")
                return {}
            
            worksheet = workbook[sheet_name]
            number_formats = {}
            
            # Step 1: Get header row from analysis data
            if not sheet_data.header_positions:
                print(f"[NUMBER_FORMATS] No header positions found for {sheet_name}")
                return {}
            
            header_row = sheet_data.header_positions[0].row
            print(f"[NUMBER_FORMATS] Header row: {header_row}")
            
            # Step 2: Use footer info from sheet_data if available, otherwise detect it
            if hasattr(sheet_data, 'footer_info') and sheet_data.footer_info:
                footer_row = sheet_data.footer_info.row
                print(f"[NUMBER_FORMATS] Using detected footer row: {footer_row}")
            else:
                # Fallback: Find footer row using SUM formula detection
                footer_row = self._find_table_footer_row(worksheet, header_row)
                if not footer_row:
                    print(f"[NUMBER_FORMATS] No footer found for {sheet_name}, using fallback")
                    footer_row = min(header_row + 20, worksheet.max_row)  # Fallback
                else:
                    print(f"[NUMBER_FORMATS] Detected footer row: {footer_row}")
            
            print(f"[NUMBER_FORMATS] Footer row: {footer_row}")
            print(f"[NUMBER_FORMATS] Data range: rows {header_row + 1} to {footer_row - 1}")
            
            # Step 3: Detect if header has 1 row or 2 rows
            header_rows = self._detect_header_row_count(worksheet, header_row)
            print(f"[NUMBER_FORMATS] Header has {header_rows} row(s)")
            
            # Step 4: Calculate first data row by skipping header rows
            first_data_row = header_row + header_rows
            print(f"[NUMBER_FORMATS] First data row: {first_data_row}")
            
            # Step 5: Get column IDs from template and map to Excel columns
            template_column_ids = self._get_template_column_ids(sheet_name)
            print(f"[NUMBER_FORMATS] Template column IDs for {sheet_name}: {template_column_ids}")
            
            # Step 6: Create comprehensive column mapping including dual-column support
            column_mappings = self._create_comprehensive_column_mapping(sheet_data, template_column_ids, worksheet, header_row)
            print(f"[NUMBER_FORMATS] Column mappings: {column_mappings}")
            
            # Step 7: Extract formats for each mapped column
            for col_id, excel_col in column_mappings.items():
                if col_id in template_column_ids:
                    # Look at the first data row cell to get actual data formatting
                    cell = worksheet.cell(row=first_data_row, column=excel_col)
                    
                    if cell.value is not None and cell.number_format:
                        # Get the actual Excel number format from first data cell
                        excel_format = cell.number_format
                        
                        # Clean and standardize the format
                        standardized_format = self._standardize_excel_format(excel_format)
                        
                        if standardized_format:
                            number_formats[col_id] = standardized_format
                            print(f"[NUMBER_FORMATS] {col_id} (data row {first_data_row}, col {excel_col}): {excel_format} -> {standardized_format}")
                    else:
                        # Fallback: try next few data rows if first data row is empty
                        for offset in [1, 2, 3]:
                            fallback_row = first_data_row + offset
                            if fallback_row < footer_row:
                                fallback_cell = worksheet.cell(row=fallback_row, column=excel_col)
                                
                                if fallback_cell.value is not None and fallback_cell.number_format:
                                    excel_format = fallback_cell.number_format
                                    standardized_format = self._standardize_excel_format(excel_format)
                                    
                                    if standardized_format:
                                        number_formats[col_id] = standardized_format
                                        print(f"[NUMBER_FORMATS] {col_id} (fallback data row {fallback_row}, col {excel_col}): {excel_format} -> {standardized_format}")
                                        break
            
            workbook.close()
            return number_formats
            
        except Exception as e:
            print(f"[NUMBER_FORMATS] Error reading Excel file: {e}")
            return {}
    
    def _find_table_footer_row(self, worksheet, header_row: int) -> Optional[int]:
        """
        Find the footer row using SUM formula detection (same as footer detector).
        
        Args:
            worksheet: The worksheet to search
            header_row: The header row number
            
        Returns:
            Footer row number if found, None otherwise
        """
        # Use same SUM formula patterns as footer detector
        formula_patterns = [r'=sum\(', r'=SUM\(', r'=Sum\(']
        
        # Search for SUM formulas below header
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_text = str(cell.value)
                    
                    # Check if cell contains a SUM formula
                    for pattern in formula_patterns:
                        if pattern.lower() in cell_text.lower():
                            print(f"[NUMBER_FORMATS] Found footer with SUM formula at row {row}")
                            return row
        
        return None
    
    def _map_header_to_column_id(self, header_keyword: str) -> Optional[str]:
        """
        Map header keyword to column ID.
        
        Args:
            header_keyword: Header text from Excel
            
        Returns:
            Column ID if mapping found, None otherwise
        """
        # Simple mapping based on common patterns
        keyword_lower = header_keyword.lower()
        
        if 'amount' in keyword_lower or 'total' in keyword_lower:
            return 'col_amount'
        elif 'unit price' in keyword_lower or 'price' in keyword_lower:
            return 'col_unit_price'
        elif 'quantity' in keyword_lower or 'qty' in keyword_lower:
            return 'col_qty_sf'
        elif 'item' in keyword_lower:
            return 'col_item'
        elif 'description' in keyword_lower:
            return 'col_desc'
        
        return None
    
    def _standardize_excel_format(self, excel_format: str) -> Optional[str]:
        """
        Standardize Excel number format to a clean pattern.
        
        Args:
            excel_format: Raw Excel number format
            
        Returns:
            Standardized format pattern or None
        """
        if not excel_format or excel_format == 'General':
            return None
        
        # Clean up the format
        format_clean = excel_format.strip()
        
        # Handle common Excel formats
        if format_clean in ['0.00', '#,##0.00', '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-']:
            return '#,##0.00'
        elif format_clean in ['0', '#,##0', '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-']:
            return '#,##0'
        elif format_clean in ['0.0000000', '#,##0.0000000']:
            return '#,##0.0000000'
        elif '0.00%' in format_clean:
            return '0.00%'
        elif '@' in format_clean:
            return '@'  # Text format
        
        # Try to detect decimal places from the format
        if '.' in format_clean:
            # Count zeros after decimal point
            decimal_part = format_clean.split('.')[-1]
            zero_count = 0
            for char in decimal_part:
                if char == '0':
                    zero_count += 1
                elif char not in ['#', ',', '_', '-', ' ', '?']:
                    break
            
            if zero_count == 2:
                return '#,##0.00'
            elif zero_count >= 7:
                return '#,##0.0000000'
            elif zero_count > 0:
                return f"#,##0.{'0' * zero_count}"
        
        # If format contains thousands separator, use integer format
        if '#,##0' in format_clean and '.' not in format_clean:
            return '#,##0'
        
        print(f"[NUMBER_FORMATS] Unknown format pattern: {excel_format}")
        return None
    
    def _detect_header_row_count(self, worksheet, header_row: int) -> int:
        """
        Detect if header has 1 row or 2 rows by analyzing cell content and formatting.
        
        Args:
            worksheet: The worksheet to analyze
            header_row: The starting header row number
            
        Returns:
            Number of header rows (1 or 2)
        """
        # Check if the row immediately after header_row also contains header-like content
        next_row = header_row + 1
        
        if next_row > worksheet.max_row:
            return 1  # Only one row possible
        
        # Count non-empty cells in header row and next row
        header_row_cells = 0
        next_row_cells = 0
        header_row_text_cells = 0
        next_row_text_cells = 0
        
        # Check first 10 columns for header patterns
        for col in range(1, min(11, worksheet.max_column + 1)):
            # Check header row
            header_cell = worksheet.cell(row=header_row, column=col)
            if header_cell.value is not None:
                header_row_cells += 1
                cell_text = str(header_cell.value).strip()
                # Count text cells (not just numbers)
                if len(cell_text) > 0 and not cell_text.replace('.', '').replace(',', '').isdigit():
                    header_row_text_cells += 1
            
            # Check next row
            next_cell = worksheet.cell(row=next_row, column=col)
            if next_cell.value is not None:
                next_row_cells += 1
                cell_text = str(next_cell.value).strip()
                # Count text cells (not just numbers)
                if len(cell_text) > 0 and not cell_text.replace('.', '').replace(',', '').isdigit():
                    next_row_text_cells += 1
        
        print(f"[HEADER_DETECTION] Row {header_row}: {header_row_cells} cells, {header_row_text_cells} text cells")
        print(f"[HEADER_DETECTION] Row {next_row}: {next_row_cells} cells, {next_row_text_cells} text cells")
        
        # Decision logic:
        # If next row has text content that looks like sub-headers, it's a 2-row header
        # Check for common sub-header patterns like "PCS", "SF", units, etc.
        
        # Check if next row contains sub-header-like text
        is_subheader = False
        if next_row_text_cells >= 1:
            # Check actual text content for sub-header patterns
            for col in range(1, min(11, worksheet.max_column + 1)):
                next_cell = worksheet.cell(row=next_row, column=col)
                if next_cell.value is not None:
                    cell_text = str(next_cell.value).strip().upper()
                    # Common sub-header patterns
                    subheader_patterns = ['PCS', 'SF', 'KGS', 'USD', 'UNIT', 'QTY', 'AMOUNT', 'TOTAL']
                    if any(pattern in cell_text for pattern in subheader_patterns):
                        is_subheader = True
                        print(f"[HEADER_DETECTION] Found sub-header pattern: {cell_text}")
                        break
        
        if next_row_text_cells >= 2 or is_subheader:
            print(f"[HEADER_DETECTION] Detected 2-row header (next row has {next_row_text_cells} text cells or sub-header patterns)")
            return 2
        else:
            print(f"[HEADER_DETECTION] Detected 1-row header (next row appears to be data)")
            return 1
    
    def _get_template_column_ids(self, sheet_name: str) -> List[str]:
        """
        Get all column IDs defined in the template for a specific sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            List of column IDs that should have number formats
        """
        column_ids = []
        
        # Extract column IDs from template config if available
        if hasattr(self, 'template_config') and self.template_config:
            data_mapping = self.template_config.get('data_mapping', {})
            sheet_config = data_mapping.get(sheet_name, {})
            
            # Get column IDs from header_to_write
            header_to_write = sheet_config.get('header_to_write', [])
            for header_entry in header_to_write:
                if isinstance(header_entry, dict) and 'id' in header_entry:
                    column_ids.append(header_entry['id'])
        
        # Fallback to hardcoded IDs if template not available
        if not column_ids:
            sheet_specific_ids = {
                'Invoice': ['col_qty_sf', 'col_unit_price', 'col_amount', 'col_item', 'col_desc'],
                'Contract': ['col_qty_sf', 'col_unit_price', 'col_amount', 'col_item', 'col_desc'],
                'Packing list': ['col_qty_pcs', 'col_qty_sf', 'col_net', 'col_gross', 'col_cbm', 'col_item', 'col_desc']
            }
            column_ids = sheet_specific_ids.get(sheet_name, [])
        
        return column_ids
    
    def _map_excel_column_to_template_ids(self, header_pos: Any, template_column_ids: List[str]) -> List[str]:
        """
        Map an Excel column to one or more template column IDs.
        
        Args:
            header_pos: Header position object with keyword and column info
            template_column_ids: List of valid template column IDs
            
        Returns:
            List of matching template column IDs
        """
        keyword = header_pos.keyword.lower()
        matched_ids = []
        
        # Map based on header keyword and template column IDs
        if 'quantity' in keyword or 'qty' in keyword:
            # Check if this is PCS or SF quantity
            if 'pcs' in keyword or header_pos.row == 22:  # Row 22 is typically PCS sub-header
                if 'col_qty_pcs' in template_column_ids:
                    matched_ids.append('col_qty_pcs')
            else:  # Default to SF quantity
                if 'col_qty_sf' in template_column_ids:
                    matched_ids.append('col_qty_sf')
        elif 'unit price' in keyword or 'price' in keyword:
            if 'col_unit_price' in template_column_ids:
                matched_ids.append('col_unit_price')
        elif 'amount' in keyword or 'total' in keyword:
            if 'col_amount' in template_column_ids:
                matched_ids.append('col_amount')
        elif 'item' in keyword:
            if 'col_item' in template_column_ids:
                matched_ids.append('col_item')
        elif 'description' in keyword:
            if 'col_desc' in template_column_ids:
                matched_ids.append('col_desc')
        elif 'n.w' in keyword or 'net' in keyword:
            if 'col_net' in template_column_ids:
                matched_ids.append('col_net')
        elif 'g.w' in keyword or 'gross' in keyword:
            if 'col_gross' in template_column_ids:
                matched_ids.append('col_gross')
        elif 'cbm' in keyword:
            if 'col_cbm' in template_column_ids:
                matched_ids.append('col_cbm')
        
        return matched_ids
    
    def _create_comprehensive_column_mapping(self, sheet_data: Any, template_column_ids: List[str], 
                                           worksheet, header_row: int) -> Dict[str, int]:
        """
        Create comprehensive column mapping that handles dual-column structures.
        
        Args:
            sheet_data: Sheet data with header positions
            template_column_ids: List of valid template column IDs
            worksheet: Excel worksheet
            header_row: Header row number
            
        Returns:
            Dictionary mapping template column IDs to Excel column numbers
        """
        column_mappings = {}
        
        # First pass: Map main headers
        for header_pos in sheet_data.header_positions:
            col = header_pos.column
            keyword = header_pos.keyword.lower()
            
            # Standard mappings
            if 'unit price' in keyword or 'price' in keyword:
                if 'col_unit_price' in template_column_ids:
                    column_mappings['col_unit_price'] = col
            elif 'amount' in keyword or 'total' in keyword:
                if 'col_amount' in template_column_ids:
                    column_mappings['col_amount'] = col
            elif 'item' in keyword:
                if 'col_item' in template_column_ids:
                    column_mappings['col_item'] = col
            elif 'description' in keyword:
                if 'col_desc' in template_column_ids:
                    column_mappings['col_desc'] = col
            elif 'n.w' in keyword or 'net' in keyword:
                if 'col_net' in template_column_ids:
                    column_mappings['col_net'] = col
            elif 'g.w' in keyword or 'gross' in keyword:
                if 'col_gross' in template_column_ids:
                    column_mappings['col_gross'] = col
            elif 'cbm' in keyword:
                if 'col_cbm' in template_column_ids:
                    column_mappings['col_cbm'] = col
        
        # Second pass: Handle dual-column quantity structures
        self._map_dual_quantity_columns(sheet_data, template_column_ids, worksheet, header_row, column_mappings)
        
        return column_mappings
    
    def _map_dual_quantity_columns(self, sheet_data: Any, template_column_ids: List[str], 
                                 worksheet, header_row: int, column_mappings: Dict[str, int]) -> None:
        """
        Handle dual-column quantity mapping (PCS and SF).
        
        Args:
            sheet_data: Sheet data with header positions
            template_column_ids: List of valid template column IDs
            worksheet: Excel worksheet
            header_row: Header row number
            column_mappings: Dictionary to update with mappings
        """
        # Look for quantity-related headers
        quantity_headers = []
        for header_pos in sheet_data.header_positions:
            if 'quantity' in header_pos.keyword.lower() or 'qty' in header_pos.keyword.lower():
                quantity_headers.append(header_pos)
        
        # Check for sub-headers in the next row (dual-column structure)
        sub_header_row = header_row + 1
        
        for qty_header in quantity_headers:
            base_col = qty_header.column
            
            # Check for PCS and SF sub-headers around the quantity column
            for col_offset in range(-1, 3):  # Check columns around the quantity header
                check_col = base_col + col_offset
                if check_col > 0:
                    sub_cell = worksheet.cell(row=sub_header_row, column=check_col)
                    
                    if sub_cell.value is not None:
                        sub_text = str(sub_cell.value).strip().upper()
                        
                        if 'PCS' in sub_text and 'col_qty_pcs' in template_column_ids:
                            column_mappings['col_qty_pcs'] = check_col
                            print(f"[DUAL_COLUMN] Found PCS at col {check_col}")
                        elif 'SF' in sub_text and 'col_qty_sf' in template_column_ids:
                            column_mappings['col_qty_sf'] = check_col
                            print(f"[DUAL_COLUMN] Found SF at col {check_col}")
        
        # Fallback: If no dual-column structure found, map quantity to col_qty_sf
        if 'col_qty_sf' in template_column_ids and 'col_qty_sf' not in column_mappings:
            for qty_header in quantity_headers:
                column_mappings['col_qty_sf'] = qty_header.column
                print(f"[DUAL_COLUMN] Fallback: Mapped quantity to col_qty_sf at col {qty_header.column}")
                break
    
    def _find_sheet_config(self, config: Dict[str, Any], sheet_name: str) -> Dict[str, Any]:
        """
        Find the configuration section for a specific sheet.
        
        Args:
            config: Full configuration dictionary
            sheet_name: Name of the sheet to find
            
        Returns:
            Sheet configuration dictionary or None if not found
        """
        # Check if this is a sheet-specific configuration
        if sheet_name in config:
            return config[sheet_name]
            
        # Check data_mapping section
        data_mapping = config.get('data_mapping', {})
        if sheet_name in data_mapping:
            return data_mapping[sheet_name]
            
        return None